from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl

from app.betting.odds_driven import OUT_DIR, calculate_ev, connect, normalize_raw_odds, write_scores_workbook


RUN_NAME = "today_4_matches_live_api_odds_probe"
SCORES_PATH = OUT_DIR / "today_4_matches_live_api_odds_value_scores_v6_strict_market_taxonomy.xlsx"
AUDIT_PATH = OUT_DIR / "today_4_matches_market_mapping_audit_v6.xlsx"
PRE_STRICT_PATH = OUT_DIR / "today_4_matches_live_api_odds_value_scores_v4_odds_normalized_fixed.xlsx"
DEFAULT_MIN_MODEL_PROB = 0.25
DEFAULT_MIN_SAMPLE_SIZE = 10


def actionable_where(alias: str = "") -> str:
    p = f"{alias}." if alias else ""
    return f"""
        {p}verdict='VALUE'
        AND {p}expected_value > 0
        AND {p}model_probability >= {DEFAULT_MIN_MODEL_PROB}
        AND {p}sample_size >= {DEFAULT_MIN_SAMPLE_SIZE}
        AND {p}probability_status='ok'
        AND {p}data_quality_status='ok'
        AND {p}market_mapping_status='OK'
        AND {p}exact_market_match=1
        AND {p}data_completeness_status='COMPLETE'
        AND COALESCE({p}model_uses_proxy,0)=0
        AND COALESCE({p}field_mapping_status,'OK') NOT IN ('WRONG','MISSING_FIELD')
        AND COALESCE({p}side_line_status,'OK')='OK'
    """


def legacy_actionable_where(alias: str = "") -> str:
    p = f"{alias}." if alias else ""
    return f"""
        {p}verdict='VALUE'
        AND {p}expected_value > 0
        AND {p}model_probability >= {DEFAULT_MIN_MODEL_PROB}
        AND {p}sample_size >= {DEFAULT_MIN_SAMPLE_SIZE}
        AND {p}probability_status='ok'
        AND {p}data_quality_status='ok'
    """


def count_pre_strict_actionable() -> int | None:
    if not PRE_STRICT_PATH.exists():
        return None
    wb = openpyxl.load_workbook(PRE_STRICT_PATH, read_only=True, data_only=True)
    if "EV Ranking" not in wb.sheetnames:
        return None
    ws = wb["EV Ranking"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return None
    index = {str(h): i for i, h in enumerate(headers)}
    count = 0
    for row in rows:
        verdict = row[index.get("verdict", -1)] if "verdict" in index else None
        dq = row[index.get("data_quality_status", -1)] if "data_quality_status" in index else None
        ps = row[index.get("probability_status", -1)] if "probability_status" in index else None
        ev = row[index.get("expected_value", -1)] if "expected_value" in index else None
        if verdict == "VALUE" and dq == "ok" and ps == "ok" and ev is not None and float(ev) > 0:
            count += 1
    return count


def rows_to_sheet(wb: openpyxl.Workbook, name: str, rows: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(["status"])
        ws.append(["no rows"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col[:200])
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 60)


def build_audit(con: sqlite3.Connection, summary: dict[str, Any]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["metric", "value"])
    for key, value in summary.items():
        ws.append([key, value])

    mapping_rows = con.execute(
        """
        SELECT match_name, raw_market_name, raw_selection_name,
               market_type AS proposed_market_type,
               market_mapping_status, market_mapping_reason,
               exact_market_match, canonical_market_type,
               statshub_field_used, field_mapping_status,
               side_line_status, data_completeness_status,
               CASE
                   WHEN market_mapping_status='OK' THEN 'allow'
                   WHEN market_mapping_status='REVIEW' THEN 'review_only'
                   ELSE 'block'
               END AS action
        FROM betting_value_scores_new
        WHERE COALESCE(market_mapping_status,'OK') != 'OK'
        ORDER BY market_mapping_status, match_name, raw_market_name, raw_selection_name
        """
    ).fetchall()
    rows_to_sheet(wb, "market_mapping_audit", mapping_rows)

    top_rows = con.execute(
        f"""
        SELECT rank, match_name, player_name, team_name, raw_market_name,
               raw_selection_name, side, line, odds_decimal,
               model_probability, sample_size, expected_value, market_mapping_status
        FROM betting_value_scores_new
        WHERE {actionable_where()}
        ORDER BY expected_value DESC
        LIMIT 20
        """
    ).fetchall()
    rows_to_sheet(wb, "top20_actionable", top_rows)
    AUDIT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(AUDIT_PATH)


def main() -> None:
    with connect() as con:
        normalize_raw_odds(con, replace=True)
        calculate_ev(con, replace=True)
        write_scores_workbook(con, SCORES_PATH)
        pre_strict = count_pre_strict_actionable()
        summary = {
            "total_normalized_rows": con.execute("SELECT COUNT(*) FROM betting_odds_normalized").fetchone()[0],
            "market_mapping_status_OK": con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_mapping_status='OK'").fetchone()[0],
            "market_mapping_status_REVIEW": con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_mapping_status='REVIEW'").fetchone()[0],
            "blocked_unverified_market": con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_mapping_status='BLOCKED_UNVERIFIED_MARKET'").fetchone()[0],
            "blocked_market_variant": con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_mapping_status='BLOCKED_MARKET_VARIANT'").fetchone()[0],
            "outside_box_sot_rows_blocked": con.execute(
                """
                SELECT COUNT(*) FROM betting_odds_normalized
                WHERE raw_market_name='Player Shots on Target Outside Box'
                  AND market_mapping_status IN ('UNSUPPORTED','BLOCKED_UNVERIFIED_MARKET','BLOCKED_MARKET_VARIANT')
                """
            ).fetchone()[0],
            "player_shots_on_target_regular_rows_allowed": con.execute(
                """
                SELECT COUNT(*) FROM betting_odds_normalized
                WHERE raw_market_name='Player Shots on Target'
                  AND market_type='player_shots_on_target'
                  AND market_mapping_status='OK'
                """
            ).fetchone()[0],
            "actionable_VALUE_before_strict_taxonomy": con.execute(
                f"SELECT COUNT(*) FROM betting_value_scores_new WHERE {legacy_actionable_where()}"
            ).fetchone()[0],
            "actionable_VALUE_after_strict_taxonomy": con.execute(
                f"SELECT COUNT(*) FROM betting_value_scores_new WHERE {actionable_where()}"
            ).fetchone()[0],
        }
        build_audit(con, summary)
        top20 = con.execute(
            f"""
            SELECT rank, match_name, COALESCE(player_name, team_name, match_name) AS subject,
                   raw_market_name, raw_selection_name, odds_decimal,
                   model_probability, expected_value
            FROM betting_value_scores_new
            WHERE {actionable_where()}
            ORDER BY expected_value DESC
            LIMIT 20
            """
        ).fetchall()
    print(json.dumps({
        "scores": str(SCORES_PATH),
        "audit": str(AUDIT_PATH),
        "summary": summary,
        "top20": [dict(row) for row in top20],
    }, indent=2))


if __name__ == "__main__":
    main()
