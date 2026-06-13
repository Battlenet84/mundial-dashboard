from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.odds_driven import calculate_ev, connect, ensure_schema, insert_raw_rows, normalize_raw_odds
from scripts.check_canada_bosnia_player_props_availability import (
    PROP_MARKER_WORDS,
    extract_bookmakers,
    find_book_by_name,
    market_is_player_prop,
)
from scripts.fetch_canada_bosnia_live_api_odds import (
    BOOKMAKER as PREVIOUS_BOOKMAKER,
    STATSHUB_EVENT_ID,
    TARGET_AWAY,
    TARGET_HOME,
    TARGET_MATCH,
    price_rows_from_outcome,
)
from scripts.probe_odds_api_io import is_configured_api_key, load_config


RUN_NAME = "canada_bosnia_1xbet_api_markets_probe"
API_EVENT_ID = "66456916"
RAW_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/canada_bosnia_1xbet_api_markets_value_scores.xlsx")
BET365_RUN_DIR = Path("data/raw/odds/canada_bosnia_all_available_api_markets_probe")
TIMEOUT_SECONDS = 20


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    safe = {key: value for key, value in params.items() if key.lower() != "apikey"}
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    return f"{url}?{urlencode(safe)}" if safe else url


def save_response(method: str, endpoint: str, params: dict[str, Any], status_code: int | None, payload: Any, error: str | None, base_url: str) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    path = RAW_DIR / f"{method.lower()}_{endpoint.replace('/', '_')}_{stamp()}.json"
    wrapper = {
        "fetched_at_utc": now_utc(),
        "run_name": RUN_NAME,
        "method": method,
        "action": endpoint,
        "url_without_api_key": public_url(base_url, endpoint, params),
        "params_without_api_key": {key: value for key, value in params.items() if key.lower() != "apikey"},
        "status_code": status_code,
        "response_json": payload,
        "error": error,
    }
    path.write_text(json.dumps(wrapper, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def api_request(method: str, endpoint: str, params: dict[str, Any], base_url: str) -> dict[str, Any]:
    status_code = None
    payload: Any = None
    error = None
    try:
        response = requests.request(method, f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}", params=params, timeout=TIMEOUT_SECONDS)
        status_code = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"response_text_preview": response.text[:1000]}
        if response.status_code >= 400:
            error = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        error = str(exc)
    raw_file = save_response(method, endpoint, params, status_code, payload, error, base_url)
    return {"method": method, "endpoint": endpoint, "params": params, "status_code": status_code, "payload": payload, "error": error, "raw_file": raw_file}


def payload_count(payload: Any) -> int:
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("data", "bookmakers", "response", "results"):
            if isinstance(payload.get(key), list):
                return len(payload[key])
    return 1 if payload else 0


def selected_bookmakers(payload: Any) -> list[str]:
    if isinstance(payload, dict) and isinstance(payload.get("bookmakers"), list):
        return extract_bookmakers(payload["bookmakers"])
    return extract_bookmakers(payload)


def odds_payloads(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict) and payload.get("bookmakers"):
        return [payload]
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    return []


def analyze_payload(payload: Any, preferred_bookmaker: str | None = None) -> dict[str, Any]:
    rows = 0
    prop_rows = 0
    team_rows = 0
    groups: dict[str, int] = {}
    prop_groups: dict[str, int] = {}
    bookmaker_names: list[str] = []
    for event_payload in odds_payloads(payload):
        bookmakers = event_payload.get("bookmakers")
        if not isinstance(bookmakers, dict):
            continue
        for bookmaker, markets in bookmakers.items():
            if preferred_bookmaker and bookmaker.lower() != preferred_bookmaker.lower():
                continue
            bookmaker_names.append(bookmaker)
            for market in markets or []:
                if not isinstance(market, dict):
                    continue
                name = str(market.get("name") or market.get("key") or "unknown")
                outcomes = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
                market_rows = 0
                for outcome in outcomes or []:
                    if isinstance(outcome, dict):
                        market_rows += len(price_rows_from_outcome(name, outcome, TARGET_HOME, TARGET_AWAY))
                groups[name] = groups.get(name, 0) + market_rows
                rows += market_rows
                if market_is_player_prop(name):
                    prop_groups[name] = prop_groups.get(name, 0) + market_rows
                    prop_rows += market_rows
                else:
                    team_rows += market_rows
    return {
        "bookmakers": sorted(set(bookmaker_names)),
        "groups": groups,
        "rows": rows,
        "prop_groups": prop_groups,
        "prop_rows": prop_rows,
        "team_rows": team_rows,
    }


def raw_rows_from_payload(payload: Any, raw_file: Path, bookmaker_filter: str) -> list[dict[str, Any]]:
    rows = []
    for event_payload in odds_payloads(payload):
        bookmakers = event_payload.get("bookmakers")
        if not isinstance(bookmakers, dict):
            continue
        for bookmaker, markets in bookmakers.items():
            if bookmaker.lower() != bookmaker_filter.lower():
                continue
            for market_index, market in enumerate(markets or []):
                if not isinstance(market, dict):
                    continue
                market_name = str(market.get("name") or market.get("key") or f"market_{market_index}")
                outcomes = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
                for odds_index, outcome in enumerate(outcomes or []):
                    if not isinstance(outcome, dict):
                        continue
                    for selection, line, odds in price_rows_from_outcome(market_name, outcome, TARGET_HOME, TARGET_AWAY):
                        rows.append(
                            {
                                "run_name": RUN_NAME,
                                "source_name": "Odds-API.io",
                                "bookmaker": bookmaker,
                                "match_name": TARGET_MATCH,
                                "event_id": API_EVENT_ID,
                                "api_event_id": API_EVENT_ID,
                                "statshub_event_id": STATSHUB_EVENT_ID,
                                "raw_market_group": market_name,
                                "raw_market_name": market_name,
                                "raw_selection_name": selection,
                                "raw_line": line,
                                "raw_odds": odds,
                                "odds_format": "decimal",
                                "captured_at": now_utc(),
                                "raw_payload": json.dumps({"market_index": market_index, "odds_index": odds_index, "market": market, "outcome": outcome}, ensure_ascii=False),
                                "source_url": str(event_payload.get("urls", {}).get(bookmaker) or ""),
                                "request_id": raw_file.stem,
                                "raw_file": str(raw_file),
                                "status": "raw",
                                "notes": RUN_NAME,
                            }
                        )
    return rows


def latest_bet365_payload() -> tuple[Any, Path | None]:
    if not BET365_RUN_DIR.exists():
        return None, None
    files = sorted(BET365_RUN_DIR.glob("odds_*.json"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not files:
        return None, None
    wrapper = json.loads(files[0].read_text(encoding="utf-8"))
    return wrapper.get("response_json"), files[0]


def add_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([row[h] for h in headers])
        elif isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(52, len(str(header)) + 4))


def table_rows(con: sqlite3.Connection, sql: str) -> tuple[list[str], list[sqlite3.Row]]:
    cursor = con.execute(sql)
    rows = cursor.fetchall()
    return (list(rows[0].keys()) if rows else [desc[0] for desc in cursor.description]), rows


def write_workbook(
    run_summary: dict[str, Any],
    selection_rows: list[dict[str, Any]],
    comparison_rows: list[dict[str, Any]],
    discovery_rows: list[dict[str, Any]],
    raw_sources: list[dict[str, Any]],
) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "run_summary", ["metric", "value"], [{"metric": k, "value": v} for k, v in run_summary.items()])
    add_sheet(wb, "bookmaker_selection", ["step", "selected_bookmakers", "status_code", "error", "raw_file", "notes"], selection_rows)
    add_sheet(wb, "market_comparison_bet365_vs_1xbet", ["bookmaker", "market_groups", "raw_rows", "player_prop_rows", "team_match_rows", "raw_file"], comparison_rows)
    add_sheet(wb, "market_discovery_1xbet", ["market_name", "rows", "is_player_prop", "detected_keywords"], discovery_rows)
    with connect() as con:
        for name, sql in [
            ("ev_ranking", "SELECT rank, match_name, bookmaker, market_type, raw_market_name, raw_selection_name, team_name, player_name, line, odds_decimal, implied_probability, model_probability, edge, expected_value, probability_method, sample_size, verdict, notes FROM betting_value_scores_new WHERE expected_value IS NOT NULL ORDER BY expected_value DESC"),
            ("raw_odds_1xbet", "SELECT * FROM betting_odds_raw ORDER BY id"),
            ("normalized_markets_1xbet", "SELECT * FROM betting_odds_normalized ORDER BY id"),
            ("unsupported_markets", "SELECT * FROM betting_value_scores_new WHERE verdict='UNSUPPORTED' ORDER BY id"),
            ("unmatched_selections", "SELECT * FROM betting_value_scores_new WHERE verdict='UNMATCHED' ORDER BY id"),
        ]:
            headers, rows = table_rows(con, sql)
            add_sheet(wb, name, headers, rows)
        prop_headers, prop_rows = table_rows(con, "SELECT * FROM betting_odds_raw WHERE raw_market_name LIKE '%Player%' OR raw_market_name LIKE '%Shots%' OR raw_market_name LIKE '%Anytime%' OR raw_market_name LIKE '%Goalscorer%' ORDER BY id")
        add_sheet(wb, "player_props_found", prop_headers, prop_rows)
    add_sheet(wb, "data_quality", ["endpoint", "status_code", "rows_or_items", "raw_file", "error"], raw_sources)
    wb.save(OUT_XLSX)
    return OUT_XLSX


def main() -> None:
    argparse.ArgumentParser(description="Test Odds-API.io Canada/Bosnia markets with 1xBet.").parse_args()
    config = load_config()
    if not is_configured_api_key(config.api_key):
        raise SystemExit("ODDS_API_IO_KEY missing; cannot run live/API 1xBet test.")

    raw_sources: list[dict[str, Any]] = []

    def request(method: str, endpoint: str, params: dict[str, Any]) -> dict[str, Any]:
        result = api_request(method, endpoint, params, config.base_url)
        raw_sources.append(
            {
                "endpoint": f"{method} {endpoint}",
                "status_code": result["status_code"],
                "rows_or_items": payload_count(result["payload"]),
                "raw_file": str(result["raw_file"]),
                "error": result["error"],
            }
        )
        return result

    available_result = request("GET", "bookmakers", {})
    selected_before_result = request("GET", "bookmakers/selected", {"apiKey": config.api_key})
    available = extract_bookmakers(available_result["payload"])
    selected_before = selected_bookmakers(selected_before_result["payload"])
    one_xbet = find_book_by_name(available, "1xbet")
    if not one_xbet:
        run_summary = {"classification": "BOOKMAKER_NOT_AVAILABLE", "1xBet available": "no"}
        workbook = write_workbook(run_summary, [], [], [], raw_sources)
        print("1xBet available: no")
        print(f"Output Excel path: {workbook}")
        return

    selection_rows = [
        {
            "step": "before",
            "selected_bookmakers": ",".join(selected_before),
            "status_code": selected_before_result["status_code"],
            "error": selected_before_result["error"],
            "raw_file": str(selected_before_result["raw_file"]),
            "notes": "Original selected bookmakers preserved before change",
        }
    ]

    target_selection = []
    if PREVIOUS_BOOKMAKER in selected_before:
        target_selection.append(PREVIOUS_BOOKMAKER)
    if one_xbet not in target_selection:
        target_selection.append(one_xbet)
    select_result = request("PUT", "bookmakers/selected/select", {"apiKey": config.api_key, "bookmakers": ",".join(target_selection)})
    selected_after_result = request("GET", "bookmakers/selected", {"apiKey": config.api_key})
    selected_after = selected_bookmakers(selected_after_result["payload"])
    if one_xbet.lower() not in {name.lower() for name in selected_after}:
        clear_result = request("PUT", "bookmakers/selected/clear", {"apiKey": config.api_key})
        only_result = request("PUT", "bookmakers/selected/select", {"apiKey": config.api_key, "bookmakers": one_xbet})
        selected_after_result = request("GET", "bookmakers/selected", {"apiKey": config.api_key})
        selected_after = selected_bookmakers(selected_after_result["payload"])
        selection_rows.extend(
            [
                {"step": "clear_for_1xbet", "selected_bookmakers": "", "status_code": clear_result["status_code"], "error": clear_result["error"], "raw_file": str(clear_result["raw_file"]), "notes": "Fallback to 1xBet only"},
                {"step": "select_1xbet_only", "selected_bookmakers": one_xbet, "status_code": only_result["status_code"], "error": only_result["error"], "raw_file": str(only_result["raw_file"]), "notes": ""},
            ]
        )
    selection_rows.append(
        {
            "step": "after_select",
            "selected_bookmakers": ",".join(selected_after),
            "status_code": selected_after_result["status_code"],
            "error": selected_after_result["error"],
            "raw_file": str(selected_after_result["raw_file"]),
            "notes": "",
        }
    )
    selected_ok = one_xbet.lower() in {name.lower() for name in selected_after}
    if not selected_ok:
        restore_status = restore_selection(request, selected_before, selection_rows)
        run_summary = {
            "classification": "BOOKMAKER_SELECTION_FAILED",
            "1xBet available": "yes",
            "1xBet selected successfully": "no",
            "restore_status": restore_status,
        }
        workbook = write_workbook(run_summary, selection_rows, [], [], raw_sources)
        print("1xBet available: yes")
        print("1xBet selected successfully: no")
        print(f"Output Excel path: {workbook}")
        return

    odds_result = request("GET", "odds", {"apiKey": config.api_key, "eventId": API_EVENT_ID, "bookmakers": one_xbet})
    one_xbet_analysis = analyze_payload(odds_result["payload"], one_xbet)
    raw_rows = raw_rows_from_payload(odds_result["payload"], Path(odds_result["raw_file"]), one_xbet)

    with connect() as con:
        ensure_schema(con)
        con.execute("DELETE FROM betting_odds_raw")
        con.execute("DELETE FROM betting_odds_normalized")
        con.execute("DELETE FROM betting_value_scores_new")
        insert_raw_rows(con, raw_rows, replace=False)
        normalize_raw_odds(con, replace=True)
        ev_summary = calculate_ev(con, replace=True)

    bet365_payload, bet365_file = latest_bet365_payload()
    bet365_analysis = analyze_payload(bet365_payload, PREVIOUS_BOOKMAKER) if bet365_payload else {"groups": {}, "rows": 0, "prop_rows": 0, "team_rows": 0}
    comparison_rows = [
        {
            "bookmaker": PREVIOUS_BOOKMAKER,
            "market_groups": ", ".join(sorted(bet365_analysis["groups"].keys())),
            "raw_rows": bet365_analysis["rows"],
            "player_prop_rows": bet365_analysis["prop_rows"],
            "team_match_rows": bet365_analysis["team_rows"],
            "raw_file": str(bet365_file or ""),
        },
        {
            "bookmaker": one_xbet,
            "market_groups": ", ".join(sorted(one_xbet_analysis["groups"].keys())),
            "raw_rows": one_xbet_analysis["rows"],
            "player_prop_rows": one_xbet_analysis["prop_rows"],
            "team_match_rows": one_xbet_analysis["team_rows"],
            "raw_file": str(odds_result["raw_file"]),
        },
    ]
    discovery_rows = []
    for market, count in sorted(one_xbet_analysis["groups"].items()):
        keywords = [word for word in PROP_MARKER_WORDS + ["Corners", "BTTS", "Double Chance", "Draw No Bet", "Team Total"] if word.lower() in market.lower()]
        discovery_rows.append(
            {
                "market_name": market,
                "rows": count,
                "is_player_prop": "yes" if market_is_player_prop(market) else "no",
                "detected_keywords": ", ".join(keywords),
            }
        )

    restore_status = restore_selection(request, selected_before, selection_rows)
    final_selected_result = request("GET", "bookmakers/selected", {"apiKey": config.api_key})
    final_selected = selected_bookmakers(final_selected_result["payload"])
    selection_rows.append(
        {
            "step": "final",
            "selected_bookmakers": ",".join(final_selected),
            "status_code": final_selected_result["status_code"],
            "error": final_selected_result["error"],
            "raw_file": str(final_selected_result["raw_file"]),
            "notes": restore_status,
        }
    )

    new_market_types = sorted(set(one_xbet_analysis["groups"]) - set(bet365_analysis["groups"]))
    run_summary = {
        "run_name": RUN_NAME,
        "1xBet available": "yes",
        "1xBet selected successfully": "yes",
        "odds fetch successful": "yes" if odds_result["status_code"] and 200 <= odds_result["status_code"] < 300 else "no",
        "raw odds rows": one_xbet_analysis["rows"],
        "player prop rows": one_xbet_analysis["prop_rows"],
        "team_match_rows": one_xbet_analysis["team_rows"],
        "EV rows calculated": ev_summary["ev_rows"],
        "new market types compared to Bet365": ", ".join(new_market_types) if new_market_types else "-",
        "restore_status": restore_status,
        "final selected bookmakers": ",".join(final_selected),
    }
    workbook = write_workbook(run_summary, selection_rows, comparison_rows, discovery_rows, raw_sources)

    print("CANADA BOSNIA 1XBET API MARKETS")
    print("1xBet available: yes")
    print("1xBet selected successfully: yes")
    print(f"Odds fetch successful: {run_summary['odds fetch successful']}")
    print(f"Market groups returned by 1xBet: {', '.join(sorted(one_xbet_analysis['groups'].keys())) if one_xbet_analysis['groups'] else '-'}")
    print(f"Raw odds rows: {one_xbet_analysis['rows']}")
    print(f"Player prop rows found: {one_xbet_analysis['prop_rows']}")
    print(f"New market types compared to Bet365: {run_summary['new market types compared to Bet365']}")
    print(f"EV rows calculated: {ev_summary['ev_rows']}")
    print("Top 20 EV rows:")
    for row in ev_summary["top20"]:
        print(f"- #{row['rank']} {row['raw_market_name']} | {row['selection']} | odds={row['odds_decimal']} | p={row['model_probability']} | EV={row['expected_value']} | {row['verdict']}")
    print(f"Whether 1xBet solves missing player props: {'yes' if one_xbet_analysis['prop_rows'] else 'no'}")
    print("Whether another bookmaker should be tried next: yes, if player props are required; try a prop-heavy book that can be selected without 403")
    print(f"Final selected bookmaker state: {', '.join(final_selected) if final_selected else '-'}")
    print(f"Output Excel path: {workbook}")


def restore_selection(request_func, original: list[str], selection_rows: list[dict[str, Any]]) -> str:
    clear = request_func("PUT", "bookmakers/selected/clear", {"apiKey": load_config().api_key})
    if original:
        restore = request_func("PUT", "bookmakers/selected/select", {"apiKey": load_config().api_key, "bookmakers": ",".join(original)})
        status = "restored" if restore["status_code"] and 200 <= restore["status_code"] < 300 else "restore_failed"
        selection_rows.append(
            {
                "step": "restore_original",
                "selected_bookmakers": ",".join(original),
                "status_code": restore["status_code"],
                "error": restore["error"],
                "raw_file": str(restore["raw_file"]),
                "notes": status,
            }
        )
        return status
    selection_rows.append(
        {
            "step": "restore_original_empty",
            "selected_bookmakers": "",
            "status_code": clear["status_code"],
            "error": clear["error"],
            "raw_file": str(clear["raw_file"]),
            "notes": "cleared selection because original was empty",
        }
    )
    return "restored_empty"


if __name__ == "__main__":
    main()
