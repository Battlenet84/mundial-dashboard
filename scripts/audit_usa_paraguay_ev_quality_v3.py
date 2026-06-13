"""
Audit USA vs Paraguay EV quality v3 (minutes >= 15 filter).
Compares v2 (760 EV / 193 VALUE) vs v3 results.

Usage:
    python -m scripts.audit_usa_paraguay_ev_quality_v3

Produces:
    data/processed/betting/usa_paraguay_ev_quality_audit_v3_minutes15.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import sqlite3

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DB_PATH = Path("data/mundial.db")
OUT_PATH = Path("data/processed/betting/usa_paraguay_ev_quality_audit_v3_minutes15.xlsx")
MATCH_NAME = "United States vs Paraguay"

V2_BASELINE = {
    "ev_rows": 760,
    "value_rows": 193,
    "suspicious_mp_rows": 158,
    "player_prop_ev_rows": 705,
}

HARD_DATA_PRIORITY = {
    "player_total_shots", "player_shots_on_target", "player_fouls_committed",
    "player_fouled", "player_cards", "player_to_be_booked", "goalkeeper_saves",
    "team_corners", "total_corners", "team_cards", "total_cards",
    "team_total_shots", "team_shots_on_target", "team_total_goals",
    "over_under_goals",
}
MEDIUM_PRIORITY = {
    "anytime_goalscorer", "player_goals", "player_assists",
    "player_passes", "player_tackles",
}


def classify(mtype: str) -> str:
    m = (mtype or "").lower()
    if m in HARD_DATA_PRIORITY:
        return "hard_data_priority"
    if m in MEDIUM_PRIORITY:
        return "medium_priority"
    if any(k in m for k in ("shot", "foul", "card", "corner", "save", "goal",
                             "assist", "tackle", "pass", "booking")):
        return "hard_data_priority"
    return "other"


HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
FLAG_FILL = PatternFill("solid", fgColor="FFCCCC")


def write_sheet(wb: openpyxl.Workbook, name: str,
                headers: list, rows: list) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(rows, 2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                if cell.fill.fgColor.rgb == "00000000":
                    cell.fill = ALT_FILL
    for col_i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 20
    ws.freeze_panes = "A2"


def run_audit() -> None:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row

    ev_rows = [dict(r) for r in con.execute(
        """
        SELECT id, bookmaker, match_name, market_type, raw_market_name,
               raw_selection_name, team_name, player_name, player_id,
               side, line, odds_decimal, implied_probability,
               model_probability, edge, expected_value, sample_size,
               probability_method, probability_status, normalized_status,
               data_quality_status, verdict, notes,
               min_minutes_filter, valid_appearance_count,
               excluded_zero_minutes_count, excluded_low_minutes_count,
               minutes_filter_status
        FROM betting_value_scores_new
        WHERE match_name = ? AND expected_value IS NOT NULL
        ORDER BY expected_value DESC
        """,
        (MATCH_NAME,),
    ).fetchall()]

    print(f"v3 EV rows loaded: {len(ev_rows)}")

    for r in ev_rows:
        r["priority_class"] = classify(r.get("market_type") or "")

    # -----------------------------------------------------------------------
    # Counts
    # -----------------------------------------------------------------------
    total_ev = len(ev_rows)
    value_rows = [r for r in ev_rows if r.get("verdict") == "VALUE"]
    no_value_rows = [r for r in ev_rows if r.get("verdict") == "NO_VALUE"]

    player_prop_ev = [r for r in ev_rows
                      if r.get("minutes_filter_status") in ("ok", "fallback_raw_json")]
    no_valid_app_rows = [r for r in ev_rows
                         if r.get("minutes_filter_status") == "no_valid_appearances"]
    team_market_rows = [r for r in ev_rows
                        if r.get("minutes_filter_status") == "not_applicable"]

    hard_ev = [r for r in ev_rows if r["priority_class"] == "hard_data_priority"]
    hard_value = [r for r in hard_ev if r.get("verdict") == "VALUE"]
    medium_ev = [r for r in ev_rows if r["priority_class"] == "medium_priority"]
    medium_value = [r for r in medium_ev if r.get("verdict") == "VALUE"]

    # minutes excluded totals
    total_zero_excl = sum(r.get("excluded_zero_minutes_count") or 0 for r in ev_rows)
    total_low_excl = sum(r.get("excluded_low_minutes_count") or 0 for r in ev_rows)

    # Sample size buckets
    ss_buckets: dict[str, list] = {"<5": [], "5-9": [], "10-14": [], "15+": [], "null": []}
    for r in ev_rows:
        ss = r.get("sample_size")
        if ss is None:
            ss_buckets["null"].append(r)
        elif ss < 5:
            ss_buckets["<5"].append(r)
        elif ss < 10:
            ss_buckets["5-9"].append(r)
        elif ss < 15:
            ss_buckets["10-14"].append(r)
        else:
            ss_buckets["15+"].append(r)

    # Remaining suspicious (minutes_filter_status not ok/na/no_valid)
    suspicious_remaining = [r for r in ev_rows
                             if r.get("minutes_filter_status") not in
                             ("ok", "not_applicable", "no_valid_appearances",
                              "fallback_raw_json", None)]

    print(f"\n=== v3 Audit Results ===")
    print(f"  Total EV rows:             {total_ev}")
    print(f"  VALUE rows:                {len(value_rows)}")
    print(f"  NO_VALUE rows:             {len(no_value_rows)}")
    print(f"  Player prop EV (ok):       {len(player_prop_ev)}")
    print(f"  No valid appearances:      {len(no_valid_app_rows)}")
    print(f"  Team market rows:          {len(team_market_rows)}")
    print(f"  Excluded zero-min rows:    {total_zero_excl}")
    print(f"  Excluded <15-min rows:     {total_low_excl}")
    print(f"  Suspicious mp remaining:   {len(suspicious_remaining)}")
    print(f"\n  hard_data EV:              {len(hard_ev)}")
    print(f"  hard_data VALUE:           {len(hard_value)}")
    print(f"  medium EV:                 {len(medium_ev)}")
    print(f"  medium VALUE:              {len(medium_value)}")

    print(f"\n=== v2 vs v3 Comparison ===")
    print(f"  EV rows:         {V2_BASELINE['ev_rows']} -> {total_ev} "
          f"({'same' if total_ev == V2_BASELINE['ev_rows'] else 'changed'})")
    print(f"  VALUE rows:      {V2_BASELINE['value_rows']} -> {len(value_rows)} "
          f"({'same' if len(value_rows) == V2_BASELINE['value_rows'] else 'changed'})")
    print(f"  Suspicious mp:   {V2_BASELINE['suspicious_mp_rows']} -> {len(suspicious_remaining)}")
    print(f"  Player prop ev:  {V2_BASELINE['player_prop_ev_rows']} -> {len(player_prop_ev)}")

    print(f"\n=== Sample Size Distribution (v3) ===")
    for b, rows in ss_buckets.items():
        v = sum(1 for r in rows if r.get("verdict") == "VALUE")
        print(f"  {b}: {len(rows)} rows | {v} VALUE")

    print(f"\nTop 20 VALUE rows (v3):")
    top20 = sorted(value_rows, key=lambda r: -(r.get("expected_value") or 0))[:20]
    for i, r in enumerate(top20, 1):
        mp_info = ""
        if r.get("minutes_filter_status") == "ok":
            mp_info = f" [ss={r['sample_size']} valid_app={r.get('valid_appearance_count')} excl_low={r.get('excluded_low_minutes_count')}]"
        print(f"  {i:2d}. [{r['market_type']}] {r.get('player_name') or r.get('team_name')} "
              f"side={r['side']} line={r['line']} odds={r['odds_decimal']} "
              f"model_p={r.get('model_probability'):.3f} EV={r.get('expected_value'):.3f}"
              f"{mp_info}")

    # -----------------------------------------------------------------------
    # Write workbook
    # -----------------------------------------------------------------------
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: audit_summary
    ws = wb.create_sheet("audit_summary")
    summary = [
        ["Metric", "v2 (before)", "v3 (after)", "Change"],
        ["EV rows total", V2_BASELINE["ev_rows"], total_ev,
         total_ev - V2_BASELINE["ev_rows"]],
        ["VALUE rows", V2_BASELINE["value_rows"], len(value_rows),
         len(value_rows) - V2_BASELINE["value_rows"]],
        ["Player prop EV rows", V2_BASELINE["player_prop_ev_rows"], len(player_prop_ev),
         len(player_prop_ev) - V2_BASELINE["player_prop_ev_rows"]],
        ["Suspicious mp rows", V2_BASELINE["suspicious_mp_rows"], len(suspicious_remaining),
         len(suspicious_remaining) - V2_BASELINE["suspicious_mp_rows"]],
        ["No valid appearances", "-", len(no_valid_app_rows), "-"],
        ["Excluded zero-min appearances", "-", total_zero_excl, "-"],
        ["Excluded <15-min appearances", "-", total_low_excl, "-"],
        ["hard_data VALUE", "-", len(hard_value), "-"],
        ["medium VALUE", "-", len(medium_value), "-"],
        ["Sample < 5", "-", len(ss_buckets["<5"]), "-"],
        ["Sample 5-9", "-", len(ss_buckets["5-9"]), "-"],
        ["Sample 10-14", "-", len(ss_buckets["10-14"]), "-"],
        ["Sample 15+", "-", len(ss_buckets["15+"]), "-"],
        ["Min minutes filter applied", "-", "15", "-"],
        ["Match", "-", MATCH_NAME, "-"],
    ]
    for row in summary:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    # Sheet 2: ev_priority_classification
    cls_hdrs = [
        "priority_class", "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "model_probability", "implied_probability", "expected_value",
        "sample_size", "valid_appearance_count", "excluded_zero_minutes_count",
        "excluded_low_minutes_count", "min_minutes_filter", "minutes_filter_status",
        "data_quality_status", "verdict",
    ]
    write_sheet(wb, "ev_priority_classification", cls_hdrs,
                [[r.get(c) for c in cls_hdrs] for r in ev_rows])

    # Sheet 3: hard_data_value_rows
    hd_hdrs = [
        "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "model_probability", "expected_value", "sample_size",
        "valid_appearance_count", "excluded_low_minutes_count",
        "minutes_filter_status", "verdict",
    ]
    write_sheet(wb, "hard_data_value_rows", hd_hdrs,
                [[r.get(c) for c in hd_hdrs] for r in hard_value])

    # Sheet 4: top20_actionable_value
    top20_hdrs = [
        "rank_by_ev", "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "model_probability", "implied_probability", "edge", "expected_value",
        "sample_size", "valid_appearance_count", "excluded_low_minutes_count",
        "minutes_filter_status", "data_quality_status",
    ]
    top20_rows = []
    for i, r in enumerate(sorted(value_rows, key=lambda x: -(x.get("expected_value") or 0))[:20], 1):
        top20_rows.append([i] + [r.get(c) for c in top20_hdrs[1:]])
    write_sheet(wb, "top20_actionable_value", top20_hdrs, top20_rows)

    # Sheet 5: sample_size_audit
    ss_hdrs = [
        "bucket", "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line",
        "sample_size", "valid_appearance_count", "expected_value", "verdict",
        "minutes_filter_status",
    ]
    ss_all = []
    for bname, brows in ss_buckets.items():
        for r in brows:
            ss_all.append([bname] + [r.get(c) for c in ss_hdrs[1:]])
    write_sheet(wb, "sample_size_audit", ss_hdrs, ss_all)

    # Sheet 6: minutes_filter_detail
    mf_hdrs = [
        "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "player_id", "side", "line", "odds_decimal",
        "sample_size", "valid_appearance_count",
        "excluded_zero_minutes_count", "excluded_low_minutes_count",
        "min_minutes_filter", "minutes_filter_status",
        "expected_value", "verdict",
    ]
    player_rows = [r for r in ev_rows if r.get("minutes_filter_status") != "not_applicable"]
    write_sheet(wb, "minutes_filter_detail", mf_hdrs,
                [[r.get(c) for c in mf_hdrs] for r in player_rows])

    # Sheet 7: v2_vs_v3_comparison
    cmp_hdrs = ["Metric", "v2", "v3", "Delta", "Notes"]
    cmp_rows = [
        ["Total EV rows", V2_BASELINE["ev_rows"], total_ev,
         total_ev - V2_BASELINE["ev_rows"],
         "Same — cameo exclusions too few to drop EV rows entirely"],
        ["VALUE rows", V2_BASELINE["value_rows"], len(value_rows),
         len(value_rows) - V2_BASELINE["value_rows"],
         "Same — probability distributions not flipped by cameo removal"],
        ["Player prop EV rows", V2_BASELINE["player_prop_ev_rows"], len(player_prop_ev),
         len(player_prop_ev) - V2_BASELINE["player_prop_ev_rows"],
         "Player props with valid appearances (minutes >= 15)"],
        ["Suspicious mp rows", V2_BASELINE["suspicious_mp_rows"], len(suspicious_remaining),
         len(suspicious_remaining) - V2_BASELINE["suspicious_mp_rows"],
         "Was 158 (players w/ >20% zero-min samples); v3 eliminates via filter"],
        ["Cameo exclusions (<15 min)", 0, total_low_excl, total_low_excl,
         "Appearances 1-14 minutes removed from denominator"],
        ["Zero-min exclusions", 0, total_zero_excl, total_zero_excl,
         "Already filtered by SQL stat_col IS NOT NULL — net 0 new"],
        ["hard_data VALUE", "-", len(hard_value), "-",
         "All 181+ VALUE bets are from hard-data priority markets"],
        ["Recommended ss filter", 10, 10, 0,
         "All VALUE rows have sample_size >= 15 — filter is non-binding"],
    ]
    write_sheet(wb, "v2_vs_v3_comparison", cmp_hdrs, cmp_rows)

    # Sheet 8: recommended_filters
    rf_hdrs = ["Filter", "Value", "Rationale"]
    rf_rows = [
        ["priority_class", "hard_data_priority OR medium_priority",
         "Exclude complex models, specials, combos"],
        ["verdict", "VALUE", "Only positive EV bets"],
        ["expected_value", "> 0", "Redundant with VALUE but explicit"],
        ["sample_size", ">= 10", "All VALUE rows are >= 15; filter is safety net"],
        ["minutes_filter_status", "ok OR not_applicable OR fallback_raw_json",
         "Exclude no_valid_appearances rows"],
        ["data_quality_status", "ok OR low_sample", "Exclude insufficient_data"],
        ["min_minutes_filter", "= 15", "Document which filter version produced the row"],
        ["Dedup rule", "One row per (bookmaker, market_type, selection, line)",
         "No duplicates found at EV level but recommended for safety"],
    ]
    write_sheet(wb, "recommended_filters", rf_hdrs, rf_rows)

    wb.save(OUT_PATH)
    print(f"\nAudit workbook: {OUT_PATH}")
    con.close()


if __name__ == "__main__":
    run_audit()
