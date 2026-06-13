from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.betting.odds_driven import calculate_ev, connect, ensure_schema, normalize_raw_odds


RUN_NAME = "canada_bosnia_full_bookmaker_coverage_resolver"
EVENT_ID = "66456916"
STATSHUB_EVENT_ID = "15186836"
MATCH_NAME = "Canada vs Bosnia and Herzegovina"
HOME = "Canada"
AWAY = "Bosnia and Herzegovina"
RAW_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/canada_bosnia_full_bookmaker_coverage_resolver.xlsx")
DEFAULT_BASE_URL = "https://api.odds-api.io/v3"

PROP_TERMS = [
    "player",
    "props",
    "shot",
    "shots on target",
    "anytime",
    "goalscorer",
    "pass",
    "assist",
    "card",
    "foul",
    "save",
]
RICH_TERMS = [
    "corner",
    "card",
    "booking",
    "foul",
    "team total",
    "btts",
    "both teams",
    "double chance",
    "draw no bet",
    "asian",
    "handicap",
    "alternate",
]
PRIORITY_BOOKS = [
    "DraftKings",
    "FanDuel",
    "BetMGM",
    "Caesars",
    "Unibet",
    "SingBet",
    "Bet365",
    "1xBet",
    "Pinnacle",
    "Bovada",
    "BetVictor",
    "Coral",
    "Ladbrokes",
]
BASIC_MARKETS = {"ml", "ml ht", "spread", "spread ht", "totals", "totals ht"}


def now_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def load_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if key not in {"ODDS_API_IO_KEY", "ODDS_API_IO_BASE_URL"}:
            continue
        values[key] = value.strip().strip('"').strip("'")
    return values


def config() -> tuple[str, str]:
    env_file = load_env_file(Path(".env"))
    api_key = os.getenv("ODDS_API_IO_KEY") or env_file.get("ODDS_API_IO_KEY") or ""
    base_url = os.getenv("ODDS_API_IO_BASE_URL") or env_file.get("ODDS_API_IO_BASE_URL") or DEFAULT_BASE_URL
    if not api_key:
        raise SystemExit(
            "ODDS_API_IO_KEY no configurada. Pegá tu key en el archivo .env o definila como variable de entorno."
        )
    return api_key, base_url.rstrip("/")


def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    clean = {k: v for k, v in params.items() if k.lower() != "apikey"}
    query = urllib.parse.urlencode(clean, doseq=True)
    return f"{base_url}/{endpoint.lstrip('/')}" + (f"?{query}" if query else "")


def safe_slug(name: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip()).strip("_")
    return slug or "unknown_bookmaker"


def save_response(
    method: str,
    action: str,
    endpoint: str,
    params: dict[str, Any],
    status_code: int | None,
    payload: Any,
    error: str | None,
    base_url: str,
    bookmaker: str | None = None,
) -> Path:
    target = RAW_DIR / (safe_slug(bookmaker) if bookmaker else "_meta")
    target.mkdir(parents=True, exist_ok=True)
    path = target / f"{action}_{stamp()}.json"
    clean_params = {k: v for k, v in params.items() if k.lower() != "apikey"}
    path.write_text(
        json.dumps(
            {
                "fetched_at_utc": now_utc(),
                "run_name": RUN_NAME,
                "method": method,
                "action": action,
                "url_without_api_key": public_url(base_url, endpoint, params),
                "params_without_api_key": clean_params,
                "status_code": status_code,
                "response_json": payload,
                "error": error,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return path


def request(method: str, endpoint: str, params: dict[str, Any], base_url: str, api_key: str, action: str, bookmaker: str | None = None) -> dict[str, Any]:
    request_params = dict(params)
    request_params["apiKey"] = api_key
    url = f"{base_url}/{endpoint.lstrip('/')}?{urllib.parse.urlencode(request_params, doseq=True)}"
    status_code: int | None = None
    payload: Any = None
    error: str | None = None
    try:
        req = urllib.request.Request(url, method=method)
        with urllib.request.urlopen(req, timeout=30) as response:
            status_code = response.status
            body = response.read().decode("utf-8")
            payload = json.loads(body) if body else None
    except urllib.error.HTTPError as exc:
        status_code = exc.code
        body = exc.read().decode("utf-8", errors="replace")
        try:
            payload = json.loads(body) if body else None
        except json.JSONDecodeError:
            payload = {"body": body[:500]}
        error = str(exc)
    except Exception as exc:  # noqa: BLE001
        error = str(exc)
    raw_file = save_response(method, action, endpoint, request_params, status_code, payload, error, base_url, bookmaker)
    return {
        "method": method,
        "endpoint": endpoint,
        "params": {k: v for k, v in params.items() if k.lower() != "apikey"},
        "status_code": status_code,
        "payload": payload,
        "error": error,
        "raw_file": str(raw_file),
    }


def item_list(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "bookmakers", "items", "results"):
            value = payload.get(key)
            if isinstance(value, list):
                return value
        return list(payload.values()) if all(isinstance(v, dict) for v in payload.values()) else []
    return []


def bookmaker_name(item: Any) -> str | None:
    if isinstance(item, str):
        return item
    if not isinstance(item, dict):
        return None
    for key in ("name", "bookmaker", "title", "key", "id", "slug"):
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def bookmaker_region(item: Any) -> str | None:
    if not isinstance(item, dict):
        return None
    for key in ("country", "region", "geo", "jurisdiction"):
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def extract_bookmakers(payload: Any) -> list[str]:
    names: list[str] = []
    for item in item_list(payload):
        name = bookmaker_name(item)
        if name and name not in names:
            names.append(name)
    return names


def normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def priority_score(name: str, selected: bool) -> int:
    norm = normalize_name(name)
    score = 10
    for index, wanted in enumerate(PRIORITY_BOOKS):
        if normalize_name(wanted) == norm:
            score += 1000 - index * 25
            break
    if selected:
        score += 150
    for term in ("sports", "bet", "book", "365"):
        if term in norm:
            score += 5
    return score


def build_candidates(available_payload: Any, selected_names: list[str]) -> list[dict[str, Any]]:
    selected_norm = {normalize_name(name) for name in selected_names}
    rows = []
    for item in item_list(available_payload):
        name = bookmaker_name(item)
        if not name:
            continue
        selected = normalize_name(name) in selected_norm
        rows.append(
            {
                "bookmaker_name": name,
                "bookmaker_id_or_key": item.get("id") or item.get("key") or item.get("slug") if isinstance(item, dict) else name,
                "country_region": bookmaker_region(item),
                "available": "yes",
                "currently_selected": "yes" if selected else "no",
                "priority_score": priority_score(name, selected),
            }
        )
    dedup: dict[str, dict[str, Any]] = {}
    for row in rows:
        dedup.setdefault(normalize_name(row["bookmaker_name"]), row)
    return sorted(dedup.values(), key=lambda row: (-int(row["priority_score"]), row["bookmaker_name"].lower()))


def odds_payloads(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("data", "events", "results"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        return [payload]
    return []


def iter_market_blocks(event: dict[str, Any]) -> list[tuple[str, dict[str, Any]]]:
    blocks: list[tuple[str, dict[str, Any]]] = []
    bookmakers = event.get("bookmakers")
    if isinstance(bookmakers, dict):
        for book_name, markets in bookmakers.items():
            if isinstance(markets, list):
                blocks.extend((str(book_name), market) for market in markets if isinstance(market, dict))
            elif isinstance(markets, dict):
                values = markets.get("markets") or markets.get("odds") or []
                if isinstance(values, list):
                    blocks.extend((str(book_name), market) for market in values if isinstance(market, dict))
    elif isinstance(bookmakers, list):
        for book in bookmakers:
            if not isinstance(book, dict):
                continue
            book_name = bookmaker_name(book) or "unknown"
            markets = book.get("markets") or book.get("odds") or []
            if isinstance(markets, list):
                blocks.extend((book_name, market) for market in markets if isinstance(market, dict))
    odds = event.get("odds")
    if isinstance(odds, list):
        blocks.extend(("unknown", market) for market in odds if isinstance(market, dict))
    return blocks


def market_label(market: dict[str, Any]) -> str:
    return str(market.get("name") or market.get("market") or market.get("group") or market.get("type") or "")


def market_rows(market: dict[str, Any]) -> list[dict[str, Any]]:
    rows = market.get("odds") or market.get("outcomes") or market.get("selections") or []
    return rows if isinstance(rows, list) else []


def count_terms(markets: list[str], terms: list[str]) -> int:
    lower = [name.lower() for name in markets]
    return sum(1 for name in lower if any(term in name for term in terms))


def classify_coverage(status_code: int | None, select_ok: bool, markets: list[str], player_prop_rows: int, rich_rows: int) -> str:
    if not select_ok:
        return "SELECTION_FAILED"
    if status_code == 403:
        return "FETCH_FORBIDDEN_403"
    if status_code is None or (status_code >= 400):
        return "FETCH_ERROR"
    if player_prop_rows > 0:
        return "PLAYER_PROPS_FOUND"
    if not markets:
        return "NO_MARKETS_FOR_EVENT"
    basic = {name.lower() for name in markets}.issubset(BASIC_MARKETS)
    if rich_rows > 0 or not basic:
        return "RICH_MARKETS_FOUND_NO_PLAYER_PROPS"
    return "BASIC_MARKETS_ONLY"


def analyze_odds(result: dict[str, Any], bookmaker: str) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    market_rows_out: list[dict[str, Any]] = []
    expanded: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []
    status_code = result.get("status_code")
    payload = result.get("payload")
    captured_at = now_utc()
    source_url = result["raw_file"]
    for event in odds_payloads(payload):
        for detected_book, market in iter_market_blocks(event):
            book = bookmaker if detected_book == "unknown" else detected_book
            market_name = market_label(market)
            outcomes = market_rows(market)
            market_rows_out.append(
                {
                    "bookmaker": book,
                    "market_name": market_name,
                    "outcome_rows": len(outcomes),
                    "is_player_prop": "yes" if count_terms([market_name], PROP_TERMS) else "no",
                    "is_expanded_team_market": "yes" if count_terms([market_name], RICH_TERMS) else "no",
                }
            )
            for outcome in outcomes:
                if not isinstance(outcome, dict):
                    continue
                line = outcome.get("hdp") or outcome.get("line") or outcome.get("points")
                selections = []
                for key in ("home", "draw", "away", "over", "under", "yes", "no"):
                    if key in outcome:
                        selections.append((key, outcome.get(key)))
                if not selections and {"name", "price"} <= set(outcome):
                    selections.append((str(outcome.get("name")), outcome.get("price")))
                for selection, price in selections:
                    try:
                        odds = float(price) if price not in (None, "") else None
                    except (TypeError, ValueError):
                        odds = None
                    row = {
                        "run_name": RUN_NAME,
                        "source_name": "Odds-API.io",
                        "bookmaker": book,
                        "match_name": MATCH_NAME,
                        "event_id": EVENT_ID,
                        "api_event_id": EVENT_ID,
                        "statshub_event_id": STATSHUB_EVENT_ID,
                        "raw_market_group": market_name,
                        "raw_market_name": market_name,
                        "raw_selection_name": selection,
                        "raw_line": line,
                        "raw_odds": odds,
                        "odds_format": "decimal",
                        "captured_at": captured_at,
                        "raw_payload": json.dumps(outcome, ensure_ascii=False),
                        "source_url": source_url,
                        "request_id": f"{RUN_NAME}:{bookmaker}:{result.get('raw_file')}",
                        "raw_file": result.get("raw_file"),
                        "status": "live_api",
                        "notes": "full_bookmaker_coverage_resolver",
                    }
                    raw_rows.append(row)
                    expanded.append(
                        {
                            "bookmaker": book,
                            "market_name": market_name,
                            "selection_name": selection,
                            "line": line,
                            "odds_decimal": odds,
                            "is_player_prop": "yes" if count_terms([market_name], PROP_TERMS) else "no",
                            "raw_file": result.get("raw_file"),
                        }
                    )
    market_names = sorted({row["market_name"] for row in market_rows_out if row["market_name"]})
    player_prop_rows = sum(1 for row in expanded if row["is_player_prop"] == "yes")
    rich_rows = sum(1 for row in market_rows_out if row["is_expanded_team_market"] == "yes")
    team_match_rows = len(expanded)
    summary = {
        "bookmaker": bookmaker,
        "status_code": status_code,
        "json_received": "yes" if payload is not None else "no",
        "market_count": len(market_names),
        "market_names": ", ".join(market_names),
        "raw_odds_rows": len(raw_rows),
        "player_prop_rows": player_prop_rows,
        "team_match_rows": team_match_rows,
        "shots_rows": count_terms(market_names, ["shot"]),
        "shots_on_target_rows": count_terms(market_names, ["shots on target"]),
        "cards_rows": count_terms(market_names, ["card", "booking"]),
        "corners_rows": count_terms(market_names, ["corner"]),
        "btts_rows": count_terms(market_names, ["btts", "both teams"]),
        "double_chance_rows": count_terms(market_names, ["double chance"]),
        "draw_no_bet_rows": count_terms(market_names, ["draw no bet"]),
        "team_total_rows": count_terms(market_names, ["team total"]),
        "raw_file": result.get("raw_file"),
    }
    return summary, market_rows_out, expanded, raw_rows


def clear_selected(base_url: str, api_key: str) -> dict[str, Any]:
    return request("PUT", "bookmakers/selected/clear", {}, base_url, api_key, "selected_clear")


def select_bookmakers(base_url: str, api_key: str, books: list[str], action: str, bookmaker: str | None = None) -> dict[str, Any]:
    return request("PUT", "bookmakers/selected/select", {"bookmakers": ",".join(books)}, base_url, api_key, action, bookmaker)


def get_selected(base_url: str, api_key: str, action: str) -> dict[str, Any]:
    return request("GET", "bookmakers/selected", {}, base_url, api_key, action)


def restore_selection(base_url: str, api_key: str, original: list[str]) -> tuple[list[str], list[dict[str, Any]]]:
    logs = [clear_selected(base_url, api_key)]
    if original:
        logs.append(select_bookmakers(base_url, api_key, original, "selected_restore"))
    final = get_selected(base_url, api_key, "selected_final")
    logs.append(final)
    return extract_bookmakers(final.get("payload")), logs


def latest_file(pattern: str) -> Path | None:
    files = sorted(RAW_DIR.glob(pattern), key=lambda path: path.stat().st_mtime, reverse=True)
    return files[0] if files else None


def load_saved_result(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {"status_code": None, "payload": None, "error": "missing cached raw", "raw_file": None}
    data = json.loads(path.read_text(encoding="utf-8"))
    return {
        "method": data.get("method"),
        "endpoint": data.get("action"),
        "params": data.get("params_without_api_key") or {},
        "status_code": data.get("status_code"),
        "payload": data.get("response_json"),
        "error": data.get("error"),
        "raw_file": str(path),
    }


def insert_raw_rows(con: sqlite3.Connection, rows: list[dict[str, Any]]) -> None:
    con.execute("DELETE FROM betting_odds_raw")
    con.execute("DELETE FROM betting_odds_normalized")
    con.execute("DELETE FROM betting_value_scores_new")
    if not rows:
        con.commit()
        return
    columns = list(rows[0])
    placeholders = ", ".join("?" for _ in columns)
    con.executemany(
        f"INSERT INTO betting_odds_raw ({', '.join(columns)}) VALUES ({placeholders})",
        [[row.get(col) for col in columns] for row in rows],
    )
    con.commit()


def table_rows(con: sqlite3.Connection, sql: str) -> tuple[list[str], list[sqlite3.Row]]:
    rows = con.execute(sql).fetchall()
    headers = [column[0] for column in con.execute(sql).description]
    return headers, rows


def add_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any]) -> None:
    def cell_value(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False)
        return value

    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([cell_value(row[h]) for h in headers])
        elif isinstance(row, dict):
            ws.append([cell_value(row.get(h)) for h in headers])
        else:
            ws.append([cell_value(value) for value in row])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = width


def write_workbook(
    run_summary: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    test_log: list[dict[str, Any]],
    coverage: list[dict[str, Any]],
    market_names: list[dict[str, Any]],
    player_props: list[dict[str, Any]],
    expanded: list[dict[str, Any]],
    best_raw_rows: list[dict[str, Any]],
    raw_sources: list[dict[str, Any]],
    con: sqlite3.Connection,
) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "run_summary", ["metric", "value"], run_summary)
    add_sheet(wb, "bookmaker_candidates", list(candidates[0]) if candidates else ["bookmaker_name"], candidates)
    add_sheet(wb, "bookmaker_test_log", list(test_log[0]) if test_log else ["bookmaker"], test_log)
    add_sheet(wb, "bookmaker_coverage_ranking", list(coverage[0]) if coverage else ["bookmaker"], coverage)
    add_sheet(wb, "market_names_by_bookmaker", list(market_names[0]) if market_names else ["bookmaker", "market_name"], market_names)
    add_sheet(wb, "player_props_found", list(player_props[0]) if player_props else ["bookmaker", "market_name"], player_props)
    add_sheet(wb, "expanded_markets_found", list(expanded[0]) if expanded else ["bookmaker", "market_name"], expanded)
    add_sheet(wb, "best_bookmaker_raw_odds", list(best_raw_rows[0]) if best_raw_rows else ["bookmaker"], best_raw_rows)
    for title, sql in [
        ("normalized_markets", "SELECT * FROM betting_odds_normalized"),
        ("ev_ranking", "SELECT * FROM betting_value_scores_new ORDER BY expected_value DESC"),
        ("unsupported_markets", "SELECT * FROM betting_odds_normalized WHERE normalized_status = 'unsupported_market'"),
        ("unmatched_selections", "SELECT * FROM betting_odds_normalized WHERE normalized_status NOT IN ('ok', 'unsupported_market')"),
    ]:
        headers, rows = table_rows(con, sql)
        add_sheet(wb, title, headers, rows)
    data_quality = [
        {"metric": "tested_books", "value": len(test_log), "notes": ""},
        {"metric": "books_with_any_markets", "value": sum(1 for row in coverage if row.get("raw_odds_rows", 0) > 0), "notes": ""},
        {"metric": "books_with_player_props", "value": sum(1 for row in coverage if row.get("player_prop_rows", 0) > 0), "notes": ""},
        {"metric": "db_raw_rows", "value": con.execute("SELECT COUNT(*) FROM betting_odds_raw").fetchone()[0], "notes": ""},
        {"metric": "db_normalized_rows", "value": con.execute("SELECT COUNT(*) FROM betting_odds_normalized").fetchone()[0], "notes": ""},
        {"metric": "db_ev_rows", "value": con.execute("SELECT COUNT(*) FROM betting_value_scores_new").fetchone()[0], "notes": ""},
    ]
    add_sheet(wb, "data_quality", ["metric", "value", "notes"], data_quality)
    add_sheet(wb, "raw_sources", list(raw_sources[0]) if raw_sources else ["action", "raw_file"], raw_sources)
    wb.save(OUT_XLSX)
    return OUT_XLSX


def finalize_outputs(
    candidates: list[dict[str, Any]],
    test_log: list[dict[str, Any]],
    coverage: list[dict[str, Any]],
    market_names: list[dict[str, Any]],
    player_props: list[dict[str, Any]],
    expanded: list[dict[str, Any]],
    all_raw_rows: list[dict[str, Any]],
    raw_sources: list[dict[str, Any]],
    original_selected: list[str],
    final_selected: list[str],
    stop_reason: str,
) -> tuple[Path, dict[str, Any], sqlite3.Connection]:
    coverage.sort(
        key=lambda row: (
            -int(row.get("player_prop_rows", 0)),
            -int(row.get("raw_odds_rows", 0)),
            -int(row.get("market_count", 0)),
            str(row.get("bookmaker", "")).lower(),
        )
    )
    best = coverage[0] if coverage else {}
    best_bookmaker = best.get("bookmaker")
    best_raw_rows = [row for row in all_raw_rows if row.get("bookmaker") == best_bookmaker] if best_bookmaker else []

    con = connect()
    ensure_schema(con)
    insert_raw_rows(con, best_raw_rows)
    normalize_stats = normalize_raw_odds(con, replace=True)
    ev_stats = calculate_ev(con, replace=True)
    workbook = write_workbook(
        [
            {"metric": "run_name", "value": RUN_NAME},
            {"metric": "event_id", "value": EVENT_ID},
            {"metric": "statshub_event_id", "value": STATSHUB_EVENT_ID},
            {"metric": "total_bookmakers_available", "value": len(candidates)},
            {"metric": "total_bookmakers_tested", "value": len(test_log)},
            {"metric": "request_count_used", "value": len(raw_sources)},
            {"metric": "rate_limit_status", "value": "hit" if stop_reason == "RATE_LIMIT_STOP" else "not_hit"},
            {"metric": "stop_reason", "value": stop_reason},
            {"metric": "original_selected_bookmakers", "value": ", ".join(original_selected)},
            {"metric": "final_selected_bookmakers", "value": ", ".join(final_selected)},
            {"metric": "books_with_any_markets", "value": sum(1 for row in coverage if row.get("raw_odds_rows", 0) > 0)},
            {"metric": "books_with_player_props", "value": sum(1 for row in coverage if row.get("player_prop_rows", 0) > 0)},
            {"metric": "best_bookmaker_overall", "value": best_bookmaker or ""},
            {"metric": "best_bookmaker_for_player_props", "value": next((row["bookmaker"] for row in coverage if row.get("player_prop_rows", 0) > 0), "")},
            {"metric": "best_bookmaker_for_team_markets", "value": best_bookmaker or ""},
            {"metric": "automatic_player_props_available", "value": "yes" if any(row.get("player_prop_rows", 0) > 0 for row in coverage) else "no"},
            {"metric": "normalized_rows", "value": normalize_stats},
            {"metric": "ev_rows", "value": ev_stats},
        ],
        candidates,
        test_log,
        coverage,
        market_names,
        player_props,
        expanded,
        best_raw_rows,
        raw_sources,
        con,
    )
    return workbook, best, con


def rebuild_from_cache(max_books: int) -> tuple[Path, dict[str, Any], sqlite3.Connection, dict[str, Any]]:
    available_result = load_saved_result(latest_file("_meta/bookmakers_*.json"))
    selected_before_result = load_saved_result(latest_file("_meta/selected_before_*.json"))
    selected_final_result = load_saved_result(latest_file("_meta/selected_final_*.json"))
    original_selected = extract_bookmakers(selected_before_result.get("payload"))
    final_selected = extract_bookmakers(selected_final_result.get("payload"))
    candidates = build_candidates(available_result.get("payload"), original_selected)
    candidate_by_slug = {safe_slug(row["bookmaker_name"]): row for row in candidates}
    odds_files = []
    for path in RAW_DIR.glob("*/odds_*.json"):
        if path.parts[-2] == "_meta":
            continue
        odds_files.append(path)
    odds_files.sort(key=lambda path: candidates.index(candidate_by_slug.get(path.parts[-2], candidates[-1])) if path.parts[-2] in candidate_by_slug else 9999)
    if max_books:
        odds_files = odds_files[:max_books]

    test_log: list[dict[str, Any]] = []
    coverage: list[dict[str, Any]] = []
    market_names: list[dict[str, Any]] = []
    expanded: list[dict[str, Any]] = []
    player_props: list[dict[str, Any]] = []
    all_raw_rows: list[dict[str, Any]] = []
    raw_sources: list[dict[str, Any]] = [
        {"action": "bookmakers", "status_code": available_result.get("status_code"), "raw_file": available_result.get("raw_file")},
        {"action": "selected_before", "status_code": selected_before_result.get("status_code"), "raw_file": selected_before_result.get("raw_file")},
        {"action": "selected_final", "status_code": selected_final_result.get("status_code"), "raw_file": selected_final_result.get("raw_file")},
    ]
    for idx, odds_file in enumerate(odds_files, start=1):
        candidate = candidate_by_slug.get(odds_file.parts[-2], {})
        book = candidate.get("bookmaker_name") or odds_file.parts[-2]
        odds_result = load_saved_result(odds_file)
        summary, per_market, per_expanded, raw_rows = analyze_odds(odds_result, book)
        rich_rows = sum(1 for row in per_market if row["is_expanded_team_market"] == "yes")
        summary["select_status_code"] = None
        summary["selection_verified"] = "yes"
        summary["result_classification"] = classify_coverage(
            summary.get("status_code"),
            True,
            summary["market_names"].split(", ") if summary["market_names"] else [],
            summary["player_prop_rows"],
            rich_rows,
        )
        coverage.append(summary)
        market_names.extend(per_market)
        expanded.extend(per_expanded)
        player_props.extend(row for row in per_expanded if row.get("is_player_prop") == "yes")
        all_raw_rows.extend(raw_rows)
        test_log.append(
            {
                "test_order": idx,
                "bookmaker": book,
                "priority_score": candidate.get("priority_score"),
                "select_status_code": None,
                "selection_verified": "yes",
                "fetch_status_code": odds_result.get("status_code"),
                "json_received": summary["json_received"],
                "raw_odds_rows": summary["raw_odds_rows"],
                "player_prop_rows": summary["player_prop_rows"],
                "result_classification": summary["result_classification"],
                "error": odds_result.get("error"),
                "raw_file": odds_result.get("raw_file"),
            }
        )
        raw_sources.append({"action": f"cached_odds:{book}", "status_code": odds_result.get("status_code"), "raw_file": odds_result.get("raw_file")})

    stop_reason = f"REBUILT_FROM_CACHE_{len(odds_files)}"
    workbook, best, con = finalize_outputs(
        candidates,
        test_log,
        coverage,
        market_names,
        player_props,
        expanded,
        all_raw_rows,
        raw_sources,
        original_selected,
        final_selected,
        stop_reason,
    )
    summary = {
        "total_bookmakers_available": len(candidates),
        "total_bookmakers_tested": len(test_log),
        "request_count_used": len(raw_sources),
        "rate_limit_status": "not_hit",
        "stop_reason": stop_reason,
        "original_selected": original_selected,
        "final_selected": final_selected,
        "books_with_any_markets": sum(1 for row in coverage if row.get("raw_odds_rows", 0) > 0),
        "books_with_player_props": sum(1 for row in coverage if row.get("player_prop_rows", 0) > 0),
        "best_bookmaker": best.get("bookmaker") or "",
    }
    return workbook, best, con, summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Resolve best Odds-API.io bookmaker coverage for Canada vs Bosnia.")
    parser.add_argument("--max-books", type=int, default=60, help="Max prioritized books to test. Use 0 for all available.")
    parser.add_argument("--sleep", type=float, default=0.15, help="Seconds between tested bookmakers.")
    parser.add_argument("--from-cache", action="store_true", help="Rebuild workbook from saved raw JSON without API calls.")
    args = parser.parse_args()

    if args.from_cache:
        workbook, best, con, summary = rebuild_from_cache(args.max_books)
        con.close()
        print("ODDS-API.IO BOOKMAKER COVERAGE RESOLVER")
        print(f"Total bookmakers available: {summary['total_bookmakers_available']}")
        print(f"Total bookmakers tested: {summary['total_bookmakers_tested']}")
        print(f"Request count used: {summary['request_count_used']} (cache rebuild, no API calls)")
        print(f"Rate-limit status: {summary['rate_limit_status']}")
        print(f"Stop reason: {summary['stop_reason']}")
        print(f"Original selected bookmakers: {', '.join(summary['original_selected']) or '-'}")
        print(f"Final selected bookmakers: {', '.join(summary['final_selected']) or '-'}")
        print(f"Books with any markets: {summary['books_with_any_markets']}")
        print(f"Books with player props: {summary['books_with_player_props']}")
        print(f"Best bookmaker found: {summary['best_bookmaker'] or '-'}")
        if not summary["books_with_player_props"]:
            print("Current Odds-API.io account/event does not expose player props through any tested selectable bookmaker.")
        print(f"Workbook: {workbook}")
        return

    api_key, base_url = config()
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    available_result = request("GET", "bookmakers", {}, base_url, api_key, "bookmakers")
    selected_before_result = get_selected(base_url, api_key, "selected_before")
    original_selected = extract_bookmakers(selected_before_result.get("payload"))
    candidates = build_candidates(available_result.get("payload"), original_selected)
    candidates_to_test = candidates if args.max_books == 0 else candidates[: args.max_books]

    test_log: list[dict[str, Any]] = []
    coverage: list[dict[str, Any]] = []
    market_names: list[dict[str, Any]] = []
    expanded: list[dict[str, Any]] = []
    player_props: list[dict[str, Any]] = []
    all_raw_rows: list[dict[str, Any]] = []
    raw_sources: list[dict[str, Any]] = [
        {"action": "bookmakers", "status_code": available_result.get("status_code"), "raw_file": available_result.get("raw_file")},
        {"action": "selected_before", "status_code": selected_before_result.get("status_code"), "raw_file": selected_before_result.get("raw_file")},
    ]
    stop_reason = "ALL_AVAILABLE_BOOKMAKERS_TESTED"
    final_selected: list[str] = []
    restore_logs: list[dict[str, Any]] = []

    try:
        for idx, candidate in enumerate(candidates_to_test, start=1):
            book = candidate["bookmaker_name"]
            clear_result = clear_selected(base_url, api_key)
            if clear_result.get("status_code") == 429:
                stop_reason = "RATE_LIMIT_STOP"
                raw_sources.append({"action": f"clear_before_select:{book}", "status_code": clear_result.get("status_code"), "raw_file": clear_result.get("raw_file")})
                break
            select_result = select_bookmakers(base_url, api_key, [book], "selected_select", book)
            verify_result = get_selected(base_url, api_key, "selected_verify")
            selected_now = extract_bookmakers(verify_result.get("payload"))
            selected_now_norm = {normalize_name(name) for name in selected_now}
            verified = normalize_name(book) in selected_now_norm and len(selected_now_norm) == 1
            odds_result: dict[str, Any]
            if not verified:
                odds_result = {
                    "status_code": None,
                    "payload": None,
                    "error": "selection verification failed",
                    "raw_file": None,
                }
            else:
                odds_result = request("GET", "odds", {"eventId": EVENT_ID, "bookmakers": book}, base_url, api_key, "odds", book)

            summary, per_market, per_expanded, raw_rows = analyze_odds(odds_result, book)
            summary["select_status_code"] = select_result.get("status_code")
            summary["selection_verified"] = "yes" if verified else "no"
            rich_rows = sum(1 for row in per_market if row["is_expanded_team_market"] == "yes")
            summary["result_classification"] = classify_coverage(summary.get("status_code"), verified, summary["market_names"].split(", ") if summary["market_names"] else [], summary["player_prop_rows"], rich_rows)
            coverage.append(summary)
            market_names.extend(per_market)
            expanded.extend(per_expanded)
            player_props.extend(row for row in per_expanded if row.get("is_player_prop") == "yes")
            all_raw_rows.extend(raw_rows)
            test_log.append(
                {
                    "test_order": idx,
                    "bookmaker": book,
                    "priority_score": candidate["priority_score"],
                    "select_status_code": select_result.get("status_code"),
                    "selection_verified": "yes" if verified else "no",
                    "fetch_status_code": odds_result.get("status_code"),
                    "json_received": summary["json_received"],
                    "raw_odds_rows": summary["raw_odds_rows"],
                    "player_prop_rows": summary["player_prop_rows"],
                    "result_classification": summary["result_classification"],
                    "error": odds_result.get("error"),
                    "raw_file": odds_result.get("raw_file"),
                }
            )
            raw_sources.extend(
                [
                    {"action": f"clear_before_select:{book}", "status_code": clear_result.get("status_code"), "raw_file": clear_result.get("raw_file")},
                    {"action": f"select:{book}", "status_code": select_result.get("status_code"), "raw_file": select_result.get("raw_file")},
                    {"action": f"verify:{book}", "status_code": verify_result.get("status_code"), "raw_file": verify_result.get("raw_file")},
                    {"action": f"odds:{book}", "status_code": odds_result.get("status_code"), "raw_file": odds_result.get("raw_file")},
                ]
            )

            if odds_result.get("status_code") == 429:
                stop_reason = "RATE_LIMIT_STOP"
                break
            if sum(1 for row in coverage if row.get("result_classification") == "PLAYER_PROPS_FOUND") >= 3:
                stop_reason = "PLAYER_PROPS_FOUND_LIMIT"
                break
            if sum(1 for row in coverage if row.get("result_classification") == "RICH_MARKETS_FOUND_NO_PLAYER_PROPS") >= 5:
                stop_reason = "RICH_MARKETS_FOUND_LIMIT"
                break
            if args.max_books and idx >= args.max_books and len(candidates) > idx:
                stop_reason = f"REQUEST_BUDGET_LIMIT_{args.max_books}"
                break
            time.sleep(args.sleep)
    finally:
        final_selected, restore_logs = restore_selection(base_url, api_key, original_selected)
        for result in restore_logs:
            raw_sources.append({"action": result.get("endpoint"), "status_code": result.get("status_code"), "raw_file": result.get("raw_file")})

    coverage = sorted(
        coverage,
        key=lambda row: (
            -int(row.get("player_prop_rows", 0)),
            -int(row.get("raw_odds_rows", 0)),
            -int(row.get("market_count", 0)),
            str(row.get("bookmaker", "")).lower(),
        ),
    )
    best = coverage[0] if coverage else {}
    best_bookmaker = best.get("bookmaker")
    best_raw_rows = [row for row in all_raw_rows if row.get("bookmaker") == best_bookmaker] if best_bookmaker else []

    con = connect()
    ensure_schema(con)
    insert_raw_rows(con, best_raw_rows)
    normalize_stats = normalize_raw_odds(con, replace=True)
    ev_stats = calculate_ev(con, replace=True)
    workbook = write_workbook(
        [
            {"metric": "run_name", "value": RUN_NAME},
            {"metric": "event_id", "value": EVENT_ID},
            {"metric": "statshub_event_id", "value": STATSHUB_EVENT_ID},
            {"metric": "total_bookmakers_available", "value": len(candidates)},
            {"metric": "total_bookmakers_tested", "value": len(test_log)},
            {"metric": "request_count_used", "value": len(raw_sources)},
            {"metric": "rate_limit_status", "value": "hit" if stop_reason == "RATE_LIMIT_STOP" else "not_hit"},
            {"metric": "stop_reason", "value": stop_reason},
            {"metric": "original_selected_bookmakers", "value": ", ".join(original_selected)},
            {"metric": "final_selected_bookmakers", "value": ", ".join(final_selected)},
            {"metric": "books_with_any_markets", "value": sum(1 for row in coverage if row.get("raw_odds_rows", 0) > 0)},
            {"metric": "books_with_player_props", "value": sum(1 for row in coverage if row.get("player_prop_rows", 0) > 0)},
            {"metric": "best_bookmaker_overall", "value": best_bookmaker or ""},
            {"metric": "best_bookmaker_for_player_props", "value": next((row["bookmaker"] for row in coverage if row.get("player_prop_rows", 0) > 0), "")},
            {"metric": "best_bookmaker_for_team_markets", "value": best_bookmaker or ""},
            {"metric": "automatic_player_props_available", "value": "yes" if any(row.get("player_prop_rows", 0) > 0 for row in coverage) else "no"},
            {"metric": "normalized_rows", "value": normalize_stats},
            {"metric": "ev_rows", "value": ev_stats},
        ],
        candidates,
        test_log,
        coverage,
        market_names,
        player_props,
        expanded,
        best_raw_rows,
        raw_sources,
        con,
    )
    con.close()

    print("ODDS-API.IO BOOKMAKER COVERAGE RESOLVER")
    print(f"Total bookmakers available: {len(candidates)}")
    print(f"Total bookmakers tested: {len(test_log)}")
    print(f"Request count used: {len(raw_sources)}")
    print(f"Rate-limit status: {'hit' if stop_reason == 'RATE_LIMIT_STOP' else 'not_hit'}")
    print(f"Stop reason: {stop_reason}")
    print(f"Original selected bookmakers: {', '.join(original_selected) or '-'}")
    print(f"Final selected bookmakers: {', '.join(final_selected) or '-'}")
    print(f"Books with any markets: {sum(1 for row in coverage if row.get('raw_odds_rows', 0) > 0)}")
    print(f"Books with player props: {sum(1 for row in coverage if row.get('player_prop_rows', 0) > 0)}")
    print(f"Best bookmaker found: {best_bookmaker or '-'}")
    if not any(row.get("player_prop_rows", 0) > 0 for row in coverage):
        print("Current Odds-API.io account/event does not expose player props through any tested selectable bookmaker.")
    print(f"Workbook: {workbook}")


if __name__ == "__main__":
    main()
