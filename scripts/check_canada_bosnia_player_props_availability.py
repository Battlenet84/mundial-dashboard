from __future__ import annotations

import argparse
import json
import math
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.odds_driven import calculate_ev, connect, ensure_schema, insert_raw_rows, normalize_raw_odds
from scripts.fetch_canada_bosnia_all_available_api_markets import payload_count
from scripts.fetch_canada_bosnia_live_api_odds import (
    BOOKMAKER as PREVIOUS_BOOKMAKER,
    STATSHUB_EVENT_ID,
    TARGET_AWAY,
    TARGET_HOME,
    TARGET_MATCH,
    price_rows_from_outcome,
)
from scripts.probe_odds_api_io import is_configured_api_key, load_config


RUN_NAME = "canada_bosnia_player_props_availability_review"
API_EVENT_ID = "66456916"
RAW_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/canada_bosnia_player_props_availability_review.xlsx")
TIMEOUT_SECONDS = 20
PROP_HEAVY_BOOKS = ["DraftKings", "FanDuel", "BetMGM", "Caesars", "Bet365", "Unibet", "Pinnacle", "SingBet"]
PROP_MARKER_WORDS = [
    "Player",
    "Props",
    "Shots",
    "Shots on Target",
    "Anytime",
    "Goalscorer",
    "Passes",
    "Assists",
    "Cards",
    "Fouls",
    "Saves",
]


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    safe = {key: value for key, value in params.items() if key.lower() != "apikey"}
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    return f"{url}?{urlencode(safe)}" if safe else url


def save_response(action: str, params: dict[str, Any], status_code: int | None, payload: Any, error: str | None, base_url: str) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    path = RAW_DIR / f"{action.replace('/', '_')}_{stamp()}.json"
    wrapper = {
        "fetched_at_utc": now_utc(),
        "run_name": RUN_NAME,
        "action": action,
        "url_without_api_key": public_url(base_url, action, params),
        "params_without_api_key": {key: value for key, value in params.items() if key.lower() != "apikey"},
        "status_code": status_code,
        "response_json": payload,
        "error": error,
    }
    path.write_text(json.dumps(wrapper, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def api_get(endpoint: str, params: dict[str, Any], base_url: str) -> dict[str, Any]:
    status_code = None
    payload: Any = None
    error = None
    try:
        response = requests.get(f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}", params=params, timeout=TIMEOUT_SECONDS)
        status_code = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"response_text_preview": response.text[:1000]}
        if response.status_code >= 400:
            error = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        error = str(exc)
    raw_file = save_response(endpoint, params, status_code, payload, error, base_url)
    return {"endpoint": endpoint, "params": params, "status_code": status_code, "payload": payload, "error": error, "raw_file": raw_file}


def item_list(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "bookmakers", "response", "results"):
            if isinstance(payload.get(key), list):
                return payload[key]
    return []


def bookmaker_name(item: Any) -> str | None:
    if isinstance(item, str):
        return item
    if isinstance(item, dict):
        for key in ("name", "title", "bookmaker", "key", "slug"):
            value = item.get(key)
            if value:
                return str(value)
    return None


def extract_bookmakers(payload: Any) -> list[str]:
    names = []
    for item in item_list(payload):
        name = bookmaker_name(item)
        if name and name not in names:
            names.append(name)
    return names


def find_book_by_name(available: list[str], target: str) -> str | None:
    target_l = target.lower()
    for name in available:
        if name.lower() == target_l:
            return name
    for name in available:
        if target_l in name.lower() or name.lower() in target_l:
            return name
    return None


def chunks(items: list[str], size: int = 30) -> list[list[str]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def market_is_player_prop(name: str) -> bool:
    lower = name.lower()
    markers = [word.lower() for word in PROP_MARKER_WORDS]
    return any(marker in lower for marker in markers)


def odds_payloads(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict) and payload.get("bookmakers"):
        return [payload]
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    return []


def analyze_odds_response(result: dict[str, Any], requested_books: list[str]) -> list[dict[str, Any]]:
    rows = []
    for event_payload in odds_payloads(result["payload"]):
        bookmakers = event_payload.get("bookmakers")
        if not isinstance(bookmakers, dict):
            continue
        for bookmaker, markets in bookmakers.items():
            if not isinstance(markets, list):
                continue
            market_names = []
            prop_markets = []
            prop_rows = 0
            total_rows = 0
            for market in markets:
                if not isinstance(market, dict):
                    continue
                name = str(market.get("name") or market.get("key") or "unknown")
                market_names.append(name)
                outcomes = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
                rows_for_market = 0
                for outcome in outcomes or []:
                    if isinstance(outcome, dict):
                        rows_for_market += len(price_rows_from_outcome(name, outcome, TARGET_HOME, TARGET_AWAY))
                total_rows += rows_for_market
                if market_is_player_prop(name):
                    prop_markets.append(name)
                    prop_rows += rows_for_market
            rows.append(
                {
                    "bookmaker": bookmaker,
                    "requested_in_batch": "yes" if bookmaker in requested_books else "returned_unrequested_alias",
                    "market_names": sorted(set(market_names)),
                    "market_count": len(set(market_names)),
                    "total_rows": total_rows,
                    "player_prop_market_names": sorted(set(prop_markets)),
                    "player_prop_market_count": len(set(prop_markets)),
                    "player_prop_rows": prop_rows,
                    "status_code": result["status_code"],
                    "raw_file": str(result["raw_file"]),
                    "error": result["error"],
                }
            )
    return rows


def raw_rows_from_player_props(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for event_payload in odds_payloads(result["payload"]):
        bookmakers = event_payload.get("bookmakers")
        if not isinstance(bookmakers, dict):
            continue
        for bookmaker, markets in bookmakers.items():
            if not isinstance(markets, list):
                continue
            for market_index, market in enumerate(markets):
                if not isinstance(market, dict):
                    continue
                market_name = str(market.get("name") or market.get("key") or f"market_{market_index}")
                if not market_is_player_prop(market_name):
                    continue
                outcomes = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
                for odds_index, outcome in enumerate(outcomes or []):
                    if not isinstance(outcome, dict):
                        continue
                    for selection, line, odds in price_rows_from_outcome(market_name, outcome, TARGET_HOME, TARGET_AWAY):
                        rows.append(
                            {
                                "run_name": RUN_NAME,
                                "source_name": "odds_api_io_live_player_props_review",
                                "bookmaker": bookmaker,
                                "match_name": TARGET_MATCH,
                                "event_id": API_EVENT_ID,
                                "api_event_id": API_EVENT_ID,
                                "statshub_event_id": STATSHUB_EVENT_ID,
                                "raw_market_group": market_name,
                                "raw_market_name": market_name,
                                "raw_selection_name": selection,
                                "raw_line": line,
                                "raw_odds": odds,
                                "odds_format": "decimal",
                                "captured_at": now_utc(),
                                "raw_payload": json.dumps({"market_index": market_index, "odds_index": odds_index, "market": market, "outcome": outcome}, ensure_ascii=False),
                                "source_url": str(event_payload.get("urls", {}).get(bookmaker) or ""),
                                "request_id": Path(result["raw_file"]).stem,
                                "raw_file": str(result["raw_file"]),
                                "status": "raw",
                                "notes": RUN_NAME,
                            }
                        )
    return rows


def add_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([row[h] for h in headers])
        elif isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(54, len(str(header)) + 4))


def table_rows(con: sqlite3.Connection, sql: str) -> tuple[list[str], list[sqlite3.Row]]:
    cursor = con.execute(sql)
    rows = cursor.fetchall()
    headers = list(rows[0].keys()) if rows else [desc[0] for desc in cursor.description]
    return headers, rows


def write_workbook(
    bookmaker_coverage: list[dict[str, Any]],
    market_names_by_bookmaker: list[dict[str, Any]],
    props_found: list[dict[str, Any]],
    diagnostics: list[dict[str, Any]],
    raw_sources: list[dict[str, Any]],
) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "bookmaker_coverage", ["bookmaker", "available", "selected", "prop_heavy_target", "previous_run_used", "notes"], bookmaker_coverage)
    add_sheet(wb, "market_names_by_bookmaker", ["bookmaker", "market_count", "market_names", "player_prop_market_count", "player_prop_market_names", "player_prop_rows", "raw_file"], market_names_by_bookmaker)
    add_sheet(wb, "player_props_found", ["bookmaker", "market_name", "raw_rows", "raw_file"], props_found)
    with connect() as con:
        if props_found:
            headers, ev_rows = table_rows(
                con,
                """
                SELECT rank, match_name, bookmaker, market_type, raw_market_name,
                       raw_selection_name, player_name, player_id, line, odds_decimal,
                       implied_probability, model_probability, edge, expected_value,
                       probability_method, sample_size, verdict, notes
                FROM betting_value_scores_new
                WHERE market_type LIKE 'player_%' AND expected_value IS NOT NULL
                ORDER BY expected_value DESC
                """,
            )
        else:
            headers, ev_rows = (
                ["rank", "match_name", "bookmaker", "market_type", "raw_market_name", "raw_selection_name", "player_name", "line", "odds_decimal", "expected_value", "verdict"],
                [],
            )
        add_sheet(wb, "ev_ranking_player_props", headers, ev_rows)
    add_sheet(wb, "no_props_diagnostics", ["classification", "reason", "detail"], diagnostics)
    add_sheet(wb, "raw_sources", ["endpoint", "status_code", "rows_or_items", "raw_file", "bookmakers_requested", "error"], raw_sources)
    wb.save(OUT_XLSX)
    return OUT_XLSX


def main() -> None:
    argparse.ArgumentParser(description="Check Canada/Bosnia player props availability across bookmakers.").parse_args()
    config = load_config()
    if not is_configured_api_key(config.api_key):
        raise SystemExit("live API odds unavailable because configuration is missing: ODDS_API_IO_KEY")

    attempted = successful = failed = skipped_duplicate = 0
    raw_sources: list[dict[str, Any]] = []

    def call(endpoint: str, params: dict[str, Any], requested: list[str] | None = None) -> dict[str, Any]:
        nonlocal attempted, successful, failed
        attempted += 1
        result = api_get(endpoint, params, config.base_url)
        if result["status_code"] and 200 <= result["status_code"] < 300 and not result["error"]:
            successful += 1
        else:
            failed += 1
        raw_sources.append(
            {
                "endpoint": endpoint,
                "status_code": result["status_code"],
                "rows_or_items": payload_count(result["payload"]),
                "raw_file": str(result["raw_file"]),
                "bookmakers_requested": ",".join(requested or []),
                "error": result["error"],
            }
        )
        return result

    bookmakers_result = call("bookmakers", {}, [])
    selected_result = call("bookmakers/selected", {"apiKey": config.api_key}, [])
    available = extract_bookmakers(bookmakers_result["payload"])
    selected = extract_bookmakers(selected_result["payload"])
    prop_heavy_available = [find_book_by_name(available, name) for name in PROP_HEAVY_BOOKS]
    prop_heavy_available = [name for name in prop_heavy_available if name]

    coverage = []
    selected_set = {name.lower() for name in selected}
    for name in sorted(set(available)):
        coverage.append(
            {
                "bookmaker": name,
                "available": "yes",
                "selected": "yes" if name.lower() in selected_set else "no",
                "prop_heavy_target": "yes" if any(name.lower() == p.lower() for p in prop_heavy_available) else "no",
                "previous_run_used": "yes" if name.lower() == PREVIOUS_BOOKMAKER.lower() else "no",
                "notes": "",
            }
        )

    batches: list[tuple[str, list[str]]] = []
    seen_batch_keys = set()

    def add_batch(label: str, names: list[str]) -> None:
        nonlocal skipped_duplicate
        clean = [name for name in names if name]
        key = tuple(sorted(name.lower() for name in clean))
        if not clean or key in seen_batch_keys:
            skipped_duplicate += 1
            return
        seen_batch_keys.add(key)
        batches.append((label, clean))

    add_batch("selected", selected[:30])
    for idx, chunk in enumerate(chunks(available, 30), start=1):
        add_batch(f"all_active_{idx}", chunk)
    add_batch("prop_heavy", prop_heavy_available[:30])

    market_rows = []
    player_prop_raw_rows = []
    for label, books in batches:
        result = call("odds", {"apiKey": config.api_key, "eventId": API_EVENT_ID, "bookmakers": ",".join(books)}, books)
        analysis = analyze_odds_response(result, books)
        for row in analysis:
            row["batch"] = label
            market_rows.append(row)
        player_prop_raw_rows.extend(raw_rows_from_player_props(result))

    props_found_summary: list[dict[str, Any]] = []
    for row in market_rows:
        for market in row["player_prop_market_names"]:
            props_found_summary.append(
                {
                    "bookmaker": row["bookmaker"],
                    "market_name": market,
                    "raw_rows": row["player_prop_rows"],
                    "raw_file": row["raw_file"],
                }
            )

    classification = "PLAYER_PROPS_FOUND" if player_prop_raw_rows else "NO_PLAYER_PROPS_FOR_EVENT"
    access_errors = [source for source in raw_sources if source["status_code"] in (401, 403)]
    rate_errors = [source for source in raw_sources if source["status_code"] == 429]
    if not player_prop_raw_rows and len(available) <= 1:
        classification = "BOOKMAKER_ACCESS_LIMITED"
    if access_errors:
        classification = "PLAN_OR_ACCESS_LIMITED"

    if player_prop_raw_rows:
        with connect() as con:
            ensure_schema(con)
            con.execute("DELETE FROM betting_odds_raw")
            con.execute("DELETE FROM betting_odds_normalized")
            con.execute("DELETE FROM betting_value_scores_new")
            insert_raw_rows(con, player_prop_raw_rows, replace=False)
            normalize_raw_odds(con, replace=True)
            calculate_ev(con, replace=True)

    diagnostics = [
        {"classification": classification, "reason": "player prop rows found" if player_prop_raw_rows else "no returned market names matched player prop markers", "detail": f"player_prop_rows={len(player_prop_raw_rows)}"},
        {"classification": classification, "reason": "bookmaker batches tested", "detail": f"batches={len(batches)}; attempted={attempted}; successful={successful}; failed={failed}; skipped_duplicate={skipped_duplicate}"},
        {"classification": classification, "reason": "access errors", "detail": json.dumps(access_errors, ensure_ascii=False)},
        {"classification": classification, "reason": "rate limit errors", "detail": json.dumps(rate_errors, ensure_ascii=False)},
    ]

    market_sheet_rows = [
        {
            "bookmaker": row["bookmaker"],
            "market_count": row["market_count"],
            "market_names": ", ".join(row["market_names"]),
            "player_prop_market_count": row["player_prop_market_count"],
            "player_prop_market_names": ", ".join(row["player_prop_market_names"]),
            "player_prop_rows": row["player_prop_rows"],
            "raw_file": row["raw_file"],
        }
        for row in market_rows
    ]
    workbook = write_workbook(coverage, market_sheet_rows, props_found_summary, diagnostics, raw_sources)

    unique_prop_books = sorted(set(row["bookmaker"] for row in props_found_summary))
    unique_prop_markets = sorted(set(row["market_name"] for row in props_found_summary))
    ev_rows = 0
    if player_prop_raw_rows:
        with connect() as con:
            ev_rows = con.execute("SELECT COUNT(*) FROM betting_value_scores_new WHERE market_type LIKE 'player_%' AND expected_value IS NOT NULL").fetchone()[0]

    print("CANADA BOSNIA PLAYER PROPS AVAILABILITY")
    print(f"Classification: {classification}")
    print(f"All active bookmakers available: {len(available)}")
    print(f"Currently selected bookmakers: {len(selected)}")
    print(f"Previous Canada/Bosnia bookmaker: {PREVIOUS_BOOKMAKER}")
    print("Target prop-heavy availability/selection:")
    for target in PROP_HEAVY_BOOKS:
        found = find_book_by_name(available, target)
        selected_flag = bool(found and found.lower() in selected_set)
        print(f"- {target}: available={'yes' if found else 'no'} selected={'yes' if selected_flag else 'no'} resolved={found or '-'}")
    print(f"Bookmaker batches tested: {len(batches)}")
    print(f"Requests: attempted={attempted} successful={successful} failed={failed} skipped_duplicate={skipped_duplicate}")
    print(f"Player props found: {'yes' if player_prop_raw_rows else 'no'}")
    print(f"Bookmakers returning props: {', '.join(unique_prop_books) if unique_prop_books else '-'}")
    print(f"Player prop market names: {', '.join(unique_prop_markets) if unique_prop_markets else '-'}")
    print(f"Player prop rows: {len(player_prop_raw_rows)}")
    print(f"EV rows calculated for player props: {ev_rows}")
    print(f"Output Excel path: {workbook}")
    print(f"Raw files dir: {RAW_DIR}")


if __name__ == "__main__":
    main()
