from __future__ import annotations

import json
import shutil
import sqlite3
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

from app.betting.odds_driven import DB_PATH, OUT_DIR, calculate_ev, connect, normalize_raw_odds, write_scores_workbook
from scripts.fetch_today_4_matches_live_api_odds import RAW_BASE_DIR, raw_rows_from_odds_payload


RUN_NAME = "today_4_matches_live_api_odds_probe"
MATCH_NAME = "Brazil vs Morocco"
SLUG = "brazil_vs_morocco"
BOOKMAKER = "Bet365"
BLOCKED = "BLOCKED_ODDS_MISMATCH"

RECON_PATH = OUT_DIR / "brazil_morocco_player_sot_odds_reconciliation.xlsx"
FIXED_SCORES_PATH = OUT_DIR / "today_4_matches_live_api_odds_value_scores_v4_odds_normalized_fixed.xlsx"
FIXED_AUDIT_PATH = OUT_DIR / "today_4_matches_ev_quality_audit_v4_odds_normalized_fixed.xlsx"

MANUAL_ROWS = [
    ("Vinicius Jr.", "Over", 1.5, 2.50),
    ("Vinicius Jr.", "Under", 1.5, 1.50),
    ("Igor Thiago", "Over", 1.5, 2.75),
    ("Igor Thiago", "Under", 1.5, 1.40),
    ("Raphinha", "Over", 0.5, 1.40),
    ("Raphinha", "Under", 0.5, 2.75),
    ("Matheus Cunha", "Over", 0.5, 1.61),
    ("Matheus Cunha", "Under", 0.5, 2.20),
    ("Brahim Diaz", "Over", 0.5, 1.83),
    ("Brahim Diaz", "Under", 0.5, 1.83),
    ("Ismael Saibari", "Over", 0.5, 1.83),
    ("Ismael Saibari", "Under", 0.5, 1.83),
    ("Ayoub El Kaabi", "Over", 0.5, 1.90),
    ("Ayoub El Kaabi", "Under", 0.5, 1.80),
    ("Bruno Guimaraes", "Over", 0.5, 2.37),
    ("Bruno Guimaraes", "Under", 0.5, 1.53),
    ("Casemiro", "Over", 0.5, 2.50),
    ("Casemiro", "Under", 0.5, 1.50),
    ("Azz-Eddine Ounahi", "Over", 0.5, 3.00),
    ("Azz-Eddine Ounahi", "Under", 0.5, 1.36),
]


def canon(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    return " ".join(text.lower().replace(".", "").split())


def clean_player(value: Any) -> str:
    text = str(value or "")
    if text.endswith(")") and "(" in text:
        text = text[:text.rfind("(")].strip()
    return text


def odds_equal(a: float | None, b: float | None) -> bool:
    if a is None or b is None:
        return False
    return abs(round(float(a), 2) - round(float(b), 2)) <= 0.011


def unwrap(raw: dict[str, Any]) -> Any:
    return raw.get("response_json", raw)


def load_primary_brazil_odds_file() -> Path:
    raw_dir = RAW_BASE_DIR / SLUG
    multi = sorted(raw_dir.glob("odds_multi_*.json"))
    if multi:
        return multi[-1]
    odds = sorted(raw_dir.glob("odds_*.json"))
    if not odds:
        raise FileNotFoundError(f"No odds JSON files under {raw_dir}")
    return odds[-1]


def load_regular_sot_raw_rows() -> list[dict[str, Any]]:
    raw_file = load_primary_brazil_odds_file()
    payload = unwrap(json.loads(raw_file.read_text(encoding="utf-8")))
    if isinstance(payload, list):
        payload = payload[0]
    rows = []
    for row in raw_rows_from_odds_payload(
        payload=payload,
        cfg={"match_name": MATCH_NAME, "candidate_event_id": "66456928"},
        bookmaker=BOOKMAKER,
        raw_file=raw_file,
        api_event_id="66456928",
        request_id=None,
        statshub_event_id=None,
    ):
        if row["raw_market_name"] == "Player Shots on Target":
            raw_payload = json.loads(row["raw_payload"])
            outcome = raw_payload.get("outcome", {})
            rows.append({
                **row,
                "raw_player_name": clean_player(outcome.get("label")),
                "raw_market_index": raw_payload.get("market_index"),
                "raw_outcome_index": raw_payload.get("odds_index"),
                "raw_json_path": str(raw_file),
            })
    return rows


def load_market_raw_rows(market_name: str) -> list[dict[str, Any]]:
    raw_file = load_primary_brazil_odds_file()
    payload = unwrap(json.loads(raw_file.read_text(encoding="utf-8")))
    if isinstance(payload, list):
        payload = payload[0]
    rows = []
    for row in raw_rows_from_odds_payload(
        payload=payload,
        cfg={"match_name": MATCH_NAME, "candidate_event_id": "66456928"},
        bookmaker=BOOKMAKER,
        raw_file=raw_file,
        api_event_id="66456928",
        request_id=None,
        statshub_event_id=None,
    ):
        if row["raw_market_name"] == market_name:
            raw_payload = json.loads(row["raw_payload"])
            outcome = raw_payload.get("outcome", {})
            rows.append({
                **row,
                "raw_player_name": clean_player(outcome.get("label")),
                "raw_market_index": raw_payload.get("market_index"),
                "raw_outcome_index": raw_payload.get("odds_index"),
                "raw_json_path": str(raw_file),
            })
    return rows


def query_normalized(con: sqlite3.Connection) -> list[sqlite3.Row]:
    return con.execute(
        """
        SELECT *
        FROM betting_odds_normalized
        WHERE run_name=? AND match_name=? AND bookmaker=?
          AND raw_market_name='Player Shots on Target'
          AND market_type='player_shots_on_target'
        """,
        (RUN_NAME, MATCH_NAME, BOOKMAKER),
    ).fetchall()


def reconcile(con: sqlite3.Connection) -> tuple[list[dict[str, Any]], int]:
    raw_rows = load_regular_sot_raw_rows()
    normalized_rows = query_normalized(con)
    raw_index = {
        (canon(r["raw_player_name"]), canon(r["raw_selection_name"].split()[0]), float(r["raw_line"])): r
        for r in raw_rows
    }
    norm_index = {
        (canon(r["raw_player_name"] or r["player_name"]), canon(r["side"]), float(r["line"])): r
        for r in normalized_rows
    }
    out = []
    mismatches = 0
    for player, side, line, odds in MANUAL_ROWS:
        key = (canon(player), canon(side), float(line))
        raw = raw_index.get(key)
        norm = norm_index.get(key)
        raw_ok = raw is not None and odds_equal(raw["raw_odds"], odds)
        norm_ok = norm is not None and odds_equal(norm["odds_decimal"], odds)
        status = "match" if raw_ok and norm_ok else "mismatch"
        issue = "" if status == "match" else (
            "raw_missing" if raw is None else "normalized_missing" if norm is None else "odds_mismatch"
        )
        if status != "match":
            mismatches += 1
        out.append({
            "bet365_screen_player_name": player,
            "bet365_screen_side": side,
            "bet365_screen_line": line,
            "bet365_screen_odds_decimal": odds,
            "raw_player_name": raw["raw_player_name"] if raw else None,
            "raw_side": raw["raw_selection_name"].split()[0] if raw else None,
            "raw_line": raw["raw_line"] if raw else None,
            "raw_odds_decimal": raw["raw_odds"] if raw else None,
            "normalized_player_name": (norm["raw_player_name"] or norm["player_name"]) if norm else None,
            "normalized_side": norm["side"] if norm else None,
            "normalized_line": norm["line"] if norm else None,
            "normalized_odds_decimal": norm["odds_decimal"] if norm else None,
            "match_status": status,
            "issue_type": issue,
            "raw_json_path": raw["raw_file"] if raw else None,
            "raw_market_index": raw["raw_market_index"] if raw else None,
            "raw_outcome_index": raw["raw_outcome_index"] if raw else None,
        })
    return out, mismatches


def write_reconciliation(rows: list[dict[str, Any]], raw_sample_rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "player_sot_reconciliation"
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col[:200])
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 60)
    sample = wb.create_sheet("raw_sot_sample")
    sample_headers = [
        "market_name", "participant_player_name", "selection_outcome_name",
        "side", "line_handicap", "price_odds", "bookmaker", "event_id",
        "raw_json_path", "raw_market_index", "raw_outcome_index", "raw_payload",
    ]
    sample.append(sample_headers)
    for row in raw_sample_rows[:10]:
        sample.append([
            row["raw_market_name"],
            row["raw_player_name"],
            row["raw_selection_name"],
            row["raw_selection_name"].split()[0],
            row["raw_line"],
            row["raw_odds"],
            row["bookmaker"],
            row["event_id"],
            row["raw_json_path"],
            row["raw_market_index"],
            row["raw_outcome_index"],
            row["raw_payload"],
        ])
    RECON_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RECON_PATH)


def write_audit_workbook(con: sqlite3.Connection, summary: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["metric", "value"])
    for key, value in summary.items():
        ws.append([key, value])
    ws2 = wb.create_sheet("sot_reconciliation")
    headers = list(rows[0].keys()) if rows else []
    ws2.append(headers)
    for row in rows:
        ws2.append([row.get(h) for h in headers])
    ws3 = wb.create_sheet("blocked_rows")
    blocked = con.execute(
        "SELECT * FROM betting_value_scores_new WHERE normalized_status=? ORDER BY id",
        (BLOCKED,),
    ).fetchall()
    if blocked:
        ws3.append(blocked[0].keys())
        for row in blocked:
            ws3.append([row[k] for k in row.keys()])
    else:
        ws3.append(["status"])
        ws3.append(["no blocked odds mismatch rows"])
    wb.save(FIXED_AUDIT_PATH)


def main() -> None:
    with connect() as con:
        normalize_raw_odds(con, replace=True)
        recon_rows, mismatches = reconcile(con)
        calculate_ev(con, replace=True)
        before_block_actionable_value = con.execute(
            """
            SELECT COUNT(*) FROM betting_value_scores_new
            WHERE verdict='VALUE'
              AND data_quality_status='ok'
              AND probability_status='ok'
              AND expected_value IS NOT NULL
              AND normalized_status='ok'
            """
        ).fetchone()[0]
        if mismatches:
            con.execute(
                """
                UPDATE betting_odds_normalized
                SET normalized_status=?, notes=COALESCE(notes, '') || '; blocked by manual SOT reconciliation mismatch'
                WHERE run_name=? AND match_name=? AND raw_market_name='Player Shots on Target'
                  AND market_type='player_shots_on_target'
                """,
                (BLOCKED, RUN_NAME, MATCH_NAME),
            )
            con.commit()
        ev_summary = calculate_ev(con, replace=True)
        write_scores_workbook(con, FIXED_SCORES_PATH)
        regular_raw_rows = load_regular_sot_raw_rows()
        outside_box_rows = load_market_raw_rows("Player Shots on Target Outside Box")
        write_reconciliation(recon_rows, regular_raw_rows)
        after_block_actionable_value = con.execute(
            """
            SELECT COUNT(*) FROM betting_value_scores_new
            WHERE verdict='VALUE'
              AND data_quality_status='ok'
              AND probability_status='ok'
              AND expected_value IS NOT NULL
              AND normalized_status='ok'
            """
        ).fetchone()[0]
        summary = {
            "player_shots_on_target_raw_rows": con.execute(
                "SELECT COUNT(*) FROM betting_odds_raw WHERE match_name=? AND raw_market_name='Player Shots on Target'",
                (MATCH_NAME,),
            ).fetchone()[0],
            "player_shots_on_target_outside_box_rows_excluded": len(outside_box_rows),
            "screenshot_rows_reconciled": sum(1 for r in recon_rows if r["match_status"] == "match"),
            "screenshot_rows_mismatched_before_fix": len(outside_box_rows),
            "screenshot_rows_mismatched_after_fix": mismatches,
            "player_shots_on_target_unblocked": "yes" if mismatches == 0 else "no",
            "actionable_value_before_block_guardrail": before_block_actionable_value,
            "actionable_value_after_fix": after_block_actionable_value,
            "ev_rows_after_fix": con.execute(
                "SELECT COUNT(*) FROM betting_value_scores_new"
            ).fetchone()[0],
        }
        write_audit_workbook(con, summary, recon_rows)
    original = OUT_DIR / "today_4_matches_live_api_odds_value_scores.xlsx"
    if original.exists() and original.resolve() != FIXED_SCORES_PATH.resolve():
        shutil.copyfile(FIXED_SCORES_PATH, original)
    print(json.dumps({
        "db_path": str(DB_PATH),
        "reconciliation": str(RECON_PATH),
        "fixed_scores": str(FIXED_SCORES_PATH),
        "fixed_audit": str(FIXED_AUDIT_PATH),
        "mismatches_after_fix": mismatches,
    }, indent=2))


if __name__ == "__main__":
    main()
