"""
StatsHub coverage completion for today's 4 World Cup matches (2026-06-13).

Matches:
  1. Qatar vs Switzerland
  2. Brazil vs Morocco
  3. Haiti vs Scotland
  4. Australia vs Turkey

Tasks:
  1. Fix Turkey/Turkiye name mismatch in DB
  2. Player ID mapping via tournamentId=16 endpoint for all 8 teams
  3. Player performance download for confirmed/skipped_existing players
  4. Coverage workbook (data/processed/statshub/today_4_matches_statshub_coverage_review.xlsx)
  5. EV rebuild from raw odds (no new Odds-API.io calls)

Usage:
    python -m scripts.fetch_today_4_matches_statshub_coverage            # dry run
    python -m scripts.fetch_today_4_matches_statshub_coverage --execute  # live StatsHub calls
    python -m scripts.fetch_today_4_matches_statshub_coverage --execute --skip-rebuild
"""
from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests

from app.db.queries import utc_now
from app.betting.odds_driven import connect

# ── Constants ────────────────────────────────────────────────────────────────

SNAPSHOT_NAME = "today_4_matches_statshub_coverage"
BASE = "https://www.statshub.com"
RAW_DIR = ROOT_DIR / "data" / "raw" / "statshub" / "snapshots" / SNAPSHOT_NAME
OUT_DIR = ROOT_DIR / "data" / "processed" / "statshub"
OUT_XLSX = OUT_DIR / "today_4_matches_statshub_coverage_review.xlsx"
RATE_DELAY = 1.5  # seconds between live API calls

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.statshub.com/",
    "Origin": "https://www.statshub.com",
}

# 8 target teams — (canonical_name, statshub_team_id, db_stored_name)
# db_stored_name is what's currently in statshub_team_players / statshub_team_performance_events
TEAMS = [
    {"name": "Qatar",       "team_id": "4792", "db_name": "Qatar"},
    {"name": "Switzerland", "team_id": "4699", "db_name": "Switzerland"},
    {"name": "Brazil",      "team_id": "4748", "db_name": "Brazil"},
    {"name": "Morocco",     "team_id": "4778", "db_name": "Morocco"},
    {"name": "Haiti",       "team_id": "7229", "db_name": "Haiti"},
    {"name": "Scotland",    "team_id": "4695", "db_name": "Scotland"},
    {"name": "Australia",   "team_id": "4741", "db_name": "Australia"},
    {"name": "Turkey",      "team_id": "4700", "db_name": "Turkiye"},  # stored as Turkiye, rename to Turkey
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normalize string for name matching."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", " ", s.lower()).strip()


def _norm_key(s: str) -> str:
    """Compact norm key for dict lookup."""
    return " ".join(_norm(s).split())


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _raw_path(endpoint: str, suffix: str = "json") -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    return RAW_DIR / f"{endpoint}.{suffix}"


def fetch(url: str, endpoint: str, execute: bool = True) -> tuple[Any | None, dict]:
    """Fetch URL with caching. Returns (payload, meta)."""
    target = _raw_path(endpoint)
    if target.exists():
        try:
            p = json.loads(target.read_text(encoding="utf-8"))
            return p, {"status": "cached", "raw_file": str(target)}
        except Exception:
            pass

    if not execute:
        return None, {"status": "dry_run", "raw_file": str(_raw_path(endpoint))}

    time.sleep(RATE_DELAY)
    try:
        r = requests.get(url, headers=HEADERS, timeout=25)
        text = r.text
        p = None
        try:
            p = json.loads(text)
        except Exception:
            pass
        suffix = "json" if p is not None else "txt"
        target = _raw_path(endpoint, suffix)
        target.write_text(text, encoding="utf-8")
        status = "ok" if r.status_code == 200 and p else f"http_{r.status_code}"
        return p, {"status": status, "raw_file": str(target), "http_code": r.status_code}
    except Exception as exc:
        _raw_path(endpoint, "txt").write_text(f"error: {exc}", encoding="utf-8")
        return None, {"status": "error", "raw_file": "", "error": str(exc)}


def _parse_endpoint_players(payload: Any) -> list[dict]:
    """Parse team players/performance endpoint → list of {player_id, player_name}."""
    if not payload:
        return []
    items: list = []
    if isinstance(payload, list):
        items = payload
    elif isinstance(payload, dict):
        for key in ("data", "players", "results", "items"):
            if isinstance(payload.get(key), list):
                items = payload[key]
                break
        if not items and "playerStatisticsEvents" in payload:
            items = payload["playerStatisticsEvents"]
    result = []
    seen = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        # Various payload shapes
        pid = str(item.get("playerId") or item.get("player_id") or
                  (item.get("player") or {}).get("id") or "").strip()
        pname = str(item.get("playerName") or item.get("player_name") or
                    (item.get("player") or {}).get("name") or "").strip()
        if not pid or not pname or pid in seen:
            continue
        seen.add(pid)
        result.append({"player_id": pid, "player_name": pname})
    return result


def _parse_player_perf(payload: Any, player_id: str, player_name: str,
                        team_name: str, endpoint: str) -> list[dict]:
    """Parse player performance endpoint → list of per-event stat dicts."""
    rows = []
    if not payload:
        return rows
    items = payload if isinstance(payload, list) else payload.get("data", payload.get("events", []))
    if not isinstance(items, list):
        return rows
    for item in items:
        if not isinstance(item, dict):
            continue
        stats = item.get("player_statistics_event") or item
        evt   = item.get("events") or {}
        row_pid = str(stats.get("playerId") or stats.get("player_id") or "")
        if row_pid and row_pid != str(player_id):
            continue
        ts = evt.get("startTimestamp")
        if ts:
            date_v = datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d")
        else:
            date_v = evt.get("date") or item.get("date") or item.get("matchDate") or ""
        tourn = evt.get("tournament") or item.get("tournament") or {}
        tournament_name = tourn.get("name", "") if isinstance(tourn, dict) else str(tourn or "")
        rows.append({
            "player_id": player_id, "player_name": player_name, "team_name": team_name,
            "endpoint_name": endpoint, "event_date": str(date_v)[:10],
            "tournament_name": tournament_name,
            "minutes_played":  _to_float(stats.get("minutesPlayed") or stats.get("minutes")),
            "goals":           _to_float(stats.get("goals") or stats.get("goalsScored")),
            "assists":         _to_float(stats.get("goalAssist") or stats.get("assists")),
            "shots":           _to_float(stats.get("shots") or stats.get("totalShots")),
            "shots_on_target": _to_float(stats.get("onTargetScoringAttempt") or stats.get("shotsOnTarget")),
            "fouls":           _to_float(stats.get("fouls") or stats.get("foulsCommitted")),
            "was_fouled":      _to_float(stats.get("wasFouled") or stats.get("foulsSuffered")),
            "yellow_cards":    _to_float(stats.get("yellowCard") or stats.get("yellowCards")),
            "red_cards":       _to_float(stats.get("redCard") or stats.get("redCards")),
            "xG":              _to_float(stats.get("expectedGoals") or stats.get("xG")),
            "xA":              _to_float(stats.get("expectedAssists") or stats.get("xA")),
            "key_passes":      _to_float(stats.get("keyPass") or stats.get("keyPasses")),
            "passes":          _to_float(stats.get("totalPass") or stats.get("passes") or stats.get("totalPasses")),
            "accurate_passes": _to_float(stats.get("accuratePass") or stats.get("accuratePasses")),
            "tackles":         _to_float(stats.get("totalTackle") or stats.get("tackles")),
            "possession_lost": _to_float(stats.get("possessionLostCtrl") or stats.get("possessionLost")),
            "raw_json": json.dumps(item)[:1000],
        })
    return rows


# ── Phase 0: Fix Turkey/Turkiye name mismatch ────────────────────────────────

def phase0_fix_turkey(con: Any) -> dict:
    """Rename 'Turkiye' → 'Turkey' in all relevant DB tables."""
    print("\n=== PHASE 0: Fix Turkey/Turkiye name mismatch ===")
    results: dict[str, int] = {}

    # Check current state
    tur_perf = con.execute(
        "SELECT COUNT(*) FROM statshub_team_performance_events WHERE team_name='Turkiye'"
    ).fetchone()[0]
    tur_players = con.execute(
        "SELECT COUNT(*) FROM statshub_team_players WHERE team_name='Turkiye'"
    ).fetchone()[0]
    turkey_in_wct = con.execute(
        "SELECT COUNT(*) FROM statshub_world_cup_teams WHERE team_name='Turkey'"
    ).fetchone()[0]

    print(f"  Turkiye rows in team_perf_events: {tur_perf}")
    print(f"  Turkiye rows in team_players:     {tur_players}")
    print(f"  Turkey rows in world_cup_teams:   {turkey_in_wct}")

    # Rename in team_performance_events
    if tur_perf > 0:
        con.execute(
            "UPDATE statshub_team_performance_events SET team_name='Turkey' WHERE team_name='Turkiye'"
        )
        print(f"  Updated {tur_perf} rows in statshub_team_performance_events: Turkiye → Turkey")
    results["team_perf_renamed"] = tur_perf

    # Rename in team_players
    if tur_players > 0:
        con.execute(
            "UPDATE statshub_team_players SET team_name='Turkey' WHERE team_name='Turkiye'"
        )
        print(f"  Updated {tur_players} rows in statshub_team_players: Turkiye → Turkey")
    results["team_players_renamed"] = tur_players

    # Ensure Turkey row in statshub_world_cup_teams
    if turkey_in_wct == 0:
        # Copy Turkiye row data to Turkey
        turkiye_row = con.execute(
            "SELECT * FROM statshub_world_cup_teams WHERE team_name='Turkiye' LIMIT 1"
        ).fetchone()
        if turkiye_row:
            cols = [d[0] for d in con.execute(
                "SELECT * FROM statshub_world_cup_teams LIMIT 0"
            ).description]
            row_dict = dict(zip(cols, turkiye_row))
            row_dict.pop("id", None)
            row_dict["team_name"] = "Turkey"
            row_dict["team_name_canonical"] = "turkey"
            row_dict["imported_at"] = utc_now()
            insert_cols = list(row_dict.keys())
            placeholders = ",".join("?" * len(insert_cols))
            con.execute(
                f"INSERT INTO statshub_world_cup_teams ({','.join(insert_cols)}) VALUES ({placeholders})",
                [row_dict[c] for c in insert_cols],
            )
            print("  Inserted Turkey row in statshub_world_cup_teams (copied from Turkiye)")
            results["turkey_row_inserted"] = 1
        else:
            print("  WARNING: No Turkiye row found to copy — inserting minimal Turkey row")
            con.execute(
                "INSERT INTO statshub_world_cup_teams (snapshot_name, statshub_team_id, team_name, team_name_canonical, imported_at) "
                "VALUES (?,?,?,?,?)",
                (SNAPSHOT_NAME, "4700", "Turkey", "turkey", utc_now()),
            )
            results["turkey_row_inserted"] = 1
    else:
        print("  Turkey already exists in statshub_world_cup_teams")
        results["turkey_row_inserted"] = 0

    # Also rename Turkiye → Turkey in statshub_world_cup_teams to keep consistent
    con.execute(
        "UPDATE statshub_world_cup_teams SET team_name='Turkey', team_name_canonical='turkey' "
        "WHERE team_name='Turkiye'"
    )

    con.commit()
    print("  Phase 0 complete.")
    return results


# ── Phase 1: Player ID mapping ────────────────────────────────────────────────

def phase1_player_id_mapping(con: Any, execute: bool) -> dict:
    """
    For each team, call GET /api/team/{tid}/players/performance?tournamentId=16&location=both
    and try to match roster players by name, confirming player_ids.
    """
    print("\n=== PHASE 1: Player ID mapping (tournamentId=16 endpoint) ===")
    total_newly_confirmed = 0
    team_results: list[dict] = []

    for team in TEAMS:
        tid   = team["team_id"]
        tname = team["name"]
        ep    = f"team_{tid}_players_wc26_tournamentId16"
        url   = f"{BASE}/api/team/{tid}/players/performance?tournamentId=16&location=both"

        print(f"\n  {tname} (id={tid})")
        payload, meta = fetch(url, ep, execute=execute)
        ep_players = _parse_endpoint_players(payload) if payload else []
        ep_by_norm = {_norm_key(p["player_name"]): p for p in ep_players}
        ep_by_id   = {p["player_id"]: p for p in ep_players}
        print(f"    Endpoint players: {len(ep_players)} (status={meta['status']})")

        # Load roster for this team (use team_id)
        roster = [dict(r) for r in con.execute("""
            SELECT id, player_name, player_id, statshub_player_id_status,
                   jersey_number, player_id_match_method
            FROM statshub_team_players
            WHERE team_id=?
            ORDER BY CAST(jersey_number AS INTEGER)
        """, (tid,)).fetchall()]

        newly_confirmed = 0
        unresolved = 0

        for rp in roster:
            rstatus  = rp["statshub_player_id_status"] or ""
            rcur_id  = rp["player_id"] or ""
            rname    = rp["player_name"] or ""
            row_id   = rp["id"]

            if rstatus in ("confirmed", "skipped_existing") and rcur_id:
                # Verify in endpoint but don't downgrade
                ep_match = ep_by_id.get(rcur_id)
                if ep_match and rstatus == "skipped_existing":
                    # Upgrade to confirmed
                    con.execute(
                        "UPDATE statshub_team_players SET statshub_player_id_status='confirmed', "
                        "player_id_match_source=?, player_id_match_method=?, "
                        "player_id_confidence_score=95.0, updated_at=? WHERE id=?",
                        ("tournamentId=16_endpoint", "id_in_endpoint", utc_now(), row_id),
                    )
                    newly_confirmed += 1
                continue

            if not ep_players:
                unresolved += 1
                continue

            # Match by: 1) existing id in endpoint, 2) normalized full name, 3) surname unique
            ep_match = None
            method = ""

            if rcur_id and rcur_id in ep_by_id:
                ep_match = ep_by_id[rcur_id]
                method = "existing_id_in_endpoint"
            else:
                norm_r = _norm_key(rname)
                if norm_r and norm_r in ep_by_norm:
                    ep_match = ep_by_norm[norm_r]
                    method = "name_exact_norm"
                else:
                    parts = rname.strip().split()
                    last = _norm(parts[-1]) if parts else ""
                    if len(last) > 3:
                        hits = [p for n, p in ep_by_norm.items() if last in n.split()]
                        if len(hits) == 1:
                            ep_match = hits[0]
                            method = f"surname_unique:{parts[-1]}"
                        elif len(hits) > 1:
                            first = _norm(parts[0]) if parts else ""
                            fl_hits = [p for n, p in ep_by_norm.items()
                                       if last in n.split() and first in n.split()]
                            if len(fl_hits) == 1:
                                ep_match = fl_hits[0]
                                method = f"first_last:{parts[0]}_{parts[-1]}"

            if ep_match:
                new_pid = ep_match["player_id"]
                con.execute(
                    "UPDATE statshub_team_players SET player_id=?, "
                    "statshub_player_id_status='confirmed', player_id_match_source=?, "
                    "player_id_match_method=?, player_id_confidence_score=90.0, "
                    "updated_at=? WHERE id=?",
                    (new_pid, "tournamentId=16_endpoint", method, utc_now(), row_id),
                )
                newly_confirmed += 1
                print(f"    [CONFIRM] #{rp['jersey_number']} {rname} → id={new_pid} [{method}]")
            else:
                unresolved += 1

        con.commit()
        print(f"    Newly confirmed: {newly_confirmed} | Unresolved: {unresolved}")
        total_newly_confirmed += newly_confirmed
        team_results.append({
            "team_name": tname, "team_id": tid,
            "endpoint_players": len(ep_players),
            "newly_confirmed": newly_confirmed,
            "unresolved": unresolved,
            "status": meta["status"],
        })

    print(f"\n  Total newly confirmed across 8 teams: {total_newly_confirmed}")
    return {"team_results": team_results, "total_newly_confirmed": total_newly_confirmed}


# ── Phase 2: Player performance download ─────────────────────────────────────

def phase2_player_performance(con: Any, execute: bool) -> dict:
    """
    For every confirmed/skipped_existing player across 8 teams,
    download GET /api/player/{pid}/performance?limit=50 and insert into
    statshub_player_performance_events.
    """
    print("\n=== PHASE 2: Player performance download ===")
    team_ids = [t["team_id"] for t in TEAMS]

    confirmed_players = [dict(r) for r in con.execute("""
        SELECT team_id, team_name, player_id, player_name, statshub_player_id_status
        FROM statshub_team_players
        WHERE team_id IN ({})
          AND statshub_player_id_status IN ('confirmed', 'skipped_existing')
          AND player_id IS NOT NULL AND player_id != ''
        ORDER BY team_name, player_name
    """.format(",".join("?" * len(team_ids))), team_ids).fetchall()]

    print(f"  Total confirmed players to download: {len(confirmed_players)}")

    downloaded = 0
    cached = 0
    failed = 0
    total_events_inserted = 0
    player_results: list[dict] = []

    for p in confirmed_players:
        pid   = p["player_id"]
        pname = p["player_name"]
        tname = p["team_name"]
        ep    = f"player_{pid}_performance_limit50_today_4_matches"
        url   = f"{BASE}/api/player/{pid}/performance?limit=50"

        # Check if already in DB (skip re-download for this run)
        existing_count = con.execute(
            "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=? AND endpoint_name=?",
            (pid, ep),
        ).fetchone()[0]
        if existing_count > 0:
            cached += 1
            player_results.append({
                "team_name": tname, "player_name": pname, "player_id": pid,
                "status": "db_cached", "events_inserted": 0, "existing_events": existing_count,
            })
            continue

        payload, meta = fetch(url, ep, execute=execute)
        if meta["status"] == "cached":
            cached += 1
        elif meta["status"] == "dry_run":
            player_results.append({
                "team_name": tname, "player_name": pname, "player_id": pid,
                "status": "dry_run", "events_inserted": 0, "existing_events": 0,
            })
            continue
        elif "error" in meta["status"] or meta["status"].startswith("http_"):
            failed += 1
            print(f"  FAIL {pname} ({tname}): {meta['status']}")
            player_results.append({
                "team_name": tname, "player_name": pname, "player_id": pid,
                "status": meta["status"], "events_inserted": 0, "existing_events": 0,
            })
            continue
        else:
            downloaded += 1

        perf_rows = _parse_player_perf(payload, pid, pname, tname, ep) if payload else []
        if perf_rows:
            con.execute(
                "DELETE FROM statshub_player_performance_events WHERE player_id=? AND endpoint_name=?",
                (pid, ep),
            )
            for row in perf_rows:
                con.execute("""
                    INSERT INTO statshub_player_performance_events
                        (snapshot_name, player_id, player_name, team_name, endpoint_name,
                         event_date, tournament_name, minutes_played, goals, assists,
                         shots, shots_on_target, fouls, was_fouled, yellow_cards, red_cards,
                         xG, xA, key_passes, passes, accurate_passes, tackles,
                         possession_lost, raw_json, imported_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (SNAPSHOT_NAME, pid, pname, tname, ep,
                      row["event_date"], row["tournament_name"], row["minutes_played"],
                      row["goals"], row["assists"], row["shots"], row["shots_on_target"],
                      row["fouls"], row["was_fouled"], row["yellow_cards"], row["red_cards"],
                      row["xG"], row["xA"], row["key_passes"], row["passes"],
                      row["accurate_passes"], row["tackles"], row["possession_lost"],
                      row["raw_json"], utc_now()))
            con.commit()
            total_events_inserted += len(perf_rows)

        print(f"  {pname} ({tname}): events={len(perf_rows)} [status={meta['status']}]")
        player_results.append({
            "team_name": tname, "player_name": pname, "player_id": pid,
            "status": meta["status"], "events_inserted": len(perf_rows), "existing_events": 0,
        })

    print(f"\n  Downloaded: {downloaded} | Cached: {cached} | Failed: {failed}")
    print(f"  Total player-event rows inserted: {total_events_inserted}")

    # Summary per team
    team_summary: dict[str, dict] = {}
    for pr in player_results:
        tn = pr["team_name"]
        if tn not in team_summary:
            team_summary[tn] = {"players": 0, "events": 0, "failed": 0}
        team_summary[tn]["players"] += 1
        team_summary[tn]["events"] += pr.get("events_inserted", 0) + pr.get("existing_events", 0)
        if pr["status"] not in ("ok", "cached", "db_cached"):
            team_summary[tn]["failed"] += 1

    return {
        "player_results": player_results,
        "team_summary": team_summary,
        "downloaded": downloaded,
        "cached": cached,
        "failed": failed,
        "total_events_inserted": total_events_inserted,
    }


# ── Phase 3: Coverage summary ─────────────────────────────────────────────────

def _coverage_summary(con: Any) -> dict:
    """Compute per-team and per-match coverage metrics."""
    team_ids = {t["team_id"]: t["name"] for t in TEAMS}

    rows = {}
    for tid, tname in team_ids.items():
        db_name = next(t["db_name"] for t in TEAMS if t["team_id"] == tid)
        # Use canonical name (Turkey, not Turkiye) after phase 0 rename
        effective_name = tname  # all renamed to canonical

        perf_rows = con.execute(
            "SELECT COUNT(*) FROM statshub_team_performance_events WHERE team_name=?",
            (effective_name,),
        ).fetchone()[0]

        confirmed = con.execute(
            "SELECT COUNT(*) FROM statshub_team_players WHERE team_id=? "
            "AND statshub_player_id_status IN ('confirmed','skipped_existing')",
            (tid,),
        ).fetchone()[0]
        total_players = con.execute(
            "SELECT COUNT(*) FROM statshub_team_players WHERE team_id=?", (tid,)
        ).fetchone()[0]
        player_events = con.execute(
            """SELECT COUNT(*) FROM statshub_player_performance_events pe
               JOIN statshub_team_players tp ON pe.player_id=tp.player_id
               WHERE tp.team_id=? AND pe.minutes_played >= 15""",
            (tid,),
        ).fetchone()[0]
        player_events_any = con.execute(
            """SELECT COUNT(*) FROM statshub_player_performance_events pe
               JOIN statshub_team_players tp ON pe.player_id=tp.player_id
               WHERE tp.team_id=?""",
            (tid,),
        ).fetchone()[0]

        rows[tname] = {
            "team_name": tname,
            "team_id": tid,
            "team_perf_rows": perf_rows,
            "confirmed_players": confirmed,
            "total_players": total_players,
            "player_events_any": player_events_any,
            "player_events_min15": player_events,
            "team_data_ok": perf_rows >= 10,
            "player_data_ok": confirmed >= 5 and player_events >= 20,
        }

    matches = [
        {"match_name": "Qatar vs Switzerland",
         "home": "Qatar", "away": "Switzerland"},
        {"match_name": "Brazil vs Morocco",
         "home": "Brazil", "away": "Morocco"},
        {"match_name": "Haiti vs Scotland",
         "home": "Haiti", "away": "Scotland"},
        {"match_name": "Australia vs Turkey",
         "home": "Australia", "away": "Turkey"},
    ]
    match_coverage = []
    for m in matches:
        hd = rows.get(m["home"], {})
        ad = rows.get(m["away"], {})
        both_team_ok = hd.get("team_data_ok", False) and ad.get("team_data_ok", False)
        both_player_ok = hd.get("player_data_ok", False) and ad.get("player_data_ok", False)
        match_coverage.append({
            "match_name": m["match_name"],
            "home_team": m["home"],
            "away_team": m["away"],
            "home_team_perf_rows": hd.get("team_perf_rows", 0),
            "away_team_perf_rows": ad.get("team_perf_rows", 0),
            "home_confirmed_players": hd.get("confirmed_players", 0),
            "away_confirmed_players": ad.get("confirmed_players", 0),
            "home_player_events_min15": hd.get("player_events_min15", 0),
            "away_player_events_min15": ad.get("player_events_min15", 0),
            "team_level_ev_ok": both_team_ok,
            "player_prop_ev_ok": both_player_ok,
            "data_completeness_status": "COMPLETE" if (both_team_ok and both_player_ok) else "PARTIAL",
        })

    return {"team_rows": list(rows.values()), "match_coverage": match_coverage}


# ── Phase 4: EV rebuild ───────────────────────────────────────────────────────

def phase4_ev_rebuild(skip: bool = False) -> dict:
    """Trigger --from-raw EV rebuild. No new Odds-API.io calls."""
    if skip:
        print("\n=== PHASE 4: EV rebuild SKIPPED (--skip-rebuild) ===")
        return {"status": "skipped"}

    print("\n=== PHASE 4: EV rebuild from raw odds ===")
    cmd = [sys.executable, "-m", "scripts.fetch_today_4_matches_live_api_odds", "--from-raw"]
    print(f"  Running: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(ROOT_DIR),
                            encoding="utf-8", errors="replace")
    if result.returncode == 0:
        print("  EV rebuild: SUCCESS")
        print(result.stdout[-2000:] if len(result.stdout) > 2000 else result.stdout)
    else:
        print(f"  EV rebuild FAILED (rc={result.returncode})")
        print(result.stderr[-1000:] if result.stderr else "")
    return {
        "status": "success" if result.returncode == 0 else "failed",
        "returncode": result.returncode,
        "stdout_tail": result.stdout[-500:] if result.stdout else "",
    }


# ── Phase 5: Excel workbook ───────────────────────────────────────────────────

def phase5_excel(
    con: Any,
    coverage: dict,
    phase1_results: dict,
    phase2_results: dict,
    ev_rebuild_result: dict,
) -> Path:
    """Write 11-sheet coverage review workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  WARNING: openpyxl not available — skipping Excel export")
        return OUT_XLSX

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.remove(ws)

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

    def _hdr(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

    def _autofit(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    # Sheet 1: match_coverage_overview
    ws1 = wb.create_sheet("match_coverage_overview")
    h1 = ["match_name", "home_team", "away_team",
          "home_team_perf_rows", "away_team_perf_rows",
          "home_confirmed_players", "away_confirmed_players",
          "home_player_events_min15", "away_player_events_min15",
          "team_level_ev_ok", "player_prop_ev_ok", "data_completeness_status"]
    _hdr(ws1, h1)
    for mc in coverage["match_coverage"]:
        row = [mc[h] for h in h1]
        ws1.append(row)
        fill = GREEN_FILL if mc["data_completeness_status"] == "COMPLETE" else YELLOW_FILL
        for cell in ws1[ws1.max_row]:
            cell.fill = fill
    _autofit(ws1)

    # Sheet 2: team_coverage_detail
    ws2 = wb.create_sheet("team_coverage_detail")
    h2 = ["team_name", "team_id", "team_perf_rows",
          "confirmed_players", "total_players",
          "player_events_any", "player_events_min15",
          "team_data_ok", "player_data_ok"]
    _hdr(ws2, h2)
    for tr in coverage["team_rows"]:
        row = [tr[h] for h in h2]
        ws2.append(row)
        fill = GREEN_FILL if tr["team_data_ok"] and tr["player_data_ok"] else RED_FILL
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
    _autofit(ws2)

    # Sheet 3: player_id_mapping_summary
    ws3 = wb.create_sheet("player_id_mapping_summary")
    h3 = ["team_name", "team_id", "endpoint_players",
          "newly_confirmed", "unresolved", "status"]
    _hdr(ws3, h3)
    for tr in phase1_results.get("team_results", []):
        ws3.append([tr[h] for h in h3])
    _autofit(ws3)

    # Sheet 4: confirmed_players_all_teams
    ws4 = wb.create_sheet("confirmed_players_all_teams")
    h4 = ["team_name", "player_name", "player_id",
          "statshub_player_id_status", "jersey_number", "player_id_match_method"]
    _hdr(ws4, h4)
    team_ids = [t["team_id"] for t in TEAMS]
    confirmed = con.execute("""
        SELECT team_name, player_name, player_id, statshub_player_id_status,
               jersey_number, player_id_match_method
        FROM statshub_team_players
        WHERE team_id IN ({})
          AND statshub_player_id_status IN ('confirmed','skipped_existing')
          AND player_id IS NOT NULL
        ORDER BY team_name, CAST(jersey_number AS INTEGER)
    """.format(",".join("?" * len(team_ids))), team_ids).fetchall()
    for r in confirmed:
        ws4.append(list(r))
    _autofit(ws4)

    # Sheet 5: unresolved_players
    ws5 = wb.create_sheet("unresolved_players")
    h5 = ["team_name", "player_name", "player_id",
          "statshub_player_id_status", "jersey_number"]
    _hdr(ws5, h5)
    unresolved = con.execute("""
        SELECT team_name, player_name, player_id, statshub_player_id_status, jersey_number
        FROM statshub_team_players
        WHERE team_id IN ({})
          AND statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        ORDER BY team_name, CAST(jersey_number AS INTEGER)
    """.format(",".join("?" * len(team_ids))), team_ids).fetchall()
    for r in unresolved:
        ws5.append(list(r))
    _autofit(ws5)

    # Sheet 6: player_performance_summary
    ws6 = wb.create_sheet("player_performance_summary")
    h6 = ["team_name", "player_name", "player_id",
          "status", "events_inserted", "existing_events"]
    _hdr(ws6, h6)
    for pr in phase2_results.get("player_results", []):
        ws6.append([pr.get(h, "") for h in h6])
    _autofit(ws6)

    # Sheet 7: player_perf_events_sample (min15)
    ws7 = wb.create_sheet("player_perf_events_min15_sample")
    h7 = ["player_name", "team_name", "event_date", "tournament_name",
          "minutes_played", "goals", "assists", "shots", "yellow_cards", "xG"]
    _hdr(ws7, h7)
    sample_rows = con.execute("""
        SELECT player_name, team_name, event_date, tournament_name,
               minutes_played, goals, assists, shots, yellow_cards, xG
        FROM statshub_player_performance_events
        WHERE player_id IN (
            SELECT player_id FROM statshub_team_players
            WHERE team_id IN ({}) AND player_id IS NOT NULL
        ) AND minutes_played >= 15
        ORDER BY team_name, player_name, event_date DESC
        LIMIT 500
    """.format(",".join("?" * len(team_ids))), team_ids).fetchall()
    for r in sample_rows:
        ws7.append(list(r))
    _autofit(ws7)

    # Sheet 8: team_performance_summary
    ws8 = wb.create_sheet("team_performance_summary")
    h8 = ["team_name", "events", "avg_goals_for", "avg_goals_against",
          "avg_corners", "avg_yellow_cards", "avg_shots"]
    _hdr(ws8, h8)
    for t in TEAMS:
        tname = t["name"]
        row = con.execute("""
            SELECT COUNT(*),
                   AVG(goals_for), AVG(goals_against),
                   AVG(corners), AVG(yellow_cards), AVG(shots)
            FROM statshub_team_performance_events
            WHERE team_name=?
        """, (tname,)).fetchone()
        if row:
            ws8.append([tname] + [round(v, 3) if v is not None else None for v in row])
    _autofit(ws8)

    # Sheet 9: ev_betting_value (from betting_value_scores_new if available)
    ws9 = wb.create_sheet("ev_betting_value")
    try:
        h9 = ["match_name", "market_type", "bet_description", "team_name",
              "player_name", "player_id", "verdict", "expected_value",
              "model_probability", "sample_size", "line", "odds_decimal",
              "market_scope", "priority_class", "minutes_filter_status"]
        _hdr(ws9, h9)
        ev_rows = con.execute("""
            SELECT match_name, market_type, bet_description, team_name,
                   player_name, player_id, verdict, expected_value,
                   model_probability, sample_size, line, odds_decimal,
                   market_scope, priority_class, minutes_filter_status
            FROM betting_value_scores_new
            WHERE run_name='today_4_matches_live_api_odds_probe'
              AND verdict='VALUE' AND expected_value > 0
            ORDER BY expected_value DESC
            LIMIT 200
        """).fetchall()
        for r in ev_rows:
            ws9.append(list(r))
    except Exception as e:
        ws9.append([f"EV data not available: {e}"])
    _autofit(ws9)

    # Sheet 10: ev_quality_audit
    ws10 = wb.create_sheet("ev_quality_audit")
    h10 = ["market_scope", "market_type", "verdict", "count",
           "avg_ev", "avg_sample_size", "minutes_filter_ok", "notes"]
    _hdr(ws10, h10)
    try:
        audit = con.execute("""
            SELECT market_scope, market_type, verdict,
                   COUNT(*) as cnt,
                   ROUND(AVG(expected_value),4) as avg_ev,
                   ROUND(AVG(sample_size),1) as avg_ss,
                   SUM(CASE WHEN minutes_filter_status='ok' THEN 1 ELSE 0 END) as mp_ok,
                   ''
            FROM betting_value_scores_new
            WHERE run_name='today_4_matches_live_api_odds_probe'
            GROUP BY market_scope, market_type, verdict
            ORDER BY cnt DESC
        """).fetchall()
        for r in audit:
            ws10.append(list(r))
    except Exception as e:
        ws10.append([f"Audit data not available: {e}"])
    _autofit(ws10)

    # Sheet 11: run_metadata
    ws11 = wb.create_sheet("run_metadata")
    h11 = ["key", "value"]
    _hdr(ws11, h11)
    ev_rb = ev_rebuild_result or {}
    metadata = [
        ("snapshot_name", SNAPSHOT_NAME),
        ("generated_at", utc_now()),
        ("match_date", "2026-06-13"),
        ("matches", "Qatar vs Switzerland, Brazil vs Morocco, Haiti vs Scotland, Australia vs Turkey"),
        ("phase0_turkey_fix", "Turkiye → Turkey renamed in team_perf_events, team_players, world_cup_teams"),
        ("phase1_newly_confirmed", str(phase1_results.get("total_newly_confirmed", 0))),
        ("phase2_downloaded", str(phase2_results.get("downloaded", 0))),
        ("phase2_cached", str(phase2_results.get("cached", 0))),
        ("phase2_failed", str(phase2_results.get("failed", 0))),
        ("phase2_events_inserted", str(phase2_results.get("total_events_inserted", 0))),
        ("ev_rebuild_status", ev_rb.get("status", "not_run")),
        ("ev_rebuild_returncode", str(ev_rb.get("returncode", ""))),
    ]
    for k, v in metadata:
        ws11.append([k, v])
    _autofit(ws11)

    wb.save(OUT_XLSX)
    print(f"\n  Workbook saved: {OUT_XLSX}")
    return OUT_XLSX


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="StatsHub coverage for today's 4 matches")
    parser.add_argument("--execute", action="store_true",
                        help="Make live StatsHub API calls (default: dry run)")
    parser.add_argument("--skip-rebuild", action="store_true",
                        help="Skip the --from-raw EV rebuild step")
    args = parser.parse_args()

    if args.execute:
        if os.environ.get("STATSHUB_ENABLED", "").lower() not in ("true", "1", "yes"):
            print("ERROR: --execute requires STATSHUB_ENABLED=true in environment", file=sys.stderr)
            sys.exit(1)
        print("Mode: LIVE (StatsHub API calls enabled)")
    else:
        print("Mode: DRY RUN (no API calls — pass --execute to fetch live data)")

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    with connect() as con:
        # Phase 0: Fix Turkey name
        phase0_results = phase0_fix_turkey(con)

        # Phase 1: Player ID mapping
        phase1_results = phase1_player_id_mapping(con, execute=args.execute)

        # Phase 2: Player performance download
        phase2_results = phase2_player_performance(con, execute=args.execute)

        # Coverage summary
        coverage = _coverage_summary(con)

        # Report
        print("\n=== Coverage Summary ===")
        for mc in coverage["match_coverage"]:
            status = mc["data_completeness_status"]
            print(f"\n  {mc['match_name']} [{status}]")
            print(f"    Home ({mc['home_team']}): "
                  f"team_perf={mc['home_team_perf_rows']} "
                  f"confirmed_players={mc['home_confirmed_players']} "
                  f"player_events_min15={mc['home_player_events_min15']}")
            print(f"    Away ({mc['away_team']}): "
                  f"team_perf={mc['away_team_perf_rows']} "
                  f"confirmed_players={mc['away_confirmed_players']} "
                  f"player_events_min15={mc['away_player_events_min15']}")
            print(f"    team_ev_ok={mc['team_level_ev_ok']} "
                  f"player_prop_ev_ok={mc['player_prop_ev_ok']}")

        # Phase 4: EV rebuild
        ev_result = phase4_ev_rebuild(skip=args.skip_rebuild)

        # Phase 5: Excel workbook
        xlsx_path = phase5_excel(con, coverage, phase1_results, phase2_results, ev_result)

    print(f"\n=== Done ===")
    print(f"Coverage workbook: {xlsx_path}")
    if not args.execute:
        print("\nTo run live StatsHub downloads:")
        print("  set STATSHUB_ENABLED=true && python -m scripts.fetch_today_4_matches_statshub_coverage --execute")


if __name__ == "__main__":
    main()
