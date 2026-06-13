"""
Audit USA vs Paraguay EV quality.

Usage:
    python -m scripts.audit_usa_paraguay_ev_quality

Produces:
    data/processed/betting/usa_paraguay_ev_quality_audit.xlsx
"""
from __future__ import annotations

import json
import re
import sqlite3
import sys
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DB_PATH = Path("data/mundial.db")
OUT_PATH = Path("data/processed/betting/usa_paraguay_ev_quality_audit.xlsx")
RUN_NAME = "usa_paraguay_live_api_odds_probe"
MATCH_NAME = "United States vs Paraguay"

# ---------------------------------------------------------------------------
# Priority classification
# ---------------------------------------------------------------------------
HARD_DATA_PRIORITY = {
    "player_total_shots", "player_shots_on_target", "player_fouls_committed",
    "player_fouled", "player_cards", "player_to_be_booked", "goalkeeper_saves",
    "team_corners", "total_corners", "team_cards", "total_cards",
    "team_total_shots", "team_shots_on_target", "team_total_goals",
    "over_under_goals",
}
MEDIUM_PRIORITY = {
    "anytime_goalscorer", "player_goals", "player_assists",
    "player_passes", "player_tackles",
}
COMPLEX_IGNORE = {
    "match_winner", "1x2", "double_chance", "draw_no_bet", "handicap",
    "correct_score", "first_half_winner", "first_half_goals",
    "first_half_corners", "first_half_cards", "bet_builder",
    "half_time_full_time", "asian_handicap", "next_goalscorer",
    "last_goalscorer", "first_goalscorer", "multi_scorers",
}


def classify_market(mtype: str) -> str:
    if not mtype:
        return "unknown_needs_review"
    m = mtype.lower()
    if m in HARD_DATA_PRIORITY:
        return "hard_data_priority"
    if m in MEDIUM_PRIORITY:
        return "medium_priority"
    if m in COMPLEX_IGNORE:
        return "complex_ignore_for_now"
    # partial keyword rules
    if any(k in m for k in ("first_half", "half_time", "ht_", "_ht", "correct_score",
                             "builder", "combo", "exotic", "asian", "handicap", "spread")):
        return "complex_ignore_for_now"
    if any(k in m for k in ("shot", "foul", "card", "corner", "save", "goal",
                             "assist", "tackle", "pass", "booking")):
        return "hard_data_priority"
    return "unknown_needs_review"


# ---------------------------------------------------------------------------
# StatsHub field mapping
# ---------------------------------------------------------------------------
FIELD_MAPPING: dict[str, dict] = {
    "player_total_shots": {
        "db_table": "statshub_player_performance_events",
        "db_field": "shots",
        "expected_meaning": "total shots (on + off target)",
        "note": "Includes shots off target; not same as shots_on_target",
    },
    "player_shots_on_target": {
        "db_table": "statshub_player_performance_events",
        "db_field": "shots_on_target",
        "expected_meaning": "shots on target only",
        "note": "OK — direct field",
    },
    "player_fouls_committed": {
        "db_table": "statshub_player_performance_events",
        "db_field": "fouls",
        "expected_meaning": "fouls committed by player",
        "note": "Verify this is COMMITTED not suffered",
    },
    "player_fouled": {
        "db_table": "statshub_player_performance_events",
        "db_field": "was_fouled",
        "expected_meaning": "times player was fouled (fouled suffered)",
        "note": "Must be was_fouled not fouls — different stat",
    },
    "player_cards": {
        "db_table": "statshub_player_performance_events",
        "db_field": "yellow_cards",
        "expected_meaning": "yellow cards received",
        "note": "Does not include red cards unless field is total_cards",
    },
    "player_to_be_booked": {
        "db_table": "statshub_player_performance_events",
        "db_field": "yellow_cards",
        "expected_meaning": "yellow cards (to be booked = any card)",
        "note": "Should include reds too if available",
    },
    "goalkeeper_saves": {
        "db_table": "statshub_team_performance_events",
        "db_field": "goalkeeper_saves",
        "expected_meaning": "team-level GK saves",
        "note": "Team proxy — not individual GK performance",
    },
    "team_corners": {
        "db_table": "statshub_team_performance_events",
        "db_field": "corners",
        "expected_meaning": "corners won by team",
        "note": "OK — direct field",
    },
    "total_corners": {
        "db_table": "statshub_team_performance_events",
        "db_field": "corners (home + away combined)",
        "expected_meaning": "total match corners",
        "note": "Paired independent events — may underestimate variance",
    },
    "team_cards": {
        "db_table": "statshub_team_performance_events",
        "db_field": "yellow_cards",
        "expected_meaning": "yellow cards for team",
        "note": "Does not include red cards",
    },
    "total_cards": {
        "db_table": "statshub_team_performance_events",
        "db_field": "yellow_cards (home + away combined)",
        "expected_meaning": "total match cards",
        "note": "Yellows only — red cards may be missing",
    },
    "team_total_shots": {
        "db_table": "statshub_team_performance_events",
        "db_field": "shots",
        "expected_meaning": "total shots by team",
        "note": "OK — direct field",
    },
    "team_shots_on_target": {
        "db_table": "statshub_team_performance_events",
        "db_field": "shots_on_target",
        "expected_meaning": "shots on target by team",
        "note": "OK — direct field",
    },
    "team_total_goals": {
        "db_table": "statshub_team_performance_events",
        "db_field": "goals_for",
        "expected_meaning": "goals scored by team",
        "note": "OK — direct field",
    },
    "over_under_goals": {
        "db_table": "statshub_team_performance_events",
        "db_field": "goals_for + goals_against / match total",
        "expected_meaning": "total match goals",
        "note": "REVIEW: may use only home goals_for, not combined match total",
    },
    "anytime_goalscorer": {
        "db_table": "statshub_player_performance_events",
        "db_field": "goals",
        "expected_meaning": "goals scored per appearance",
        "note": "OK — direct field, medium priority",
    },
    "player_goals": {
        "db_table": "statshub_player_performance_events",
        "db_field": "goals",
        "expected_meaning": "goals scored per appearance",
        "note": "OK — direct field, medium priority",
    },
    "player_assists": {
        "db_table": "statshub_player_performance_events",
        "db_field": "assists",
        "expected_meaning": "assists per appearance",
        "note": "OK — medium priority",
    },
    "player_passes": {
        "db_table": "statshub_player_performance_events",
        "db_field": "passes",
        "expected_meaning": "passes per appearance",
        "note": "Medium priority",
    },
    "player_tackles": {
        "db_table": "statshub_player_performance_events",
        "db_field": "tackles",
        "expected_meaning": "tackles per appearance",
        "note": "Medium priority",
    },
}

FIELD_REVIEW_FLAGS: dict[str, str] = {
    "player_fouled": "REVIEW",  # was_fouled vs fouls — different stat, verify
    "player_cards": "REVIEW",   # only yellows, not total bookings
    "player_to_be_booked": "REVIEW",  # only yellows
    "total_cards": "REVIEW",    # only yellows, not reds
    "team_cards": "REVIEW",     # only yellows
    "goalkeeper_saves": "REVIEW",  # team proxy, not individual
    "over_under_goals": "REVIEW",  # may not be true match total
    "total_corners": "REVIEW",  # independence assumption
}


def get_field_mapping_status(mtype: str) -> str:
    if mtype in FIELD_REVIEW_FLAGS:
        return FIELD_REVIEW_FLAGS[mtype]
    if mtype in FIELD_MAPPING:
        return "OK"
    return "MISSING_FIELD"


# ---------------------------------------------------------------------------
# Side/line audit helpers
# ---------------------------------------------------------------------------
def audit_side_line(row: dict) -> tuple[str, str]:
    """Returns (status, note) for side/line parsing."""
    side = (row.get("side") or "").lower().strip()
    line = row.get("line")
    sel = (row.get("raw_selection_name") or "").lower()
    mtype = (row.get("market_type") or "").lower()
    notes = []

    # Check side is populated
    if not side:
        return "SUSPICIOUS", "side is empty"

    if side not in ("over", "under", "yes", "no", "home", "away", "draw",
                    "team", "player", "none"):
        notes.append(f"unexpected side='{side}'")

    # For over/under markets, line must be numeric
    if side in ("over", "under"):
        if line is None or line == "":
            notes.append("line is null for over/under market")
        else:
            try:
                float(line)
            except (TypeError, ValueError):
                notes.append(f"line='{line}' not numeric")

    # Check selection name contains matching keyword
    if side == "over" and "under" in sel and "over" not in sel:
        notes.append(f"side=over but selection='{row.get('raw_selection_name')}'")
    if side == "under" and "over" in sel and "under" not in sel:
        notes.append(f"side=under but selection='{row.get('raw_selection_name')}'")

    # For player markets, line 0 might be valid (anytime scorer) but suspicious for others
    if line == 0 and mtype in ("player_total_shots", "player_shots_on_target",
                                "player_fouls_committed", "player_fouled",
                                "player_cards"):
        notes.append("line=0 unusual for this market type")

    if notes:
        return "SUSPICIOUS", "; ".join(notes)
    return "OK", ""


# ---------------------------------------------------------------------------
# Minutes played check
# ---------------------------------------------------------------------------
def check_minutes_played(con: sqlite3.Connection, player_id: str, mtype: str) -> dict:
    if not player_id or mtype not in {
        "player_total_shots", "player_shots_on_target", "player_fouls_committed",
        "player_fouled", "player_cards", "player_to_be_booked", "player_goals",
        "player_assists", "player_passes", "player_tackles", "anytime_goalscorer",
    }:
        return {"checked": False, "reason": "not a player prop"}

    rows = con.execute(
        "SELECT COUNT(*) as total, "
        "SUM(CASE WHEN (minutes_played IS NULL OR minutes_played = 0) THEN 1 ELSE 0 END) as zero_min "
        "FROM statshub_player_performance_events WHERE player_id=?",
        (str(player_id),),
    ).fetchone()
    if not rows or rows[0] == 0:
        return {"checked": True, "total": 0, "zero_min": 0,
                "status": "NO_DATA", "note": "no events in DB for player"}

    total = rows[0] or 0
    zero_min = rows[1] or 0
    pct_zero = round(zero_min / total * 100, 1) if total > 0 else 0
    status = "SUSPICIOUS" if pct_zero > 20 else "OK"
    return {
        "checked": True,
        "total": total,
        "zero_min": zero_min,
        "pct_zero": pct_zero,
        "status": status,
        "note": f"{zero_min}/{total} rows have 0/null minutes ({pct_zero}%)",
    }


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------
DUP_KEY_COLS = [
    "run_name", "bookmaker", "match_name", "market_type",
    "raw_market_name", "raw_selection_name", "team_name",
    "player_name", "side", "line", "odds_decimal",
]


def make_dup_key(row: dict) -> tuple:
    return tuple(str(row.get(c) or "").strip().lower() for c in DUP_KEY_COLS)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
FLAG_FILL = PatternFill("solid", fgColor="FFCCCC")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")


def write_sheet(wb: openpyxl.Workbook, name: str,
                headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                if cell.fill.fgColor.rgb == "00000000":
                    cell.fill = ALT_FILL
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "A2"


def flag_row(ws, row_idx: int, status: str, ncols: int) -> None:
    fill = FLAG_FILL if status == "SUSPICIOUS" else (
        OK_FILL if status == "OK" else WARN_FILL
    )
    for col in range(1, ncols + 1):
        ws.cell(row=row_idx, column=col).fill = fill


# ---------------------------------------------------------------------------
# Main audit
# ---------------------------------------------------------------------------
def run_audit() -> None:
    if not DB_PATH.exists():
        print(f"ERROR: DB not found at {DB_PATH}", file=sys.stderr)
        sys.exit(1)

    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row

    # -----------------------------------------------------------------------
    # Load all EV rows for this match
    # -----------------------------------------------------------------------
    ev_rows = [dict(r) for r in con.execute(
        """
        SELECT id, bookmaker, match_name, market_type, raw_market_name,
               raw_selection_name, team_name, player_name, player_id,
               side, line, odds_decimal, implied_probability,
               model_probability, edge, expected_value, sample_size,
               probability_method, probability_status, normalized_status,
               data_quality_status, verdict, notes
        FROM betting_value_scores_new
        WHERE match_name = ? AND expected_value IS NOT NULL
        ORDER BY expected_value DESC
        """,
        (MATCH_NAME,),
    ).fetchall()]

    print(f"Loaded {len(ev_rows)} EV rows for {MATCH_NAME}")

    # -----------------------------------------------------------------------
    # Task 1: Priority classification
    # -----------------------------------------------------------------------
    for r in ev_rows:
        r["priority_class"] = classify_market(r.get("market_type") or "")
        r["field_mapping_status"] = get_field_mapping_status(r.get("market_type") or "")

    counts_by_priority: dict[str, int] = {}
    value_by_priority: dict[str, int] = {}
    no_value_by_priority: dict[str, int] = {}
    for r in ev_rows:
        pc = r["priority_class"]
        counts_by_priority[pc] = counts_by_priority.get(pc, 0) + 1
        if r.get("verdict") == "VALUE":
            value_by_priority[pc] = value_by_priority.get(pc, 0) + 1
        elif r.get("verdict") == "NO_VALUE":
            no_value_by_priority[pc] = no_value_by_priority.get(pc, 0) + 1

    print("\n--- Task 1: Priority Classification ---")
    for pc, cnt in sorted(counts_by_priority.items()):
        v = value_by_priority.get(pc, 0)
        nv = no_value_by_priority.get(pc, 0)
        print(f"  {pc}: {cnt} EV rows | {v} VALUE | {nv} NO_VALUE")

    # -----------------------------------------------------------------------
    # Task 2: Duplicate audit
    # -----------------------------------------------------------------------
    # Normalized table duplicates
    norm_rows = [dict(r) for r in con.execute(
        """
        SELECT run_name, bookmaker, match_name, market_type, raw_market_name,
               raw_selection_name, team_name, player_name, side, line, odds_decimal
        FROM betting_odds_normalized
        WHERE match_name = ?
        """,
        (MATCH_NAME,),
    ).fetchall()]

    norm_key_counts: dict[tuple, int] = {}
    for r in norm_rows:
        k = make_dup_key(r)
        norm_key_counts[k] = norm_key_counts.get(k, 0) + 1
    norm_dups = {k: v for k, v in norm_key_counts.items() if v > 1}

    ev_key_counts: dict[tuple, int] = {}
    for r in ev_rows:
        k = make_dup_key(r)
        ev_key_counts[k] = ev_key_counts.get(k, 0) + 1
    ev_dups = {k: v for k, v in ev_key_counts.items() if v > 1}

    value_ev_rows = [r for r in ev_rows if r.get("verdict") == "VALUE"]
    val_key_counts: dict[tuple, int] = {}
    for r in value_ev_rows:
        k = make_dup_key(r)
        val_key_counts[k] = val_key_counts.get(k, 0) + 1
    val_dups = {k: v for k, v in val_key_counts.items() if v > 1}

    dup_ev_row_count = sum(v - 1 for v in ev_key_counts.values() if v > 1)
    dup_val_row_count = sum(v - 1 for v in val_key_counts.values() if v > 1)

    print(f"\n--- Task 2: Duplicates ---")
    print(f"  Normalized dup groups: {len(norm_dups)}")
    print(f"  EV dup groups: {len(ev_dups)} ({dup_ev_row_count} extra rows)")
    print(f"  VALUE dup groups: {len(val_dups)} ({dup_val_row_count} extra rows)")

    # Sample duplicates
    dup_examples = []
    for k, cnt in sorted(ev_dups.items(), key=lambda x: -x[1])[:10]:
        matching = [r for r in ev_rows if make_dup_key(r) == k]
        for r in matching[:2]:
            dup_examples.append({**r, "_dup_count": cnt})

    # Dedup proposal: keep highest EV row per key
    dedup_ev_rows: dict[tuple, dict] = {}
    for r in ev_rows:
        k = make_dup_key(r)
        existing = dedup_ev_rows.get(k)
        if existing is None or (r.get("expected_value") or -999) > (existing.get("expected_value") or -999):
            dedup_ev_rows[k] = r
    deduped = list(dedup_ev_rows.values())
    deduped.sort(key=lambda r: -(r.get("expected_value") or 0))
    print(f"  After dedup: {len(deduped)} unique EV rows")

    # -----------------------------------------------------------------------
    # Task 3: Over/Under audit
    # -----------------------------------------------------------------------
    side_results = {"OK": 0, "SUSPICIOUS": 0}
    side_audit_rows = []
    for r in ev_rows:
        status, note = audit_side_line(r)
        r["side_line_status"] = status
        r["side_line_note"] = note
        side_results[status] = side_results.get(status, 0) + 1
        if status == "SUSPICIOUS":
            side_audit_rows.append(r)

    print(f"\n--- Task 3: Over/Under Audit ---")
    print(f"  OK: {side_results.get('OK', 0)}")
    print(f"  SUSPICIOUS: {side_results.get('SUSPICIOUS', 0)}")
    for r in side_audit_rows[:5]:
        print(f"    [{r['market_type']}] sel='{r['raw_selection_name']}' "
              f"side={r['side']} line={r['line']} → {r['side_line_note']}")

    # -----------------------------------------------------------------------
    # Task 4: Sample size audit
    # -----------------------------------------------------------------------
    buckets = {"<5": [], "5-9": [], "10-14": [], "15+": [], "null": []}
    for r in ev_rows:
        ss = r.get("sample_size")
        if ss is None:
            buckets["null"].append(r)
        elif ss < 5:
            buckets["<5"].append(r)
        elif ss < 10:
            buckets["5-9"].append(r)
        elif ss < 15:
            buckets["10-14"].append(r)
        else:
            buckets["15+"].append(r)

    print(f"\n--- Task 4: Sample Size ---")
    for b, rows in buckets.items():
        val_cnt = sum(1 for r in rows if r.get("verdict") == "VALUE")
        print(f"  {b}: {len(rows)} EV rows | {val_cnt} VALUE")

    low_sample_value = [r for r in (buckets["<5"] + buckets["5-9"])
                        if r.get("verdict") == "VALUE"]
    print(f"  VALUE rows with sample_size < 10: {len(low_sample_value)}")
    for r in low_sample_value[:5]:
        print(f"    [{r['market_type']}] {r.get('player_name') or r.get('team_name')} "
              f"ss={r['sample_size']} EV={r.get('expected_value'):.3f}")

    # -----------------------------------------------------------------------
    # Task 5: Minutes played audit
    # -----------------------------------------------------------------------
    mp_audit_rows = []
    mp_issues = 0
    for r in ev_rows:
        mtype = r.get("market_type") or ""
        pid = r.get("player_id")
        if mtype not in {
            "player_total_shots", "player_shots_on_target", "player_fouls_committed",
            "player_fouled", "player_cards", "player_to_be_booked", "player_goals",
            "player_assists", "player_passes", "player_tackles", "anytime_goalscorer",
        }:
            r["mp_status"] = "N/A"
            r["mp_note"] = "not a player prop"
            continue
        if not pid:
            r["mp_status"] = "NO_PLAYER_ID"
            r["mp_note"] = "no player_id"
            mp_issues += 1
            mp_audit_rows.append(r)
            continue
        result = check_minutes_played(con, str(pid), mtype)
        r["mp_status"] = result.get("status", "UNKNOWN")
        r["mp_note"] = result.get("note", "")
        if result.get("status") in ("SUSPICIOUS", "NO_DATA", "NO_PLAYER_ID"):
            mp_issues += 1
            mp_audit_rows.append(r)

    player_prop_ev_count = sum(1 for r in ev_rows if r.get("mp_status") not in ("N/A", None))
    no_data_count = sum(1 for r in ev_rows if r.get("mp_status") == "NO_DATA")
    no_id_count = sum(1 for r in ev_rows if r.get("mp_status") == "NO_PLAYER_ID")
    suspicious_mp_count = sum(1 for r in ev_rows if r.get("mp_status") == "SUSPICIOUS")
    true_mp_issues = suspicious_mp_count + no_id_count
    mp_issues = true_mp_issues  # reassign for summary

    print(f"\n--- Task 5: Minutes Played ---")
    print(f"  Player prop EV rows: {player_prop_ev_count}")
    print(f"  No player_id: {no_id_count}")
    print(f"  Player not found in statshub: {no_data_count}")
    print(f"  Suspicious (>20%% zero-min rows): {suspicious_mp_count}")
    print(f"  True issues (no_id + suspicious): {true_mp_issues}")
    for r in [x for x in mp_audit_rows if x.get("mp_status") == "SUSPICIOUS"][:5]:
        print(f"    [{r['market_type']}] {r.get('player_name')} pid={r.get('player_id')} "
              f"-> {r['mp_status']}: {r['mp_note']}")

    # -----------------------------------------------------------------------
    # Task 6: StatsHub field mapping
    # -----------------------------------------------------------------------
    mtype_counts: dict[str, dict] = {}
    for r in ev_rows:
        mt = r.get("market_type") or "UNKNOWN"
        if mt not in mtype_counts:
            mtype_counts[mt] = {"count": 0, "value": 0, "sample_sizes": []}
        mtype_counts[mt]["count"] += 1
        if r.get("verdict") == "VALUE":
            mtype_counts[mt]["value"] += 1
        ss = r.get("sample_size")
        if ss is not None:
            mtype_counts[mt]["sample_sizes"].append(ss)

    # Verify DB columns actually exist
    player_cols_exist: set[str] = set()
    team_cols_exist: set[str] = set()
    try:
        pinfo = con.execute("PRAGMA table_info(statshub_player_performance_events)").fetchall()
        player_cols_exist = {row[1] for row in pinfo}
    except Exception:
        pass
    try:
        tinfo = con.execute("PRAGMA table_info(statshub_team_performance_events)").fetchall()
        team_cols_exist = {row[1] for row in tinfo}
    except Exception:
        pass

    field_mapping_audit_rows = []
    for mt, info in sorted(mtype_counts.items(), key=lambda x: -x[1]["count"]):
        fm = FIELD_MAPPING.get(mt, {})
        db_field = fm.get("db_field", "UNKNOWN")
        db_table = fm.get("db_table", "UNKNOWN")
        base_field = db_field.split(" ")[0].split("(")[0]
        if "player" in db_table:
            exists = base_field in player_cols_exist
        elif "team" in db_table:
            exists = base_field in team_cols_exist
        else:
            exists = False
        status = get_field_mapping_status(mt)
        avg_ss = round(sum(info["sample_sizes"]) / len(info["sample_sizes"]), 1) if info["sample_sizes"] else None
        field_mapping_audit_rows.append({
            "market_type": mt,
            "db_table": db_table or "UNKNOWN",
            "db_field": db_field or "UNKNOWN",
            "field_exists": "YES" if exists else ("UNKNOWN" if not db_table else "NO"),
            "expected_meaning": fm.get("expected_meaning", ""),
            "ev_count": info["count"],
            "value_count": info["value"],
            "avg_sample_size": avg_ss,
            "mapping_status": status,
            "note": fm.get("note", ""),
            "priority_class": classify_market(mt),
        })

    wrong_or_review = [r for r in field_mapping_audit_rows
                       if r["mapping_status"] in ("REVIEW", "WRONG", "MISSING_FIELD")]
    print(f"\n--- Task 6: Field Mapping ---")
    print(f"  Market types with EV: {len(field_mapping_audit_rows)}")
    print(f"  REVIEW/WRONG: {len(wrong_or_review)}")
    for r in wrong_or_review:
        print(f"    [{r['mapping_status']}] {r['market_type']} -> {r['db_field']}: {r['note']}")

    # -----------------------------------------------------------------------
    # Final actionable EV rows
    # -----------------------------------------------------------------------
    actionable = [
        r for r in deduped
        if r.get("priority_class") in ("hard_data_priority", "medium_priority")
        and r.get("side_line_status") == "OK"
        and (r.get("sample_size") or 0) >= 10
        and r.get("field_mapping_status") not in ("WRONG", "MISSING_FIELD")
    ]
    actionable_value = [r for r in actionable if r.get("verdict") == "VALUE"]
    actionable_hard = [r for r in actionable if r.get("priority_class") == "hard_data_priority"]
    actionable_hard_value = [r for r in actionable_hard if r.get("verdict") == "VALUE"]

    print(f"\n--- Final Actionable ---")
    print(f"  Total EV rows: {len(ev_rows)}")
    print(f"  After dedup: {len(deduped)}")
    print(f"  Actionable (priority + side OK + ss>=10 + field OK): {len(actionable)}")
    print(f"  Actionable VALUE: {len(actionable_value)}")
    print(f"  Actionable hard-data: {len(actionable_hard)}")
    print(f"  Actionable hard-data VALUE: {len(actionable_hard_value)}")

    print("\nTop 20 actionable EV rows:")
    for i, r in enumerate(sorted(actionable_value, key=lambda x: -(x.get("expected_value") or 0))[:20], 1):
        print(f"  {i:2d}. [{r['market_type']}] {r.get('player_name') or r.get('team_name')} "
              f"side={r['side']} line={r['line']} odds={r['odds_decimal']} "
              f"model_p={r.get('model_probability'):.3f} EV={r.get('expected_value'):.3f} "
              f"ss={r['sample_size']}")

    # -----------------------------------------------------------------------
    # Task 7: Write audit workbook
    # -----------------------------------------------------------------------
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Sheet 1: audit_summary
    ws_sum = wb.create_sheet("audit_summary")
    summary_data = [
        ["Metric", "Value"],
        ["Total EV rows", len(ev_rows)],
        ["After dedup EV rows", len(deduped)],
        ["hard_data_priority EV", counts_by_priority.get("hard_data_priority", 0)],
        ["medium_priority EV", counts_by_priority.get("medium_priority", 0)],
        ["complex_ignore EV", counts_by_priority.get("complex_ignore_for_now", 0)],
        ["unknown EV", counts_by_priority.get("unknown_needs_review", 0)],
        ["VALUE (hard data)", value_by_priority.get("hard_data_priority", 0)],
        ["VALUE (medium)", value_by_priority.get("medium_priority", 0)],
        ["Duplicate EV groups", len(ev_dups)],
        ["Duplicate extra rows", dup_ev_row_count],
        ["VALUE dup groups", len(val_dups)],
        ["Side/line SUSPICIOUS rows", side_results.get("SUSPICIOUS", 0)],
        ["Sample < 5", len(buckets["<5"])],
        ["Sample 5-9", len(buckets["5-9"])],
        ["Sample 10-14", len(buckets["10-14"])],
        ["Sample 15+", len(buckets["15+"])],
        ["Sample null", len(buckets["null"])],
        ["VALUE rows with ss < 10", len(low_sample_value)],
        ["Minutes played issues", mp_issues],
        ["Field mapping REVIEW/WRONG", len(wrong_or_review)],
        ["Actionable EV rows (all filters)", len(actionable)],
        ["Actionable VALUE rows", len(actionable_value)],
        ["Actionable hard-data VALUE", len(actionable_hard_value)],
        ["Recommended min sample_size", 10],
        ["Run name", RUN_NAME],
        ["Match", MATCH_NAME],
    ]
    for row in summary_data:
        ws_sum.append(row)
    ws_sum[1][0].font = Font(bold=True)
    ws_sum[1][1].font = Font(bold=True)
    for cell in ws_sum[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 20

    # Sheet 2: ev_priority_classification
    ev_class_hdrs = [
        "priority_class", "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "model_probability", "implied_probability", "expected_value",
        "sample_size", "data_quality_status", "verdict",
        "side_line_status", "side_line_note", "field_mapping_status", "mp_status",
    ]
    ev_class_rows = []
    for r in ev_rows:
        ev_class_rows.append([r.get(c) for c in ev_class_hdrs])
    write_sheet(wb, "ev_priority_classification", ev_class_hdrs, ev_class_rows)

    # Sheet 3: hard_data_ev_rows
    hard_rows = [r for r in ev_rows if r.get("priority_class") == "hard_data_priority"]
    hard_hdrs = [
        "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "model_probability", "implied_probability", "expected_value",
        "sample_size", "verdict", "side_line_status", "field_mapping_status",
    ]
    write_sheet(wb, "hard_data_ev_rows", hard_hdrs,
                [[r.get(c) for c in hard_hdrs] for r in hard_rows])

    # Sheet 4: value_rows_hard_data
    hard_value_rows = [r for r in hard_rows if r.get("verdict") == "VALUE"]
    hard_value_rows.sort(key=lambda x: -(x.get("expected_value") or 0))
    write_sheet(wb, "value_rows_hard_data", hard_hdrs,
                [[r.get(c) for c in hard_value_rows[0].keys() if c in hard_hdrs]
                 for r in hard_value_rows]
                if hard_value_rows else [])
    # Redo properly
    wb.remove(wb["value_rows_hard_data"])
    write_sheet(wb, "value_rows_hard_data", hard_hdrs,
                [[r.get(c) for c in hard_hdrs] for r in hard_value_rows])

    # Sheet 5: duplicates_audit
    dup_hdrs = [
        "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "expected_value", "verdict", "_dup_count",
    ]
    write_sheet(wb, "duplicates_audit", dup_hdrs,
                [[r.get(c) for c in dup_hdrs] for r in dup_examples])

    # Sheet 6: over_under_audit
    ou_hdrs = [
        "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "model_probability", "expected_value", "verdict",
        "side_line_status", "side_line_note",
    ]
    suspicious_rows = [r for r in ev_rows if r.get("side_line_status") == "SUSPICIOUS"]
    write_sheet(wb, "over_under_audit", ou_hdrs,
                [[r.get(c) for c in ou_hdrs] for r in suspicious_rows])

    # Sheet 7: sample_size_audit
    ss_hdrs = [
        "sample_size_bucket", "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "team_name", "side", "line", "odds_decimal",
        "expected_value", "sample_size", "verdict",
    ]
    ss_rows_all = []
    for bucket_name, bucket_list in buckets.items():
        for r in bucket_list:
            row_out = [bucket_name] + [r.get(c) for c in ss_hdrs[1:]]
            ss_rows_all.append(row_out)
    write_sheet(wb, "sample_size_audit", ss_hdrs, ss_rows_all)

    # Sheet 8: minutes_played_audit
    mp_hdrs = [
        "market_type", "raw_market_name", "raw_selection_name",
        "player_name", "player_id", "side", "line",
        "expected_value", "sample_size", "verdict",
        "mp_status", "mp_note",
    ]
    mp_rows_all = [r for r in ev_rows if r.get("mp_status") not in ("N/A", None)]
    write_sheet(wb, "minutes_played_audit", mp_hdrs,
                [[r.get(c) for c in mp_hdrs] for r in mp_rows_all])

    # Sheet 9: statshub_field_mapping_audit
    fm_hdrs = [
        "market_type", "priority_class", "db_table", "db_field", "field_exists",
        "expected_meaning", "ev_count", "value_count", "avg_sample_size",
        "mapping_status", "note",
    ]
    write_sheet(wb, "statshub_field_mapping_audit", fm_hdrs,
                [[r.get(c) for c in fm_hdrs] for r in field_mapping_audit_rows])

    # Sheet 10: excluded_complex_markets
    complex_rows = [r for r in ev_rows
                    if r.get("priority_class") in ("complex_ignore_for_now", "unknown_needs_review")]
    cx_hdrs = [
        "priority_class", "market_type", "raw_market_name", "raw_selection_name",
        "side", "line", "odds_decimal", "expected_value", "verdict",
    ]
    write_sheet(wb, "excluded_complex_markets", cx_hdrs,
                [[r.get(c) for c in cx_hdrs] for r in complex_rows])

    # Sheet 11: recommended_dashboard_filters
    rec_hdrs = ["Filter", "Value", "Rationale"]
    rec_rows = [
        ["priority_class", "hard_data_priority OR medium_priority",
         "Exclude complex models, specials, combos"],
        ["expected_value", "> 0", "Only positive EV bets"],
        ["sample_size", ">= 10", "Minimum statistical reliability (prefer >= 15 for publish)"],
        ["side_line_status", "OK", "No suspicious side/line parsing"],
        ["field_mapping_status", "OK or REVIEW", "Exclude WRONG / MISSING_FIELD"],
        ["verdict", "VALUE", "Show only actionable bets"],
        ["data_quality_status", "OK or low_sample", "Exclude insufficient_data"],
        ["dedup_rule", "Keep highest EV row per (bookmaker + market + selection + line)",
         "Prevents /odds and /odds/multi double-counting"],
        ["NOTE", "REVIEW rows are included but flagged",
         "User decides whether to act on REVIEW field mappings"],
    ]
    write_sheet(wb, "recommended_dashboard_filters", rec_hdrs, rec_rows)

    wb.save(OUT_PATH)
    print(f"\nAudit workbook saved: {OUT_PATH}")

    con.close()


if __name__ == "__main__":
    run_audit()
