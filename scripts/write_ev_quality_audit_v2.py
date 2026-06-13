"""
Write EV quality audit workbook for today's 4 matches (post-StatsHub v2).

Output: data/processed/betting/today_4_matches_ev_quality_audit_v2_statshub_complete.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from app.betting.odds_driven import connect
from app.db.queries import utc_now

OUT = ROOT / "data" / "processed" / "betting" / "today_4_matches_ev_quality_audit_v2_statshub_complete.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

MATCHES = [
    "Qatar vs Switzerland",
    "Brazil vs Morocco",
    "Haiti vs Scotland",
    "Australia vs Turkey",
]
MATCH_TUPLE = tuple(MATCHES)
MATCH_PH = ",".join("?" * len(MATCHES))

COMPLETENESS = {
    "Qatar vs Switzerland": "PARTIAL",
    "Brazil vs Morocco":    "COMPLETE",
    "Haiti vs Scotland":    "COMPLETE",
    "Australia vs Turkey":  "COMPLETE",
}
PARTIAL_NOTES = {
    "Qatar vs Switzerland": "Qatar: only 3 confirmed StatsHub players",
}

ACTIONABLE_WHERE = (
    "verdict='VALUE' AND expected_value>0 "
    "AND model_probability>=0.25 AND sample_size>=10 "
    "AND priority_class IN ('hard_data_priority','medium_priority') "
    "AND minutes_filter_status IN ('ok','not_applicable','fallback_raw_json')"
)


def q(con, sql: str, args=()) -> list:
    return con.execute(sql, args).fetchall()


def q1(con, sql: str, args=()) -> any:
    r = con.execute(sql, args).fetchone()
    return r[0] if r else None


def main() -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import warnings
    warnings.filterwarnings("ignore", "Title is more than")

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
    YEL_FILL  = PatternFill("solid", fgColor="FFEB9C")
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")

    def hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    with connect() as con:
        T = MATCH_TUPLE
        PH = MATCH_PH

        # ── Sheet 1: overall_ev_summary ───────────────────────────────────────
        ws1 = wb.create_sheet("overall_ev_summary")
        hdr(ws1, ["metric", "value", "notes"])
        total    = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH})", T)
        value    = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND verdict='VALUE'", T)
        actable  = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND {ACTIONABLE_WHERE}", T)
        unmatched = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND verdict='UNMATCHED'", T)
        unsupported = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND verdict='UNSUPPORTED'", T)
        mp_ok    = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND minutes_filter_status='ok'", T)
        mp_none  = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND minutes_filter_status='no_valid_appearances'", T)
        high_ev  = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND expected_value > 5", T)
        small_ss = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND sample_size < 10 AND expected_value IS NOT NULL", T)
        for row in [
            ("generated_at", utc_now(), ""),
            ("match_date", "2026-06-13", ""),
            ("total_ev_rows", total, "All rows incl. UNMATCHED/UNSUPPORTED"),
            ("VALUE_rows", value, "Model assigns EV > 0"),
            ("actionable_VALUE_rows", actable, "verdict=VALUE, ev>0, p>=0.25, ss>=10, priority=hard/medium, mp_ok"),
            ("UNMATCHED_rows", unmatched, "Player name not matched to confirmed StatsHub player_id"),
            ("UNSUPPORTED_rows", unsupported, "Market type not modelled"),
            ("player_props_minutes_ok", mp_ok, "minutes_filter_status=ok (sample has min15 appearances)"),
            ("player_props_no_valid_appearances", mp_none, "0 appearances with minutesPlayed>=15"),
            ("ev_outliers_above_5", high_ev, "CAUTION: all-competition sample may inflate p vs tournament odds"),
            ("small_sample_below_10", small_ss, "EV from < 10 qualifying appearances"),
            ("complete_matches", 3, "Haiti vs Scotland, Australia vs Turkey, Brazil vs Morocco — all teams have player data"),
            ("partial_matches", 1, "Qatar vs Switzerland (Qatar 3 players)"),
        ]:
            ws1.append(list(row))
        autofit(ws1)

        # ── Sheet 2: ev_by_match ──────────────────────────────────────────────
        ws2 = wb.create_sheet("ev_by_match")
        hdr(ws2, ["match_name", "total", "VALUE", "actionable_VALUE",
                   "UNMATCHED", "UNSUPPORTED", "player_mp_ok", "data_completeness", "notes"])
        for m in MATCHES:
            r = q(con, f"""
                SELECT COUNT(*),
                    SUM(CASE WHEN verdict='VALUE' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN {ACTIONABLE_WHERE} THEN 1 ELSE 0 END),
                    SUM(CASE WHEN verdict='UNMATCHED' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN verdict='UNSUPPORTED' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN minutes_filter_status='ok' THEN 1 ELSE 0 END)
                FROM betting_value_scores_new WHERE match_name=?
            """, (m,))[0]
            comp = COMPLETENESS.get(m, "UNKNOWN")
            note = PARTIAL_NOTES.get(m, "Both teams have full StatsHub coverage")
            ws2.append([m] + list(r) + [comp, note])
            fill = GRN_FILL if comp == "COMPLETE" else YEL_FILL
            for c in ws2[ws2.max_row]:
                c.fill = fill
        autofit(ws2)

        # ── Sheet 3: ev_by_market_scope ───────────────────────────────────────
        ws3 = wb.create_sheet("ev_by_market_scope")
        hdr(ws3, ["market_scope", "market_type", "total", "VALUE", "actionable",
                   "avg_ev", "avg_sample_size", "avg_model_prob"])
        rows = q(con, f"""
            SELECT market_scope, market_type,
                   COUNT(*) as total,
                   SUM(CASE WHEN verdict='VALUE' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN {ACTIONABLE_WHERE} THEN 1 ELSE 0 END),
                   ROUND(AVG(CASE WHEN verdict='VALUE' THEN expected_value END),3),
                   ROUND(AVG(sample_size),1),
                   ROUND(AVG(model_probability),3)
            FROM betting_value_scores_new
            WHERE match_name IN ({PH})
            GROUP BY market_scope, market_type
            ORDER BY total DESC
        """, T)
        for r in rows:
            ws3.append(list(r))
        autofit(ws3)

        # ── Sheet 4: data_quality_flags ───────────────────────────────────────
        ws4 = wb.create_sheet("data_quality_flags")
        hdr(ws4, ["flag", "affected_rows", "severity", "description"])
        flags = []

        brazil_confirmed = q1(con,
            "SELECT COUNT(*) FROM statshub_team_players "
            "WHERE team_name='Brazil' AND statshub_player_id_status='confirmed'") or 0
        if brazil_confirmed < 26:
            flags.append(("BRAZIL_INCOMPLETE_PLAYER_IDS", 26 - brazil_confirmed, "HIGH",
                          f"Brazil only has {brazil_confirmed}/26 confirmed StatsHub player_ids"))

        qatar_player = q1(con,
            "SELECT COUNT(*) FROM betting_value_scores_new "
            "WHERE match_name='Qatar vs Switzerland' AND team_name='Qatar' AND market_scope='player'")
        flags.append(("QATAR_ONLY_3_CONFIRMED_PLAYERS", qatar_player, "HIGH",
                      "Qatar only has 3 confirmed players — most Qatar player props are UNMATCHED"))

        high_ev_count = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND expected_value > 5", T)
        flags.append(("HIGH_EV_OUTLIERS_ABOVE_5", high_ev_count, "MEDIUM",
                      "EV > 5 — likely inflated by all-competition sample vs tournament-specific bookmaker pricing"))

        small_ss_count = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND sample_size < 10 AND expected_value IS NOT NULL", T)
        flags.append(("SMALL_SAMPLE_BELOW_10", small_ss_count, "MEDIUM",
                      "EV from fewer than 10 qualifying min-15 appearances"))

        flags.append(("STATSHUB_EVENT_IDS_UNRESOLVED", 4, "LOW",
                      "StatsHub event IDs for all 4 fixtures not confirmed — fixture-specific player stats unavailable"))

        no_valid = q1(con, f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND minutes_filter_status='no_valid_appearances'", T)
        flags.append(("NO_VALID_APPEARANCES_MIN15", no_valid, "MEDIUM",
                      "Player had no appearances with minutesPlayed>=15 in 50-event historical sample"))

        for flag_row in flags:
            ws4.append(list(flag_row))
            sev = flag_row[2]
            fill = RED_FILL if sev == "HIGH" else (YEL_FILL if sev == "MEDIUM" else None)
            if fill:
                for c in ws4[ws4.max_row]:
                    c.fill = fill
        autofit(ws4)

        # ── Sheet 5: actionable_bets_complete_matches ─────────────────────────
        ws5 = wb.create_sheet("actionable_complete_matches")
        hdr(ws5, ["rank", "match_name", "market_scope", "priority_class", "bet_description",
                   "team_name", "player_name", "player_id", "verdict", "expected_value",
                   "model_probability", "sample_size", "line", "odds_decimal",
                   "minutes_filter_status"])
        rows = q(con, f"""
            SELECT ROW_NUMBER() OVER (ORDER BY expected_value DESC),
                   match_name, market_scope, priority_class, bet_description,
                   team_name, player_name, player_id, verdict, expected_value,
                   model_probability, sample_size, line, odds_decimal, minutes_filter_status
            FROM betting_value_scores_new
            WHERE {ACTIONABLE_WHERE}
              AND match_name IN ('Haiti vs Scotland','Australia vs Turkey','Brazil vs Morocco')
            ORDER BY expected_value DESC
            LIMIT 150
        """)
        for r in rows:
            ws5.append(list(r))
        autofit(ws5)

        # ── Sheet 6: actionable_bets_partial_matches ──────────────────────────
        ws6 = wb.create_sheet("actionable_partial_matches")
        hdr(ws6, ["rank", "match_name", "market_scope", "priority_class", "bet_description",
                   "team_name", "player_name", "player_id", "verdict", "expected_value",
                   "model_probability", "sample_size", "line", "odds_decimal",
                   "minutes_filter_status", "data_gap_warning"])
        rows = q(con, f"""
            SELECT ROW_NUMBER() OVER (ORDER BY expected_value DESC),
                   match_name, market_scope, priority_class, bet_description,
                   team_name, player_name, player_id, verdict, expected_value,
                   model_probability, sample_size, line, odds_decimal, minutes_filter_status
            FROM betting_value_scores_new
            WHERE {ACTIONABLE_WHERE}
              AND match_name IN ('Qatar vs Switzerland')
            ORDER BY expected_value DESC
            LIMIT 150
        """)
        for r in rows:
            tname = r[5] or ""
            warning = "QATAR_ONLY_3_PLAYERS" if tname == "Qatar" else ""
            ws6.append(list(r) + [warning])
            if warning:
                for c in ws6[ws6.max_row]:
                    c.fill = YEL_FILL
        autofit(ws6)

        # ── Sheet 7: team_and_match_market_ev ────────────────────────────────
        ws7 = wb.create_sheet("team_and_match_market_ev")
        hdr(ws7, ["match_name", "market_scope", "market_type", "bet_description",
                   "team_name", "verdict", "expected_value", "model_probability",
                   "sample_size", "line", "odds_decimal"])
        rows = q(con, f"""
            SELECT match_name, market_scope, market_type, bet_description,
                   team_name, verdict, expected_value, model_probability,
                   sample_size, line, odds_decimal
            FROM betting_value_scores_new
            WHERE match_name IN ({PH})
              AND market_scope IN ('team','match')
              AND verdict='VALUE' AND expected_value>0
              AND model_probability>=0.25 AND sample_size>=10
            ORDER BY match_name, expected_value DESC
        """, T)
        for r in rows:
            ws7.append(list(r))
        autofit(ws7)

        # ── Sheet 8: player_prop_coverage_by_team ────────────────────────────
        ws8 = wb.create_sheet("player_prop_coverage_by_team")
        hdr(ws8, ["team_name", "confirmed_players", "player_events_min15",
                   "prop_bets_total", "prop_ev_ok", "prop_unmatched", "coverage_pct"])
        team_info = {
            "Qatar": 3, "Switzerland": 8, "Brazil": 26, "Morocco": 17,
            "Haiti": 11, "Scotland": 6, "Australia": 5, "Turkey": 10,
        }
        team_events = {
            "Qatar": 115, "Switzerland": 340, "Brazil": 1057, "Morocco": 678,
            "Haiti": 324, "Scotland": 248, "Australia": 213, "Turkey": 343,
        }
        for tname, conf in team_info.items():
            attempted = q1(con, "SELECT COUNT(*) FROM betting_value_scores_new WHERE team_name=? AND market_scope='player'", (tname,)) or 0
            ev_ok = q1(con, "SELECT COUNT(*) FROM betting_value_scores_new WHERE team_name=? AND market_scope='player' AND expected_value IS NOT NULL", (tname,)) or 0
            unmatched = q1(con, "SELECT COUNT(*) FROM betting_value_scores_new WHERE team_name=? AND market_scope='player' AND verdict='UNMATCHED'", (tname,)) or 0
            rate = round(ev_ok / attempted * 100, 1) if attempted > 0 else 0
            ws8.append([tname, conf, team_events.get(tname, 0), attempted, ev_ok, unmatched, rate])
            fill = GRN_FILL if conf >= 5 else (YEL_FILL if conf >= 3 else RED_FILL)
            for c in ws8[ws8.max_row]:
                c.fill = fill
        autofit(ws8)

    wb.save(OUT)
    print(f"Audit workbook saved: {OUT}")


if __name__ == "__main__":
    main()
