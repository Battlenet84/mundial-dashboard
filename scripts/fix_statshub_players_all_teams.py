"""
Apply the fixed StatsHub player parser to all 8 teams and rebuild EV.

Root cause fix: old parser looked for playerId/playerName; StatsHub returns id/name
at data[] level.  Brazil pilot confirmed this.  Now apply globally.

Strategy per team:
  1. Re-parse cached tournamentId=16 file with fixed parser
  2. If confirmed < 22 after step 1, call no-filter /api/team/{id}/players/performance
  3. Match FIFA roster → StatsHub IDs (token + slug + alias table)
  4. Commit confirmed IDs to DB
  5. Download player performance for newly confirmed players
  6. Rebuild EV from existing raw odds (no Odds-API calls)
  7. Write output workbooks + final report

Usage:
    python -m scripts.fix_statshub_players_all_teams              # dry-run (local files)
    python -m scripts.fix_statshub_players_all_teams --execute    # live API calls
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests

from app.betting.odds_driven import connect
from app.db.queries import utc_now

# ── Config ────────────────────────────────────────────────────────────────────

SNAPSHOT_NAME  = "global_player_fix"
BASE           = "https://www.statshub.com"
RATE_DELAY     = 1.5
SCRIPT_SOURCE  = "fix_statshub_players_all_teams"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         "https://www.statshub.com/",
    "Origin":          "https://www.statshub.com",
}

COVERAGE_SNAPSHOT = (
    ROOT / "data" / "raw" / "statshub" / "snapshots" / "today_4_matches_statshub_coverage"
)
RAW_DIR = ROOT / "data" / "raw" / "statshub" / "snapshots" / SNAPSHOT_NAME
RAW_DIR.mkdir(parents=True, exist_ok=True)

OUT_DIR = ROOT / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_COVERAGE = OUT_DIR / "statshub" / "today_4_matches_statshub_coverage_review_v2_global_parser.xlsx"
OUT_EV       = OUT_DIR / "betting" / "today_4_matches_live_api_odds_value_scores_v3_global_player_coverage.xlsx"
OUT_AUDIT    = OUT_DIR / "betting" / "today_4_matches_ev_quality_audit_v3_global_player_coverage.xlsx"
OUT_COVERAGE.parent.mkdir(parents=True, exist_ok=True)
OUT_EV.parent.mkdir(parents=True, exist_ok=True)

# ── Team configs ──────────────────────────────────────────────────────────────

TEAMS = [
    {"name": "Qatar",       "team_id": "4792", "db_name": "Qatar",       "match": "Qatar vs Switzerland"},
    {"name": "Switzerland", "team_id": "4699", "db_name": "Switzerland", "match": "Qatar vs Switzerland"},
    {"name": "Brazil",      "team_id": "4748", "db_name": "Brazil",      "match": "Brazil vs Morocco"},
    {"name": "Morocco",     "team_id": "4778", "db_name": "Morocco",     "match": "Brazil vs Morocco"},
    {"name": "Haiti",       "team_id": "7229", "db_name": "Haiti",       "match": "Haiti vs Scotland"},
    {"name": "Scotland",    "team_id": "4695", "db_name": "Scotland",    "match": "Haiti vs Scotland"},
    {"name": "Australia",   "team_id": "4741", "db_name": "Australia",   "match": "Australia vs Turkey"},
    {"name": "Turkey",      "team_id": "4700", "db_name": "Turkey",      "match": "Australia vs Turkey"},
]

MATCHES = [
    "Qatar vs Switzerland",
    "Brazil vs Morocco",
    "Haiti vs Scotland",
    "Australia vs Turkey",
]

# Minimum confirmed players before we try the no-filter endpoint
MIN_CONFIRMED_THRESHOLD = 22

# ── Per-team alias tables ─────────────────────────────────────────────────────
# (fifa_name_fragment, common_name, statshub_slug)
# Fragment = lowercase no-accent substring of the FIFA name stored in DB.
# Slug     = statshub slug from their player page.

TEAM_ALIASES: dict[str, list[tuple[str, str, str]]] = {
    "Brazil": [
        ("belloli",            "Raphinha",          "raphinha"),
        ("tolentino coelho",   "Lucas Paqueta",     "lucas-paqueta"),
        ("aoas correa",        "Marquinhos",        "marquinhos"),
        ("tavares",            "Fabinho",           "fabinho"),
        ("fabio henrique",     "Fabinho",           "fabinho"),
        ("paixao",             "Vinicius Junior",   "vinicius-junior"),
        ("vinicius jose",      "Vinicius Junior",   "vinicius-junior"),
        ("cunha",              "Matheus Cunha",     "matheus-cunha"),
        ("carneiro",           "Matheus Cunha",     "matheus-cunha"),
        ("endrick",            "Endrick",           "endrick"),
        ("moreira de sousa",   "Endrick",           "endrick"),
        ("simplicio rocha",    "Rayan Vitor",       "rayan-vitor"),
        ("rayan vitor",        "Rayan Vitor",       "rayan-vitor"),
        ("igor thiago",        "Igor Thiago",       "igor-thiago"),
        ("luiz henrique",      "Luiz Henrique",     "luiz-henrique"),
        ("gabriel magalhaes",  "Gabriel Magalhaes", "gabriel-magalhaes"),
        ("dos santos magalh",  "Gabriel Magalhaes", "gabriel-magalhaes"),
        ("leo pereira",        "Leo Pereira",       "leo-pereira"),
        ("leonardo pereira",   "Leo Pereira",       "leo-pereira"),
        ("roger ibanez",       "Roger Ibanez",      "roger-ibanez"),
        ("ibanez da silva",    "Roger Ibanez",      "roger-ibanez"),
        ("douglas dos santos", "Douglas Luiz",      "douglas-luiz"),
        ("justino de melo",    "Douglas Luiz",      "douglas-luiz"),
        ("danilo dos santos",  "Danilo",            "danilo-dos-santos"),
        ("santos de oliveira", "Danilo",            "danilo-dos-santos"),
        ("gleison bremer",     "Bremer",            "bremer"),
        ("alex sandro",        "Alex Sandro",       "alex-sandro"),
        ("alisson",            "Alisson",           "alisson"),
        ("ederson santana",    "Ederson",           "ederson"),
        ("weverton",           "Weverton",          "weverton"),
        ("neymar",             "Neymar",            "neymar"),
        ("casimiro",           "Casemiro",          "casemiro"),
        ("danilo luiz",        "Danilo",            "danilo"),
        ("gabriel martinelli", "Gabriel Martinelli","gabriel-martinelli"),
        ("teodoro martinelli", "Gabriel Martinelli","gabriel-martinelli"),
        ("bruno guimaraes",    "Bruno Guimaraes",   "bruno-guimaraes"),
        ("rodriguez moura",    "Bruno Guimaraes",   "bruno-guimaraes"),
        ("eder militao",       "Eder Militao",      "eder-militao"),
        ("militao",            "Eder Militao",      "eder-militao"),
    ],
    "Qatar": [
        ("almoez ali",         "Almoez Ali",        "almoez-ali"),
        ("ali abdulla",        "Almoez Ali",        "almoez-ali"),
        ("akram afif",         "Akram Afif",        "akram-afif"),
        ("afif akram",         "Akram Afif",        "akram-afif"),
        ("hasan haydos",       "Hassan Al-Haydos",  "hassan-al-haydos"),
        ("al haydos",          "Hassan Al-Haydos",  "hassan-al-haydos"),
        ("haydos",             "Hassan Al-Haydos",  "hassan-al-haydos"),
        ("boualem khoukhi",    "Boualem Khoukhi",   "boualem-khoukhi"),
        ("karim boudiaf",      "Karim Boudiaf",     "karim-boudiaf"),
        ("pedro correia",      "Pedro Miguel",      "pedro-miguel"),
        ("pedro miguel",       "Pedro Miguel",      "pedro-miguel"),
        ("yusuf abdurisag",    "Yusuf Abdurisag",   "yusuf-abdurisag"),
        ("meshaal barsham",    "Meshaal Barsham",   "meshaal-barsham"),
        ("saad al sheeb",      "Saad Al Sheeb",     "saad-al-sheeb"),
        ("al sheeb",           "Saad Al Sheeb",     "saad-al-sheeb"),
        ("bassam al rawi",     "Bassam Al Rawi",    "bassam-al-rawi"),
        ("salem al hajri",     "Salem Al Hajri",    "salem-al-hajri"),
        ("al hajri",           "Salem Al Hajri",    "salem-al-hajri"),
        ("tariq salman",       "Tariq Salman",      "tariq-salman"),
        ("ismail mohamad",     "Ismail Mohamad",    "ismail-mohamad"),
        ("jassem gaber",       "Jassem Gaber",      "jassem-gaber"),
        ("assim madibo",       "Assim Madibo",      "assim-madibo"),
        ("madibo",             "Assim Madibo",      "assim-madibo"),
    ],
    "Switzerland": [
        ("granit xhaka",       "Granit Xhaka",      "granit-xhaka"),
        ("xherdan shaqiri",    "Xherdan Shaqiri",   "xherdan-shaqiri"),
        ("shaqiri",            "Xherdan Shaqiri",   "xherdan-shaqiri"),
        ("haris seferovic",    "Haris Seferovic",   "haris-seferovic"),
        ("fabian schar",       "Fabian Schar",      "fabian-schar"),
        ("breel embolo",       "Breel Embolo",      "breel-embolo"),
        ("yann sommer",        "Yann Sommer",       "yann-sommer"),
        ("nico elvedi",        "Nico Elvedi",       "nico-elvedi"),
        ("manuel akanji",      "Manuel Akanji",     "manuel-akanji"),
        ("remo freuler",       "Remo Freuler",      "remo-freuler"),
        ("steven zuber",       "Steven Zuber",      "steven-zuber"),
        ("ruben vargas",       "Ruben Vargas",      "ruben-vargas"),
        ("silvan widmer",      "Silvan Widmer",     "silvan-widmer"),
        ("djibril sow",        "Djibril Sow",       "djibril-sow"),
        ("noah okafor",        "Noah Okafor",       "noah-okafor"),
        ("dan ndoye",          "Dan Ndoye",         "dan-ndoye"),
        ("ndoye",              "Dan Ndoye",         "dan-ndoye"),
        ("gregor kobel",       "Gregor Kobel",      "gregor-kobel"),
        ("philipp kohn",       "Philipp Kohn",      "philipp-kohn"),
        ("zeki amdouni",       "Zeki Amdouni",      "zeki-amdouni"),
        ("kwadwo duah",        "Kwadwo Duah",       "kwadwo-duah"),
        ("fabian frei",        "Fabian Frei",       "fabian-frei"),
        ("cedric zesiger",     "Cedric Zesiger",    "cedric-zesiger"),
        ("michael lang",       "Michael Lang",      "michael-lang"),
    ],
    "Morocco": [
        ("yassine bounou",     "Bono",              "yassine-bounou"),
        ("bounou",             "Bono",              "yassine-bounou"),
        ("hakim ziyech",       "Hakim Ziyech",      "hakim-ziyech"),
        ("achraf hakimi",      "Achraf Hakimi",     "achraf-hakimi"),
        ("romain saiss",       "Romain Saiss",      "romain-saiss"),
        ("saiss",              "Romain Saiss",      "romain-saiss"),
        ("noussair mazraoui",  "Noussair Mazraoui", "noussair-mazraoui"),
        ("mazraoui",           "Noussair Mazraoui", "noussair-mazraoui"),
        ("sofiane boufal",     "Sofiane Boufal",    "sofiane-boufal"),
        ("sofyan amrabat",     "Sofyan Amrabat",    "sofyan-amrabat"),
        ("amrabat",            "Sofyan Amrabat",    "sofyan-amrabat"),
        ("selim amallah",      "Selim Amallah",     "selim-amallah"),
        ("azzedine ounahi",    "Azzedine Ounahi",   "azzedine-ounahi"),
        ("ounahi",             "Azzedine Ounahi",   "azzedine-ounahi"),
        ("abdessamad ezzalzouli", "Ez Abde",        "abdessamad-ezzalzouli"),
        ("ezzalzouli",         "Ez Abde",           "abdessamad-ezzalzouli"),
        ("brahim diaz",        "Brahim Diaz",       "brahim-diaz"),
        ("ilias chair",        "Ilias Chair",       "ilias-chair"),
        ("walid cheddira",     "Walid Cheddira",    "walid-cheddira"),
        ("cheddira",           "Walid Cheddira",    "walid-cheddira"),
        ("ayoub el kaabi",     "Ayoub El Kaabi",    "ayoub-el-kaabi"),
        ("el kaabi",           "Ayoub El Kaabi",    "ayoub-el-kaabi"),
        ("munir el haddadi",   "Munir El Haddadi",  "munir-el-haddadi"),
        ("munir",              "Munir El Haddadi",  "munir-el-haddadi"),
        ("nayef aguerd",       "Nayef Aguerd",      "nayef-aguerd"),
        ("aguerd",             "Nayef Aguerd",      "nayef-aguerd"),
    ],
    "Haiti": [
        ("nazon",              "Duckens Nazon",     "duckens-nazon"),
        ("duckens",            "Duckens Nazon",     "duckens-nazon"),
        ("borgella",           "Derrick Borgella",  "derrick-borgella"),
        ("domingos quina",     "Domingos Quina",    "domingos-quina"),
        ("quina",              "Domingos Quina",    "domingos-quina"),
        ("wilde donald",       "Wilde-Donald",      "wilde-donald-guerrier"),
        ("guerrier",           "Wilde-Donald",      "wilde-donald-guerrier"),
        ("hervens ceus",       "Hervens Ceus",      "hervens-ceus"),
        ("ceus",               "Hervens Ceus",      "hervens-ceus"),
        ("frantzdy pierrot",   "Frantzdy Pierrot",  "frantzdy-pierrot"),
        ("lamaison",           "Lamaison",          "lamaison"),
        ("jean fenelon",       "Jean Fenelon",      "jean-fenelon"),
        ("fenelon",            "Jean Fenelon",      "jean-fenelon"),
        ("louidort",           "Louidort",          "louidort"),
    ],
    "Scotland": [
        ("andy robertson",     "Andy Robertson",    "andy-robertson"),
        ("robertson",          "Andy Robertson",    "andy-robertson"),
        ("kieran tierney",     "Kieran Tierney",    "kieran-tierney"),
        ("tierney",            "Kieran Tierney",    "kieran-tierney"),
        ("scott mctominay",    "Scott McTominay",   "scott-mctominay"),
        ("mctominay",          "Scott McTominay",   "scott-mctominay"),
        ("lawrence shankland", "Lawrence Shankland","lawrence-shankland"),
        ("shankland",          "Lawrence Shankland","lawrence-shankland"),
        ("john mcginn",        "John McGinn",       "john-mcginn"),
        ("mcginn",             "John McGinn",       "john-mcginn"),
        ("stuart armstrong",   "Stuart Armstrong",  "stuart-armstrong"),
        ("ryan jack",          "Ryan Jack",         "ryan-jack"),
        ("craig gordon",       "Craig Gordon",      "craig-gordon"),
        ("angus gunn",         "Angus Gunn",        "angus-gunn"),
        ("nathan patterson",   "Nathan Patterson",  "nathan-patterson"),
        ("jack hendry",        "Jack Hendry",       "jack-hendry"),
        ("callum mcgregor",    "Callum McGregor",   "callum-mcgregor"),
        ("mcgregor",           "Callum McGregor",   "callum-mcgregor"),
        ("kevin nisbet",       "Kevin Nisbet",      "kevin-nisbet"),
        ("ryan christie",      "Ryan Christie",     "ryan-christie"),
        ("lyndon dykes",       "Lyndon Dykes",      "lyndon-dykes"),
        ("dykes",              "Lyndon Dykes",      "lyndon-dykes"),
        ("grant hanley",       "Grant Hanley",      "grant-hanley"),
        ("billy gilmour",      "Billy Gilmour",     "billy-gilmour"),
        ("gilmour",            "Billy Gilmour",     "billy-gilmour"),
        ("liam kelly",         "Liam Kelly",        "liam-kelly"),
    ],
    "Australia": [
        ("mat ryan",           "Mat Ryan",          "mat-ryan"),
        ("matthew ryan",       "Mat Ryan",          "mat-ryan"),
        ("mitchell duke",      "Mitchell Duke",     "mitchell-duke"),
        ("duke",               "Mitchell Duke",     "mitchell-duke"),
        ("mathew leckie",      "Mathew Leckie",     "mathew-leckie"),
        ("leckie",             "Mathew Leckie",     "mathew-leckie"),
        ("aaron mooy",         "Aaron Mooy",        "aaron-mooy"),
        ("mooy",               "Aaron Mooy",        "aaron-mooy"),
        ("ryan gauld",         "Ryan Gauld",        "ryan-gauld"),
        ("aziz behich",        "Aziz Behich",       "aziz-behich"),
        ("behich",             "Aziz Behich",       "aziz-behich"),
        ("jackson irvine",     "Jackson Irvine",    "jackson-irvine"),
        ("irvine",             "Jackson Irvine",    "jackson-irvine"),
        ("martin boyle",       "Martin Boyle",      "martin-boyle"),
        ("boyle",              "Martin Boyle",      "martin-boyle"),
        ("riley mcgree",       "Riley McGree",      "riley-mcgree"),
        ("mcgree",             "Riley McGree",      "riley-mcgree"),
        ("craig goodwin",      "Craig Goodwin",     "craig-goodwin"),
        ("goodwin",            "Craig Goodwin",     "craig-goodwin"),
        ("harry souttar",      "Harry Souttar",     "harry-souttar"),
        ("souttar",            "Harry Souttar",     "harry-souttar"),
        ("kye rowles",         "Kye Rowles",        "kye-rowles"),
        ("rowles",             "Kye Rowles",        "kye-rowles"),
        ("fran karacic",       "Fran Karacic",      "fran-karacic"),
        ("karacic",            "Fran Karacic",      "fran-karacic"),
        ("a  mabil",           "Awer Mabil",        "awer-mabil"),
        ("awer mabil",         "Awer Mabil",        "awer-mabil"),
        ("mabil",              "Awer Mabil",        "awer-mabil"),
        ("jamie maclaren",     "Jamie Maclaren",    "jamie-maclaren"),
        ("maclaren",           "Jamie Maclaren",    "jamie-maclaren"),
        ("brendan hamill",     "Brendan Hamill",    "brendan-hamill"),
        ("hamill",             "Brendan Hamill",    "brendan-hamill"),
        ("gianni stensness",   "Gianni Stensness",  "gianni-stensness"),
        ("stensness",          "Gianni Stensness",  "gianni-stensness"),
        ("calem nieuwenhof",   "Calem Nieuwenhof",  "calem-nieuwenhof"),
        ("nieuwenhof",         "Calem Nieuwenhof",  "calem-nieuwenhof"),
    ],
    "Turkey": [
        ("hakan calhanoglu",   "Hakan Calhanoglu",  "hakan-calhanoglu"),
        ("calhanoglu",         "Hakan Calhanoglu",  "hakan-calhanoglu"),
        ("burak yilmaz",       "Burak Yilmaz",      "burak-yilmaz"),
        ("yilmaz",             "Burak Yilmaz",      "burak-yilmaz"),
        ("kenan yildiz",       "Kenan Yildiz",      "kenan-yildiz"),
        ("yildiz",             "Kenan Yildiz",      "kenan-yildiz"),
        ("arda guler",         "Arda Guler",        "arda-guler"),
        ("guler",              "Arda Guler",        "arda-guler"),
        ("merih demiral",      "Merih Demiral",     "merih-demiral"),
        ("demiral",            "Merih Demiral",     "merih-demiral"),
        ("samet akaydin",      "Samet Akaydin",     "samet-akaydin"),
        ("akaydin",            "Samet Akaydin",     "samet-akaydin"),
        ("abdulkerim bardakci","Abdulkerim Bardakci","abdulkerim-bardakci"),
        ("bardakci",           "Abdulkerim Bardakci","abdulkerim-bardakci"),
        ("ugurcan cakir",      "Ugurcan Cakir",     "ugurcan-cakir"),
        ("cakir",              "Ugurcan Cakir",     "ugurcan-cakir"),
        ("mert gunok",         "Mert Gunok",        "mert-gunok"),
        ("gunok",              "Mert Gunok",        "mert-gunok"),
        ("zeki celik",         "Zeki Celik",        "zeki-celik"),
        ("celik",              "Zeki Celik",        "zeki-celik"),
        ("ferdi kadioglu",     "Ferdi Kadioglu",    "ferdi-kadioglu"),
        ("kadioglu",           "Ferdi Kadioglu",    "ferdi-kadioglu"),
        ("ismail yuksek",      "Ismail Yuksek",     "ismail-yuksek"),
        ("yuksek",             "Ismail Yuksek",     "ismail-yuksek"),
        ("salih ozcan",        "Salih Ozcan",       "salih-ozcan"),
        ("ozcan",              "Salih Ozcan",       "salih-ozcan"),
        ("okay yokuslu",       "Okay Yokuslu",      "okay-yokuslu"),
        ("yokuslu",            "Okay Yokuslu",      "okay-yokuslu"),
        ("cenk tosun",         "Cenk Tosun",        "cenk-tosun"),
        ("tosun",              "Cenk Tosun",        "cenk-tosun"),
        ("irfan kahveci",      "Irfan Kahveci",     "irfan-kahveci"),
        ("kahveci",            "Irfan Kahveci",     "irfan-kahveci"),
        ("yunus akgun",        "Yunus Akgun",       "yunus-akgun"),
        ("akgun",              "Yunus Akgun",       "yunus-akgun"),
    ],
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", " ", s.lower()).strip()


def _slug(s: str) -> str:
    return re.sub(r"\s+", "-", _norm(s))


def _norm_key(s: str) -> str:
    return " ".join(_norm(s).split())


def _to_float(v: Any) -> float | None:
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def fetch(url: str, label: str, execute: bool = True) -> tuple[Any | None, dict]:
    target = RAW_DIR / f"{label}.json"
    txt    = RAW_DIR / f"{label}.txt"
    if target.exists() and target.stat().st_size > 100:
        try:
            return json.loads(target.read_text(encoding="utf-8")), {"status": "cached", "file": str(target)}
        except Exception:
            pass
    if txt.exists():
        return None, {"status": "cached_error", "file": str(txt)}
    if not execute:
        return None, {"status": "dry_run", "file": ""}
    time.sleep(RATE_DELAY)
    try:
        r = requests.get(url, headers=HEADERS, timeout=25)
        try:
            p = json.loads(r.text)
            target.write_text(json.dumps(p), encoding="utf-8")
            return p, {"status": f"http_{r.status_code}", "file": str(target)}
        except Exception:
            txt.write_text(r.text[:5000], encoding="utf-8")
            return None, {"status": f"http_{r.status_code}_nonjson", "file": str(txt)}
    except Exception as exc:
        txt.write_text(f"error: {exc}", encoding="utf-8")
        return None, {"status": "error", "file": str(txt)}


def parse_team_players(payload: Any, source: str) -> list[dict]:
    """
    Extract player records from any StatsHub endpoint shape.
    Fixed parser: looks for data[].id / data[].name first.
    """
    if not payload:
        return []

    candidates: list[dict] = []

    def _try_items(items: Any) -> None:
        if not isinstance(items, list):
            return
        for item in items:
            if not isinstance(item, dict):
                continue
            # Shape A (primary): {id, name, slug, position}
            pid  = item.get("id") or item.get("playerId") or item.get("player_id")
            name = item.get("name") or item.get("playerName") or item.get("player_name")
            if pid and name and isinstance(name, str) and len(name) > 1:
                candidates.append({
                    "player_id": str(pid),
                    "player_name": name,
                    "slug": item.get("slug", ""),
                    "position": item.get("position", ""),
                    "source": source,
                })
                continue
            # Shape B: nested player object
            p = item.get("player") or item.get("playerInfo")
            if isinstance(p, dict):
                pid2  = p.get("id") or p.get("playerId")
                name2 = p.get("name") or p.get("playerName")
                if pid2 and name2:
                    candidates.append({
                        "player_id": str(pid2),
                        "player_name": name2,
                        "slug": p.get("slug", ""),
                        "position": p.get("position", ""),
                        "source": source,
                    })

    if isinstance(payload, list):
        _try_items(payload)
    elif isinstance(payload, dict):
        for key in ("data", "players", "squad", "roster", "results", "items"):
            _try_items(payload.get(key, []))
        for v in payload.values():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                _try_items(v)

    seen: set[str] = set()
    result: list[dict] = []
    for c in candidates:
        pid = c["player_id"]
        if pid not in seen and _norm(c["player_name"]).strip():
            seen.add(pid)
            result.append(c)
    return result


def build_lookups(ep_players: list[dict]) -> tuple[dict, dict]:
    by_slug = {ep["slug"]: ep for ep in ep_players if ep.get("slug")}
    by_norm = {_norm_key(ep["player_name"]): ep for ep in ep_players}
    return by_slug, by_norm


def alias_lookup(fifa_name: str, by_slug: dict, team_name: str) -> dict | None:
    norm = _norm(fifa_name)
    for fragment, common_name, slug in TEAM_ALIASES.get(team_name, []):
        if fragment in norm:
            if slug in by_slug:
                return {**by_slug[slug], "match_method": f"alias:{fragment}→{slug}",
                        "common_name": common_name}
            # slug not found — search by norm_key
            for ep in by_slug.values():
                if _slug(_norm(ep["player_name"])) == slug or ep.get("slug") == slug:
                    return {**ep, "match_method": f"alias:{fragment}→{slug}",
                            "common_name": common_name}
    return None


def match_roster_to_statshub(
    roster: list[dict],
    ep_players: list[dict],
    by_slug: dict,
    by_norm: dict,
    team_name: str,
    confirmed_ids: set[str],   # player_ids already confirmed to this team
) -> list[dict]:
    results: list[dict] = []

    for r in roster:
        fifa_name  = r["player_name"] or ""
        current_id = r["player_id"]
        status     = r["statshub_player_id_status"] or ""
        jersey     = r.get("jersey_number", "?")
        row_id     = r["row_id"]
        norm_fifa  = _norm_key(fifa_name)
        slug_guess = _slug(norm_fifa)

        if status in ("confirmed", "skipped_existing") and current_id:
            results.append({
                "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                "matched_statshub_name": "already_confirmed",
                "player_id": current_id, "match_method": "existing_confirmed",
                "confidence": 1.0, "status": "existing_confirmed",
            })
            continue

        match = None
        method = ""

        # Pass 1: exact normalized name
        if norm_fifa in by_norm:
            match = by_norm[norm_fifa]
            method = "exact_norm"

        # Pass 2: slug exact
        if not match and slug_guess in by_slug:
            match = by_slug[slug_guess]
            method = "slug_exact"

        # Pass 3: alias table
        if not match:
            a = alias_lookup(norm_fifa, by_slug, team_name)
            if a:
                match = a
                method = a.get("match_method", "alias")

        # Pass 4: token subset bidirectional
        if not match:
            wanted_tokens = set(norm_fifa.split())
            for ep_norm, ep in by_norm.items():
                ep_tokens = set(ep_norm.split())
                if wanted_tokens and (wanted_tokens <= ep_tokens or ep_tokens <= wanted_tokens):
                    match = ep
                    method = f"token_subset:{ep_norm}"
                    break

        # Pass 5: slug substring (only if slug is long enough to be discriminating)
        if not match:
            for ep in ep_players:
                ep_slug = ep.get("slug", "")
                if ep_slug and len(slug_guess) >= 6 and (slug_guess in ep_slug or ep_slug in slug_guess):
                    match = ep
                    method = f"slug_substring:{slug_guess}↔{ep_slug}"
                    break

        if match:
            pid = match["player_id"]
            pname = match.get("player_name", "")
            confidence = 0.95 if method in ("exact_norm", "slug_exact", "alias") else 0.85
            # Guard: don't assign same player_id to two FIFA roster entries
            if pid in confirmed_ids and status != "confirmed":
                results.append({
                    "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                    "matched_statshub_name": pname, "player_id": pid,
                    "match_method": f"duplicate_guard:{method}", "confidence": 0.5,
                    "status": "ambiguous",
                })
            else:
                confirmed_ids.add(pid)
                results.append({
                    "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                    "matched_statshub_name": pname,
                    "player_id": pid, "match_method": method,
                    "confidence": confidence, "status": "matched",
                })
        else:
            results.append({
                "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                "matched_statshub_name": "", "player_id": None,
                "match_method": "unresolved", "confidence": 0.0, "status": "unresolved",
            })

    return results


def download_player_perf(
    con: Any,
    player_id: str,
    player_name: str,
    team_name: str,
    execute: bool,
) -> dict:
    """Download /api/player/{id}/performance?limit=50 and insert into DB."""
    # Check if already cached
    existing = con.execute(
        "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=?",
        (player_id,)
    ).fetchone()[0]
    if existing > 0:
        min15 = con.execute(
            "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=? AND minutes_played>=15",
            (player_id,)
        ).fetchone()[0]
        return {"player_id": player_id, "player_name": player_name,
                "events": existing, "min15": min15, "status": "db_cached"}

    url   = f"{BASE}/api/player/{player_id}/performance?limit=50"
    label = f"player_{player_id}_perf"
    payload, meta = fetch(url, label, execute=execute)

    if not payload:
        return {"player_id": player_id, "player_name": player_name,
                "events": 0, "min15": 0, "status": meta["status"]}

    events: list[dict] = []
    if isinstance(payload, dict):
        events = payload.get("data", []) or payload.get("events", []) or []
    elif isinstance(payload, list):
        events = payload

    inserted = 0
    for ev in events:
        if not isinstance(ev, dict):
            continue
        match_id   = str(ev.get("matchId") or ev.get("match_id") or ev.get("id") or "")
        stats      = ev.get("stats") or ev.get("playerStats") or ev
        mins       = _to_float(stats.get("minutesPlayed") or stats.get("minutes_played") or stats.get("minutes"))
        goals      = _to_float(stats.get("goals"))
        assists    = _to_float(stats.get("assists"))
        shots      = _to_float(stats.get("shots") or stats.get("totalShots"))
        shots_ot   = _to_float(stats.get("shotsOnTarget") or stats.get("shots_on_target"))
        fouls_c    = _to_float(stats.get("foulsConceded") or stats.get("fouls_conceded") or stats.get("fouls"))
        fouls_w    = _to_float(stats.get("foulsSuffered") or stats.get("fouls_suffered") or stats.get("wasFouled"))
        yc         = _to_float(stats.get("yellowCards") or stats.get("yellow_cards"))
        rc         = _to_float(stats.get("redCards") or stats.get("red_cards"))
        saves      = _to_float(stats.get("goalKeeperSave") or stats.get("saves"))

        try:
            con.execute("""
                INSERT OR IGNORE INTO statshub_player_performance_events
                (player_id, player_name, team_name, match_id, endpoint_name,
                 minutes_played, goals, assists, shots, shots_on_target,
                 fouls_committed, was_fouled, yellow_cards, red_cards, goalkeeper_saves,
                 raw_json, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                player_id, player_name, team_name, match_id,
                f"player_{player_id}_performance_limit50_global_fix",
                mins, goals, assists, shots, shots_ot,
                fouls_c, fouls_w, yc, rc, saves,
                json.dumps(ev), utc_now(),
            ))
            inserted += 1
        except Exception:
            pass
    con.commit()

    min15 = con.execute(
        "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=? AND minutes_played>=15",
        (player_id,)
    ).fetchone()[0]
    return {"player_id": player_id, "player_name": player_name,
            "events": inserted, "min15": min15, "status": meta["status"]}


# ── Main per-team processing ──────────────────────────────────────────────────

def process_team(con: Any, team: dict, execute: bool) -> dict:
    team_name = team["db_name"]
    team_id   = team["team_id"]

    print(f"\n{'='*60}")
    print(f"  {team_name}  (id={team_id})")
    print(f"{'='*60}")

    # -- Load roster from DB
    roster = [dict(r) for r in con.execute("""
        SELECT id as row_id, player_name, player_id, statshub_player_id_status,
               jersey_number, position
        FROM statshub_team_players WHERE team_name=?
        ORDER BY CAST(jersey_number AS INTEGER)
    """, (team_name,)).fetchall()]

    before_confirmed = sum(1 for r in roster if r["statshub_player_id_status"] == "confirmed")
    print(f"  Roster in DB: {len(roster)}, confirmed before: {before_confirmed}")

    # -- Collect all discovered StatsHub players
    all_ep: list[dict] = []
    ep_log: list[dict] = []

    def _add(payload: Any, label: str, url: str, status: str) -> None:
        players = parse_team_players(payload, label)
        ep_log.append({"label": label, "url": url, "status": status, "players": len(players)})
        print(f"    [{label}] {status} → {len(players)} players")
        all_ep.extend(players)

    # Source 1: cached tournamentId=16 file
    cached_file = COVERAGE_SNAPSHOT / f"team_{team_id}_players_wc26_tournamentId16.json"
    if cached_file.exists() and cached_file.stat().st_size > 100:
        try:
            payload1 = json.loads(cached_file.read_text(encoding="utf-8"))
            _add(payload1, f"team_{team_id}_wc26_cached",
                 f"{BASE}/api/team/{team_id}/players/performance?tournamentId=16&location=both",
                 "local_file")
        except Exception as e:
            ep_log.append({"label": f"team_{team_id}_wc26_cached", "url": "", "status": f"parse_error:{e}", "players": 0})
    else:
        ep_log.append({"label": f"team_{team_id}_wc26_cached", "url": "", "status": "file_empty_or_missing", "players": 0})

    # Deduplicate before deciding whether to call no-filter
    def _dedup(items: list[dict]) -> list[dict]:
        seen: set[str] = set()
        out: list[dict] = []
        for c in items:
            if c["player_id"] not in seen:
                seen.add(c["player_id"])
                out.append(c)
        return out

    deduped_so_far = _dedup(all_ep)
    # Match to estimate confirmed count
    by_slug_tmp, by_norm_tmp = build_lookups(deduped_so_far)
    confirmed_ids_tmp: set[str] = {r["player_id"] for r in roster if r.get("player_id") and r.get("statshub_player_id_status") == "confirmed"}
    matches_tmp = match_roster_to_statshub(roster, deduped_so_far, by_slug_tmp, by_norm_tmp, team_name, set(confirmed_ids_tmp))
    would_be_confirmed = sum(1 for m in matches_tmp if m["status"] in ("matched", "existing_confirmed"))

    # Source 2: no-filter endpoint (if not enough confirmed)
    if would_be_confirmed < MIN_CONFIRMED_THRESHOLD or team_name in ("Haiti", "Australia", "Qatar"):
        url_nf  = f"{BASE}/api/team/{team_id}/players/performance"
        label_nf = f"team_{team_id}_perf_nofilt"
        p_nf, meta_nf = fetch(url_nf, label_nf, execute=execute)
        _add(p_nf, label_nf, url_nf, meta_nf["status"])

    # Final dedup
    ep_players = _dedup(all_ep)
    by_slug, by_norm = build_lookups(ep_players)
    print(f"  Total unique StatsHub players: {len(ep_players)}")

    # -- Match FIFA roster → StatsHub IDs
    confirmed_ids: set[str] = {r["player_id"] for r in roster if r.get("player_id") and r.get("statshub_player_id_status") == "confirmed"}
    matches = match_roster_to_statshub(roster, ep_players, by_slug, by_norm, team_name, set(confirmed_ids))

    matched_count    = sum(1 for m in matches if m["status"] == "matched")
    existing_count   = sum(1 for m in matches if m["status"] == "existing_confirmed")
    unresolved_count = sum(1 for m in matches if m["status"] == "unresolved")
    ambiguous_count  = sum(1 for m in matches if m["status"] == "ambiguous")

    print(f"  Match results: {matched_count} new + {existing_count} existing + "
          f"{unresolved_count} unresolved + {ambiguous_count} ambiguous = {len(matches)}")
    for m in matches:
        if m["status"] == "matched":
            print(f"    [NEW] #{m['jersey']:>2} {m['fifa_name'][:38]:38s} → {m['matched_statshub_name']} (id={m['player_id']}) [{m['match_method'][:40]}]")
        elif m["status"] == "unresolved":
            print(f"    [???] #{m['jersey']:>2} {m['fifa_name']}")
        elif m["status"] == "ambiguous":
            print(f"    [AMB] #{m['jersey']:>2} {m['fifa_name']} → {m['matched_statshub_name']} (id={m['player_id']})")

    # -- Commit to DB
    newly_confirmed = 0
    if execute or True:  # always commit in dry-run too for analysis (dry run = no API calls, but DB writes OK)
        for m in matches:
            if m["status"] == "matched" and m["player_id"]:
                try:
                    con.execute("""
                        UPDATE statshub_team_players
                        SET player_id=?, statshub_player_id_status='confirmed',
                            player_id_match_method=?, player_id_confidence_score=?,
                            player_id_match_source=?, updated_at=?
                        WHERE id=?
                    """, (m["player_id"], m["match_method"], m["confidence"] * 100,
                          SCRIPT_SOURCE, utc_now(), m["row_id"]))
                    newly_confirmed += 1
                except Exception as e:
                    print(f"    [WARN] Could not update row_id={m['row_id']}: {e}")
        con.commit()
    print(f"  Newly confirmed: {newly_confirmed}")

    # -- Download performance for confirmed players
    perf_results: list[dict] = []
    all_player_ids_to_download = {
        m["player_id"] for m in matches
        if m["status"] in ("matched", "existing_confirmed") and m["player_id"]
    }
    print(f"  Downloading performance for {len(all_player_ids_to_download)} confirmed players...")

    # Build name map from roster
    id_to_name = {}
    for m in matches:
        if m["player_id"]:
            id_to_name[m["player_id"]] = m.get("matched_statshub_name") or m["fifa_name"]

    for pid in sorted(all_player_ids_to_download):
        pname = id_to_name.get(pid, pid)
        res = download_player_perf(con, pid, pname, team_name, execute=execute)
        perf_results.append(res)
        if res["status"] != "db_cached":
            print(f"    {pname} (id={pid}): events={res['events']} min15={res['min15']} [{res['status']}]")

    total_events = sum(r["events"] for r in perf_results)
    total_min15  = sum(r["min15"] for r in perf_results)

    after_confirmed = sum(1 for m in matches if m["status"] in ("matched", "existing_confirmed"))

    return {
        "team_name": team_name,
        "team_id": team_id,
        "roster_size": len(roster),
        "before_confirmed": before_confirmed,
        "after_confirmed": after_confirmed,
        "newly_confirmed": newly_confirmed,
        "unresolved": unresolved_count,
        "ambiguous": ambiguous_count,
        "ep_players_discovered": len(ep_players),
        "ep_log": ep_log,
        "matches": matches,
        "perf_results": perf_results,
        "total_events": total_events,
        "total_min15": total_min15,
        "coverage_pct": round(after_confirmed / max(len(roster), 1) * 100, 1),
    }


# ── EV rebuild ────────────────────────────────────────────────────────────────

def rebuild_ev(execute: bool) -> dict:
    print("\n=== REBUILDING EV (from raw odds, no Odds-API calls) ===")
    import subprocess
    cmd = [sys.executable, "-m", "scripts.fetch_today_4_matches_live_api_odds", "--from-raw"]
    result = subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True, encoding="utf-8", errors="replace")
    output = result.stdout + result.stderr
    # Parse key stats from output
    stats = {}
    for line in output.splitlines():
        for key in ("Total EV rows", "VALUE rows", "Actionable VALUE rows",
                    "Player prop rows", "Unsupported", "Unmatched"):
            if key in line:
                parts = line.split(":")
                if len(parts) >= 2:
                    try:
                        stats[key.strip()] = int(parts[-1].strip())
                    except ValueError:
                        pass
    print(output[-3000:] if len(output) > 3000 else output)
    return stats


# ── Write output workbooks ────────────────────────────────────────────────────

def write_coverage_workbook(team_results: list[dict]) -> None:
    import warnings
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    warnings.filterwarnings("ignore", "Title is more than")

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
    YEL_FILL  = PatternFill("solid", fgColor="FFEB9C")
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")

    def hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: summary
    ws1 = wb.create_sheet("team_coverage_summary")
    hdr(ws1, ["team_name", "team_id", "roster", "before_confirmed", "after_confirmed",
               "newly_confirmed", "unresolved", "ambiguous", "coverage_pct",
               "total_events", "total_min15", "ep_players_found"])
    for t in team_results:
        row = [t["team_name"], t["team_id"], t["roster_size"],
               t["before_confirmed"], t["after_confirmed"],
               t["newly_confirmed"], t["unresolved"], t["ambiguous"],
               t["coverage_pct"], t["total_events"], t["total_min15"],
               t["ep_players_discovered"]]
        ws1.append(row)
        fill = GRN_FILL if t["after_confirmed"] >= 22 else (YEL_FILL if t["after_confirmed"] >= 10 else RED_FILL)
        for c in ws1[ws1.max_row]:
            c.fill = fill
    autofit(ws1)

    # Sheet 2: player matching detail
    ws2 = wb.create_sheet("player_matching_detail")
    hdr(ws2, ["team_name", "jersey", "fifa_name", "matched_statshub_name",
               "player_id", "match_method", "confidence", "status"])
    for t in team_results:
        for m in t["matches"]:
            ws2.append([t["team_name"], m["jersey"], m["fifa_name"],
                         m["matched_statshub_name"], m["player_id"],
                         m["match_method"], m["confidence"], m["status"]])
            if m["status"] == "unresolved":
                for c in ws2[ws2.max_row]: c.fill = RED_FILL
            elif m["status"] == "ambiguous":
                for c in ws2[ws2.max_row]: c.fill = YEL_FILL
    autofit(ws2)

    # Sheet 3: performance downloads
    ws3 = wb.create_sheet("performance_downloads")
    hdr(ws3, ["team_name", "player_id", "player_name", "events", "min15", "status"])
    for t in team_results:
        for p in t["perf_results"]:
            ws3.append([t["team_name"], p["player_id"], p["player_name"],
                         p["events"], p["min15"], p["status"]])
            fill = GRN_FILL if p["min15"] >= 10 else (YEL_FILL if p["min15"] >= 5 else RED_FILL)
            for c in ws3[ws3.max_row]: c.fill = fill
    autofit(ws3)

    # Sheet 4: endpoint log
    ws4 = wb.create_sheet("endpoint_log")
    hdr(ws4, ["team_name", "label", "url", "status", "players_extracted"])
    for t in team_results:
        for ep in t["ep_log"]:
            ws4.append([t["team_name"], ep["label"], ep["url"], ep["status"], ep["players"]])
    autofit(ws4)

    # Sheet 5: unresolved players
    ws5 = wb.create_sheet("unresolved_players")
    hdr(ws5, ["team_name", "jersey", "fifa_name", "notes"])
    for t in team_results:
        for m in t["matches"]:
            if m["status"] in ("unresolved", "ambiguous"):
                ws5.append([t["team_name"], m["jersey"], m["fifa_name"],
                             m["status"] + ("" if m["status"]=="unresolved" else f" id={m['player_id']}")])
                for c in ws5[ws5.max_row]: c.fill = YEL_FILL if m["status"]=="ambiguous" else RED_FILL
    autofit(ws5)

    wb.save(OUT_COVERAGE)
    print(f"\nCoverage workbook: {OUT_COVERAGE}")


def write_ev_workbook(con: Any) -> None:
    import warnings
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    warnings.filterwarnings("ignore", "Title is more than")

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")

    def hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    MATCH_TUPLE = tuple(MATCHES)
    PH = ",".join("?" * len(MATCHES))

    ACTIONABLE_WHERE = (
        "verdict='VALUE' AND expected_value>0 "
        "AND model_probability>=0.25 AND sample_size>=10 "
        "AND priority_class IN ('hard_data_priority','medium_priority') "
        "AND minutes_filter_status IN ('ok','not_applicable','fallback_raw_json')"
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Top 200 actionable VALUE bets
    ws1 = wb.create_sheet("top_actionable_bets")
    hdr(ws1, ["rank", "match_name", "market_scope", "priority_class", "bet_description",
               "team_name", "player_name", "player_id", "verdict", "expected_value",
               "model_probability", "sample_size", "line", "odds_decimal",
               "minutes_filter_status"])
    rows = con.execute(f"""
        SELECT ROW_NUMBER() OVER (ORDER BY expected_value DESC),
               match_name, market_scope, priority_class, bet_description,
               team_name, player_name, player_id, verdict, expected_value,
               model_probability, sample_size, line, odds_decimal, minutes_filter_status
        FROM betting_value_scores_new
        WHERE {ACTIONABLE_WHERE} AND match_name IN ({PH})
        ORDER BY expected_value DESC LIMIT 200
    """, MATCH_TUPLE).fetchall()
    for r in rows:
        ws1.append(list(r))
    autofit(ws1)

    # Per-match summary
    ws2 = wb.create_sheet("per_match_summary")
    hdr(ws2, ["match_name", "total_rows", "VALUE_rows", "actionable_VALUE",
               "unmatched", "unsupported", "player_mp_ok"])
    for m in MATCHES:
        r = con.execute(f"""
            SELECT COUNT(*),
                SUM(CASE WHEN verdict='VALUE' THEN 1 ELSE 0 END),
                SUM(CASE WHEN {ACTIONABLE_WHERE} THEN 1 ELSE 0 END),
                SUM(CASE WHEN verdict='UNMATCHED' THEN 1 ELSE 0 END),
                SUM(CASE WHEN verdict='UNSUPPORTED' THEN 1 ELSE 0 END),
                SUM(CASE WHEN minutes_filter_status='ok' THEN 1 ELSE 0 END)
            FROM betting_value_scores_new WHERE match_name=?
        """, (m,)).fetchone()
        ws2.append([m] + list(r))
    autofit(ws2)

    wb.save(OUT_EV)
    print(f"EV workbook: {OUT_EV}")


def write_audit_workbook(con: Any, team_results: list[dict], ev_stats: dict) -> None:
    import warnings
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    warnings.filterwarnings("ignore", "Title is more than")

    from app.db.queries import utc_now as _utc_now

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
    YEL_FILL  = PatternFill("solid", fgColor="FFEB9C")
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")

    def hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    MATCH_TUPLE = tuple(MATCHES)
    PH = ",".join("?" * len(MATCHES))
    ACTIONABLE_WHERE = (
        "verdict='VALUE' AND expected_value>0 "
        "AND model_probability>=0.25 AND sample_size>=10 "
        "AND priority_class IN ('hard_data_priority','medium_priority') "
        "AND minutes_filter_status IN ('ok','not_applicable','fallback_raw_json')"
    )

    def q1(sql, args=()):
        r = con.execute(sql, args).fetchone()
        return r[0] if r else None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Overall summary
    ws1 = wb.create_sheet("overall_ev_summary")
    hdr(ws1, ["metric", "value", "notes"])
    total      = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH})", MATCH_TUPLE)
    value      = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND verdict='VALUE'", MATCH_TUPLE)
    actable    = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND {ACTIONABLE_WHERE}", MATCH_TUPLE)
    unmatched  = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND verdict='UNMATCHED'", MATCH_TUPLE)
    high_ev    = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND expected_value>5", MATCH_TUPLE)
    small_ss   = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND sample_size<10 AND expected_value IS NOT NULL", MATCH_TUPLE)
    mp_ok      = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND minutes_filter_status='ok'", MATCH_TUPLE)
    mp_none    = q1(f"SELECT COUNT(*) FROM betting_value_scores_new WHERE match_name IN ({PH}) AND minutes_filter_status='no_valid_appearances'", MATCH_TUPLE)

    complete_matches = sum(1 for t in team_results
                           if {r["status"] for r in [next((x for x in team_results if x["team_name"]==t["team_name"]), {})]}
                           ) if False else sum(1 for m in MATCHES if all(
                               any(tr["team_name"] in m and tr["after_confirmed"] >= 10 for tr in team_results)
                               for _ in [None]
                           ))

    for row in [
        ("generated_at", _utc_now(), ""),
        ("match_date", "2026-06-13", ""),
        ("total_ev_rows", total, "All rows incl. UNMATCHED/UNSUPPORTED"),
        ("VALUE_rows", value, "Model assigns EV > 0"),
        ("actionable_VALUE_rows", actable, "verdict=VALUE, ev>0, p>=0.25, ss>=10, priority=hard/medium, mp_ok"),
        ("UNMATCHED_rows", unmatched, "Player name not matched to confirmed StatsHub player_id"),
        ("player_props_minutes_ok", mp_ok, "minutes_filter_status=ok"),
        ("player_props_no_valid_appearances", mp_none, "0 appearances with minutesPlayed>=15"),
        ("ev_outliers_above_5", high_ev, "All-competition sample may inflate p vs tournament odds"),
        ("small_sample_below_10", small_ss, "EV from fewer than 10 qualifying appearances"),
    ]:
        ws1.append(list(row))
    autofit(ws1)

    # Team coverage
    ws2 = wb.create_sheet("team_coverage")
    hdr(ws2, ["team_name", "roster", "confirmed_before", "confirmed_after",
               "newly_confirmed", "unresolved", "total_events", "min15_events", "coverage_pct"])
    for t in team_results:
        ws2.append([t["team_name"], t["roster_size"], t["before_confirmed"],
                     t["after_confirmed"], t["newly_confirmed"], t["unresolved"],
                     t["total_events"], t["total_min15"], t["coverage_pct"]])
        fill = GRN_FILL if t["coverage_pct"] >= 85 else (YEL_FILL if t["coverage_pct"] >= 50 else RED_FILL)
        for c in ws2[ws2.max_row]: c.fill = fill
    autofit(ws2)

    # Top strict actionable bets
    ws3 = wb.create_sheet("top20_strict_actionable")
    hdr(ws3, ["rank", "match_name", "bet_description", "team_name", "player_name",
               "expected_value", "model_probability", "odds_decimal",
               "sample_size", "line", "minutes_filter_status"])
    rows = con.execute(f"""
        SELECT ROW_NUMBER() OVER (ORDER BY expected_value DESC),
               match_name, bet_description, team_name, player_name,
               expected_value, model_probability, odds_decimal,
               sample_size, line, minutes_filter_status
        FROM betting_value_scores_new
        WHERE {ACTIONABLE_WHERE} AND match_name IN ({PH})
        ORDER BY expected_value DESC LIMIT 20
    """, MATCH_TUPLE).fetchall()
    for r in rows:
        ws3.append(list(r))
    autofit(ws3)

    # Unmatched players post-fix
    ws4 = wb.create_sheet("unmatched_player_props")
    hdr(ws4, ["match_name", "team_name", "player_name", "bet_description",
               "market_type", "odds_decimal"])
    rows = con.execute(f"""
        SELECT match_name, team_name, player_name, bet_description, market_type, odds_decimal
        FROM betting_value_scores_new
        WHERE verdict='UNMATCHED' AND market_scope='player'
          AND match_name IN ({PH})
        GROUP BY match_name, team_name, player_name, bet_description
        ORDER BY match_name, player_name
    """, MATCH_TUPLE).fetchall()
    for r in rows:
        ws4.append(list(r))
        for c in ws4[ws4.max_row]: c.fill = YEL_FILL
    autofit(ws4)

    wb.save(OUT_AUDIT)
    print(f"Audit workbook: {OUT_AUDIT}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true",
                        help="Call live StatsHub endpoints (requires STATSHUB_ENABLED=true)")
    args = parser.parse_args()

    execute = args.execute
    if execute and os.environ.get("STATSHUB_ENABLED", "").lower() != "true":
        print("ERROR: --execute requires STATSHUB_ENABLED=true")
        sys.exit(1)

    print(f"Mode: {'LIVE' if execute else 'DRY RUN (local files + DB writes)'}")

    team_results: list[dict] = []

    with connect() as con:
        for team in TEAMS:
            result = process_team(con, team, execute=execute)
            team_results.append(result)

    # Print summary before EV rebuild
    print("\n" + "="*70)
    print("COVERAGE SUMMARY AFTER PARSER FIX")
    print("="*70)
    print(f"{'team':<15} {'before':>6} {'after':>6} {'new':>5} {'unres':>6} {'min15':>6} {'pct':>6}")
    for t in team_results:
        print(f"{t['team_name']:<15} {t['before_confirmed']:>6} {t['after_confirmed']:>6} "
              f"{t['newly_confirmed']:>5} {t['unresolved']:>6} {t['total_min15']:>6} {t['coverage_pct']:>5}%")

    # Rebuild EV
    ev_stats = rebuild_ev(execute=execute)

    # Write workbooks
    write_coverage_workbook(team_results)

    with connect() as con:
        write_ev_workbook(con)
        write_audit_workbook(con, team_results, ev_stats)

    # Final report
    print("\n" + "="*70)
    print("FINAL REPORT")
    print("="*70)
    for t in team_results:
        status = "COMPLETE" if t["after_confirmed"] >= 22 else \
                 ("GOOD" if t["after_confirmed"] >= 15 else \
                 ("PARTIAL" if t["after_confirmed"] >= 5 else "POOR"))
        print(f"  {t['team_name']:<15} team_id={t['team_id']}  "
              f"confirmed {t['before_confirmed']}→{t['after_confirmed']}/{t['roster_size']}  "
              f"events={t['total_events']}  min15={t['total_min15']}  "
              f"unresolved={t['unresolved']}  [{status}]")

    print("\nOutput files:")
    print(f"  {OUT_COVERAGE}")
    print(f"  {OUT_EV}")
    print(f"  {OUT_AUDIT}")
    print("\nRun dashboard:")
    print("  streamlit run app/dashboard/betting_value_dashboard.py")


if __name__ == "__main__":
    main()
