"""
Fetch live/API odds for today's 4 World Cup matches from Odds-API.io.

Run name: today_4_matches_live_api_odds_probe

Matches:
  1. Qatar vs Switzerland          (candidate event_id: 66456918)
  2. Brazil vs Morocco             (candidate event_id: 66456928)
  3. Haiti vs Scotland             (candidate event_id: 66456930)
  4. Australia vs Turkey           (candidate event_id: 66456942)

Usage:
    python -m scripts.fetch_today_4_matches_live_api_odds
    python -m scripts.fetch_today_4_matches_live_api_odds --from-raw
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.odds_driven import (
    MARKET_SCOPE_MAP,
    PRIORITY_CLASS_MAP,
    TARGET_MATCHES,
    calculate_ev,
    canonical,
    connect,
    ensure_schema,
    insert_raw_rows,
    normalize_raw_odds,
)
from scripts.probe_odds_api_io import is_configured_api_key, load_config

# ---------------------------------------------------------------------------
# Run configuration
# ---------------------------------------------------------------------------

RUN_NAME = "today_4_matches_live_api_odds_probe"
SPORT = "football"
PREFERRED_BOOKMAKER = "Bet365"
FALLBACK_BOOKMAKER = "DraftKings"
TIMEOUT_SECONDS = 20

RAW_BASE_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/today_4_matches_live_api_odds_value_scores.xlsx")
OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

# Match definitions — candidate event_ids are verified against the API
MATCH_CONFIGS: list[dict[str, Any]] = [
    {
        "match_name": "Qatar vs Switzerland",
        "home_team": "Qatar",
        "away_team": "Switzerland",
        "candidate_event_id": "66456918",
        "home_aliases": ("qatar", "qat"),
        "away_aliases": ("switzerland", "sui", "suisse", "schweiz"),
        "slug": "qatar_vs_switzerland",
    },
    {
        "match_name": "Brazil vs Morocco",
        "home_team": "Brazil",
        "away_team": "Morocco",
        "candidate_event_id": "66456928",
        "home_aliases": ("brazil", "brasil", "bra"),
        "away_aliases": ("morocco", "mar", "maroc"),
        "slug": "brazil_vs_morocco",
    },
    {
        "match_name": "Haiti vs Scotland",
        "home_team": "Haiti",
        "away_team": "Scotland",
        "candidate_event_id": "66456930",
        "home_aliases": ("haiti", "haïti", "hai"),
        "away_aliases": ("scotland", "sco"),
        "slug": "haiti_vs_scotland",
    },
    {
        "match_name": "Australia vs Turkey",
        "home_team": "Australia",
        "away_team": "Turkey",
        "candidate_event_id": "66456942",
        "home_aliases": ("australia", "aus", "socceroos"),
        "away_aliases": ("turkey", "tur", "turkiye", "türkiye"),
        "slug": "australia_vs_turkey",
    },
]


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def sqlite_safe(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _str_field(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    if isinstance(v, dict):
        return v.get("name") or v.get("slug") or json.dumps(v)
    return str(v)


def text_has(text: Any, aliases: tuple[str, ...]) -> bool:
    value = str(text or "").lower()
    return any(alias in value for alias in aliases)


def payload_count(payload: Any) -> int:
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("data", "events", "response", "results", "bookmakers", "leagues"):
            if isinstance(payload.get(key), list):
                return len(payload[key])
    return 1 if payload else 0


def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    safe = {k: v for k, v in params.items() if k.lower() != "apikey"}
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    return f"{url}?{urlencode(safe)}" if safe else url


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def _save_response(
    slug: str, action: str, params: dict[str, Any],
    status_code: int | None, payload: Any, error: str | None, base_url: str,
) -> Path:
    raw_dir = RAW_BASE_DIR / slug
    raw_dir.mkdir(parents=True, exist_ok=True)
    safe_action = action.replace("/", "_")
    path = raw_dir / f"{safe_action}_{stamp()}.json"
    wrapper = {
        "fetched_at_utc": now_utc(),
        "run_name": RUN_NAME,
        "action": action,
        "slug": slug,
        "url_without_api_key": public_url(base_url, action, params),
        "params_without_api_key": {k: v for k, v in params.items() if k.lower() != "apikey"},
        "status_code": status_code,
        "response_json": payload,
        "error": error,
    }
    path.write_text(json.dumps(wrapper, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def api_get(
    endpoint: str, params: dict[str, Any], base_url: str, slug: str = "shared",
) -> dict[str, Any]:
    status_code = None
    payload: Any = None
    error = None
    try:
        response = requests.get(
            f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}",
            params=params,
            timeout=TIMEOUT_SECONDS,
        )
        status_code = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"response_text_preview": response.text[:1000]}
        if response.status_code >= 400:
            error = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        error = str(exc)
    raw_file = _save_response(slug, endpoint, params, status_code, payload, error, base_url)
    return {
        "endpoint": endpoint, "params": params,
        "status_code": status_code, "payload": payload,
        "error": error, "raw_file": raw_file,
    }


# ---------------------------------------------------------------------------
# Event matching
# ---------------------------------------------------------------------------

def find_match_in_events(
    events: Any, cfg: dict[str, Any],
) -> tuple[dict[str, Any] | None, str, float]:
    """
    Search events list for the target match.
    Returns (event_dict, api_event_id, confidence).
    confidence: 1.0=team-name match, 0.5=candidate-id-only, 0.0=not found.
    """
    items = (
        events if isinstance(events, list)
        else events.get("data", []) if isinstance(events, dict)
        else []
    )
    home_aliases = cfg["home_aliases"]
    away_aliases = cfg["away_aliases"]
    candidate_id = cfg["candidate_event_id"]

    team_match: dict[str, Any] | None = None
    id_match: dict[str, Any] | None = None

    for event in items:
        home = event.get("home") or event.get("homeTeam") or event.get("home_team")
        away = event.get("away") or event.get("awayTeam") or event.get("away_team")
        eid = str(event.get("id") or event.get("eventId") or event.get("event_id") or "")

        name_ok = (
            text_has(home, home_aliases) and text_has(away, away_aliases)
        ) or (
            text_has(home, away_aliases) and text_has(away, home_aliases)
        )
        if name_ok and team_match is None:
            team_match = event
        if eid == candidate_id and id_match is None:
            id_match = event

    if team_match is not None:
        eid = str(
            team_match.get("id") or team_match.get("eventId")
            or team_match.get("event_id") or candidate_id
        )
        return team_match, eid, 1.0

    if id_match is not None:
        eid = str(
            id_match.get("id") or id_match.get("eventId")
            or id_match.get("event_id") or candidate_id
        )
        return id_match, eid, 0.5

    return None, candidate_id, 0.0


# ---------------------------------------------------------------------------
# Odds parsing — match-context-aware
# ---------------------------------------------------------------------------

def _price_rows_from_outcome(
    market_name: str, outcome: dict[str, Any], home: str, away: str
) -> list[tuple[str, float | None, float]]:
    rows: list[tuple[str, float | None, float]] = []
    line = (
        outcome.get("line") or outcome.get("point")
        or outcome.get("handicap") or outcome.get("hdp")
    )
    for key, label in (("home", home), ("away", away), ("draw", "Draw")):
        if outcome.get(key) not in (None, ""):
            try:
                rows.append((label, line, float(outcome[key])))
            except (TypeError, ValueError):
                pass
    sel_label = outcome.get("label") or outcome.get("name") or outcome.get("selection")
    for side_key, side_str in (("over", "Over"), ("under", "Under")):
        if outcome.get(side_key) not in (None, ""):
            selection = (
                sel_label if sel_label and "chance" in market_name.lower()
                else f"{side_str} {line}".strip()
            )
            try:
                rows.append((selection, line, float(outcome[side_key])))
            except (TypeError, ValueError):
                pass
    if not rows:
        for key in ("price", "odds", "decimal"):
            if outcome.get(key) not in (None, ""):
                try:
                    rows.append((
                        str(sel_label or outcome.get("team") or outcome.get("player") or market_name),
                        line, float(outcome[key]),
                    ))
                    break
                except (TypeError, ValueError):
                    pass
    return rows


def _markets_for_bookmaker(payload: dict[str, Any], bookmaker: str) -> list[Any]:
    bookmakers = payload.get("bookmakers", {})
    if isinstance(bookmakers, dict):
        return bookmakers.get(bookmaker, [])
    if isinstance(bookmakers, list):
        for book in bookmakers:
            if isinstance(book, dict):
                name = str(book.get("name") or book.get("key") or "")
                if name.lower() == bookmaker.lower():
                    return book.get("markets", book.get("odds", []))
    return []


def raw_rows_from_odds_payload(
    payload: dict[str, Any],
    cfg: dict[str, Any],
    bookmaker: str,
    api_event_id: str,
    statshub_event_id: str | None,
    raw_file: Path,
    request_id: str,
) -> list[dict[str, Any]]:
    home = str(payload.get("home") or cfg["home_team"])
    away = str(payload.get("away") or cfg["away_team"])
    match_name = cfg["match_name"]
    captured_at = now_utc()
    markets = _markets_for_bookmaker(payload, bookmaker)
    rows: list[dict[str, Any]] = []
    for mi, market in enumerate(markets):
        if not isinstance(market, dict):
            continue
        market_name = market.get("name") or market.get("key") or f"market_{mi}"
        outcomes = (
            market.get("odds") if isinstance(market.get("odds"), list)
            else market.get("outcomes", [])
        )
        for oi, outcome in enumerate(outcomes or []):
            if not isinstance(outcome, dict):
                continue
            for selection, line, odds_val in _price_rows_from_outcome(
                str(market_name), outcome, home, away
            ):
                rows.append({
                    "run_name": RUN_NAME,
                    "source_name": "Odds-API.io",
                    "bookmaker": bookmaker,
                    "match_name": match_name,
                    "event_id": api_event_id,
                    "api_event_id": api_event_id,
                    "statshub_event_id": statshub_event_id,
                    "raw_market_group": str(market_name),
                    "raw_market_name": str(market_name),
                    "raw_selection_name": str(selection),
                    "raw_line": line,
                    "raw_odds": odds_val,
                    "odds_format": "decimal",
                    "captured_at": captured_at,
                    "raw_payload": json.dumps({
                        "market_index": mi, "odds_index": oi,
                        "market": market, "outcome": outcome,
                    }, ensure_ascii=False),
                    "source_url": str(payload.get("urls", {}).get(bookmaker, "")),
                    "request_id": request_id,
                    "raw_file": str(raw_file),
                    "status": "raw",
                    "notes": RUN_NAME,
                })
    return rows


def extract_market_groups(
    payload: dict[str, Any], cfg: dict[str, Any], bookmaker: str,
) -> list[dict[str, Any]]:
    home = str(payload.get("home") or cfg["home_team"])
    away = str(payload.get("away") or cfg["away_team"])
    markets = _markets_for_bookmaker(payload, bookmaker)
    groups = []
    for market in markets:
        if not isinstance(market, dict):
            continue
        name = market.get("name") or market.get("key") or "unknown"
        odds_list = (
            market.get("odds") if isinstance(market.get("odds"), list)
            else market.get("outcomes", [])
        )
        row_count = sum(
            len(_price_rows_from_outcome(str(name), o, home, away))
            for o in (odds_list or []) if isinstance(o, dict)
        )
        groups.append({
            "match_name": cfg["match_name"],
            "market_key": str(name),
            "market_name": str(name),
            "rows_returned": row_count,
        })
    return groups


# ---------------------------------------------------------------------------
# StatsHub ID resolution from DB
# ---------------------------------------------------------------------------

def resolve_statshub_ids(con: sqlite3.Connection, cfg: dict[str, Any]) -> dict[str, Any]:
    """Attempt to resolve StatsHub team/player IDs from local DB."""
    home = cfg["home_team"]
    away = cfg["away_team"]

    def team_id(team: str) -> str | None:
        row = con.execute(
            "SELECT statshub_team_id FROM statshub_world_cup_teams WHERE team_name=? LIMIT 1",
            (team,),
        ).fetchone()
        return str(row[0]) if row and row[0] is not None else None

    def has_team_perf(team: str) -> int:
        try:
            return con.execute(
                "SELECT COUNT(*) FROM statshub_team_performance_events WHERE team_name=? LIMIT 1",
                (team,),
            ).fetchone()[0]
        except Exception:
            return 0

    def has_players(team: str) -> int:
        try:
            return con.execute(
                "SELECT COUNT(*) FROM statshub_team_players WHERE team_name=? LIMIT 1",
                (team,),
            ).fetchone()[0]
        except Exception:
            return 0

    home_id = team_id(home)
    away_id = team_id(away)

    # Try to find a StatsHub event ID from any fixture table
    statshub_event_id: str | None = None
    for table in ("statshub_raw_events", "statshub_worldcup_events", "fixtures"):
        try:
            cols = {r[1] for r in con.execute(f"PRAGMA table_info({table})").fetchall()}
            if "event_id" in cols:
                # Search by team name if possible
                if "home_team" in cols and "away_team" in cols:
                    row = con.execute(
                        f"SELECT event_id FROM {table} WHERE home_team=? AND away_team=? LIMIT 1",
                        (home, away),
                    ).fetchone()
                    if row:
                        statshub_event_id = str(row[0])
                        break
        except Exception:
            continue

    return {
        "home_team_id": home_id,
        "away_team_id": away_id,
        "statshub_event_id": statshub_event_id,
        "home_team_perf_rows": has_team_perf(home),
        "away_team_perf_rows": has_team_perf(away),
        "home_player_rows": has_players(home),
        "away_player_rows": has_players(away),
    }


# ---------------------------------------------------------------------------
# Bookmaker selection
# ---------------------------------------------------------------------------

def get_selected_bookmakers(api_key: str, base_url: str) -> dict[str, Any]:
    result = api_get("bookmakers/selected", {"apiKey": api_key}, base_url, slug="shared")
    books: list[str] = []
    payload = result["payload"]
    if isinstance(payload, list):
        books = [
            str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
            for b in payload
        ]
    elif isinstance(payload, dict):
        items = payload.get("data") or payload.get("bookmakers") or []
        if isinstance(items, list):
            books = [
                str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
                for b in items
            ]
    return {
        "selected_bookmakers": books,
        "status_code": result["status_code"],
        "error": result["error"],
        "raw_file": result["raw_file"],
    }


def pick_bookmaker(selected: list[str]) -> str:
    lowers = [b.lower() for b in selected]
    if any("bet365" in b for b in lowers):
        return PREFERRED_BOOKMAKER
    if any("draftkings" in b for b in lowers):
        return FALLBACK_BOOKMAKER
    return PREFERRED_BOOKMAKER


# ---------------------------------------------------------------------------
# Workbook writer — 12 sheets
# ---------------------------------------------------------------------------

def _add_sheet(
    wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any],
) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([row[h] for h in headers])
        elif isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(
            12, min(50, len(str(header)) + 4)
        )


def _table_rows(
    con: sqlite3.Connection, sql: str, params: tuple = (),
) -> tuple[list[str], list[sqlite3.Row]]:
    cursor = con.execute(sql, params)
    rows = cursor.fetchall()
    headers = list(rows[0].keys()) if rows else [d[0] for d in cursor.description]
    return headers, rows


def write_workbook(
    con: sqlite3.Connection,
    run_summary_rows: list[dict[str, Any]],
    match_resolution: list[dict[str, Any]],
    market_coverage: list[dict[str, Any]],
    raw_sources: list[dict[str, Any]],
) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. run_summary
    _add_sheet(wb, "run_summary", ["metric", "value"],
               [{"metric": k, "value": str(v)} for row in run_summary_rows for k, v in row.items()])

    # 2. match_resolution
    mr_headers = [
        "match_name", "odds_api_event_id", "statshub_event_id",
        "home_team", "away_team", "home_team_id", "away_team_id",
        "confidence", "status", "home_team_perf_rows", "away_team_perf_rows",
        "home_player_rows", "away_player_rows", "odds_found", "raw_rows",
    ]
    _add_sheet(wb, "match_resolution", mr_headers, match_resolution)

    # 3. market_coverage_by_match
    mc_headers = ["match_name", "market_key", "market_name", "rows_returned", "status"]
    _add_sheet(wb, "market_coverage_by_match", mc_headers, market_coverage)

    # 4. ev_ranking_all_matches
    h, rows = _table_rows(con,
        """SELECT rank, match_name, bookmaker, market_type, raw_market_name, raw_selection_name,
                  market_scope, priority_class, bet_description,
                  team_name, player_name, side, line, odds_decimal,
                  implied_probability, model_probability, edge, expected_value,
                  probability_method, sample_size, valid_appearance_count,
                  minutes_filter_status, data_quality_status, verdict, notes
           FROM betting_value_scores_new
           WHERE expected_value IS NOT NULL
           ORDER BY expected_value DESC"""
    )
    _add_sheet(wb, "ev_ranking_all_matches", h, rows)

    # 5. ev_ranking_actionable
    h, rows = _table_rows(con,
        """SELECT rank, match_name, bookmaker, market_type, raw_market_name, raw_selection_name,
                  market_scope, priority_class, bet_description,
                  team_name, player_name, side, line, odds_decimal,
                  implied_probability, model_probability, edge, expected_value,
                  probability_method, sample_size, valid_appearance_count,
                  minutes_filter_status, verdict, notes
           FROM betting_value_scores_new
           WHERE verdict='VALUE'
             AND expected_value > 0
             AND model_probability >= 0.25
             AND sample_size >= 10
             AND priority_class IN ('hard_data_priority','medium_priority')
             AND minutes_filter_status IN ('ok','not_applicable','fallback_raw_json')
           ORDER BY expected_value DESC"""
    )
    _add_sheet(wb, "ev_ranking_actionable", h, rows)

    # 6. raw_odds
    h, rows = _table_rows(con, "SELECT * FROM betting_odds_raw ORDER BY id")
    _add_sheet(wb, "raw_odds", h, rows)

    # 7. normalized_markets
    h, rows = _table_rows(con, "SELECT * FROM betting_odds_normalized ORDER BY id")
    _add_sheet(wb, "normalized_markets", h, rows)

    # 8. player_props_found
    h, rows = _table_rows(con,
        """SELECT match_name, bookmaker, market_type, raw_market_name, raw_selection_name,
                  player_name, player_id, side, line, odds_decimal,
                  normalized_status, match_confidence, notes
           FROM betting_odds_normalized WHERE market_type LIKE 'player_%' ORDER BY match_name, raw_market_name"""
    )
    _add_sheet(wb, "player_props_found", h, rows)

    # 9. unmatched_selections
    h, rows = _table_rows(con,
        """SELECT match_name, market_type, raw_market_name, raw_selection_name,
                  team_name, player_name, side, line, odds_decimal,
                  verdict, normalized_status, notes
           FROM betting_value_scores_new WHERE verdict='UNMATCHED' ORDER BY match_name"""
    )
    _add_sheet(wb, "unmatched_selections", h, rows)

    # 10. unsupported_markets
    h, rows = _table_rows(con,
        """SELECT match_name, market_type, raw_market_name, raw_selection_name,
                  side, line, odds_decimal, verdict, probability_status, notes
           FROM betting_value_scores_new WHERE verdict='UNSUPPORTED' ORDER BY market_type"""
    )
    _add_sheet(wb, "unsupported_markets", h, rows)

    # 11. data_quality
    h, rows = _table_rows(con,
        """SELECT match_name, probability_status, normalized_status, data_quality_status,
                  verdict, priority_class, minutes_filter_status, COUNT(*) as rows
           FROM betting_value_scores_new
           GROUP BY match_name, probability_status, normalized_status, data_quality_status,
                    verdict, priority_class, minutes_filter_status
           ORDER BY match_name, rows DESC"""
    )
    _add_sheet(wb, "data_quality", h, rows)

    # 12. raw_sources
    rs_headers = ["slug", "raw_file", "endpoint", "status_code", "rows_or_items", "error", "captured_at"]
    _add_sheet(wb, "raw_sources", rs_headers, raw_sources)

    wb.save(OUT_XLSX)
    return OUT_XLSX


# ---------------------------------------------------------------------------
# Load from existing raw files (--from-raw mode)
# ---------------------------------------------------------------------------

def _unwrap(raw: dict[str, Any]) -> Any:
    return raw.get("response_json", raw)


def load_raw_files_for_slug(slug: str) -> dict[str, Any]:
    raw_dir = RAW_BASE_DIR / slug
    if not raw_dir.exists():
        return {}
    result: dict[str, Any] = {}
    patterns = {
        "bookmakers_selected": "bookmakers_selected",
        "events": "events_",
        "odds_multi": "odds_multi",
        "odds": "odds_",
    }
    for key, prefix in patterns.items():
        matches = sorted(raw_dir.glob(f"{prefix}*.json"), reverse=True)
        if key == "odds":
            matches = [f for f in matches if "multi" not in f.name]
        if not matches:
            continue
        path = matches[0]
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            result[key] = {"status_code": None, "payload": None, "raw_file": path, "error": str(exc)}
            continue
        payload = _unwrap(data)
        result[key] = {
            "status_code": data.get("status_code", 200) or 200,
            "payload": payload,
            "raw_file": path,
            "error": data.get("error"),
            "fetched_at_utc": data.get("fetched_at_utc", now_utc()),
        }
    return result


# ---------------------------------------------------------------------------
# Core multi-match pipeline
# ---------------------------------------------------------------------------

def process_match(
    cfg: dict[str, Any],
    events_payload: Any,
    odds_payload: dict[str, Any],
    bookmaker: str,
    odds_raw_file: Path,
    con: sqlite3.Connection,
    raw_sources: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    Process one match: find event, parse raw odds rows, resolve StatsHub IDs.
    Returns match_resolution dict.
    """
    match_name = cfg["match_name"]
    candidate_id = cfg["candidate_event_id"]
    slug = cfg["slug"]

    event, api_event_id, confidence = find_match_in_events(events_payload, cfg)

    status = (
        "found" if confidence == 1.0
        else "candidate_id_only" if confidence == 0.5
        else "not_found"
    )
    print(f"  [{match_name}] status={status} event_id={api_event_id} conf={confidence}")

    # Resolve StatsHub IDs from DB
    sh = resolve_statshub_ids(con, cfg)
    statshub_event_id = sh["statshub_event_id"]

    # Build raw rows (even if event not found by name — use payload we fetched)
    raw_rows: list[dict[str, Any]] = []
    market_groups: list[dict[str, Any]] = []

    # Try preferred bookmaker, then fallback variant
    if odds_payload.get("bookmakers") or not odds_payload:
        for bm_try in [bookmaker, f"{bookmaker} (no latency)"]:
            rows = raw_rows_from_odds_payload(
                odds_payload, cfg, bm_try, api_event_id, statshub_event_id,
                odds_raw_file, odds_raw_file.stem,
            )
            if rows:
                raw_rows = rows
                market_groups = extract_market_groups(odds_payload, cfg, bm_try)
                break

    print(f"    raw_rows={len(raw_rows)}, market_groups={len(market_groups)}")

    resolution = {
        "match_name": match_name,
        "odds_api_event_id": api_event_id,
        "statshub_event_id": statshub_event_id or "not_resolved",
        "home_team": cfg["home_team"],
        "away_team": cfg["away_team"],
        "home_team_id": sh["home_team_id"] or "not_found",
        "away_team_id": sh["away_team_id"] or "not_found",
        "confidence": confidence,
        "status": status,
        "home_team_perf_rows": sh["home_team_perf_rows"],
        "away_team_perf_rows": sh["away_team_perf_rows"],
        "home_player_rows": sh["home_player_rows"],
        "away_player_rows": sh["away_player_rows"],
        "odds_found": len(raw_rows) > 0,
        "raw_rows": len(raw_rows),
    }
    return resolution, raw_rows, market_groups


def _db_counts(con: sqlite3.Connection) -> dict[str, int]:
    def n(sql: str) -> int:
        return con.execute(sql).fetchone()[0]
    return {
        "raw": n("SELECT COUNT(*) FROM betting_odds_raw"),
        "normalized": n("SELECT COUNT(*) FROM betting_odds_normalized"),
        "ev_total": n("SELECT COUNT(*) FROM betting_value_scores_new WHERE expected_value IS NOT NULL"),
        "value": n("SELECT COUNT(*) FROM betting_value_scores_new WHERE verdict='VALUE'"),
        "no_value": n("SELECT COUNT(*) FROM betting_value_scores_new WHERE verdict='NO_VALUE'"),
        "unsupported": n("SELECT COUNT(*) FROM betting_value_scores_new WHERE verdict='UNSUPPORTED'"),
        "unmatched": n("SELECT COUNT(*) FROM betting_value_scores_new WHERE verdict='UNMATCHED'"),
        "player_props": n(
            "SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type LIKE 'player_%'"
        ),
        "player_ev_ok": n(
            "SELECT COUNT(*) FROM betting_value_scores_new "
            "WHERE expected_value IS NOT NULL AND minutes_filter_status='ok'"
        ),
        "actionable": n(
            "SELECT COUNT(*) FROM betting_value_scores_new "
            "WHERE verdict='VALUE' AND expected_value>0 AND model_probability>=0.25 "
            "AND sample_size>=10 AND priority_class IN ('hard_data_priority','medium_priority') "
            "AND minutes_filter_status IN ('ok','not_applicable','fallback_raw_json')"
        ),
    }


def run_pipeline(
    all_match_data: list[tuple[dict, list, list]],
    raw_sources: list[dict[str, Any]],
    selected_before: list[str],
    bookmaker: str,
    bookmaker_check_status: str,
    from_raw: bool,
    allow_partial: bool = False,
) -> None:
    """Insert all match raw rows, normalize, calculate EV, write workbook."""
    # Safety guard: refuse to overwrite a multi-match DB with single-match data.
    matches_with_data = [r for r, rows, _ in all_match_data if rows]
    if len(matches_with_data) == 1 and not allow_partial:
        match_name = matches_with_data[0]["match_name"]
        raise SystemExit(
            f"SAFETY GUARD: only 1 match has raw odds data ({match_name!r}).\n"
            "This would overwrite the full DB with a single-match run.\n"
            "Pass --allow-partial to override, or ensure all match raw files are present."
        )

    with connect() as con:
        ensure_schema(con)
        # Replace all existing data for this pipeline run
        con.execute("DELETE FROM betting_odds_raw")
        con.execute("DELETE FROM betting_odds_normalized")
        con.execute("DELETE FROM betting_value_scores_new")

        all_raw_rows: list[dict[str, Any]] = []
        all_market_coverage: list[dict[str, Any]] = []
        match_resolution_list: list[dict[str, Any]] = []

        for resolution, raw_rows, market_groups in all_match_data:
            match_resolution_list.append(resolution)
            all_raw_rows.extend(raw_rows)
            for mg in market_groups:
                mg["status"] = "returned_data" if mg["rows_returned"] else "empty"
                all_market_coverage.append(mg)

        insert_raw_rows(con, all_raw_rows, replace=False)
        norm = normalize_raw_odds(con, replace=True)
        ev = calculate_ev(con, replace=True)
        counts = _db_counts(con)

        run_summary_rows = [
            {
                "run_name": RUN_NAME,
                "mode": "from_raw" if from_raw else "live",
                "date_utc": now_utc()[:10],
                "bookmakers_selected": ", ".join(selected_before) or "(unknown)",
                "bookmaker_used": bookmaker,
                "bookmaker_check_status": bookmaker_check_status,
                "matches_processed": len([r for r in match_resolution_list if r["odds_found"]]),
                "matches_target": len(MATCH_CONFIGS),
                "total_raw_odds_rows": counts["raw"],
                "total_normalized_rows": counts["normalized"],
                "total_ev_rows": counts["ev_total"],
                "value_rows": counts["value"],
                "no_value_rows": counts["no_value"],
                "unsupported_rows": counts["unsupported"],
                "unmatched_rows": counts["unmatched"],
                "player_prop_rows": counts["player_props"],
                "player_ev_ok_rows": counts["player_ev_ok"],
                "actionable_value_rows": counts["actionable"],
            }
        ]

        workbook = write_workbook(
            con, run_summary_rows, match_resolution_list,
            all_market_coverage, raw_sources,
        )

    # --- Final report ---
    print("\n" + "=" * 70)
    print(f"TODAY'S 4 MATCHES — LIVE API ODDS FINAL REPORT  [{RUN_NAME}]")
    print("=" * 70)
    print(f"Mode: {'rebuild from raw files' if from_raw else 'live API'}")
    print(f"Bookmaker: {bookmaker}  (selected: {selected_before or '(unknown)'})")
    print(f"\nPer-match results:")
    for r in match_resolution_list:
        print(f"  {r['match_name']}")
        print(f"    target_found:         {r['status']}")
        print(f"    odds_api_event_id:    {r['odds_api_event_id']}")
        print(f"    statshub_event_id:    {r['statshub_event_id']}")
        print(f"    raw_odds_rows:        {r['raw_rows']}")
        print(f"    team_perf_rows:       {r['home_team_perf_rows']} / {r['away_team_perf_rows']}")
        print(f"    player_rows:          {r['home_player_rows']} / {r['away_player_rows']}")
    print(f"\nOverall:")
    print(f"  Total raw odds rows:      {counts['raw']}")
    print(f"  Total normalized rows:    {counts['normalized']}")
    print(f"  Total EV rows:            {counts['ev_total']}")
    print(f"  VALUE rows:               {counts['value']}")
    print(f"  Actionable VALUE rows:    {counts['actionable']}")
    print(f"  Player prop rows:         {counts['player_props']}")
    print(f"  Unsupported:              {counts['unsupported']}")
    print(f"  Unmatched:                {counts['unmatched']}")

    # Top 20 actionable
    with connect() as con:
        top20 = con.execute(
            """SELECT rank, match_name, bet_description, odds_decimal,
                      model_probability, expected_value, verdict
               FROM betting_value_scores_new
               WHERE verdict='VALUE' AND expected_value>0 AND model_probability>=0.25
                 AND sample_size>=10 AND priority_class IN ('hard_data_priority','medium_priority')
                 AND minutes_filter_status IN ('ok','not_applicable','fallback_raw_json')
               ORDER BY expected_value DESC LIMIT 20"""
        ).fetchall()

    print("\nTop 20 actionable VALUE rows:")
    for row in top20:
        print(
            f"  #{(row['rank'] or 0):3d} [{str(row['match_name'] or '')[:20]:<20}] "
            f"{str(row['bet_description'] or '')[:40]:<40} | "
            f"odds={row['odds_decimal']} | p={row['model_probability']} | "
            f"EV={row['expected_value']}"
        )

    print(f"\nOutput Excel: {workbook}")
    print("\nRun dashboard:")
    print("  streamlit run app/dashboard/betting_value_dashboard.py")


# ---------------------------------------------------------------------------
# Live mode
# ---------------------------------------------------------------------------

def run_live(allow_partial: bool = False) -> None:
    config = load_config()
    if not is_configured_api_key(config.api_key):
        raise SystemExit("live API odds unavailable: ODDS_API_IO_KEY not configured in .env")
    if not config.base_url:
        raise SystemExit("live API odds unavailable: ODDS_API_IO_BASE_URL not configured")

    raw_sources: list[dict[str, Any]] = []
    attempted = successful = failed = 0
    rate_limit_hit = False

    def call(endpoint: str, params: dict[str, Any], slug: str = "shared") -> dict[str, Any]:
        nonlocal attempted, successful, failed, rate_limit_hit
        attempted += 1
        result = api_get(endpoint, params, config.base_url, slug)
        ok = result["status_code"] and 200 <= result["status_code"] < 300 and not result["error"]
        if ok:
            successful += 1
        else:
            failed += 1
            if result["status_code"] == 429:
                rate_limit_hit = True
        raw_sources.append({
            "slug": slug,
            "raw_file": str(result["raw_file"]),
            "endpoint": endpoint,
            "status_code": result["status_code"],
            "rows_or_items": payload_count(result["payload"]),
            "error": result["error"],
            "captured_at": now_utc(),
        })
        return result

    # Task 1: bookmaker state
    print("Task 1: Checking bookmaker selection state...")
    bm_state = get_selected_bookmakers(config.api_key, config.base_url)
    attempted += 1
    selected_before: list[str] = []
    bookmaker_check_status = "unknown"
    if bm_state["status_code"] and 200 <= bm_state["status_code"] < 300:
        successful += 1
        selected_before = bm_state["selected_bookmakers"]
        bookmaker_check_status = "ok"
    elif bm_state["status_code"] == 429:
        failed += 1
        rate_limit_hit = True
        bookmaker_check_status = "rate_limited"
    else:
        failed += 1
        bookmaker_check_status = f"error_{bm_state['status_code']}"
    raw_sources.append({
        "slug": "shared",
        "raw_file": str(bm_state["raw_file"]),
        "endpoint": "bookmakers/selected",
        "status_code": bm_state["status_code"],
        "rows_or_items": len(selected_before),
        "error": bm_state["error"],
        "captured_at": now_utc(),
    })

    bookmaker = pick_bookmaker(selected_before)
    bet365_selected = any("bet365" in b.lower() for b in selected_before)
    dk_selected = any("draftkings" in b.lower() for b in selected_before)
    print(f"  Selected bookmakers: {selected_before}")
    print(f"  Bet365 selected: {'YES' if bet365_selected else 'NO'}")
    print(f"  DraftKings selected: {'YES' if dk_selected else 'NO'}")
    print(f"  Using: {bookmaker}")

    if rate_limit_hit:
        raise SystemExit("Rate limit hit on bookmaker check")

    # Task 2: Fetch events list once (all football pending events)
    print("\nTask 2: Fetching events list...")
    events_result = call(
        "events",
        {"apiKey": config.api_key, "sport": SPORT, "bookmaker": bookmaker,
         "status": "pending", "limit": 200},
        slug="shared",
    )
    if events_result["status_code"] == 429:
        raise SystemExit("Rate limit hit on /events")
    print(f"  Events returned: {payload_count(events_result['payload'])}")

    # Task 3-4: For each match, fetch odds
    print("\nTask 3-4: Resolving events and fetching odds...")
    with connect() as con:
        ensure_schema(con)
        all_match_data: list[tuple[dict, list, list]] = []

        for cfg in MATCH_CONFIGS:
            slug = cfg["slug"]
            match_name = cfg["match_name"]
            print(f"\n  Processing: {match_name}")

            # Determine event_id to use for /odds call
            event, api_event_id, confidence = find_match_in_events(
                events_result["payload"], cfg
            )

            # Fetch /odds for this event
            odds_result = call(
                "odds",
                {"apiKey": config.api_key, "eventId": api_event_id, "bookmakers": bookmaker},
                slug=slug,
            )
            if rate_limit_hit:
                print(f"  Rate limit hit on /odds for {match_name}; using empty payload")

            # Fetch /odds/multi as supplemental
            odds_multi_result = call(
                "odds/multi",
                {"apiKey": config.api_key, "eventIds": api_event_id, "bookmakers": bookmaker},
                slug=slug,
            )

            # Determine best payload
            odds_payload = (
                odds_result["payload"] if isinstance(odds_result["payload"], dict)
                else {}
            )
            odds_raw_file = odds_result["raw_file"]

            if not odds_payload.get("bookmakers"):
                if isinstance(odds_multi_result["payload"], list) and odds_multi_result["payload"]:
                    odds_payload = odds_multi_result["payload"][0]
                    odds_raw_file = odds_multi_result["raw_file"]
                elif isinstance(odds_multi_result["payload"], dict):
                    odds_payload = odds_multi_result["payload"]
                    odds_raw_file = odds_multi_result["raw_file"]

            resolution, raw_rows, market_groups = process_match(
                cfg, events_result["payload"], odds_payload, bookmaker,
                odds_raw_file, con, raw_sources,
            )
            all_match_data.append((resolution, raw_rows, market_groups))

    run_pipeline(
        all_match_data, raw_sources, selected_before,
        bookmaker, bookmaker_check_status, from_raw=False, allow_partial=allow_partial,
    )


# ---------------------------------------------------------------------------
# Rebuild from raw files
# ---------------------------------------------------------------------------

def run_from_raw(allow_partial: bool = False) -> None:
    print("=" * 70)
    print("MODE: rebuild from existing raw files")

    raw_sources: list[dict[str, Any]] = []
    selected_before: list[str] = []
    bookmaker = PREFERRED_BOOKMAKER

    # Try to get bookmaker state from any slug's raw files
    for cfg in MATCH_CONFIGS:
        files = load_raw_files_for_slug(cfg["slug"])
        bm_info = files.get("bookmakers_selected")
        if bm_info and bm_info.get("payload"):
            payload = bm_info["payload"]
            if isinstance(payload, list):
                selected_before = [
                    str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
                    for b in payload
                ]
            elif isinstance(payload, dict):
                items = payload.get("data") or payload.get("bookmakers") or []
                selected_before = [
                    str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
                    for b in (items if isinstance(items, list) else [])
                ]
            if selected_before:
                bookmaker = pick_bookmaker(selected_before)
                break

    print(f"  Bookmaker from raw: {selected_before or '(unknown)'}  using={bookmaker}")

    # Check shared events file first
    shared_files = load_raw_files_for_slug("shared")
    events_payload: Any = None
    if shared_files.get("events"):
        events_payload = shared_files["events"]["payload"]
        raw_sources.append({
            "slug": "shared",
            "raw_file": str(shared_files["events"]["raw_file"]),
            "endpoint": "events",
            "status_code": shared_files["events"].get("status_code"),
            "rows_or_items": payload_count(events_payload),
            "error": shared_files["events"].get("error"),
            "captured_at": shared_files["events"].get("fetched_at_utc", now_utc()),
        })

    with connect() as con:
        ensure_schema(con)
        all_match_data: list[tuple[dict, list, list]] = []

        for cfg in MATCH_CONFIGS:
            slug = cfg["slug"]
            match_name = cfg["match_name"]
            print(f"\n  Loading raw files for: {match_name}")
            files = load_raw_files_for_slug(slug)

            if not files:
                print(f"    No raw files found for {slug}; recording as not_found")
                sh = resolve_statshub_ids(con, cfg)
                resolution = {
                    "match_name": match_name,
                    "odds_api_event_id": cfg["candidate_event_id"],
                    "statshub_event_id": sh["statshub_event_id"] or "not_resolved",
                    "home_team": cfg["home_team"],
                    "away_team": cfg["away_team"],
                    "home_team_id": sh["home_team_id"] or "not_found",
                    "away_team_id": sh["away_team_id"] or "not_found",
                    "confidence": 0.0,
                    "status": "no_raw_files",
                    "home_team_perf_rows": sh["home_team_perf_rows"],
                    "away_team_perf_rows": sh["away_team_perf_rows"],
                    "home_player_rows": sh["home_player_rows"],
                    "away_player_rows": sh["away_player_rows"],
                    "odds_found": False,
                    "raw_rows": 0,
                }
                all_match_data.append((resolution, [], []))
                continue

            # Use per-slug events or fall back to shared
            evts_payload = (
                files["events"]["payload"] if files.get("events")
                else events_payload
            )
            if files.get("events"):
                raw_sources.append({
                    "slug": slug,
                    "raw_file": str(files["events"]["raw_file"]),
                    "endpoint": "events",
                    "status_code": files["events"].get("status_code"),
                    "rows_or_items": payload_count(files["events"]["payload"]),
                    "error": files["events"].get("error"),
                    "captured_at": files["events"].get("fetched_at_utc", now_utc()),
                })

            odds_info = files.get("odds")
            if not odds_info:
                print(f"    No /odds raw file for {slug}")
                sh = resolve_statshub_ids(con, cfg)
                resolution = {
                    "match_name": match_name,
                    "odds_api_event_id": cfg["candidate_event_id"],
                    "statshub_event_id": sh["statshub_event_id"] or "not_resolved",
                    "home_team": cfg["home_team"],
                    "away_team": cfg["away_team"],
                    "home_team_id": sh["home_team_id"] or "not_found",
                    "away_team_id": sh["away_team_id"] or "not_found",
                    "confidence": 0.0,
                    "status": "no_odds_file",
                    "home_team_perf_rows": sh["home_team_perf_rows"],
                    "away_team_perf_rows": sh["away_team_perf_rows"],
                    "home_player_rows": sh["home_player_rows"],
                    "away_player_rows": sh["away_player_rows"],
                    "odds_found": False,
                    "raw_rows": 0,
                }
                all_match_data.append((resolution, [], []))
                continue

            odds_payload = odds_info["payload"] if isinstance(odds_info["payload"], dict) else {}
            odds_raw_file = odds_info["raw_file"]

            if not odds_payload.get("bookmakers"):
                multi_info = files.get("odds_multi")
                if multi_info:
                    if isinstance(multi_info["payload"], list) and multi_info["payload"]:
                        odds_payload = multi_info["payload"][0]
                        odds_raw_file = multi_info["raw_file"]
                    elif isinstance(multi_info["payload"], dict):
                        odds_payload = multi_info["payload"]
                        odds_raw_file = multi_info["raw_file"]

            raw_sources.append({
                "slug": slug,
                "raw_file": str(odds_raw_file),
                "endpoint": "odds",
                "status_code": odds_info.get("status_code"),
                "rows_or_items": payload_count(odds_info["payload"]),
                "error": odds_info.get("error"),
                "captured_at": odds_info.get("fetched_at_utc", now_utc()),
            })

            resolution, raw_rows, market_groups = process_match(
                cfg, evts_payload or {}, odds_payload, bookmaker,
                odds_raw_file, con, raw_sources,
            )
            all_match_data.append((resolution, raw_rows, market_groups))

    run_pipeline(
        all_match_data, raw_sources, selected_before,
        bookmaker, "from_raw_files", from_raw=True, allow_partial=allow_partial,
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fetch live API odds for today's 4 World Cup matches."
    )
    parser.add_argument(
        "--from-raw", action="store_true",
        help="Rebuild pipeline from existing raw JSON files (no API calls)",
    )
    parser.add_argument(
        "--allow-partial", action="store_true",
        help="Allow overwriting DB even when only 1 match has raw data (bypasses safety guard)",
    )
    args = parser.parse_args()
    if args.from_raw:
        run_from_raw(allow_partial=args.allow_partial)
    else:
        run_live(allow_partial=args.allow_partial)


if __name__ == "__main__":
    main()
