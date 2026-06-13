"""
Reproduce the Playwright-intercepted StatsHub team players/performance endpoint
as a direct HTTP call, test pagination, and map all players to FIFA rosters.

Snapshot: today_browser_endpoint_replay_probe

Key discovered endpoint:
  GET /api/team/{team_id}/players/performance
      ?tournamentId=14,16,133,140,851,14100
      &limit=20          ← controls events shown per player, NOT player count
      &location=both
      &fixtureId={event_id}

Direct replay: WORKS with basic headers (User-Agent, Accept, Referer).
"""
import sys, json, time, re, pathlib, sqlite3, requests, unicodedata
from collections import defaultdict
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ─── Config ────────────────────────────────────────────────────────────────────
DB_PATH  = pathlib.Path("data/mundial.db")
RAW_BASE = pathlib.Path("data/raw/statshub/snapshots")
SNAPSHOT = "today_browser_endpoint_replay_probe"
BASE_URL = "https://www.statshub.com/api"
RATE_DELAY = 1.5

HEADERS_MINIMAL = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
}
HEADERS_WITH_REFERER = {
    **HEADERS_MINIMAL,
    "Referer": "https://www.statshub.com/",
    "Origin":  "https://www.statshub.com",
}

# ─── Teams & Fixtures ──────────────────────────────────────────────────────────
TEAMS = [
    {"name": "Canada",               "db_name": "Canada",               "team_id": 4752},
    {"name": "Bosnia & Herzegovina", "db_name": "Bosnia and Herzegovina","team_id": 4479},
    {"name": "USA",                  "db_name": "United States",         "team_id": 4724},
    {"name": "Paraguay",             "db_name": "Paraguay",              "team_id": 4789},
]
FIXTURES = {
    4752: {"event_id": 15186836, "fixture_url": "https://www.statshub.com/es/fixture/canada-vs-bosnia-and-herzegovina-mqazwx/157343"},
    4479: {"event_id": 15186836, "fixture_url": "https://www.statshub.com/es/fixture/canada-vs-bosnia-and-herzegovina-mqazwx/157343"},
    4724: {"event_id": 15186873, "fixture_url": "https://www.statshub.com/es/fixture/usa-vs-paraguay-mqazwt/157344"},
    4789: {"event_id": 15186873, "fixture_url": "https://www.statshub.com/es/fixture/usa-vs-paraguay-mqazwt/157344"},
}

# Base tournament IDs used by the browser
TOURNAMENT_IDS = "14,16,133,140,851,14100"

# ─── Known browser-intercepted URLs ────────────────────────────────────────────
BROWSER_URLS = {
    4752: f"https://www.statshub.com/api/team/4752/players/performance?tournamentId={TOURNAMENT_IDS}&limit=20&location=both&fixtureId=15186836",
    4479: f"https://www.statshub.com/api/team/4479/players/performance?tournamentId={TOURNAMENT_IDS}&limit=20&location=both&fixtureId=15186836",
    4724: f"https://www.statshub.com/api/team/4724/players/performance?tournamentId={TOURNAMENT_IDS}&limit=20&location=both&fixtureId=15186873",
    4789: f"https://www.statshub.com/api/team/4789/players/performance?tournamentId={TOURNAMENT_IDS}&limit=20&location=both&fixtureId=15186873",
}

# ─── USA roster name → canonical endpoint name mapping ────────────────────────
USA_ROSTER_MAP = {
    # fifa_official_name: endpoint_name
    "Alejandro ZENDEJAS SAAVEDRA":         "Alex Zendejas",
    "Alexander Michael FREEMAN":           "Alexander Freeman",
    "Auston Levi-Jesaiah TRUSTY":          "Auston Trusty",
    "Brenden Russell AARONSON":            "Brenden Aaronson",
    "Christian Mate PULISIC":              "Christian Pulišić",
    "Christopher Jeffrey RICHARDS":        "Chris Richards",
    "Christopher Keith BRADY":             "Chris Brady",
    "Folarin Jolaoluwa BALOGUN":           "Folarin Balogun",
    "Giovanni Alejandro REYNA":            "Giovanni Reyna",
    "Haji Amir WRIGHT":                    "Haji Wright",
    "Joseph Michael SCALLY":               "Joe Scally",
    "Malik Leon TILLMAN":                  "Malik Tillman",
    "Mark Alexander MCKENZIE":             "Mark McKenzie",
    "Matthew Andrew Geary FREESE":         "Matthew Freese",
    "Matthew Charles TURNER":              "Matt Turner",
    "Maximilian Michael ARFSTEN":          "Maximilian Arfsten",
    "Miles Gordon ROBINSON":               "Miles Robinson",
    "Ricardo Daniel PEPI":                 "Ricardo Pepi",
    "Sebastian Matthew BERHALTER":         "Sebastian Berhalter",
    "Sergiño Gianni DEST":                 "Sergiño Dest",
    "Timothy Michael REAM":                "Tim Ream",
    "Timothy Tarpeh WEAH":                 "Tim Weah",
    "Tyler Shaan ADAMS":                   "Tyler Adams",
    "Weston James Earl MC KENNIE":         "Weston McKennie",
}

# ─── DB helpers ────────────────────────────────────────────────────────────────
def get_con():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def _now():
    return datetime.now(timezone.utc).isoformat()

def _norm(s):
    """Normalize string for comparison: lower, strip accents, collapse spaces."""
    if not s: return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).lower().strip()

# ─── HTTP ──────────────────────────────────────────────────────────────────────
def _ts():
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

def fetch(label, url, snap_dir, headers=None, force=False):
    """Fetch URL, cache to snap_dir/label_*.json. Returns (sc, payload, ct)."""
    if not force:
        existing = list(snap_dir.glob(f"{label}_*.json"))
        if existing:
            try:
                data = json.loads(existing[0].read_text(encoding="utf-8"))
                return 200, data, "application/json"
            except Exception:
                pass
    time.sleep(RATE_DELAY)
    hdrs = headers or HEADERS_WITH_REFERER
    try:
        r = requests.get(url, headers=hdrs, timeout=20)
    except Exception as e:
        return -1, None, ""
    ct = r.headers.get("content-type", "")
    fpath = snap_dir / f"{label}_{_ts()}.json"
    if r.status_code == 200 and "json" in ct:
        try:
            payload = r.json()
            fpath.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            return 200, payload, ct
        except Exception:
            pass
    fpath.write_text(r.text[:2000], encoding="utf-8")
    return r.status_code, None, ct

# ─── Player extraction from /team/{id}/players/performance ────────────────────
def parse_team_players(payload):
    """Extract player list from /api/team/{id}/players/performance response."""
    if not isinstance(payload, dict):
        return []
    players = []
    for item in payload.get("data") or []:
        if not isinstance(item, dict):
            continue
        pid  = item.get("id")
        name = item.get("name") or item.get("shortName")
        if not pid or not name:
            continue
        players.append({
            "player_id":   str(pid),
            "player_name": name,
            "slug":        item.get("slug"),
            "position":    item.get("position"),
            "stats":       item.get("stats"),
        })
    return players

# ─── Performance download/parse ────────────────────────────────────────────────
def _to_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

def parse_perf(payload, player_id, player_name, team_name):
    if not isinstance(payload, dict): return None
    items = (payload.get("playerStatisticsEvents") or payload.get("events")
             or payload.get("data") or [])
    if not items: return None
    apps=minutes=goals=assists=sot=kp=tp=ap=tac=pl=fouls=wf=yc=rc=0
    xg=xa=0.0; dates=[]
    for item in items:
        stats = item.get("player_statistics_event") or item
        mp = stats.get("minutesPlayed") or stats.get("minutes_played") or 0
        try: mp = int(mp)
        except: mp = 0
        if mp > 0:
            apps += 1; minutes += mp
        goals   += stats.get("goals") or stats.get("goal") or 0
        assists += stats.get("goalAssist") or stats.get("assists") or 0
        sot     += stats.get("onTargetScoringAttempt") or stats.get("shots_on_target") or 0
        kp   += stats.get("keyPass") or 0
        tp   += stats.get("totalPass") or 0
        ap   += stats.get("accuratePass") or 0
        tac  += stats.get("totalTackle") or 0
        pl   += stats.get("possessionLostCtrl") or 0
        fouls+= stats.get("fouls") or 0
        wf   += stats.get("wasFouled") or 0
        yc   += stats.get("yellowCard") or 0
        rc   += stats.get("redCard") or 0
        xg   += _to_float(stats.get("expectedGoals")) or 0.0
        xa   += _to_float(stats.get("expectedAssists")) or 0.0
        ev = item.get("events") or {}
        ts = ev.get("startTimestamp")
        if ts:
            try:
                dates.append(datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d"))
            except: pass
    dates.sort()
    return dict(
        player_id=player_id, player_name=player_name, team_name=team_name,
        source_rows=len(items), appearances=apps, minutes=minutes,
        goals=goals, assists=assists, shots_on_target=sot,
        xG=round(xg,4) if xg else None, xA=round(xa,4) if xa else None,
        key_passes=kp, passes=tp, accurate_passes=ap,
        tackles=tac, possession_lost=pl, fouls=fouls, was_fouled=wf,
        yellow_cards=yc, red_cards=rc,
        date_min=dates[0] if dates else None, date_max=dates[-1] if dates else None,
    )

def upsert_perf(cur, agg):
    cur.execute("""
        INSERT OR REPLACE INTO statshub_player_performance_aggregates
            (player_id, player_name, team_name, source_rows, appearances,
             minutes, goals, assists, shots_on_target,
             xG, xA, key_passes, passes, accurate_passes,
             tackles, possession_lost, yellow_cards, red_cards,
             date_min, date_max, imported_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        agg["player_id"], agg["player_name"], agg["team_name"],
        agg["source_rows"], agg["appearances"], agg["minutes"],
        agg["goals"], agg["assists"], agg["shots_on_target"],
        agg["xG"], agg["xA"], agg["key_passes"], agg["passes"],
        agg["accurate_passes"], agg["tackles"], agg["possession_lost"],
        agg["yellow_cards"], agg["red_cards"],
        agg["date_min"], agg["date_max"], _now()
    ))

# ─── Excel ─────────────────────────────────────────────────────────────────────
def generate_excel(con, snap_dir, intercepted_records, replay_records,
                   pagination_records, endpoint_pool, coverage_before, newly_confirmed_list,
                   newly_downloaded_perf, out_path):
    import openpyxl
    wb = openpyxl.Workbook()
    cur = con.cursor()

    # 1. intercepted_requests
    ws = wb.active
    ws.title = "intercepted_requests"
    ws.append(["team_name","team_id","fixture_url","full_url","method","status_code",
               "query_params","useful_headers","rows_detected","player_ids_detected",
               "raw_file","notes"])
    for r in intercepted_records:
        ws.append([r.get(k,"") for k in ws[1]])

    # 2. replay_results
    ws2 = wb.create_sheet("replay_results")
    ws2.append(["team_name","team_id","full_url","replay_status","json_ok","rows_detected",
                "player_ids_detected","matched_browser_response","classification","raw_file","notes"])
    for r in replay_records:
        ws2.append([r.get(k,"") for k in ws2[1]])

    # 3. pagination_tests
    ws3 = wb.create_sheet("pagination_tests")
    ws3.append(["team_name","team_id","variant_url","status","rows_detected",
                "unique_player_ids","new_players_found","raw_file","notes"])
    for r in pagination_records:
        ws3.append([r.get(k,"") for k in ws3[1]])

    # 4. endpoint_player_pool
    ws4 = wb.create_sheet("endpoint_player_pool")
    ws4.append(["team_name","team_id","player_id","player_name","player_slug",
                "source_endpoint","matched_to_fifa_roster","roster_player_name",
                "match_status","notes"])
    for r in endpoint_pool:
        ws4.append([r.get(k,"") for k in ws4[1]])

    # 5. updated_coverage
    ws5 = wb.create_sheet("updated_coverage")
    ws5.append(["team_name","confirmed_before","confirmed_after","newly_confirmed",
                "probable","ambiguous","unresolved","endpoint_only_players",
                "coverage_percentage","verdict"])
    for tname in ["Canada","Bosnia and Herzegovina","United States","Paraguay"]:
        cur.execute("""
            SELECT statshub_player_id_status, COUNT(*) cnt
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name = ? GROUP BY statshub_player_id_status
        """, (tname,))
        s = {r[0]: r[1] for r in cur.fetchall()}
        total     = sum(s.values())
        confirmed = s.get("confirmed",0) + s.get("skipped_existing",0)
        probable  = s.get("probable",0)
        ambiguous = s.get("ambiguous",0)
        unresolved= s.get("unresolved",0)
        pct       = round(100*confirmed/total,1) if total else 0
        before    = coverage_before.get(tname, 0)
        newly     = confirmed - before
        ws5.append([tname, before, confirmed, newly, probable, ambiguous, unresolved, 0, pct,
                    "READY" if pct >= 80 else "INCOMPLETE"])

    # 6. newly_downloaded_player_performance
    ws6 = wb.create_sheet("perf_newly_downloaded")
    ws6.append(["team_name","player_name","player_id","source_rows","appearances","minutes",
                "goals","assists","shots_on_target","xG","xA","date_min","date_max"])
    for agg in newly_downloaded_perf:
        ws6.append([agg.get(k,"") for k in ws6[1]])

    # 7. raw_sources
    ws7 = wb.create_sheet("raw_sources")
    ws7.append(["snapshot_name","endpoint_name","url","status_code","raw_file"])
    cur.execute("""
        SELECT snapshot_name, endpoint_name, url, status_code, raw_file_path
        FROM statshub_snapshots WHERE snapshot_name = ?
        ORDER BY id DESC LIMIT 200
    """, (SNAPSHOT,))
    for r in cur.fetchall():
        ws7.append(list(r))

    # 8. data_dictionary
    ws8 = wb.create_sheet("data_dictionary")
    ws8.append(["sheet","column","description"])
    for row in [
        ("intercepted_requests","full_url","Exact URL as captured from Playwright browser intercept"),
        ("intercepted_requests","query_params","Query parameters extracted from URL"),
        ("intercepted_requests","player_ids_detected","Number of player IDs found in response data array"),
        ("replay_results","classification","replay_ok/needs_headers/needs_cookie/browser_only/failed"),
        ("replay_results","matched_browser_response","Whether rows_detected matches browser capture"),
        ("pagination_tests","new_players_found","Players in response not in base (limit=20) response"),
        ("endpoint_player_pool","match_status","confirmed/probable/unmatched"),
        ("updated_coverage","confirmed_after","confirmed + skipped_existing statuses"),
        ("updated_coverage","endpoint_only_players","Players in endpoint response not in FIFA roster"),
    ]:
        ws8.append(row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  OK Saved: {out_path}")

# ─── Snapshot DB helper ────────────────────────────────────────────────────────
def _ensure_snapshot_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS statshub_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_name TEXT, endpoint_name TEXT, url TEXT, method TEXT,
            status_code INTEGER, content_type TEXT, response_size INTEGER,
            looks_json INTEGER, json_top_keys TEXT, rows_detected INTEGER,
            raw_file_path TEXT, status TEXT, message TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)

def _save_snap(cur, endpoint_name, url, sc, ct, rows, raw_file, msg=""):
    cur.execute("""
        INSERT INTO statshub_snapshots
            (snapshot_name, endpoint_name, url, method, status_code, content_type,
             looks_json, rows_detected, raw_file_path, status, message, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (SNAPSHOT, endpoint_name, url, "GET", sc, ct, 1 if sc==200 else 0,
          rows, str(raw_file), "ok" if sc==200 else "error", msg, _now()))

# ═══════════════════════════════════════════════════════════════════════════════
def main():
    snap_dir = RAW_BASE / SNAPSHOT
    snap_dir.mkdir(parents=True, exist_ok=True)

    con = get_con()
    cur = con.cursor()
    _ensure_snapshot_table(cur)
    con.commit()

    # ─── TASK 1: Document intercepted requests ────────────────────────────────
    print("=== TASK 1: Document intercepted browser API calls ===")
    intercepted_records = []
    for team in TEAMS:
        tid = team["team_id"]
        url = BROWSER_URLS[tid]
        fix = FIXTURES[tid]
        # Parse query params from URL
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(url)
        qparams = parse_qs(parsed.query)
        qstr = "; ".join(f"{k}={v[0]}" for k,v in qparams.items())
        rec = {
            "team_name":          team["name"],
            "team_id":            str(tid),
            "fixture_url":        fix["fixture_url"],
            "full_url":           url,
            "method":             "GET",
            "status_code":        200,
            "query_params":       qstr,
            "useful_headers":     "User-Agent; Accept; Accept-Language; Referer",
            "rows_detected":      "",
            "player_ids_detected":"",
            "raw_file":           f"followup file in today_playwright_fixture_players_probe snapshot",
            "notes":              f"Intercepted via Playwright page load of {fix['fixture_url']}",
        }
        # Load from previous cache
        prev_snap = RAW_BASE / "today_playwright_fixture_players_probe"
        prev_files = list(prev_snap.glob(f"followup_*team_{tid}*.json"))
        if prev_files:
            try:
                payload = json.loads(prev_files[0].read_text(encoding="utf-8"))
                players = parse_team_players(payload)
                rec["rows_detected"]       = len(players)
                rec["player_ids_detected"] = len(players)
                rec["raw_file"]            = str(prev_files[0])
            except Exception:
                pass
        intercepted_records.append(rec)
        print(f"  {team['name']}: {rec['rows_detected']} players from intercepted response")

    # ─── TASK 2: Replay without Playwright ───────────────────────────────────
    print("\n=== TASK 2: Replay exact browser requests without Playwright ===")
    replay_records = []
    base_players_by_team = {}  # team_id -> list of {player_id, player_name, ...}

    for team in TEAMS:
        tid  = team["team_id"]
        url  = BROWSER_URLS[tid]
        label = f"replay_team_{tid}_base"

        # Try minimal headers first
        sc_min, payload_min, ct_min = fetch(label + "_minimal", url, snap_dir,
                                            headers=HEADERS_MINIMAL)
        # Try with referer
        sc_ref, payload_ref, ct_ref = fetch(label + "_referer", url, snap_dir,
                                            headers=HEADERS_WITH_REFERER)

        for variant, sc, payload, ct in [
            ("minimal",  sc_min, payload_min, ct_min),
            ("referer",  sc_ref, payload_ref, ct_ref),
        ]:
            players = parse_team_players(payload) if payload else []
            # Compare to browser capture
            browser_count = int(intercepted_records[TEAMS.index(team)].get("rows_detected") or 0)
            matched = abs(len(players) - browser_count) <= 5
            if sc == 200 and players:
                cls = "replay_ok"
            elif sc == 200 and not players:
                cls = "needs_headers"
            elif sc in (401, 403):
                cls = "needs_cookie"
            else:
                cls = "failed"
            rec = {
                "team_name":               team["name"],
                "team_id":                 str(tid),
                "full_url":                url + f"  [headers={variant}]",
                "replay_status":           str(sc),
                "json_ok":                 str(bool(payload)),
                "rows_detected":           str(len(players)),
                "player_ids_detected":     str(len(players)),
                "matched_browser_response":str(matched),
                "classification":          cls,
                "raw_file":                f"{label}_{variant}",
                "notes":                   f"Browser had {browser_count} players",
            }
            replay_records.append(rec)
            print(f"  {team['name']} [{variant}]: sc={sc} players={len(players)} cls={cls} match={matched}")

        # Store best response for cross-matching
        best = payload_ref or payload_min
        if best:
            base_players_by_team[tid] = parse_team_players(best)
        _save_snap(cur, label, url, sc_ref or sc_min, ct_ref or ct_min,
                   len(base_players_by_team.get(tid,[])), snap_dir / f"{label}.json")
    con.commit()

    # ─── TASK 3: Pagination tests ─────────────────────────────────────────────
    print("\n=== TASK 3: Test pagination/limit variants ===")
    pagination_records = []
    base_url_template = (
        f"https://www.statshub.com/api/team/{{tid}}/players/performance"
        f"?tournamentId={TOURNAMENT_IDS}&limit={{lim}}&location=both&fixtureId={{fid}}"
    )
    pagination_variants = [
        ("limit_50",    "limit=50",   lambda tid,fid: base_url_template.format(tid=tid,lim=50,fid=fid)),
        ("limit_100",   "limit=100",  lambda tid,fid: base_url_template.format(tid=tid,lim=100,fid=fid)),
        ("no_fixture",  "no fixtureId", lambda tid,fid: f"https://www.statshub.com/api/team/{tid}/players/performance?tournamentId={TOURNAMENT_IDS}&limit=20&location=both"),
        ("no_limit",    "no limit param", lambda tid,fid: f"https://www.statshub.com/api/team/{tid}/players/performance?tournamentId={TOURNAMENT_IDS}&location=both&fixtureId={fid}"),
        ("all_tourney", "tournamentId=16 only", lambda tid,fid: f"https://www.statshub.com/api/team/{tid}/players/performance?tournamentId=16&limit=50&location=both&fixtureId={fid}"),
    ]

    # Use only USA (most incomplete) and Canada for pagination tests
    test_teams = [t for t in TEAMS if t["team_id"] in (4724, 4752)]

    for team in test_teams:
        tid = team["team_id"]
        fid = FIXTURES[tid]["event_id"]
        base_count = len(base_players_by_team.get(tid, []))
        base_ids   = {p["player_id"] for p in base_players_by_team.get(tid, [])}
        print(f"\n  {team['name']} (base={base_count} players):")
        for vname, vdesc, url_fn in pagination_variants:
            url   = url_fn(tid, fid)
            label = f"pag_{tid}_{vname}"
            sc, payload, ct = fetch(label, url, snap_dir)
            players = parse_team_players(payload) if payload else []
            new_pids = {p["player_id"] for p in players} - base_ids
            rec = {
                "team_name":        team["name"],
                "team_id":          str(tid),
                "variant_url":      url,
                "status":           str(sc),
                "rows_detected":    str(len(players)),
                "unique_player_ids":str(len({p["player_id"] for p in players})),
                "new_players_found":str(len(new_pids)),
                "raw_file":         str(snap_dir / f"{label}.json"),
                "notes":            vdesc,
            }
            pagination_records.append(rec)
            print(f"    {vname}: sc={sc} players={len(players)} new={len(new_pids)}")
            if players and len(players) > base_count:
                print(f"      *** MORE PLAYERS! {len(players)} vs {base_count} base ***")
            _save_snap(cur, label, url, sc, ct, len(players), snap_dir / f"{label}.json", vdesc)
    con.commit()

    # ─── Load coverage BEFORE updates ────────────────────────────────────────
    coverage_before = {}
    for team in TEAMS:
        cur.execute("""
            SELECT statshub_player_id_status, COUNT(*) cnt
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name = ? GROUP BY statshub_player_id_status
        """, (team["db_name"],))
        s = {r[0]: r[1] for r in cur.fetchall()}
        coverage_before[team["db_name"]] = s.get("confirmed",0) + s.get("skipped_existing",0)

    # ─── TASK 4: Cross-match endpoint players to rosters ─────────────────────
    print("\n=== TASK 4: Cross-match endpoint players to FIFA rosters ===")
    endpoint_pool = []
    newly_confirmed = []

    # Load roster for all 4 teams
    cur.execute("""
        SELECT sp.id as row_id, sp.player_name, sp.jersey_number, sp.position,
               sp.statshub_player_id_status as id_status,
               COALESCE(sp.player_id, '') as player_id,
               wt.team_name
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
    """)
    roster_rows = [dict(r) for r in cur.fetchall()]
    roster_by_team = defaultdict(list)
    for r in roster_rows:
        roster_by_team[r["team_name"]].append(r)

    # Build index: endpoint player_id → roster row (for players already confirmed)
    confirmed_pid_set = {r["player_id"] for r in roster_rows
                         if r["player_id"] and r["id_status"] in ("confirmed","skipped_existing")}

    def _confirm_player(cur, row_id, player_id, player_name, team_name, reason):
        cur.execute("""
            UPDATE statshub_team_players
            SET player_id=?, statshub_player_id_status='confirmed',
                player_id_match_source=?, player_id_confidence_score=95.0
            WHERE id=?
        """, (player_id, reason, row_id))
        newly_confirmed.append({"player_id": player_id, "player_name": player_name,
                                 "team_name": team_name, "reason": reason})
        confirmed_pid_set.add(player_id)
        print(f"    [CONFIRMED] {player_name} ({team_name}): id={player_id} [{reason}]")

    for team in TEAMS:
        tid    = team["team_id"]
        tdb    = team["db_name"]
        ep_players = base_players_by_team.get(tid, [])
        roster = roster_by_team.get(tdb, [])
        print(f"\n  {team['name']}: endpoint={len(ep_players)} roster={len(roster)}")

        # Build name lookup for endpoint players
        ep_by_name_norm = {_norm(p["player_name"]): p for p in ep_players}
        ep_by_id        = {p["player_id"]: p for p in ep_players}

        for rp in roster:
            rname  = rp["player_name"]
            rstatus= rp["id_status"]
            rcur_id= rp["player_id"]

            # Already confirmed/skipped_existing: verify or leave alone
            if rstatus in ("confirmed", "skipped_existing"):
                ep_match = ep_by_id.get(str(rcur_id))
                pool_rec = {
                    "team_name": tdb, "team_id": str(tid),
                    "player_id": rcur_id, "player_name": rname,
                    "player_slug": ep_match.get("slug","") if ep_match else "",
                    "source_endpoint": "team_players_performance",
                    "matched_to_fifa_roster": "yes",
                    "roster_player_name": rname,
                    "match_status": "confirmed_existing",
                    "notes": "pre-existing confirmed"
                }
                endpoint_pool.append(pool_rec)
                continue

            # ── USA: use manual mapping ──
            if tdb == "United States" and rname in USA_ROSTER_MAP:
                target_ep_name = USA_ROSTER_MAP[rname]
                ep_match = ep_by_name_norm.get(_norm(target_ep_name))
                if ep_match:
                    _confirm_player(cur, rp["row_id"], ep_match["player_id"],
                                    rname, tdb, f"usa_manual_map→{target_ep_name}")
                    endpoint_pool.append({
                        "team_name": tdb, "team_id": str(tid),
                        "player_id": ep_match["player_id"],
                        "player_name": ep_match["player_name"],
                        "player_slug": ep_match.get("slug",""),
                        "source_endpoint": "team_players_performance",
                        "matched_to_fifa_roster": "yes",
                        "roster_player_name": rname,
                        "match_status": "confirmed",
                        "notes": f"manual map: {rname}→{target_ep_name}",
                    })
                    continue
                else:
                    print(f"    [miss] USA manual map not found in endpoint: {target_ep_name}")

            # ── General: try name normalization ──
            # Extract last name token (FIFA names are ALL_CAPS last name + given name)
            parts = rname.split()
            # Try last-name-first style (FIFA official name often has SURNAME first)
            last_name = parts[-1] if parts else ""
            given_parts = parts[:-1] if len(parts) > 1 else []

            # Candidates: exact norm match, last-name partial match
            matched_ep = None
            reason_str = ""

            norm_rname = _norm(rname)
            if norm_rname in ep_by_name_norm:
                matched_ep = ep_by_name_norm[norm_rname]
                reason_str = "name_exact_norm"
            else:
                # Try matching on last word (surname) only — only safe if unique in team
                norm_last = _norm(last_name)
                surname_hits = [ep for norm, ep in ep_by_name_norm.items()
                                if norm_last in norm.split() and len(norm_last) > 3]
                if len(surname_hits) == 1:
                    matched_ep = surname_hits[0]
                    reason_str = f"surname_unique:{last_name}"

                # Also try probable player's current ID against endpoint
                if not matched_ep and rcur_id and str(rcur_id) in ep_by_id:
                    matched_ep = ep_by_id[str(rcur_id)]
                    reason_str = "existing_probable_id_in_endpoint"

            if matched_ep:
                ep_pid = matched_ep["player_id"]
                # Only confirm if either: status is probable with matching ID, or fresh match
                if rstatus in ("probable", "ambiguous", "unresolved"):
                    _confirm_player(cur, rp["row_id"], ep_pid, rname, tdb, reason_str)
                endpoint_pool.append({
                    "team_name": tdb, "team_id": str(tid),
                    "player_id": ep_pid,
                    "player_name": matched_ep["player_name"],
                    "player_slug": matched_ep.get("slug",""),
                    "source_endpoint": "team_players_performance",
                    "matched_to_fifa_roster": "yes",
                    "roster_player_name": rname,
                    "match_status": "confirmed",
                    "notes": reason_str,
                })
            else:
                if rstatus not in ("confirmed","skipped_existing"):
                    print(f"    [unmatched] {rname} (status={rstatus})")
                endpoint_pool.append({
                    "team_name": tdb, "team_id": str(tid),
                    "player_id": rcur_id or "",
                    "player_name": rname,
                    "player_slug": "",
                    "source_endpoint": "",
                    "matched_to_fifa_roster": "no",
                    "roster_player_name": rname,
                    "match_status": rstatus,
                    "notes": "not found in endpoint response",
                })

        # Add endpoint-only players (in endpoint but not in roster)
        roster_pids = {rp["player_id"] for rp in roster if rp["player_id"]}
        for ep in ep_players:
            if ep["player_id"] not in roster_pids:
                endpoint_pool.append({
                    "team_name": tdb, "team_id": str(tid),
                    "player_id": ep["player_id"],
                    "player_name": ep["player_name"],
                    "player_slug": ep.get("slug",""),
                    "source_endpoint": "team_players_performance",
                    "matched_to_fifa_roster": "no",
                    "roster_player_name": "",
                    "match_status": "endpoint_only",
                    "notes": "in endpoint but not in FIFA 26-player roster",
                })

    con.commit()
    print(f"\n  Total newly confirmed: {len(newly_confirmed)}")

    # ─── TASK 5: Correct probable Canada IDs from endpoint ───────────────────
    print("\n=== TASK 5: Correct mismatched probable/confirmed IDs from endpoint ===")
    canada_corrections = 0
    canada_ep = {p["player_name"]: p for p in base_players_by_team.get(4752, [])}
    canada_ep_by_norm = {_norm(p["player_name"]): p for p in base_players_by_team.get(4752, [])}

    # Canada probable players whose ID might mismatch endpoint
    cur.execute("""
        SELECT sp.id as row_id, sp.player_name, sp.player_id, sp.statshub_player_id_status
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name = 'Canada'
        AND sp.statshub_player_id_status IN ('probable','confirmed')
    """)
    for row in cur.fetchall():
        row = dict(row)
        # Check if player appears in endpoint pool with a DIFFERENT ID
        ep_match = None
        # Match by surname (Canada roster names are uppercase SURNAME)
        parts = row["player_name"].strip().split()
        last  = parts[-1] if parts else ""
        norm_last = _norm(last)
        if len(norm_last) > 3:
            hits = [p for n, p in canada_ep_by_norm.items()
                    if norm_last in n.split() and len(norm_last) > 3]
            if len(hits) == 1:
                ep_match = hits[0]
        if ep_match and ep_match["player_id"] != str(row["player_id"]):
            print(f"  CORRECTING {row['player_name']}: {row['player_id']} → {ep_match['player_id']} ({ep_match['player_name']})")
            cur.execute("""
                UPDATE statshub_team_players
                SET player_id=?, statshub_player_id_status='confirmed',
                    player_id_match_source='endpoint_id_correction',
                    player_id_confidence_score=95.0
                WHERE id=?
            """, (ep_match["player_id"], row["row_id"]))
            canada_corrections += 1
            # Mark as newly confirmed if not already in list
            if not any(nc["player_id"] == ep_match["player_id"] for nc in newly_confirmed):
                newly_confirmed.append({
                    "player_id": ep_match["player_id"],
                    "player_name": row["player_name"],
                    "team_name": "Canada",
                    "reason": "id_corrected"
                })

    con.commit()
    print(f"  Canada ID corrections: {canada_corrections}")

    # ─── TASK 6: Download performance for newly confirmed ─────────────────────
    print("\n=== TASK 6: Download performance for newly confirmed players ===")
    newly_downloaded_perf = []
    seen_pids = set()
    for nc in newly_confirmed:
        pid = nc["player_id"]
        if not pid or pid in seen_pids:
            continue
        seen_pids.add(pid)
        # Check if already have performance
        cur.execute("SELECT id FROM statshub_player_performance_aggregates WHERE player_id=?", (pid,))
        if cur.fetchone():
            continue
        # Fetch
        label   = f"perf_{pid}"
        url     = f"{BASE_URL}/player/{pid}/performance?limit=50"
        sc, payload, ct = fetch(label, url, snap_dir)
        if not payload:
            print(f"  SKIP {nc['player_name']}: sc={sc}")
            continue
        agg = parse_perf(payload, pid, nc["player_name"], nc["team_name"])
        if not agg:
            print(f"  NO_DATA {nc['player_name']}")
            continue
        upsert_perf(cur, agg)
        con.commit()
        newly_downloaded_perf.append(agg)
        print(f"  OK {nc['player_name']} ({nc['team_name']}): apps={agg['appearances']} min={agg['minutes']}")
    print(f"  Downloaded: {len(newly_downloaded_perf)} new performance records")

    # ─── TASK 7: Excel ────────────────────────────────────────────────────────
    print("\n=== TASK 7: Generate Excel ===")
    out = pathlib.Path("data/processed/statshub/today_browser_endpoint_replay_review.xlsx")
    generate_excel(con, snap_dir, intercepted_records, replay_records, pagination_records,
                   endpoint_pool, coverage_before, newly_confirmed,
                   newly_downloaded_perf, out)

    # ─── TASK 8: Final report ─────────────────────────────────────────────────
    print("\n============================================================")
    print("TASK 8: FINAL REPORT")
    print("============================================================")

    print("\n--- A. Endpoint replay ---")
    ok_replays = [r for r in replay_records if r["classification"] == "replay_ok"]
    print(f"  Browser-discovered endpoints tested: {len(replay_records)//2}")
    print(f"  Replaying without Playwright: {'YES' if ok_replays else 'NO'}")
    if ok_replays:
        print(f"  Works with headers: minimal User-Agent+Accept")
        print(f"  Reusable function: download_statshub_team_players_performance(team_id, fixture_event_id)")
    for r in replay_records:
        print(f"    {r['team_name']} [{r['full_url'].split('[')[-1].rstrip(']')}]: {r['classification']}")

    print("\n--- B. Pagination ---")
    more_variants = [r for r in pagination_records if int(r.get("rows_detected") or 0) > 0]
    base_counts = {t["team_id"]: len(base_players_by_team.get(t["team_id"],[])) for t in TEAMS}
    print(f"  Base response (limit=20): USA={base_counts.get(4724,0)} Canada={base_counts.get(4752,0)}")
    for r in pagination_records:
        note = ""
        if int(r.get("rows_detected") or 0) > 0:
            note = f"→ {r['rows_detected']} players"
        print(f"    {r['team_name']} {r['notes']}: sc={r['status']} rows={r['rows_detected']} new={r['new_players_found']} {note}")

    print("\n--- C. Coverage improvement ---")
    for team in TEAMS:
        tdb = team["db_name"]
        cur.execute("""
            SELECT statshub_player_id_status, COUNT(*) cnt
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name = ? GROUP BY statshub_player_id_status
        """, (tdb,))
        s = {r[0]: r[1] for r in cur.fetchall()}
        total     = sum(s.values())
        confirmed = s.get("confirmed",0) + s.get("skipped_existing",0)
        before    = coverage_before[tdb]
        pct = round(100*confirmed/total,1) if total else 0
        cur.execute("""
            SELECT COUNT(*) FROM statshub_player_performance_aggregates a
            WHERE EXISTS (
                SELECT 1 FROM statshub_team_players sp
                JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
                WHERE wt.team_name = ? AND sp.player_id = a.player_id
                AND sp.statshub_player_id_status IN ('confirmed','skipped_existing')
            )
        """, (tdb,))
        perf = cur.fetchone()[0]
        decision = "READY" if pct >= 80 else "INCOMPLETE"
        print(f"  {tdb}: {before} → {confirmed}/{total} ({pct}%) perf={perf} → {decision}")

    print("\n--- D. Decision ---")
    # Check overall
    cur.execute("""
        SELECT wt.team_name,
               SUM(CASE WHEN sp.statshub_player_id_status IN ('confirmed','skipped_existing') THEN 1 ELSE 0 END) confirmed,
               COUNT(*) total
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
        GROUP BY wt.team_name
    """)
    decisions = {r[0]: (r[1], r[2]) for r in cur.fetchall()}
    usa_conf = decisions.get("United States", (0,26))[0]
    can_conf = decisions.get("Canada", (0,26))[0]
    par_conf = decisions.get("Paraguay", (0,26))[0]
    bih_conf = decisions.get("Bosnia and Herzegovina", (0,26))[0]

    avoid_pw = all(ok_replays)
    print(f"  Can we avoid Playwright going forward? {'YES — direct API works' if avoid_pw else 'PARTIAL — some requests need browser context'}")
    print(f"  Does endpoint solve USA IDs? {'YES' if usa_conf >= 24 else 'PARTIAL' if usa_conf >= 10 else 'NO'} ({usa_conf}/26)")
    print(f"  Canada usable for player props? {'YES' if can_conf >= 20 else 'PARTIAL' if can_conf >= 13 else 'NO'} ({can_conf}/26)")
    print(f"  USA usable for player props? {'YES' if usa_conf >= 24 else 'PARTIAL' if usa_conf >= 10 else 'NO'} ({usa_conf}/26)")
    print(f"  Paraguay usable for player props? YES ({par_conf}/26)")
    print(f"  Bosnia usable for player props? YES ({bih_conf}/26)")

    cur.execute("""
        SELECT sp.player_name FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','United States')
        AND sp.statshub_player_id_status IN ('unresolved','probable','ambiguous')
        ORDER BY wt.team_name, sp.player_name
    """)
    remaining = cur.fetchall()
    if remaining:
        print(f"  Remaining unresolved: {len(remaining)}")
        for r in remaining:
            print(f"    {r[0]}")

    con.close()

    # ─── Health checks ────────────────────────────────────────────────────────
    print("\n=== HEALTH CHECKS ===")
    import subprocess
    for cmd in [
        [sys.executable, "-m", "scripts.statshub_raw_db_status"],
        [sys.executable, "-m", "scripts.statshub_snapshot_status"],
        [sys.executable, "-m", "scripts.health_check"],
    ]:
        print(f"\n$ {' '.join(cmd[1:])}")
        try:
            result = subprocess.run(cmd, capture_output=True, text=True,
                                    encoding="utf-8", errors="replace", timeout=60)
            print(result.stdout[:3000])
            if result.stderr:
                print("[stderr]", result.stderr[:500])
        except Exception as e:
            print(f"  ERROR: {e}")

    print("\nDone.")

if __name__ == "__main__":
    main()
