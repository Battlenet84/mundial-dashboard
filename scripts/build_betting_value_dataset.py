"""
Build betting model probabilities and value scores from StatsHub per-game data.

Steps:
1. Create DB tables: betting_odds_input, betting_model_probabilities, betting_value_scores
2. Parse raw player performance JSON → per-game prop frequencies
3. Parse team_performance_events → per-game team prop frequencies
4. Write model probabilities to DB
5. Join with betting_odds_input (if populated) → compute EV
6. Write value scores to DB + Excel
7. Write odds input template Excel
"""
import sys, json, pathlib, sqlite3
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DB_PATH   = pathlib.Path("data/mundial.db")
OUT_DIR   = pathlib.Path("data/processed/betting")
OUT_DIR.mkdir(parents=True, exist_ok=True)

NOW       = datetime.now(timezone.utc).isoformat()
MATCHES   = {
    "Canada vs Bosnia and Herzegovina": {
        "home": "Canada", "away": "Bosnia and Herzegovina",
        "home_id": "4752",  "away_id": "4479",
    },
    "United States vs Paraguay": {
        "home": "United States", "away": "Paraguay",
        "home_id": "4724",       "away_id": "4789",
    },
}
TEAMS = ["Canada", "Bosnia and Herzegovina", "United States", "Paraguay"]

SNAP_DIRS = [
    pathlib.Path("data/raw/statshub/snapshots/today_final_match_stats_probe"),
    pathlib.Path("data/raw/statshub/snapshots/today_playwright_fixture_players_probe"),
    pathlib.Path("data/raw/statshub/snapshots/today_browser_endpoint_replay_probe"),
]

# ── Player prop market definitions ────────────────────────────────────────────
# (market_type, stat_key_in_event, lines_to_generate, display_name)
PLAYER_MARKETS = [
    ("player_shots_on_target",  "onTargetScoringAttempt", [0.5, 1.5, 2.5],   "Shots on Target"),
    ("player_total_shots",      "shots",                  [0.5, 1.5, 2.5, 3.5], "Total Shots"),
    ("player_goals",            "goals",                  [0.5],             "To Score"),
    ("player_assists",          "goalAssist",             [0.5],             "To Assist"),
    ("player_fouls_committed",  "fouls",                  [0.5, 1.5],        "Fouls Committed"),
    ("player_fouled",           "wasFouled",              [0.5, 1.5],        "Times Fouled"),
    ("player_yellow_cards",     "yellowCard",             [0.5],             "Yellow Card"),
    ("player_key_passes",       "keyPass",                [0.5, 1.5],        "Key Passes"),
]

# ── Team prop market definitions ──────────────────────────────────────────────
# (market_type, col_in_events, lines, display_name, is_total_match)
TEAM_MARKETS = [
    ("team_goals",              "goals_for",        [0.5, 1.5, 2.5],   "Team Goals",        False),
    ("team_yellow_cards",       "yellow_cards",     [0.5, 1.5, 2.5, 3.5, 4.5], "Team Yellow Cards", False),
    ("team_corners",            "corners",          [3.5, 4.5, 5.5, 6.5], "Team Corners",   False),
    ("team_shots",              "shots",            [9.5, 12.5, 14.5],  "Team Total Shots",  False),
    ("team_shots_on_target",    "shots_on_target",  [2.5, 3.5, 4.5, 5.5], "Team Shots on Target", False),
    ("over_under_goals",        "goals_for",        [1.5, 2.5, 3.5, 4.5], "Match Total Goals", True),
]


def _to_float(v):
    try: return float(v)
    except: return None


def load_raw_player_events(pid):
    """Load per-game events from raw JSON; returns list of stat dicts."""
    for d in SNAP_DIRS:
        hits = list(d.glob(f"perf_{pid}*.json"))
        if hits:
            try:
                payload = json.loads(hits[0].read_text(encoding="utf-8"))
                items = (payload.get("playerStatisticsEvents") or
                         payload.get("events") or payload.get("data") or [])
                result = []
                for item in items:
                    stats = item.get("player_statistics_event") or item
                    mp = int(stats.get("minutesPlayed") or 0)
                    if mp > 0:
                        result.append(stats)
                return result
            except Exception:
                pass
    return []


# ── Schema creation ───────────────────────────────────────────────────────────
DDL = """
CREATE TABLE IF NOT EXISTS betting_odds_input (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    match_name      TEXT,
    market_type     TEXT,
    market_name     TEXT,
    selection       TEXT,
    team_name       TEXT,
    player_name     TEXT,
    player_id       TEXT,
    line            REAL,
    odds_decimal    REAL,
    bookmaker       TEXT,
    captured_at     TEXT,
    notes           TEXT,
    created_at      TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS betting_model_probabilities (
    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
    match_name           TEXT,
    market_type          TEXT,
    market_name          TEXT,
    selection            TEXT,
    team_name            TEXT,
    player_name          TEXT,
    player_id            TEXT,
    line                 REAL,
    model_probability    REAL,
    probability_method   TEXT,
    sample_size          INTEGER,
    data_quality_status  TEXT,
    notes                TEXT,
    computed_at          TEXT
);

CREATE TABLE IF NOT EXISTS betting_value_scores (
    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
    rank                 INTEGER,
    match_name           TEXT,
    market_type          TEXT,
    market_name          TEXT,
    selection            TEXT,
    team_name            TEXT,
    player_name          TEXT,
    player_id            TEXT,
    line                 REAL,
    odds_decimal         REAL,
    implied_probability  REAL,
    model_probability    REAL,
    edge                 REAL,
    expected_value       REAL,
    probability_method   TEXT,
    sample_size          INTEGER,
    data_quality_status  TEXT,
    verdict              TEXT,
    notes                TEXT,
    computed_at          TEXT
);
"""


def _match_for_team(team_name):
    for mn, md in MATCHES.items():
        if team_name in (md["home"], md["away"]):
            return mn
    return "Unknown"


# ── Compute model probabilities ───────────────────────────────────────────────
def compute_player_probabilities(cur):
    rows = []
    for team in TEAMS:
        match_name = _match_for_team(team)
        cur.execute("""
            SELECT sp.player_name, sp.player_id, sp.jersey_number
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id=wt.statshub_team_id
            WHERE wt.team_name=?
              AND sp.statshub_player_id_status IN ('confirmed','skipped_existing')
              AND sp.player_id IS NOT NULL
            ORDER BY CAST(sp.jersey_number AS INTEGER)
        """, (team,))
        players = cur.fetchall()

        for pname, pid, jersey in players:
            events = load_raw_player_events(pid)
            n_apps = len(events)
            if n_apps == 0:
                continue

            for mtype, stat_key, lines, disp in PLAYER_MARKETS:
                # collect stat values
                vals = []
                stat_missing = 0
                for ev in events:
                    v = ev.get(stat_key)
                    if v is None:
                        stat_missing += 1
                        vals.append(0)
                    else:
                        vals.append(_to_float(v) or 0)

                effective_n = n_apps - stat_missing
                if effective_n < 3:
                    dq = "insufficient_data"
                elif effective_n < 10:
                    dq = "low_sample"
                else:
                    dq = "ok"

                for line in lines:
                    hits = sum(1 for v in vals if v >= line)
                    prob = round(hits / n_apps, 4) if n_apps > 0 else None
                    selection = f"Over {line}" if line != 0.5 or mtype not in ("player_goals","player_assists","player_yellow_cards") else "Yes"
                    notes_parts = []
                    if stat_missing > 0:
                        notes_parts.append(f"{stat_missing} events missing {stat_key}")
                    rows.append({
                        "match_name": match_name,
                        "market_type": mtype,
                        "market_name": f"{pname} - {disp} {line}+",
                        "selection": selection,
                        "team_name": team,
                        "player_name": pname,
                        "player_id": str(pid),
                        "line": line,
                        "model_probability": prob,
                        "probability_method": "historical_frequency",
                        "sample_size": n_apps,
                        "data_quality_status": dq,
                        "notes": "; ".join(notes_parts) if notes_parts else None,
                    })
    return rows


def compute_team_probabilities(cur):
    rows = []
    for team in TEAMS:
        match_name = _match_for_team(team)
        cur.execute("""
            SELECT goals_for, goals_against, yellow_cards, corners,
                   shots, shots_on_target
            FROM statshub_team_performance_events
            WHERE team_name=?
            ORDER BY event_date DESC
            LIMIT 50
        """, (team,))
        events = cur.fetchall()
        n = len(events)
        if n == 0:
            continue

        for mtype, col, lines, disp, is_total in TEAM_MARKETS:
            idx_map = {"goals_for":0,"goals_against":1,"yellow_cards":2,
                       "corners":3,"shots":4,"shots_on_target":5}
            idx = idx_map[col]

            if is_total:
                # over/under goals uses home+away combined — we only have per-team
                # approximate as 2x team avg (note this in quality)
                vals = []
                for ev in events:
                    gf = ev[0] or 0
                    ga = ev[1] or 0
                    vals.append(gf + ga)
                dq = "warning"
                method = "aggregate_estimate"
                note_base = "Match total estimated from single-team perspective (home+away goals in team games)"
            else:
                vals = [ev[idx] or 0 for ev in events]
                dq = "ok" if n >= 20 else "low_sample"
                method = "historical_frequency"
                note_base = None

            for line in lines:
                hits = sum(1 for v in vals if v >= line)
                prob = round(hits / n, 4)
                selection_label = f"Over {line}"
                notes_parts = [note_base] if note_base else []
                rows.append({
                    "match_name": match_name,
                    "market_type": mtype,
                    "market_name": f"{team} - {disp} {line}+",
                    "selection": selection_label,
                    "team_name": team,
                    "player_name": None,
                    "player_id": None,
                    "line": line,
                    "model_probability": prob,
                    "probability_method": method,
                    "sample_size": n,
                    "data_quality_status": dq,
                    "notes": notes_parts[0] if notes_parts else None,
                })
    return rows


def compute_data_quality_warnings(cur):
    """Return records for unresolved/probable players as quality warnings."""
    rows = []
    for team in TEAMS:
        match_name = _match_for_team(team)
        cur.execute("""
            SELECT sp.player_name, sp.player_id, sp.statshub_player_id_status, sp.jersey_number
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id=wt.statshub_team_id
            WHERE wt.team_name=?
              AND sp.statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        """, (team,))
        for pname, pid, status, jersey in cur.fetchall():
            rows.append({
                "match_name": match_name,
                "market_type": "data_quality",
                "market_name": f"{pname} - Status: {status}",
                "selection": "N/A",
                "team_name": team,
                "player_name": pname,
                "player_id": str(pid) if pid else None,
                "line": None,
                "model_probability": None,
                "probability_method": None,
                "sample_size": None,
                "data_quality_status": f"excluded:{status}",
                "notes": f"Player excluded from props; id_status={status}",
            })
    return rows


def insert_model_probabilities(cur, rows):
    cur.execute("DELETE FROM betting_model_probabilities")
    cur.executemany("""
        INSERT INTO betting_model_probabilities
            (match_name, market_type, market_name, selection, team_name,
             player_name, player_id, line, model_probability, probability_method,
             sample_size, data_quality_status, notes, computed_at)
        VALUES (:match_name,:market_type,:market_name,:selection,:team_name,
                :player_name,:player_id,:line,:model_probability,:probability_method,
                :sample_size,:data_quality_status,:notes,:computed_at)
    """, [{**r, "computed_at": NOW} for r in rows])


def compute_value_scores(cur):
    """Join model_probabilities with odds_input, compute EV."""
    cur.execute("""
        SELECT mp.id, mp.match_name, mp.market_type, mp.market_name, mp.selection,
               mp.team_name, mp.player_name, mp.player_id, mp.line,
               mp.model_probability, mp.probability_method,
               mp.sample_size, mp.data_quality_status, mp.notes,
               oi.odds_decimal, oi.bookmaker
        FROM betting_model_probabilities mp
        LEFT JOIN betting_odds_input oi
            ON oi.match_name = mp.match_name
           AND oi.market_type = mp.market_type
           AND oi.line = mp.line
           AND (
               (oi.player_id IS NOT NULL AND oi.player_id = mp.player_id)
               OR (oi.player_id IS NULL AND mp.player_id IS NULL
                   AND oi.team_name = mp.team_name)
           )
        WHERE mp.market_type != 'data_quality'
        ORDER BY mp.match_name, mp.market_type, mp.player_name, mp.line
    """)
    score_rows = []
    for r in cur.fetchall():
        (_, match_name, mtype, mname, selection, team_name, player_name, player_id,
         line, model_prob, method, sample_size, dq, notes, odds, bookmaker) = r

        if odds is None or model_prob is None:
            score_rows.append({
                "match_name": match_name, "market_type": mtype, "market_name": mname,
                "selection": selection, "team_name": team_name, "player_name": player_name,
                "player_id": player_id, "line": line, "odds_decimal": None,
                "implied_probability": None, "model_probability": model_prob,
                "edge": None, "expected_value": None,
                "probability_method": method, "sample_size": sample_size,
                "data_quality_status": dq, "verdict": "NO_ODDS",
                "notes": notes,
            })
            continue

        implied_prob = round(1.0 / odds, 4)
        edge         = round(model_prob - implied_prob, 4)
        ev           = round(model_prob * odds - 1.0, 4)

        if dq in ("warning", "partial", "insufficient_data"):
            verdict = "REVIEW"
        elif ev > 0:
            verdict = "VALUE"
        else:
            verdict = "NO_VALUE"

        score_rows.append({
            "match_name": match_name, "market_type": mtype, "market_name": mname,
            "selection": selection, "team_name": team_name, "player_name": player_name,
            "player_id": player_id, "line": line, "odds_decimal": odds,
            "implied_probability": implied_prob, "model_probability": model_prob,
            "edge": edge, "expected_value": ev,
            "probability_method": method, "sample_size": sample_size,
            "data_quality_status": dq, "verdict": verdict,
            "notes": notes,
        })

    # rank by ev descending (None ev goes last)
    def ev_key(r):
        ev = r["expected_value"]
        return ev if ev is not None else -999
    score_rows.sort(key=ev_key, reverse=True)
    for i, r in enumerate(score_rows, 1):
        r["rank"] = i

    cur.execute("DELETE FROM betting_value_scores")
    cur.executemany("""
        INSERT INTO betting_value_scores
            (rank, match_name, market_type, market_name, selection, team_name,
             player_name, player_id, line, odds_decimal, implied_probability,
             model_probability, edge, expected_value, probability_method,
             sample_size, data_quality_status, verdict, notes, computed_at)
        VALUES (:rank,:match_name,:market_type,:market_name,:selection,:team_name,
                :player_name,:player_id,:line,:odds_decimal,:implied_probability,
                :model_probability,:edge,:expected_value,:probability_method,
                :sample_size,:data_quality_status,:verdict,:notes,:computed_at)
    """, [{**r, "computed_at": NOW} for r in score_rows])
    return score_rows


# ── Excel output ──────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F497D")
HDR_FONT  = Font(color="FFFFFF", bold=True)
VALUE_FILL  = PatternFill("solid", fgColor="C6EFCE")
REVIEW_FILL = PatternFill("solid", fgColor="FFEB9C")
NO_VAL_FILL = PatternFill("solid", fgColor="FFC7CE")

def _hdr(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = Alignment(horizontal="center")


def write_odds_template(wb, cur):
    ws = wb.create_sheet("OddsInput")
    cols = ["match_name","market_type","market_name","selection","team_name",
            "player_name","player_id","line","odds_decimal","bookmaker",
            "captured_at","notes"]
    _hdr(ws, cols)
    # Pre-populate with all model probability rows as placeholders
    cur.execute("""
        SELECT match_name, market_type, market_name, selection, team_name,
               player_name, player_id, line
        FROM betting_model_probabilities
        WHERE market_type != 'data_quality'
        ORDER BY match_name, team_name, market_type, player_name, line
    """)
    for r in cur.fetchall():
        ws.append(list(r) + [None, None, None, None])
    # Column widths
    col_widths = [35,25,55,12,25,35,12,8,14,20,25,30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_value_scores_sheet(wb, score_rows):
    ws = wb.create_sheet("ValueScores")
    cols = ["rank","match_name","market_type","selection","team_name","player_name",
            "line","model_probability","odds_decimal","implied_probability",
            "edge","expected_value","sample_size","data_quality_status","verdict","notes"]
    _hdr(ws, cols)
    for r in score_rows:
        row = [
            r["rank"], r["match_name"], r["market_type"], r["selection"],
            r["team_name"], r.get("player_name"), r["line"],
            r["model_probability"], r["odds_decimal"], r["implied_probability"],
            r["edge"], r["expected_value"], r["sample_size"],
            r["data_quality_status"], r["verdict"], r["notes"],
        ]
        ws.append(row)
        v = r["verdict"]
        fill = VALUE_FILL if v == "VALUE" else (REVIEW_FILL if v == "REVIEW" else None)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    # Freeze header and set widths
    ws.freeze_panes = "A2"
    widths = [6,35,25,12,25,30,8,16,14,18,10,14,12,20,12,40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_model_probs_sheet(wb, cur):
    ws = wb.create_sheet("ModelProbabilities")
    cols = ["match_name","market_type","market_name","selection","team_name",
            "player_name","player_id","line","model_probability","probability_method",
            "sample_size","data_quality_status","notes"]
    _hdr(ws, cols)
    cur.execute("""
        SELECT match_name, market_type, market_name, selection, team_name,
               player_name, player_id, line, model_probability, probability_method,
               sample_size, data_quality_status, notes
        FROM betting_model_probabilities
        WHERE market_type != 'data_quality'
        ORDER BY match_name, team_name, market_type, player_name, line
    """)
    for r in cur.fetchall():
        ws.append(list(r))
    ws.freeze_panes = "A2"
    widths = [35,25,55,12,25,35,12,8,16,22,12,20,30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_data_quality_sheet(wb, cur):
    ws = wb.create_sheet("DataQuality")
    _hdr(ws, ["match_name","team_name","player_name","player_id",
               "data_quality_status","notes"])
    cur.execute("""
        SELECT match_name, team_name, player_name, player_id,
               data_quality_status, notes
        FROM betting_model_probabilities
        WHERE market_type = 'data_quality'
        ORDER BY match_name, team_name
    """)
    for r in cur.fetchall():
        ws.append(list(r))
    # Also add players with no raw events
    for team in TEAMS:
        cur.execute("""
            SELECT sp.player_name, sp.player_id, sp.statshub_player_id_status
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id=wt.statshub_team_id
            WHERE wt.team_name=?
              AND sp.statshub_player_id_status IN ('confirmed','skipped_existing')
              AND sp.player_id IS NOT NULL
        """, (team,))
        match_name = _match_for_team(team)
        for pname, pid, status in cur.fetchall():
            events = load_raw_player_events(pid)
            if len(events) == 0:
                ws.append([match_name, team, pname, str(pid),
                           "no_events", "0 qualifying game appearances in raw data"])


def write_referee_sheet(wb, cur):
    ws = wb.create_sheet("Referees")
    _hdr(ws, ["match_name","referee_name","games","yellow_cards","red_cards",
               "avg_cards_per_game","avg_yc_per_game","avg_rc_per_game",
               "o35_cards_pct","o45_cards_pct","both_teams_card_pct"])
    cur.execute("SELECT match_name, referee_name, notes FROM statshub_match_referees")
    for row in cur.fetchall():
        mname, rname, notes_str = row
        # Pull from raw_referees
        cur.execute("SELECT raw_json FROM statshub_raw_referees WHERE referee_name=? LIMIT 1", (rname,))
        rr = cur.fetchone()
        if rr and rr[0]:
            try:
                d = json.loads(rr[0])
                ws.append([mname, rname,
                    d.get("games"), d.get("yellow_cards"), d.get("red_cards"),
                    d.get("avg_cards_per_game"), d.get("avg_yellow_cards_per_game"),
                    d.get("avg_red_cards_per_game"),
                    d.get("o35_cards_pct"), d.get("o45_cards_pct"),
                    d.get("both_teams_card_pct")])
            except Exception:
                ws.append([mname, rname] + [None]*9)
        else:
            # parse inline notes
            ws.append([mname, rname] + [None]*9)


def main():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # ── Create tables ─────────────────────────────────────────────────────────
    print("=== Step 1: Create betting tables ===")
    for stmt in DDL.split(";\n\n"):
        stmt = stmt.strip()
        if stmt:
            con.execute(stmt)
    con.commit()
    print("  Tables: betting_odds_input, betting_model_probabilities, betting_value_scores ✓")

    # ── Compute model probabilities ────────────────────────────────────────────
    print("\n=== Step 2: Compute player prop probabilities ===")
    player_rows = compute_player_probabilities(cur)
    print(f"  {len(player_rows)} player-prop probability rows")

    print("\n=== Step 3: Compute team prop probabilities ===")
    team_rows = compute_team_probabilities(cur)
    print(f"  {len(team_rows)} team-prop probability rows")

    print("\n=== Step 4: Data quality entries ===")
    dq_rows = compute_data_quality_warnings(cur)
    print(f"  {len(dq_rows)} data quality warning rows")

    all_rows = player_rows + team_rows + dq_rows
    insert_model_probabilities(cur, all_rows)
    con.commit()
    print(f"  Total: {len(all_rows)} rows inserted into betting_model_probabilities")

    # ── Value scores ──────────────────────────────────────────────────────────
    print("\n=== Step 5: Compute value scores (join with odds) ===")
    score_rows = compute_value_scores(cur)
    con.commit()
    odds_rows   = sum(1 for r in score_rows if r["odds_decimal"] is not None)
    no_odds     = sum(1 for r in score_rows if r["odds_decimal"] is None)
    value_rows  = sum(1 for r in score_rows if r["verdict"] == "VALUE")
    print(f"  {len(score_rows)} total rows | {odds_rows} with odds | {no_odds} awaiting odds")
    print(f"  {value_rows} VALUE opportunities")

    # ── Excel output ──────────────────────────────────────────────────────────
    print("\n=== Step 6: Write Excel ===")

    # Input template
    wb_in = openpyxl.Workbook()
    wb_in.remove(wb_in.active)
    write_odds_template(wb_in, cur)
    write_model_probs_sheet(wb_in, cur)
    write_referee_sheet(wb_in, cur)
    write_data_quality_sheet(wb_in, cur)
    template_path = OUT_DIR / "today_value_dashboard_input.xlsx"
    wb_in.save(template_path)
    print(f"  Template: {template_path}")

    # Value scores output
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    write_value_scores_sheet(wb_out, score_rows)
    write_model_probs_sheet(wb_out, cur)
    write_referee_sheet(wb_out, cur)
    write_data_quality_sheet(wb_out, cur)
    scores_path = OUT_DIR / "today_value_scores.xlsx"
    wb_out.save(scores_path)
    print(f"  Scores:   {scores_path}")

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n============================================================")
    print("BETTING DATASET SUMMARY")
    print("============================================================")
    for mn in MATCHES:
        cur.execute("""
            SELECT COUNT(DISTINCT player_name) players, COUNT(*) total_rows
            FROM betting_model_probabilities
            WHERE match_name=? AND market_type != 'data_quality'
        """, (mn,))
        row = cur.fetchone()
        print(f"  {mn}: {row[0]} players, {row[1]} model probability rows")

    cur.execute("SELECT COUNT(*) FROM betting_odds_input")
    print(f"  Odds loaded: {cur.fetchone()[0]} (fill OddsInput sheet and re-run to get EV)")

    con.close()
    print("\nDone. Next step: fill OddsInput sheet in today_value_dashboard_input.xlsx, then run again.")


if __name__ == "__main__":
    main()
