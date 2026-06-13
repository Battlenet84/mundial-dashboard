"""
Fetch live/API odds for United States vs Paraguay from Odds-API.io.

Live mode (default):
    python -m scripts.fetch_usa_paraguay_live_api_odds

Rebuild from existing raw files (no API calls):
    python -m scripts.fetch_usa_paraguay_live_api_odds --from-raw

Run name: usa_paraguay_live_api_odds_probe
"""
from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.odds_driven import (
    calculate_ev,
    connect,
    ensure_schema,
    insert_raw_rows,
    normalize_raw_odds,
)
from scripts.probe_odds_api_io import is_configured_api_key, load_config


RUN_NAME = "usa_paraguay_live_api_odds_probe"
STATSHUB_EVENT_ID = "15186873"
TARGET_HOME = "United States"
TARGET_HOME_ID = "4724"
TARGET_AWAY = "Paraguay"
TARGET_AWAY_ID = "4789"
TARGET_MATCH = f"{TARGET_HOME} vs {TARGET_AWAY}"
SPORT = "football"
PREFERRED_BOOKMAKER = "Bet365"
FALLBACK_BOOKMAKER = "DraftKings"
CANDIDATE_EVENT_ID = "66456940"
TIMEOUT_SECONDS = 20

RAW_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/usa_paraguay_live_api_odds_value_scores.xlsx")

# API returns "USA" — aliases must cover it
HOME_ALIASES = ("united states", "usa", "usmnt", "eeuu", "estados unidos", "u.s.")
AWAY_ALIASES = ("paraguay", "par")

EXPECTED_MARKETS = [
    "ML", "ML HT", "spread", "handicap", "spread HT", "totals", "over under goals",
    "totals HT", "alternate spreads", "alternate totals", "BTTS", "double chance",
    "draw no bet", "team totals", "cards", "corners", "shots", "shots on target",
    "player props", "anytime goalscorer", "player shots", "player shots on target",
    "player assists", "player cards", "player fouls", "goalkeeper saves",
]


# ---------------------------------------------------------------------------
# Serialization helpers
# ---------------------------------------------------------------------------

def sqlite_safe(value: Any) -> Any:
    """Convert dict/list to JSON string so SQLite can bind it."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


# ---------------------------------------------------------------------------
# API helpers (live mode only)
# ---------------------------------------------------------------------------

def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    safe = {k: v for k, v in params.items() if k.lower() != "apikey"}
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    return f"{url}?{urlencode(safe)}" if safe else url


def save_response(
    action: str,
    params: dict[str, Any],
    status_code: int | None,
    payload: Any,
    error: str | None,
    base_url: str,
) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    safe_action = action.replace("/", "_")
    path = RAW_DIR / f"{safe_action}_{stamp()}.json"
    wrapper = {
        "fetched_at_utc": now_utc(),
        "run_name": RUN_NAME,
        "action": action,
        "url_without_api_key": public_url(base_url, action, params),
        "params_without_api_key": {k: v for k, v in params.items() if k.lower() != "apikey"},
        "status_code": status_code,
        "response_json": payload,
        "error": error,
    }
    path.write_text(json.dumps(wrapper, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def api_get(endpoint: str, params: dict[str, Any], base_url: str) -> dict[str, Any]:
    status_code = None
    payload: Any = None
    error = None
    try:
        response = requests.get(
            f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}",
            params=params,
            timeout=TIMEOUT_SECONDS,
        )
        status_code = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"response_text_preview": response.text[:1000]}
        if response.status_code >= 400:
            error = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        error = str(exc)
    raw_file = save_response(endpoint, params, status_code, payload, error, base_url)
    return {
        "endpoint": endpoint,
        "params": params,
        "status_code": status_code,
        "payload": payload,
        "error": error,
        "raw_file": raw_file,
    }


def payload_count(payload: Any) -> int:
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("data", "events", "response", "results", "bookmakers", "leagues"):
            if isinstance(payload.get(key), list):
                return len(payload[key])
    return 1 if payload else 0


# ---------------------------------------------------------------------------
# Load from existing raw files
# ---------------------------------------------------------------------------

def _unwrap(raw: dict[str, Any]) -> Any:
    """Extract response_json from our file wrapper, or return raw as-is."""
    if "response_json" in raw:
        return raw["response_json"]
    return raw


def load_raw_files() -> dict[str, Any]:
    """
    Scan RAW_DIR for known file patterns and return a dict:
      {action_key: {status_code, payload, raw_file, error}}
    where action_key is one of: bookmakers_selected, events, odds, odds_multi
    """
    if not RAW_DIR.exists():
        return {}
    result: dict[str, Any] = {}
    patterns = {
        "bookmakers_selected": "bookmakers_selected",
        "events": "events_",
        "odds_multi": "odds_multi",
        "odds": "odds_",
    }
    for key, prefix in patterns.items():
        matches = sorted(RAW_DIR.glob(f"{prefix}*.json"), reverse=True)
        # exclude odds_multi from odds matches
        if key == "odds":
            matches = [f for f in matches if "multi" not in f.name]
        if not matches:
            continue
        path = matches[0]  # most recent
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            result[key] = {"status_code": None, "payload": None, "raw_file": path, "error": str(exc)}
            continue
        payload = _unwrap(data)
        status_code = data.get("status_code", 200)
        result[key] = {
            "status_code": status_code or 200,
            "payload": payload,
            "raw_file": path,
            "error": data.get("error"),
            "fetched_at_utc": data.get("fetched_at_utc", now_utc()),
        }
    return result


# ---------------------------------------------------------------------------
# Event matching
# ---------------------------------------------------------------------------

def text_has(text: Any, aliases: tuple[str, ...]) -> bool:
    value = str(text or "").lower()
    return any(alias in value for alias in aliases)


def _str_field(v: Any) -> str | None:
    """Extract string from a field that might be a dict (e.g. sport/league)."""
    if v is None:
        return None
    if isinstance(v, str):
        return v
    if isinstance(v, dict):
        return v.get("name") or v.get("slug") or json.dumps(v, ensure_ascii=False)
    return str(v)


def find_target_event(
    events: Any,
) -> tuple[dict[str, Any] | None, list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (best_match, accepted_brief_list, rejected_brief_list)."""
    items = (
        events
        if isinstance(events, list)
        else events.get("data", [])
        if isinstance(events, dict)
        else []
    )
    candidates: list[tuple[dict[str, Any], dict[str, Any]]] = []
    rejected: list[dict[str, Any]] = []

    for event in items:
        home = event.get("home") or event.get("homeTeam") or event.get("home_team")
        away = event.get("away") or event.get("awayTeam") or event.get("away_team")
        eid = event.get("id") or event.get("eventId") or event.get("event_id")
        id_match = str(eid) == CANDIDATE_EVENT_ID
        is_match = (text_has(home, HOME_ALIASES) and text_has(away, AWAY_ALIASES)) or (
            text_has(home, AWAY_ALIASES) and text_has(away, HOME_ALIASES)
        )
        # sport and league may be dicts — always serialize to str for storage
        brief: dict[str, Any] = {
            "api_event_id": str(eid) if eid is not None else None,
            "event_name": f"{home} vs {away}",
            "home_team": str(home or ""),
            "away_team": str(away or ""),
            "start_time": sqlite_safe(
                event.get("date") or event.get("eventDate") or event.get("startTime")
            ),
            "sport": sqlite_safe(_str_field(event.get("sport"))),
            "league": sqlite_safe(_str_field(event.get("league"))),
            "confidence": 1.0 if is_match else (0.5 if id_match else 0.0),
            "selected": False,
            "notes": "",
        }
        if is_match:
            brief["notes"] = (
                "team names matched + candidate event_id confirmed"
                if id_match
                else "team names matched"
            )
            candidates.append((event, brief))
        elif any(
            text_has(v, HOME_ALIASES + AWAY_ALIASES)
            for v in (home, away, _str_field(event.get("league")))
        ):
            brief["notes"] = "near-miss: partial team/league name match"
            rejected.append(brief)

    if candidates:
        candidates[0][1]["selected"] = True
    return (
        candidates[0][0] if candidates else None,
        [c[1] for c in candidates],
        rejected,
    )


# ---------------------------------------------------------------------------
# Odds parsing
# ---------------------------------------------------------------------------

def price_rows_from_outcome(
    market_name: str, outcome: dict[str, Any], home: str, away: str
) -> list[tuple[str, float | None, float]]:
    rows: list[tuple[str, float | None, float]] = []
    line = (
        outcome.get("line")
        or outcome.get("point")
        or outcome.get("handicap")
        or outcome.get("hdp")
    )
    for key, label in (("home", home), ("away", away), ("draw", "Draw")):
        if outcome.get(key) not in (None, ""):
            try:
                rows.append((label, line, float(outcome[key])))
            except (TypeError, ValueError):
                pass
    sel_label = outcome.get("label") or outcome.get("name") or outcome.get("selection")
    for side_key, side_str in (("over", "Over"), ("under", "Under")):
        if outcome.get(side_key) not in (None, ""):
            selection = (
                sel_label
                if sel_label and "chance" in market_name.lower()
                else f"{side_str} {line}".strip()
            )
            try:
                rows.append((selection, line, float(outcome[side_key])))
            except (TypeError, ValueError):
                pass
    # Generic price/odds/decimal fallback
    if not rows:
        for key in ("price", "odds", "decimal"):
            if outcome.get(key) not in (None, ""):
                try:
                    rows.append((
                        str(
                            sel_label
                            or outcome.get("team")
                            or outcome.get("player")
                            or market_name
                        ),
                        line,
                        float(outcome[key]),
                    ))
                    break
                except (TypeError, ValueError):
                    pass
    return rows


def _markets_for_bookmaker(payload: dict[str, Any], bookmaker: str) -> list[Any]:
    bookmakers = payload.get("bookmakers", {})
    if isinstance(bookmakers, dict):
        return bookmakers.get(bookmaker, [])
    if isinstance(bookmakers, list):
        for book in bookmakers:
            if isinstance(book, dict):
                name = str(book.get("name") or book.get("key") or "")
                if name.lower() == bookmaker.lower():
                    return book.get("markets", book.get("odds", []))
    return []


def raw_rows_from_odds_payload(
    payload: dict[str, Any],
    bookmaker: str,
    raw_file: Path,
    request_id: str,
) -> list[dict[str, Any]]:
    home = str(payload.get("home") or TARGET_HOME)
    away = str(payload.get("away") or TARGET_AWAY)
    api_event_id = str(payload.get("id") or CANDIDATE_EVENT_ID)
    captured_at = now_utc()
    markets = _markets_for_bookmaker(payload, bookmaker)
    rows: list[dict[str, Any]] = []
    for market_index, market in enumerate(markets):
        if not isinstance(market, dict):
            continue
        market_name = market.get("name") or market.get("key") or f"market_{market_index}"
        outcomes = (
            market.get("odds")
            if isinstance(market.get("odds"), list)
            else market.get("outcomes", [])
        )
        for odds_index, outcome in enumerate(outcomes or []):
            if not isinstance(outcome, dict):
                continue
            for selection, line, odds_val in price_rows_from_outcome(market_name, outcome, home, away):
                rows.append({
                    "run_name": RUN_NAME,
                    "source_name": "Odds-API.io",
                    "bookmaker": bookmaker,
                    "match_name": TARGET_MATCH,
                    "event_id": api_event_id,
                    "api_event_id": api_event_id,
                    "statshub_event_id": STATSHUB_EVENT_ID,
                    "raw_market_group": str(market_name),
                    "raw_market_name": str(market_name),
                    "raw_selection_name": str(selection),
                    "raw_line": line,
                    "raw_odds": odds_val,
                    "odds_format": "decimal",
                    "captured_at": captured_at,
                    "raw_payload": json.dumps(
                        {
                            "market_index": market_index,
                            "odds_index": odds_index,
                            "market": market,
                            "outcome": outcome,
                        },
                        ensure_ascii=False,
                    ),
                    "source_url": str(payload.get("urls", {}).get(bookmaker) or ""),
                    "request_id": str(request_id),
                    "raw_file": str(raw_file),
                    "status": "raw",
                    "notes": RUN_NAME,
                })
    return rows


def extract_market_groups(payload: dict[str, Any], bookmaker: str) -> list[dict[str, Any]]:
    home = str(payload.get("home") or TARGET_HOME)
    away = str(payload.get("away") or TARGET_AWAY)
    markets = _markets_for_bookmaker(payload, bookmaker)
    groups = []
    for market in markets:
        if not isinstance(market, dict):
            continue
        name = market.get("name") or market.get("key") or "unknown"
        odds_list = (
            market.get("odds")
            if isinstance(market.get("odds"), list)
            else market.get("outcomes", [])
        )
        row_count = sum(
            len(price_rows_from_outcome(str(name), outcome, home, away))
            for outcome in (odds_list or [])
            if isinstance(outcome, dict)
        )
        groups.append({"market_key": str(name), "market_name": str(name), "rows_returned": row_count})
    return groups


# ---------------------------------------------------------------------------
# Bookmaker selection (live mode only)
# ---------------------------------------------------------------------------

def get_selected_bookmakers(api_key: str, base_url: str) -> dict[str, Any]:
    result = api_get("bookmakers/selected", {"apiKey": api_key}, base_url)
    books: list[str] = []
    payload = result["payload"]
    if isinstance(payload, list):
        books = [
            str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
            for b in payload
        ]
    elif isinstance(payload, dict):
        items = payload.get("data") or payload.get("bookmakers") or []
        if isinstance(items, list):
            books = [
                str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
                for b in items
            ]
    return {
        "selected_bookmakers": books,
        "status_code": result["status_code"],
        "error": result["error"],
        "raw_file": result["raw_file"],
    }


def pick_bookmaker(selected: list[str]) -> str:
    canonical = [b.lower() for b in selected]
    if any("bet365" in b for b in canonical):
        return PREFERRED_BOOKMAKER
    if any("draftkings" in b for b in canonical):
        return FALLBACK_BOOKMAKER
    return PREFERRED_BOOKMAKER  # default when state unknown


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def ensure_event_candidates_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS betting_event_candidates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_name TEXT,
            api_event_id TEXT,
            event_name TEXT,
            home_team TEXT,
            away_team TEXT,
            start_time TEXT,
            sport TEXT,
            league TEXT,
            confidence REAL,
            selected INTEGER,
            notes TEXT,
            captured_at TEXT
        )
    """)
    con.commit()


def store_event_candidates(
    con: sqlite3.Connection,
    candidates: list[dict[str, Any]],
    rejected: list[dict[str, Any]],
) -> None:
    ensure_event_candidates_table(con)
    con.execute("DELETE FROM betting_event_candidates WHERE run_name=?", (RUN_NAME,))
    all_items = candidates + rejected
    captured_at = now_utc()
    for c in all_items:
        con.execute(
            """INSERT INTO betting_event_candidates
               (run_name, api_event_id, event_name, home_team, away_team, start_time,
                sport, league, confidence, selected, notes, captured_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                RUN_NAME,
                sqlite_safe(c.get("api_event_id")),
                sqlite_safe(c.get("event_name")),
                sqlite_safe(c.get("home_team")),
                sqlite_safe(c.get("away_team")),
                sqlite_safe(c.get("start_time")),
                sqlite_safe(c.get("sport")),
                sqlite_safe(c.get("league")),
                float(c.get("confidence") or 0.0),
                1 if c.get("selected") else 0,
                sqlite_safe(c.get("notes")),
                captured_at,
            ),
        )
    con.commit()


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def table_rows(
    con: sqlite3.Connection, sql: str, params: tuple = ()
) -> tuple[list[str], list[sqlite3.Row]]:
    cursor = con.execute(sql, params)
    rows = cursor.fetchall()
    headers = list(rows[0].keys()) if rows else [desc[0] for desc in cursor.description]
    return headers, rows


def add_sheet(
    wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any]
) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([row[h] for h in headers])
        elif isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(
            12, min(48, len(str(header)) + 4)
        )


def write_workbook(
    con: sqlite3.Connection,
    run_summary: dict[str, Any],
    event_candidates: list[dict[str, Any]],
    market_discovery: list[dict[str, Any]],
    raw_sources: list[dict[str, Any]],
    failed_requests: list[dict[str, Any]],
) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb, "run_summary", ["metric", "value"],
        [{"metric": k, "value": str(v)} for k, v in run_summary.items()],
    )

    ec_headers = [
        "api_event_id", "event_name", "home_team", "away_team", "start_time",
        "sport", "league", "confidence", "selected", "notes",
    ]
    add_sheet(wb, "event_candidates", ec_headers, event_candidates)

    md_headers = ["bookmaker", "raw_market_group", "raw_market_name", "rows_returned", "status", "notes"]
    add_sheet(wb, "market_discovery", md_headers, market_discovery)

    for sheet_title, sql in [
        (
            "ev_ranking",
            "SELECT rank, match_name, bookmaker, market_type, raw_market_name, "
            "raw_selection_name, team_name, player_name, line, odds_decimal, "
            "implied_probability, model_probability, edge, expected_value, "
            "probability_method, sample_size, verdict, notes "
            "FROM betting_value_scores_new "
            "WHERE expected_value IS NOT NULL ORDER BY expected_value DESC",
        ),
        ("raw_odds", "SELECT * FROM betting_odds_raw ORDER BY id"),
        ("normalized_markets", "SELECT * FROM betting_odds_normalized ORDER BY id"),
        (
            "player_props_found",
            "SELECT * FROM betting_odds_normalized WHERE market_type LIKE 'player_%' ORDER BY id",
        ),
        (
            "unsupported_markets",
            "SELECT * FROM betting_value_scores_new WHERE verdict='UNSUPPORTED' ORDER BY id",
        ),
        (
            "unmatched_selections",
            "SELECT * FROM betting_value_scores_new WHERE verdict='UNMATCHED' ORDER BY id",
        ),
    ]:
        headers, rows = table_rows(con, sql)
        add_sheet(wb, sheet_title, headers, rows)

    unresolved = con.execute(
        """
        SELECT team_name, player_name, player_id, statshub_player_id_status, notes
        FROM statshub_team_players
        WHERE team_name IN (?, ?)
          AND statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        ORDER BY team_name, player_name
        """,
        (TARGET_HOME, TARGET_AWAY),
    ).fetchall()
    dq_rows: list[dict[str, Any]] = [dict(r) for r in unresolved]
    dq_rows.extend({"issue": "failed_api_request", **item} for item in failed_requests)
    dq_headers = sorted({k for row in dq_rows for k in row}) if dq_rows else ["issue"]
    add_sheet(wb, "data_quality", dq_headers, dq_rows)

    rs_headers = ["raw_file", "endpoint", "status_code", "rows_or_items", "error", "captured_at"]
    add_sheet(wb, "raw_sources", rs_headers, raw_sources)

    wb.save(OUT_XLSX)
    return OUT_XLSX


# ---------------------------------------------------------------------------
# Core pipeline (shared by live and --from-raw modes)
# ---------------------------------------------------------------------------

def run_pipeline(
    events_payload: Any,
    odds_payload: dict[str, Any],
    bookmaker: str,
    odds_raw_file: Path,
    raw_sources: list[dict[str, Any]],
    selected_before: list[str],
    selected_after: list[str],
    attempted: int,
    successful: int,
    failed: int,
    rate_limit_hit: bool,
    bookmaker_check_status: str,
    from_raw: bool,
) -> None:
    target_event, candidates, rejected = find_target_event(events_payload)

    if not target_event:
        print("  Target event NOT FOUND in events payload")
        with connect() as con:
            ensure_schema(con)
            store_event_candidates(con, candidates, rejected)
            run_summary = {
                "target_match": TARGET_MATCH,
                "target_match_found": "no",
                "source_name": "Odds-API.io",
                "mode": "from_raw" if from_raw else "live",
                "selected_bookmakers_before": ", ".join(selected_before) or "(unknown)",
                "selected_bookmakers_after": ", ".join(selected_after) or "(same)",
                "bookmaker_used": bookmaker,
                "api_event_id": "not_found",
                "statshub_event_id": STATSHUB_EVENT_ID,
                "raw odds rows ingested": 0,
                "requests attempted/successful/failed": f"{attempted}/{successful}/{failed}",
                "rate_limit_status": "429_hit" if rate_limit_hit else "ok",
                "stop_reason": "target match not found in events payload",
            }
            write_workbook(con, run_summary, candidates + rejected, [], raw_sources, [])
        raise SystemExit(f"Target match {TARGET_MATCH} not found in events payload.")

    api_event_id = str(
        target_event.get("id")
        or target_event.get("eventId")
        or target_event.get("event_id")
        or CANDIDATE_EVENT_ID
    )
    print(f"  Target event: {target_event.get('home')} vs {target_event.get('away')} (id={api_event_id})")

    request_id = odds_raw_file.stem
    raw_rows = raw_rows_from_odds_payload(odds_payload, bookmaker, odds_raw_file, request_id)
    market_groups = extract_market_groups(odds_payload, bookmaker)

    # If preferred bookmaker returned nothing, try Bet365 (no latency) variant
    if not raw_rows and bookmaker == PREFERRED_BOOKMAKER:
        alt = "Bet365 (no latency)"
        alt_rows = raw_rows_from_odds_payload(odds_payload, alt, odds_raw_file, request_id)
        if alt_rows:
            raw_rows = alt_rows
            market_groups = extract_market_groups(odds_payload, alt)
            bookmaker = alt
            print(f"  Using bookmaker variant: {alt}")

    print(f"  Raw rows: {len(raw_rows)}, market groups: {len(market_groups)}")

    with connect() as con:
        ensure_schema(con)
        ensure_event_candidates_table(con)
        con.execute("DELETE FROM betting_odds_raw")
        con.execute("DELETE FROM betting_odds_normalized")
        con.execute("DELETE FROM betting_value_scores_new")
        store_event_candidates(con, candidates, rejected)
        insert_raw_rows(con, raw_rows, replace=False)
        norm = normalize_raw_odds(con, replace=True)
        ev = calculate_ev(con, replace=True)

        player_rows_count = con.execute(
            "SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type LIKE 'player_%'"
        ).fetchone()[0]
        team_rows_count = con.execute(
            "SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type NOT LIKE 'player_%'"
        ).fetchone()[0]

        market_discovery: list[dict[str, Any]] = []
        seen_keys: set[str] = set()
        for item in market_groups:
            seen_keys.add(item["market_key"])
            market_discovery.append({
                "bookmaker": bookmaker,
                "raw_market_group": item["market_key"],
                "raw_market_name": item["market_name"],
                "rows_returned": item["rows_returned"],
                "status": "returned_data" if item["rows_returned"] else "empty",
                "notes": "returned by /odds endpoint",
            })
        for key in EXPECTED_MARKETS:
            if key not in seen_keys:
                market_discovery.append({
                    "bookmaker": bookmaker,
                    "raw_market_group": key,
                    "raw_market_name": key,
                    "rows_returned": 0,
                    "status": "not_returned_by_api",
                    "notes": "expected but not returned by /odds for this event/bookmaker",
                })

        run_summary: dict[str, Any] = {
            "target_match": TARGET_MATCH,
            "target_match_found": "yes",
            "source_name": "Odds-API.io",
            "mode": "from_raw" if from_raw else "live",
            "selected_bookmakers_before": ", ".join(selected_before) or "(unknown)",
            "selected_bookmakers_after": ", ".join(selected_after) or "(same)",
            "bookmaker_used": bookmaker,
            "api_event_id": api_event_id,
            "statshub_event_id": STATSHUB_EVENT_ID,
            "raw odds rows ingested": len(raw_rows),
            "normalized rows": norm["normalized"],
            "supported rows": ev["supported"],
            "matched rows": norm["supported"],
            "unmatched rows": norm["unmatched"],
            "unsupported rows": ev["unsupported"],
            "EV rows calculated": ev["ev_rows"],
            "market_groups_found": len(market_groups),
            "player_prop_rows": player_rows_count,
            "team_match_prop_rows": team_rows_count,
            "requests attempted/successful/failed": f"{attempted}/{successful}/{failed}",
            "rate_limit_status": "429_hit" if rate_limit_hit else "ok",
            "bookmaker_check_status": bookmaker_check_status,
            "stop_reason": "completed",
            "raw_odds_file": str(odds_raw_file),
        }
        workbook = write_workbook(con, run_summary, candidates, market_discovery, raw_sources, [])

    # Final report
    print("\n" + "=" * 60)
    print("USA vs PARAGUAY — LIVE API ODDS FINAL REPORT")
    print("=" * 60)
    print(f"Mode: {'rebuild from raw files' if from_raw else 'live API'}")
    print(f"Target match found: yes")
    print(f"api_event_id: {api_event_id}")
    print(f"statshub_event_id: {STATSHUB_EVENT_ID}")
    print(f"Selected bookmakers BEFORE run: {selected_before or '(unknown)'}")
    print(f"Selected bookmakers AFTER run: {selected_after or '(same)'}")
    print(f"Bookmaker used: {bookmaker}")
    print(f"Raw odds rows: {len(raw_rows)}")
    print(f"Market groups found: {len(market_groups)}")
    for item in market_groups:
        print(f"  - {item['market_name']}: {item['rows_returned']} rows")
    print(f"Player prop rows: {player_rows_count}")
    print(f"Team/match prop rows: {team_rows_count}")
    print(f"Normalized rows: {norm['normalized']}")
    print(f"EV rows calculated: {ev['ev_rows']}")
    print("Top 20 EV rows:")
    for row in ev["top20"]:
        print(
            f"  #{row['rank']:3d} {str(row['raw_market_name'] or '')[:28]:<28} | "
            f"{str(row['selection'] or '')[:24]:<24} | "
            f"odds={row['odds_decimal']} | p={row['model_probability']} | "
            f"EV={row['expected_value']} | {row['verdict']}"
        )
    print(f"Unsupported markets: {ev['unsupported']}")
    print(f"Unmatched selections: {ev['unmatched']}")
    print(f"API requests: {attempted} attempted, {successful} successful, {failed} failed")
    print(f"Rate-limit status: {'429 HIT' if rate_limit_hit else 'ok'}")
    print(f"Output Excel: {workbook}")
    print("\nRun dashboard:")
    print("  streamlit run app/dashboard/betting_value_dashboard.py")


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------

def run_from_raw() -> None:
    """Rebuild pipeline entirely from existing raw JSON files. No API calls."""
    print("=" * 60)
    print("MODE: rebuild from existing raw files")
    files = load_raw_files()
    if not files:
        raise SystemExit(f"No raw files found in {RAW_DIR}")

    raw_sources: list[dict[str, Any]] = []
    for key, info in files.items():
        raw_sources.append({
            "raw_file": str(info["raw_file"]),
            "endpoint": key.replace("_", "/", 1) if key.startswith("bookmakers") else key.replace("_", "/"),
            "status_code": info.get("status_code"),
            "rows_or_items": payload_count(info.get("payload")),
            "error": info.get("error"),
            "captured_at": info.get("fetched_at_utc", now_utc()),
        })
        print(f"  Loaded: {info['raw_file'].name} (rows={payload_count(info.get('payload'))})")

    # Bookmaker state from raw file
    selected_before: list[str] = []
    bm_info = files.get("bookmakers_selected")
    if bm_info and bm_info.get("payload"):
        payload = bm_info["payload"]
        if isinstance(payload, list):
            selected_before = [
                str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
                for b in payload
            ]
        elif isinstance(payload, dict):
            items = payload.get("data") or payload.get("bookmakers") or []
            selected_before = [
                str(b.get("name") or b.get("key") or b) if isinstance(b, dict) else str(b)
                for b in (items if isinstance(items, list) else [])
            ]
    print(f"  Bookmakers from raw: {selected_before or '(file missing or empty)'}")

    bookmaker = pick_bookmaker(selected_before) if selected_before else PREFERRED_BOOKMAKER

    # Events payload
    events_info = files.get("events")
    if not events_info:
        raise SystemExit("events raw file not found")
    events_payload = events_info["payload"]

    # Odds payload — prefer /odds, supplement with /odds_multi
    odds_info = files.get("odds")
    if not odds_info:
        raise SystemExit("odds raw file not found")
    odds_payload = odds_info["payload"] if isinstance(odds_info["payload"], dict) else {}
    odds_raw_file = odds_info["raw_file"]

    if not odds_payload.get("bookmakers"):
        multi_info = files.get("odds_multi")
        if multi_info and isinstance(multi_info["payload"], list) and multi_info["payload"]:
            odds_payload = multi_info["payload"][0]
            odds_raw_file = multi_info["raw_file"]
        elif multi_info and isinstance(multi_info["payload"], dict):
            odds_payload = multi_info["payload"]
            odds_raw_file = multi_info["raw_file"]

    run_pipeline(
        events_payload=events_payload,
        odds_payload=odds_payload,
        bookmaker=bookmaker,
        odds_raw_file=odds_raw_file,
        raw_sources=raw_sources,
        selected_before=selected_before,
        selected_after=selected_before,
        attempted=0,
        successful=0,
        failed=0,
        rate_limit_hit=False,
        bookmaker_check_status="from_raw_file",
        from_raw=True,
    )


def run_live() -> None:
    """Fetch live odds from Odds-API.io."""
    config = load_config()
    if not is_configured_api_key(config.api_key):
        print("ODDS_API_IO_KEY_MISSING")
        raise SystemExit("live API odds unavailable: ODDS_API_IO_KEY not configured in .env")
    if not config.base_url:
        raise SystemExit("live API odds unavailable: ODDS_API_IO_BASE_URL not configured")

    attempted = successful = failed = 0
    rate_limit_hit = False
    raw_sources: list[dict[str, Any]] = []

    def call(endpoint: str, params: dict[str, Any], notes: str = "") -> dict[str, Any]:
        nonlocal attempted, successful, failed, rate_limit_hit
        attempted += 1
        result = api_get(endpoint, params, config.base_url)
        ok = result["status_code"] and 200 <= result["status_code"] < 300 and not result["error"]
        if ok:
            successful += 1
        else:
            failed += 1
            if result["status_code"] == 429:
                rate_limit_hit = True
        raw_sources.append({
            "raw_file": str(result["raw_file"]),
            "endpoint": endpoint,
            "status_code": result["status_code"],
            "rows_or_items": payload_count(result["payload"]),
            "error": result["error"],
            "captured_at": now_utc(),
        })
        return result

    # Task 0: bookmaker state
    print("Checking bookmaker selection state...")
    selected_before: list[str] = []
    bookmaker_check_status = "unknown"
    bm_state = get_selected_bookmakers(config.api_key, config.base_url)
    attempted += 1
    if bm_state["status_code"] and 200 <= bm_state["status_code"] < 300:
        successful += 1
        selected_before = bm_state["selected_bookmakers"]
        bookmaker_check_status = "ok"
    elif bm_state["status_code"] == 429:
        failed += 1
        rate_limit_hit = True
        bookmaker_check_status = "rate_limited"
    else:
        failed += 1
        bookmaker_check_status = f"error_{bm_state['status_code']}"
    raw_sources.append({
        "raw_file": str(bm_state["raw_file"]),
        "endpoint": "bookmakers/selected",
        "status_code": bm_state["status_code"],
        "rows_or_items": len(selected_before),
        "error": bm_state["error"],
        "captured_at": now_utc(),
    })

    bookmaker = pick_bookmaker(selected_before)
    print(f"  selected_before={selected_before}, using={bookmaker}")

    # Task 2: find event
    events_result = call(
        "events",
        {"apiKey": config.api_key, "sport": SPORT, "bookmaker": bookmaker, "status": "pending", "limit": 100},
    )
    if events_result["status_code"] == 429:
        raise SystemExit("Rate limit hit on /events")

    # Task 3: odds
    _, candidates_tmp, _ = find_target_event(events_result["payload"])
    ev_id = candidates_tmp[0]["api_event_id"] if candidates_tmp else CANDIDATE_EVENT_ID
    odds_result = call("odds", {"apiKey": config.api_key, "eventId": ev_id, "bookmakers": bookmaker})
    if odds_result["status_code"] == 429:
        raise SystemExit("Rate limit hit on /odds")
    odds_multi_result = call("odds/multi", {"apiKey": config.api_key, "eventIds": ev_id, "bookmakers": bookmaker})

    odds_payload = odds_result["payload"] if isinstance(odds_result["payload"], dict) else {}
    if not odds_payload.get("bookmakers"):
        if isinstance(odds_multi_result["payload"], list) and odds_multi_result["payload"]:
            odds_payload = odds_multi_result["payload"][0]

    run_pipeline(
        events_payload=events_result["payload"],
        odds_payload=odds_payload,
        bookmaker=bookmaker,
        odds_raw_file=odds_result["raw_file"],
        raw_sources=raw_sources,
        selected_before=selected_before,
        selected_after=selected_before,
        attempted=attempted,
        successful=successful,
        failed=failed,
        rate_limit_hit=rate_limit_hit,
        bookmaker_check_status=bookmaker_check_status,
        from_raw=False,
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fetch live/API Odds-API.io odds for USA vs Paraguay."
    )
    parser.add_argument(
        "--from-raw",
        action="store_true",
        help="Rebuild pipeline from existing raw JSON files — no API calls",
    )
    args = parser.parse_args()

    if args.from_raw:
        run_from_raw()
    else:
        run_live()


if __name__ == "__main__":
    main()
