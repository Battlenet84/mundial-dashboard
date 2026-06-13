from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.odds_driven import (
    RAW_COLUMNS,
    calculate_ev,
    connect,
    ensure_schema,
    insert_raw_rows,
    normalize_raw_odds,
)
from scripts.probe_odds_api_io import is_configured_api_key, load_config


RUN_NAME = "canada_bosnia_live_api_odds_probe"
RAW_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/canada_bosnia_live_api_odds_value_scores.xlsx")
STATSHUB_EVENT_ID = "15186836"
TARGET_HOME = "Canada"
TARGET_AWAY = "Bosnia and Herzegovina"
TARGET_MATCH = f"{TARGET_HOME} vs {TARGET_AWAY}"
SPORT = "football"
BOOKMAKER = "Bet365"
TIMEOUT_SECONDS = 20

HOME_ALIASES = ("canada", "can")
AWAY_ALIASES = ("bosnia", "bosnia and herzegovina", "bosnia & herzegovina", "bosnia-herzegovina", "bih")


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    safe = {key: value for key, value in params.items() if key.lower() != "apikey"}
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    return f"{url}?{urlencode(safe)}" if safe else url


def save_response(action: str, params: dict[str, Any], status_code: int | None, payload: Any, error: str | None, base_url: str) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    path = RAW_DIR / f"{action}_{stamp()}.json"
    wrapper = {
        "fetched_at_utc": now_utc(),
        "run_name": RUN_NAME,
        "action": action,
        "url_without_api_key": public_url(base_url, action, params),
        "params_without_api_key": {key: value for key, value in params.items() if key.lower() != "apikey"},
        "status_code": status_code,
        "response_json": payload,
        "error": error,
    }
    path.write_text(json.dumps(wrapper, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def api_get(endpoint: str, params: dict[str, Any], base_url: str) -> tuple[int | None, Any, str | None, Path]:
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    status_code = None
    payload = None
    error = None
    try:
        response = requests.get(url, params=params, timeout=TIMEOUT_SECONDS)
        status_code = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"response_text_preview": response.text[:1000]}
        if response.status_code >= 400:
            error = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        error = str(exc)
    raw_file = save_response(endpoint.strip("/"), params, status_code, payload, error, base_url)
    return status_code, payload, error, raw_file


def text_has(text: Any, aliases: tuple[str, ...]) -> bool:
    value = str(text or "").lower()
    return any(alias in value for alias in aliases)


def find_target_event(events: Any) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    items = events if isinstance(events, list) else events.get("data", []) if isinstance(events, dict) else []
    candidates = []
    rejected = []
    for event in items:
        home = event.get("home") or event.get("homeTeam") or event.get("home_team")
        away = event.get("away") or event.get("awayTeam") or event.get("away_team")
        is_match = (text_has(home, HOME_ALIASES) and text_has(away, AWAY_ALIASES)) or (
            text_has(home, AWAY_ALIASES) and text_has(away, HOME_ALIASES)
        )
        brief = {
            "api_event_id": event.get("id") or event.get("eventId") or event.get("event_id"),
            "home": home,
            "away": away,
            "date": event.get("date") or event.get("eventDate") or event.get("startTime"),
            "status": event.get("status"),
            "sport": event.get("sport"),
            "league": event.get("league"),
            "confidence": 1.0 if is_match else 0.0,
        }
        if is_match:
            candidates.append((event, brief))
        elif any(text_has(value, HOME_ALIASES + AWAY_ALIASES) for value in (home, away, event.get("league"))):
            rejected.append(brief)
    return (candidates[0][0] if candidates else None), rejected


def price_rows_from_outcome(market_name: str, outcome: dict[str, Any], home: str, away: str) -> list[tuple[str, float | None, float]]:
    rows: list[tuple[str, float | None, float]] = []
    line = outcome.get("line") or outcome.get("point") or outcome.get("handicap") or outcome.get("hdp")
    if outcome.get("home") not in (None, ""):
        rows.append((home, line, float(outcome["home"])))
    if outcome.get("draw") not in (None, ""):
        rows.append(("Draw", line, float(outcome["draw"])))
    if outcome.get("away") not in (None, ""):
        rows.append((away, line, float(outcome["away"])))
    label = outcome.get("label") or outcome.get("name") or outcome.get("selection")
    if outcome.get("over") not in (None, ""):
        selection = label if label and "chance" in market_name.lower() else f"Over {line}".strip()
        rows.append((selection, line, float(outcome["over"])))
    if outcome.get("under") not in (None, ""):
        selection = label if label and "chance" in market_name.lower() else f"Under {line}".strip()
        rows.append((selection, line, float(outcome["under"])))
    for key in ("price", "odds", "decimal"):
        if outcome.get(key) not in (None, ""):
            rows.append((str(label or outcome.get("team") or outcome.get("player") or market_name), line, float(outcome[key])))
            break
    return rows


def raw_rows_from_live_odds(payload: dict[str, Any], raw_file: Path, request_id: str) -> list[dict[str, Any]]:
    home = payload.get("home") or TARGET_HOME
    away = payload.get("away") or TARGET_AWAY
    api_event_id = str(payload.get("id") or "")
    captured_at = now_utc()
    bookmakers = payload.get("bookmakers") if isinstance(payload, dict) else {}
    markets = bookmakers.get(BOOKMAKER, []) if isinstance(bookmakers, dict) else []
    rows: list[dict[str, Any]] = []
    for market_index, market in enumerate(markets):
        if not isinstance(market, dict):
            continue
        market_name = market.get("name") or market.get("key") or f"market_{market_index}"
        outcomes = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
        for odds_index, outcome in enumerate(outcomes or []):
            if not isinstance(outcome, dict):
                continue
            for selection, line, odds in price_rows_from_outcome(market_name, outcome, home, away):
                raw_payload = {
                    "market_index": market_index,
                    "odds_index": odds_index,
                    "market": market,
                    "outcome": outcome,
                }
                rows.append(
                    {
                        "run_name": RUN_NAME,
                        "source_name": "odds_api_io_live",
                        "bookmaker": BOOKMAKER,
                        "match_name": TARGET_MATCH,
                        "event_id": api_event_id,
                        "api_event_id": api_event_id,
                        "statshub_event_id": STATSHUB_EVENT_ID,
                        "raw_market_group": market_name,
                        "raw_market_name": market_name,
                        "raw_selection_name": selection,
                        "raw_line": line,
                        "raw_odds": odds,
                        "odds_format": "decimal",
                        "captured_at": captured_at,
                        "raw_payload": json.dumps(raw_payload, ensure_ascii=False),
                        "source_url": str(payload.get("urls", {}).get(BOOKMAKER) or ""),
                        "request_id": request_id,
                        "raw_file": str(raw_file),
                        "status": "raw",
                        "notes": RUN_NAME,
                    }
                )
    return rows


def table_rows(con: sqlite3.Connection, sql: str) -> tuple[list[str], list[sqlite3.Row]]:
    rows = con.execute(sql).fetchall()
    if rows:
        return list(rows[0].keys()), rows
    cursor = con.execute(sql)
    return [desc[0] for desc in cursor.description], []


def add_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([row[h] for h in headers])
        elif isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(42, len(str(header)) + 4))


def write_workbook(con: sqlite3.Connection, run_summary: dict[str, Any], failed_requests: list[dict[str, Any]]) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "run_summary", ["metric", "value"], [{"metric": k, "value": v} for k, v in run_summary.items()])
    sheets = [
        ("ev_ranking", "SELECT rank, match_name, bookmaker, market_type, raw_market_name, raw_selection_name, team_name, player_name, line, odds_decimal, implied_probability, model_probability, edge, expected_value, probability_method, sample_size, verdict, notes FROM betting_value_scores_new WHERE expected_value IS NOT NULL ORDER BY expected_value DESC"),
        ("raw_odds", "SELECT * FROM betting_odds_raw ORDER BY id"),
        ("normalized_markets", "SELECT * FROM betting_odds_normalized ORDER BY id"),
        ("unsupported_markets", "SELECT * FROM betting_value_scores_new WHERE verdict='UNSUPPORTED' ORDER BY id"),
        ("unmatched_selections", "SELECT * FROM betting_value_scores_new WHERE verdict='UNMATCHED' ORDER BY id"),
    ]
    for name, sql in sheets:
        headers, rows = table_rows(con, sql)
        add_sheet(wb, name, headers, rows)
    unresolved = con.execute(
        """
        SELECT sp.team_name, sp.player_name, sp.player_id, sp.statshub_player_id_status, sp.notes
        FROM statshub_team_players sp
        WHERE sp.team_name IN (?, ?)
          AND sp.statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        ORDER BY sp.team_name, sp.player_name
        """,
        (TARGET_HOME, TARGET_AWAY),
    ).fetchall()
    data_quality_rows = [dict(row) for row in unresolved]
    data_quality_rows.extend({"issue": "failed_api_request", **item} for item in failed_requests)
    headers = sorted({key for row in data_quality_rows for key in row.keys()}) if data_quality_rows else ["issue"]
    add_sheet(wb, "data_quality", headers, data_quality_rows)
    wb.save(OUT_XLSX)
    return OUT_XLSX


def main() -> None:
    parser = argparse.ArgumentParser(description="Fetch live/API Odds-API.io odds for Canada vs Bosnia.")
    parser.parse_args()
    config = load_config()
    if not is_configured_api_key(config.api_key):
        raise SystemExit("live API odds unavailable because configuration is missing: ODDS_API_IO_KEY")
    if not config.base_url:
        raise SystemExit("live API odds unavailable because configuration is missing: ODDS_API_IO_BASE_URL")

    requests_attempted = 0
    requests_successful = 0
    requests_failed = 0
    failed_requests: list[dict[str, Any]] = []

    params = {"apiKey": config.api_key, "sport": SPORT, "bookmaker": BOOKMAKER, "status": "pending", "limit": 100}
    requests_attempted += 1
    events_status, events_payload, events_error, events_file = api_get("events", params, config.base_url)
    if events_status and 200 <= events_status < 300 and not events_error:
        requests_successful += 1
    else:
        requests_failed += 1
        failed_requests.append({"endpoint": "events", "status": events_status, "error": events_error, "raw_file": str(events_file)})
    target_event, rejected = find_target_event(events_payload)
    if not target_event:
        run_summary = {
            "target_match": TARGET_MATCH,
            "odds_source": "Odds-API.io",
            "API configured yes/no": "yes",
            "target_match_found": "no",
            "requests attempted/successful/failed": f"{requests_attempted}/{requests_successful}/{requests_failed}",
            "stop reason": "target match not found in live API events response",
        }
        with connect() as con:
            ensure_schema(con)
            write_workbook(con, run_summary, failed_requests)
        raise SystemExit("Target match not found in live API events response.")

    api_event_id = str(target_event.get("id") or target_event.get("eventId") or target_event.get("event_id"))
    odds_params = {"apiKey": config.api_key, "eventId": api_event_id, "bookmakers": BOOKMAKER}
    requests_attempted += 1
    odds_status, odds_payload, odds_error, odds_file = api_get("odds", odds_params, config.base_url)
    if odds_status and 200 <= odds_status < 300 and not odds_error:
        requests_successful += 1
    else:
        requests_failed += 1
        failed_requests.append({"endpoint": "odds", "status": odds_status, "error": odds_error, "raw_file": str(odds_file)})
    if not isinstance(odds_payload, dict):
        raise SystemExit("Odds response was not a JSON object.")

    request_id = odds_file.stem
    raw_rows = raw_rows_from_live_odds(odds_payload, odds_file, request_id)
    with connect() as con:
        ensure_schema(con)
        # This run must be live/API-only for Canada vs Bosnia, not cached CSV.
        con.execute("DELETE FROM betting_odds_raw")
        con.execute("DELETE FROM betting_odds_normalized")
        con.execute("DELETE FROM betting_value_scores_new")
        insert_raw_rows(con, raw_rows, replace=False)
        norm_summary = normalize_raw_odds(con, replace=True)
        ev_summary = calculate_ev(con, replace=True)
        market_groups = con.execute("SELECT raw_market_name, COUNT(*) n FROM betting_odds_raw GROUP BY raw_market_name ORDER BY n DESC").fetchall()
        player_rows = con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type LIKE 'player_%'").fetchone()[0]
        team_rows = con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type NOT LIKE 'player_%'").fetchone()[0]
        run_summary = {
            "target_match": TARGET_MATCH,
            "odds_source": "Odds-API.io",
            "API configured yes/no": "yes",
            "api_event_id": api_event_id,
            "statshub_event_id": STATSHUB_EVENT_ID,
            "raw odds rows ingested": len(raw_rows),
            "normalized rows": norm_summary["normalized"],
            "supported rows": ev_summary["supported"],
            "matched rows": norm_summary["supported"],
            "unmatched rows": norm_summary["unmatched"],
            "unsupported rows": ev_summary["unsupported"],
            "EV rows calculated": ev_summary["ev_rows"],
            "requests attempted/successful/failed": f"{requests_attempted}/{requests_successful}/{requests_failed}",
            "API budget used/remaining": "2 requests used / remaining not exposed by API response",
            "stop reason": "completed",
            "raw events file": str(events_file),
            "raw odds file": str(odds_file),
        }
        workbook = write_workbook(con, run_summary, failed_requests)

    print("CANADA BOSNIA LIVE API ODDS")
    print("API/source used: Odds-API.io")
    print("Target match found: yes")
    print(f"API event_id: {api_event_id}")
    print(f"Raw odds rows ingested: {len(raw_rows)}")
    print("Market groups found:")
    for row in market_groups[:30]:
        print(f"- {row['raw_market_name']}: {row['n']}")
    if len(market_groups) > 30:
        print(f"- ... {len(market_groups) - 30} more")
    print(f"Player prop rows found: {player_rows}")
    print(f"Team/match prop rows found: {team_rows}")
    print(f"EV rows calculated: {ev_summary['ev_rows']}")
    print("Top 20 EV rows:")
    for row in ev_summary["top20"]:
        print(
            f"- #{row['rank']} {row['raw_market_name']} | {row['selection']} | "
            f"odds={row['odds_decimal']} | p={row['model_probability']} | EV={row['expected_value']} | {row['verdict']}"
        )
    print(f"Unsupported market count: {ev_summary['unsupported']}")
    print(f"Unmatched selection count: {ev_summary['unmatched']}")
    print(f"Request budget used: {requests_attempted} attempted, {requests_successful} successful, {requests_failed} failed")
    print(f"Output Excel path: {workbook}")
    print("Dashboard status: updated to read odds-driven tables; current DB contents are this live Canada/Bosnia run")
    print(f"Rejected/ambiguous candidates: {len(rejected)}")


if __name__ == "__main__":
    main()
