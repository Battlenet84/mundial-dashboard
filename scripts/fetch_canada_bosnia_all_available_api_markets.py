from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.odds_driven import calculate_ev, connect, ensure_schema, insert_raw_rows, normalize_raw_odds
from scripts.fetch_canada_bosnia_live_api_odds import (
    AWAY_ALIASES,
    BOOKMAKER,
    HOME_ALIASES,
    SPORT,
    STATSHUB_EVENT_ID,
    TARGET_AWAY,
    TARGET_HOME,
    TARGET_MATCH,
    find_target_event,
    price_rows_from_outcome,
)
from scripts.probe_odds_api_io import is_configured_api_key, load_config


RUN_NAME = "canada_bosnia_all_available_api_markets_probe"
API_EVENT_ID = "66456916"
RAW_DIR = Path("data/raw/odds") / RUN_NAME
OUT_XLSX = Path("data/processed/betting/canada_bosnia_all_available_api_markets_value_scores.xlsx")
TIMEOUT_SECONDS = 20


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def public_url(base_url: str, endpoint: str, params: dict[str, Any]) -> str:
    safe = {key: value for key, value in params.items() if key.lower() != "apikey"}
    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    return f"{url}?{urlencode(safe)}" if safe else url


def save_response(action: str, params: dict[str, Any], status_code: int | None, payload: Any, error: str | None, base_url: str) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    path = RAW_DIR / f"{action.replace('/', '_')}_{stamp()}.json"
    wrapper = {
        "fetched_at_utc": now_utc(),
        "run_name": RUN_NAME,
        "action": action,
        "url_without_api_key": public_url(base_url, action, params),
        "params_without_api_key": {key: value for key, value in params.items() if key.lower() != "apikey"},
        "status_code": status_code,
        "response_json": payload,
        "error": error,
    }
    path.write_text(json.dumps(wrapper, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def api_get(endpoint: str, params: dict[str, Any], base_url: str) -> dict[str, Any]:
    status_code = None
    payload: Any = None
    error = None
    try:
        response = requests.get(f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}", params=params, timeout=TIMEOUT_SECONDS)
        status_code = response.status_code
        try:
            payload = response.json()
        except ValueError:
            payload = {"response_text_preview": response.text[:1000]}
        if response.status_code >= 400:
            error = f"HTTP {response.status_code}"
    except requests.RequestException as exc:
        error = str(exc)
    raw_file = save_response(endpoint, params, status_code, payload, error, base_url)
    return {"endpoint": endpoint, "params": params, "status_code": status_code, "payload": payload, "error": error, "raw_file": raw_file}


def payload_count(payload: Any) -> int:
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("data", "events", "response", "results", "bookmakers", "leagues"):
            if isinstance(payload.get(key), list):
                return len(payload[key])
    return 1 if payload else 0


def extract_market_groups(payload: dict[str, Any]) -> list[dict[str, Any]]:
    bookmakers = payload.get("bookmakers") if isinstance(payload, dict) else {}
    markets = bookmakers.get(BOOKMAKER, []) if isinstance(bookmakers, dict) else []
    groups = []
    for market in markets:
        if not isinstance(market, dict):
            continue
        name = market.get("name") or market.get("key") or "unknown"
        odds = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
        row_count = 0
        for outcome in odds or []:
            if isinstance(outcome, dict):
                row_count += len(price_rows_from_outcome(name, outcome, payload.get("home") or TARGET_HOME, payload.get("away") or TARGET_AWAY))
        groups.append({"market_key": name, "market_name": name, "rows_returned": row_count})
    return groups


def raw_rows_from_odds_payload(payload: dict[str, Any], raw_file: Path, request_id: str) -> list[dict[str, Any]]:
    home = payload.get("home") or TARGET_HOME
    away = payload.get("away") or TARGET_AWAY
    api_event_id = str(payload.get("id") or API_EVENT_ID)
    captured_at = now_utc()
    bookmakers = payload.get("bookmakers") if isinstance(payload, dict) else {}
    markets = bookmakers.get(BOOKMAKER, []) if isinstance(bookmakers, dict) else []
    rows = []
    for market_index, market in enumerate(markets):
        if not isinstance(market, dict):
            continue
        market_name = market.get("name") or market.get("key") or f"market_{market_index}"
        outcomes = market.get("odds") if isinstance(market.get("odds"), list) else market.get("outcomes", [])
        for odds_index, outcome in enumerate(outcomes or []):
            if not isinstance(outcome, dict):
                continue
            for selection, line, odds in price_rows_from_outcome(market_name, outcome, home, away):
                rows.append(
                    {
                        "run_name": RUN_NAME,
                        "source_name": "odds_api_io_live_all_available",
                        "bookmaker": BOOKMAKER,
                        "match_name": TARGET_MATCH,
                        "event_id": api_event_id,
                        "api_event_id": api_event_id,
                        "statshub_event_id": STATSHUB_EVENT_ID,
                        "raw_market_group": market_name,
                        "raw_market_name": market_name,
                        "raw_selection_name": selection,
                        "raw_line": line,
                        "raw_odds": odds,
                        "odds_format": "decimal",
                        "captured_at": captured_at,
                        "raw_payload": json.dumps({"market_index": market_index, "odds_index": odds_index, "market": market, "outcome": outcome}, ensure_ascii=False),
                        "source_url": str(payload.get("urls", {}).get(BOOKMAKER) or ""),
                        "request_id": request_id,
                        "raw_file": str(raw_file),
                        "status": "raw",
                        "notes": RUN_NAME,
                    }
                )
    return rows


def table_rows(con: sqlite3.Connection, sql: str) -> tuple[list[str], list[sqlite3.Row]]:
    cursor = con.execute(sql)
    rows = cursor.fetchall()
    headers = list(rows[0].keys()) if rows else [desc[0] for desc in cursor.description]
    return headers, rows


def add_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[Any]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        if isinstance(row, sqlite3.Row):
            ws.append([row[h] for h in headers])
        elif isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(46, len(str(header)) + 4))


def write_workbook(
    con: sqlite3.Connection,
    run_summary: dict[str, Any],
    discovery: list[dict[str, Any]],
    raw_responses: list[dict[str, Any]],
) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "run_summary", ["metric", "value"], [{"metric": k, "value": v} for k, v in run_summary.items()])
    add_sheet(wb, "market_discovery", ["market_key", "market_name", "requested", "returned_data", "rows_returned", "status", "notes"], discovery)
    add_sheet(wb, "raw_api_responses", ["endpoint", "status_code", "rows_or_items", "raw_file", "error", "notes"], raw_responses)
    for name, sql in [
        ("ev_ranking", "SELECT rank, match_name, bookmaker, market_type, raw_market_name, raw_selection_name, team_name, player_name, line, odds_decimal, implied_probability, model_probability, edge, expected_value, probability_method, sample_size, verdict, notes FROM betting_value_scores_new WHERE expected_value IS NOT NULL ORDER BY expected_value DESC"),
        ("raw_odds", "SELECT * FROM betting_odds_raw ORDER BY id"),
        ("normalized_markets", "SELECT * FROM betting_odds_normalized ORDER BY id"),
        ("unsupported_markets", "SELECT * FROM betting_value_scores_new WHERE verdict='UNSUPPORTED' ORDER BY id"),
        ("unmatched_selections", "SELECT * FROM betting_value_scores_new WHERE verdict='UNMATCHED' ORDER BY id"),
    ]:
        headers, rows = table_rows(con, sql)
        add_sheet(wb, name, headers, rows)
    quality = con.execute(
        """
        SELECT team_name, player_name, player_id, statshub_player_id_status, notes
        FROM statshub_team_players
        WHERE team_name IN (?, ?)
          AND statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        ORDER BY team_name, player_name
        """,
        (TARGET_HOME, TARGET_AWAY),
    ).fetchall()
    headers = list(quality[0].keys()) if quality else ["issue"]
    add_sheet(wb, "data_quality", headers, quality)
    wb.save(OUT_XLSX)
    return OUT_XLSX


def response_row(result: dict[str, Any], notes: str = "") -> dict[str, Any]:
    return {
        "endpoint": result["endpoint"],
        "status_code": result["status_code"],
        "rows_or_items": payload_count(result["payload"]),
        "raw_file": str(result["raw_file"]),
        "error": result["error"],
        "notes": notes,
    }


def main() -> None:
    argparse.ArgumentParser(description="Fetch all available API markets for Canada vs Bosnia.").parse_args()
    config = load_config()
    if not is_configured_api_key(config.api_key):
        raise SystemExit("live API odds unavailable because configuration is missing: ODDS_API_IO_KEY")

    attempted = successful = failed = skipped_duplicate = 0
    raw_responses: list[dict[str, Any]] = []

    def call(endpoint: str, params: dict[str, Any], notes: str = "") -> dict[str, Any]:
        nonlocal attempted, successful, failed
        attempted += 1
        result = api_get(endpoint, params, config.base_url)
        if result["status_code"] and 200 <= result["status_code"] < 300 and not result["error"]:
            successful += 1
        else:
            failed += 1
        raw_responses.append(response_row(result, notes))
        return result

    # Metadata/discovery calls. These discover API capabilities; /odds itself returns event market groups.
    sports = call("sports", {}, "supported sports; no api key required")
    bookmakers = call("bookmakers", {}, "supported bookmakers")
    leagues = call("leagues", {"apiKey": config.api_key, "sport": SPORT}, "football leagues")
    event_detail = call(f"events/{API_EVENT_ID}", {"apiKey": config.api_key}, "event-specific metadata")
    odds_multi = call("odds/multi", {"apiKey": config.api_key, "eventIds": API_EVENT_ID, "bookmakers": BOOKMAKER}, "event-specific all markets via multi endpoint")
    odds = call("odds", {"apiKey": config.api_key, "eventId": API_EVENT_ID, "bookmakers": BOOKMAKER}, "event-specific all markets")

    target_event = event_detail["payload"] if isinstance(event_detail["payload"], dict) else None
    if not target_event or str(target_event.get("id")) != API_EVENT_ID:
        # Fallback discovery if event detail shape differs.
        events = call("events", {"apiKey": config.api_key, "sport": SPORT, "bookmaker": BOOKMAKER, "status": "pending", "limit": 100}, "fallback target search")
        target_event, _ = find_target_event(events["payload"])
    if not target_event:
        raise SystemExit("Target match not found in live API.")

    odds_payload = odds["payload"] if isinstance(odds["payload"], dict) else {}
    if not odds_payload.get("bookmakers") and isinstance(odds_multi["payload"], list) and odds_multi["payload"]:
        odds_payload = odds_multi["payload"][0]
    market_groups = extract_market_groups(odds_payload)
    raw_rows = raw_rows_from_odds_payload(odds_payload, odds["raw_file"], odds["raw_file"].stem)

    discovery: list[dict[str, Any]] = []
    discovered_api_markets = sorted({item["market_key"] for item in market_groups})
    for item in market_groups:
        discovery.append(
            {
                "market_key": item["market_key"],
                "market_name": item["market_name"],
                "requested": "yes",
                "returned_data": "yes" if item["rows_returned"] else "no",
                "rows_returned": item["rows_returned"],
                "status": "returned_data" if item["rows_returned"] else "empty",
                "notes": "Discovered from /odds response; OpenAPI exposes no market-key request parameter for /odds.",
            }
        )
    for key in ("player props", "team shots", "team shots on target", "cards", "corners", "BTTS", "double chance", "draw no bet"):
        if key not in {d["market_key"] for d in discovery}:
            discovery.append(
                {
                    "market_key": key,
                    "market_name": key,
                    "requested": "no",
                    "returned_data": "no",
                    "rows_returned": 0,
                    "status": "not_exposed_by_event_odds",
                    "notes": "Not returned by /odds for this event/bookmaker; API docs do not support requesting arbitrary market keys.",
                }
            )

    with connect() as con:
        ensure_schema(con)
        con.execute("DELETE FROM betting_odds_raw")
        con.execute("DELETE FROM betting_odds_normalized")
        con.execute("DELETE FROM betting_value_scores_new")
        insert_raw_rows(con, raw_rows, replace=False)
        norm = normalize_raw_odds(con, replace=True)
        ev = calculate_ev(con, replace=True)
        player_rows = con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type LIKE 'player_%'").fetchone()[0]
        team_rows = con.execute("SELECT COUNT(*) FROM betting_odds_normalized WHERE market_type NOT LIKE 'player_%'").fetchone()[0]
        returned_keys = ",".join(discovered_api_markets)
        empty_keys = ",".join([d["market_key"] for d in discovery if d["returned_data"] == "no"])
        run_summary = {
            "target_match": TARGET_MATCH,
            "odds_source": "Odds-API.io",
            "api_event_id": API_EVENT_ID,
            "statshub_event_id": STATSHUB_EVENT_ID,
            "API configured yes/no": "yes",
            "market_keys_discovered": returned_keys,
            "market_keys_requested": "all event markets via /odds and /odds/multi; API has no per-market request param",
            "market_keys_returned_data": returned_keys,
            "market_keys_empty": empty_keys,
            "raw odds rows ingested": len(raw_rows),
            "normalized rows": norm["normalized"],
            "supported rows": ev["supported"],
            "matched rows": norm["supported"],
            "unmatched rows": norm["unmatched"],
            "unsupported rows": ev["unsupported"],
            "EV rows calculated": ev["ev_rows"],
            "requests attempted/successful/failed": f"{attempted}/{successful}/{failed}",
            "requests_skipped_duplicate": skipped_duplicate,
            "API budget used/remaining": f"{attempted} requests used / remaining not exposed by API response",
            "rate-limit errors": len([r for r in raw_responses if r["status_code"] == 429]),
            "stop reason": "completed; /odds returned all event markets exposed for Bet365",
        }
        workbook = write_workbook(con, run_summary, discovery, raw_responses)

    print("CANADA BOSNIA ALL AVAILABLE API MARKETS")
    print("API/source used: Odds-API.io")
    print("Target match found: yes")
    print(f"Market keys discovered: {returned_keys}")
    print("Market keys requested: all event markets via /odds and /odds/multi")
    print(f"Market keys that returned data: {returned_keys}")
    print(f"Market keys empty/unsupported by API: {empty_keys}")
    print(f"Raw odds rows ingested: {len(raw_rows)}")
    print(f"Player prop rows found: {player_rows}")
    print(f"Team/match prop rows found: {team_rows}")
    print(f"EV rows calculated: {ev['ev_rows']}")
    print("Top 20 EV rows:")
    for row in ev["top20"]:
        print(f"- #{row['rank']} {row['raw_market_name']} | {row['selection']} | odds={row['odds_decimal']} | p={row['model_probability']} | EV={row['expected_value']} | {row['verdict']}")
    print(f"Unsupported market count: {ev['unsupported']}")
    print(f"Unmatched selection count: {ev['unmatched']}")
    print(f"Request budget used: {attempted} attempted, {successful} successful, {failed} failed, {skipped_duplicate} skipped duplicate")
    print("Stop reason: completed; no market-key endpoint/parameter available beyond event all-markets response")
    print(f"Output Excel path: {workbook}")
    print("Dashboard status: odds-driven dashboard reads current live/API tables")


if __name__ == "__main__":
    main()
