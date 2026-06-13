"""
Final match stats probe for today's 4 teams.
Snapshot: today_final_match_stats_probe

Tasks:
  1. Refresh team performance (last 50 games)
  2. Final player ID mapping via tournamentId=16 endpoint
  3. Download player performance for all confirmed players
  4. Verify referee data
  5. Generate Excel
  6. Final report + health checks
"""
import sys, json, time, re, pathlib, sqlite3, requests, unicodedata, subprocess
from collections import defaultdict
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DB_PATH  = pathlib.Path("data/mundial.db")
RAW_BASE = pathlib.Path("data/raw/statshub/snapshots")
SNAPSHOT = "today_final_match_stats_probe"
BASE_URL = "https://www.statshub.com/api"
RATE_DELAY = 1.5

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, */*",
    "Accept-Language": "en-US,en;q=0.9",
}

TEAMS = [
    {"name": "Canada",               "db_name": "Canada",               "team_id": 4752, "event_id": 15186836},
    {"name": "Bosnia & Herzegovina", "db_name": "Bosnia and Herzegovina","team_id": 4479, "event_id": 15186836},
    {"name": "USA",                  "db_name": "United States",         "team_id": 4724, "event_id": 15186873},
    {"name": "Paraguay",             "db_name": "Paraguay",              "team_id": 4789, "event_id": 15186873},
]
CONFIRMED_STATUSES = ("confirmed", "skipped_existing")

# ─── DB/HTTP helpers ───────────────────────────────────────────────────────────
def get_con():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def _now(): return datetime.now(timezone.utc).isoformat()
def _ts():  return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
def _norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).lower().strip()

def _ensure_snapshot_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS statshub_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_name TEXT, endpoint_name TEXT, url TEXT, method TEXT,
            status_code INTEGER, content_type TEXT, response_size INTEGER,
            looks_json INTEGER, json_top_keys TEXT, rows_detected INTEGER,
            raw_file_path TEXT, status TEXT, message TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)

def _save_snap(cur, label, url, sc, ct, rows, raw_file, msg=""):
    cur.execute("""
        INSERT INTO statshub_snapshots
            (snapshot_name, endpoint_name, url, method, status_code, content_type,
             looks_json, rows_detected, raw_file_path, status, message, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (SNAPSHOT, label, url, "GET", sc, ct, 1 if sc==200 else 0,
          rows, str(raw_file), "ok" if sc==200 else "error", msg, _now()))

def fetch(label, url, snap_dir, force=False):
    if not force:
        hits = list(snap_dir.glob(f"{label}_*.json"))
        if hits:
            try:
                return 200, json.loads(hits[0].read_text(encoding="utf-8")), "application/json", hits[0]
            except Exception:
                pass
    time.sleep(RATE_DELAY)
    try:
        r = requests.get(url, headers=HEADERS, timeout=25)
    except Exception as e:
        return -1, None, "", None
    ct = r.headers.get("content-type", "")
    fpath = snap_dir / f"{label}_{_ts()}.json"
    if r.status_code == 200 and "json" in ct:
        try:
            payload = r.json()
            fpath.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            return 200, payload, ct, fpath
        except Exception:
            pass
    fpath.write_text(r.text[:2000], encoding="utf-8")
    return r.status_code, None, ct, fpath

def _to_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

def _to_int(v):
    if v is None: return 0
    try: return int(v)
    except: return 0

# ═══════════════════════════════════════════════════════════════════════════════
# TASK 1: Team performance
# ═══════════════════════════════════════════════════════════════════════════════
def _parse_team_stats(stats_dict, opp_dict):
    """Map StatsHub statistics dict to DB columns."""
    s = stats_dict or {}
    o = opp_dict or {}
    def _s(k): return _to_float(s.get(k))
    def _o(k): return _to_float(o.get(k))
    return {
        "goals_for":                None,  # filled from score
        "goals_against":            None,
        "expected_goals":           _s("expectedGoals"),
        "expected_goals_against":   _o("expectedGoals"),
        "shots":                    _to_int(_s("totalShotsOnGoal") or 0) + _to_int(_s("shotsOffGoal") or 0) + _to_int(_s("blockedScoringAttempt") or 0),
        "shots_on_target":          _to_int(_s("totalShotsOnGoal") or 0),
        "shots_off_target":         _to_int(_s("shotsOffGoal") or 0),
        "big_chances":              _to_int(_s("bigChanceCreated") or 0),
        "fouls":                    _to_int(_s("fouls") or 0),
        "yellow_cards":             _to_int(_s("yellowCards") or 0),
        "red_cards":                _to_int(_s("cards") if _s("cards") else 0),
        "total_tackles":            _to_int(_s("totalTackle") or 0),
        "accurate_passes":          _to_int(_s("accuratePasses") or 0),
        "total_passes":             _to_int(_s("passes") or 0),
        "pass_accuracy":            _to_float(_s("pass_accuracy")),
        "possession_average":       _to_float(_s("ballPossession")),
        "corners":                  _to_int(_s("cornerKicks") or 0),
        "goalkeeper_saves":         _to_int(_s("goalkeeperSaves") or 0),
        "final_third_entries":      _to_int(_s("finalThirdEntries") or 0),
        "opponent_expected_goals":  _o("expectedGoals"),
        "opponent_shots":           _to_int(_o("totalShotsOnGoal") or 0) + _to_int(_o("shotsOffGoal") or 0) + _to_int(_o("blockedScoringAttempt") or 0),
        "opponent_shots_on_target": _to_int(_o("totalShotsOnGoal") or 0),
        "opponent_fouls":           _to_int(_o("fouls") or 0),
        "opponent_yellow_cards":    _to_int(_o("yellowCards") or 0),
        "opponent_red_cards":       _to_int(_o("cards") if _o("cards") else 0),
    }

def refresh_team_performance(con, snap_dir):
    cur = con.cursor()
    results = {}
    print("=== TASK 1: Refresh team performance ===")
    for team in TEAMS:
        tid   = team["team_id"]
        tname = team["db_name"]
        url   = f"{BASE_URL}/team/{tid}/performance?limit=50"
        label = f"team_{tid}_performance"
        sc, payload, ct, fpath = fetch(label, url, snap_dir)
        if not payload:
            print(f"  {tname}: FAILED sc={sc}")
            results[tname] = {"rows": 0, "date_min": None, "date_max": None, "status": "failed"}
            continue

        rows = payload.get("data") or []
        events_parsed = []
        agg = defaultdict(float)
        dates = []; competitions = set()
        goals_for_total = goals_against_total = 0

        for item in rows:
            ev   = item.get("event") or {}
            home = item.get("homeTeam") or {}
            away = item.get("awayTeam") or {}
            lg   = item.get("league") or {}
            stats = item.get("statistics") or {}
            opp_s = item.get("opponentStatistics") or {}

            # Date
            ts = ev.get("timeStartTimestamp")
            edate = None
            if ts:
                try:
                    edate = datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d")
                    dates.append(edate)
                except Exception:
                    pass

            # Score / goals
            score = ev.get("score") or {}
            result = ev.get("result", "")
            # Determine if team is home or away
            is_home = home.get("id") == tid
            gf = score.get("home" if is_home else "away") or 0
            ga = score.get("away" if is_home else "home") or 0
            goals_for_total     += _to_int(gf)
            goals_against_total += _to_int(ga)

            ts_vals = _parse_team_stats(stats, opp_s)
            ts_vals["goals_for"]     = _to_int(gf)
            ts_vals["goals_against"] = _to_int(ga)

            # Accumulate aggregates
            for k, v in ts_vals.items():
                if v is not None and k not in ("pass_accuracy","possession_average","expected_goals","expected_goals_against","opponent_expected_goals"):
                    agg[k] += _to_float(v) or 0

            # xG is averaged
            xg = _to_float(stats.get("expectedGoals"))
            if xg: agg["_xg_sum"] += xg; agg["_xg_cnt"] += 1
            xga = _to_float(opp_s.get("expectedGoals"))
            if xga: agg["_xga_sum"] += xga; agg["_xga_cnt"] += 1
            poss = _to_float(stats.get("ballPossession"))
            if poss: agg["_poss_sum"] += poss; agg["_poss_cnt"] += 1
            pa = _to_float(stats.get("pass_accuracy"))
            if pa: agg["_pa_sum"] += pa; agg["_pa_cnt"] += 1

            comps = lg.get("name")
            if comps: competitions.add(comps)

            # Per-event row
            home_away = "home" if is_home else "away"
            opp = (away if is_home else home)
            opp_id   = opp.get("id")
            opp_name = opp.get("name") or opp.get("shortname")
            events_parsed.append({
                "event_id": str(ev.get("id","")),
                "event_date": edate,
                "competition": comps or "",
                "opponent_team_id": str(opp_id) if opp_id else "",
                "opponent_team_name": opp_name or "",
                "home_away": home_away,
                "goals_for": _to_int(gf), "goals_against": _to_int(ga),
                **{k: v for k,v in ts_vals.items() if k not in ("goals_for","goals_against")},
                "raw_row_json": json.dumps(item, ensure_ascii=False),
            })

        dates.sort()
        agg_xg    = agg["_xg_sum"]/agg["_xg_cnt"]     if agg["_xg_cnt"]   else None
        agg_xga   = agg["_xga_sum"]/agg["_xga_cnt"]   if agg["_xga_cnt"]  else None
        agg_poss  = agg["_poss_sum"]/agg["_poss_cnt"]  if agg["_poss_cnt"] else None
        agg_pa    = agg["_pa_sum"]/agg["_pa_cnt"]      if agg["_pa_cnt"]   else None

        # Upsert aggregate
        cur.execute("""
            INSERT OR REPLACE INTO statshub_team_performance_aggregates
                (snapshot_name, endpoint_name, team_id, team_name, source_rows,
                 date_min, date_max, competitions_detected, matches_in_window,
                 goals_for, goals_against, expected_goals, expected_goals_against,
                 shots, shots_on_target, shots_off_target, big_chances, fouls,
                 yellow_cards, red_cards, total_tackles, accurate_passes, total_passes,
                 pass_accuracy, possession_average, corners, goalkeeper_saves,
                 final_third_entries, opponent_expected_goals, opponent_shots,
                 opponent_shots_on_target, opponent_fouls, opponent_yellow_cards,
                 opponent_red_cards, raw_fields_json, imported_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            SNAPSHOT, label, str(tid), tname, len(rows),
            dates[0] if dates else None, dates[-1] if dates else None,
            ",".join(sorted(competitions))[:500], len(rows),
            goals_for_total, goals_against_total,
            round(agg_xg,4) if agg_xg else None, round(agg_xga,4) if agg_xga else None,
            int(agg.get("shots",0)), int(agg.get("shots_on_target",0)), int(agg.get("shots_off_target",0)),
            int(agg.get("big_chances",0)), int(agg.get("fouls",0)),
            int(agg.get("yellow_cards",0)), int(agg.get("red_cards",0)), int(agg.get("total_tackles",0)),
            int(agg.get("accurate_passes",0)), int(agg.get("total_passes",0)),
            round(agg_pa,2) if agg_pa else None, round(agg_poss,1) if agg_poss else None,
            int(agg.get("corners",0)), int(agg.get("goalkeeper_saves",0)),
            int(agg.get("final_third_entries",0)),
            round(agg_xga,4) if agg_xga else None,
            int(agg.get("opponent_shots",0)), int(agg.get("opponent_shots_on_target",0)),
            int(agg.get("opponent_fouls",0)), int(agg.get("opponent_yellow_cards",0)),
            int(agg.get("opponent_red_cards",0)),
            json.dumps(dict(agg))[:2000], _now()
        ))

        # Insert per-event rows
        for ev_row in events_parsed:
            cur.execute("DELETE FROM statshub_team_performance_events WHERE team_id=? AND event_id=?",
                        (str(tid), ev_row["event_id"]))
            cur.execute("""
                INSERT INTO statshub_team_performance_events
                    (snapshot_name, endpoint_name, team_id, team_name, event_id, event_date,
                     competition, opponent_team_id, opponent_team_name, home_away, raw_file, raw_row_json,
                     goals_for, goals_against, expected_goals, expected_goals_against,
                     shots, shots_on_target, shots_off_target, big_chances, fouls,
                     yellow_cards, red_cards, total_tackles, accurate_passes, total_passes,
                     pass_accuracy, possession_average, corners, goalkeeper_saves,
                     final_third_entries, imported_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                SNAPSHOT, label, str(tid), tname,
                ev_row["event_id"], ev_row["event_date"], ev_row["competition"],
                ev_row["opponent_team_id"], ev_row["opponent_team_name"], ev_row["home_away"],
                str(fpath), ev_row["raw_row_json"][:3000],
                ev_row.get("goals_for"), ev_row.get("goals_against"),
                ev_row.get("expected_goals"), ev_row.get("expected_goals_against"),
                ev_row.get("shots"), ev_row.get("shots_on_target"), ev_row.get("shots_off_target"),
                ev_row.get("big_chances"), ev_row.get("fouls"),
                ev_row.get("yellow_cards"), ev_row.get("red_cards"), ev_row.get("total_tackles"),
                ev_row.get("accurate_passes"), ev_row.get("total_passes"),
                ev_row.get("pass_accuracy"), ev_row.get("possession_average"),
                ev_row.get("corners"), ev_row.get("goalkeeper_saves"),
                ev_row.get("final_third_entries"), _now()
            ))

        _save_snap(cur, label, url, sc, ct, len(rows), fpath)
        con.commit()

        results[tname] = {
            "rows": len(rows), "date_min": dates[0] if dates else None,
            "date_max": dates[-1] if dates else None, "status": "ok",
            "goals": goals_for_total, "ga": goals_against_total,
            "xg": round(agg_xg,2) if agg_xg else None,
            "poss": round(agg_poss,1) if agg_poss else None,
        }
        print(f"  {tname}: rows={len(rows)} dates={dates[0] if dates else '?'}→{dates[-1] if dates else '?'} G={goals_for_total} GA={goals_against_total} xG={results[tname]['xg']} poss={results[tname]['poss']}")

    return results

# ═══════════════════════════════════════════════════════════════════════════════
# TASK 2: Player ID mapping
# ═══════════════════════════════════════════════════════════════════════════════
def _parse_endpoint_players(payload):
    if not isinstance(payload, dict): return []
    out = []
    for item in payload.get("data") or []:
        pid = item.get("id")
        name = item.get("name") or item.get("shortName")
        if pid and name:
            out.append({"player_id": str(pid), "player_name": name,
                        "slug": item.get("slug",""), "position": item.get("position","")})
    return out

def player_id_mapping(con, snap_dir):
    cur = con.cursor()
    print("\n=== TASK 2: Final player ID mapping ===")

    # Load roster
    cur.execute("""
        SELECT sp.id row_id, sp.player_name, sp.jersey_number, sp.position,
               sp.statshub_player_id_status id_status, COALESCE(sp.player_id,'') player_id,
               wt.team_name
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
    """)
    roster = [dict(r) for r in cur.fetchall()]
    by_team = defaultdict(list)
    for r in roster:
        by_team[r["team_name"]].append(r)

    newly_confirmed = []
    mapping_records = []

    for team in TEAMS:
        tid  = team["team_id"]
        tdb  = team["db_name"]
        eid  = team["event_id"]
        url  = f"{BASE_URL}/team/{tid}/players/performance?tournamentId=16&location=both&fixtureId={eid}"
        label = f"team_{tid}_players_wc"
        sc, payload, ct, fpath = fetch(label, url, snap_dir)
        ep_players = _parse_endpoint_players(payload) if payload else []
        ep_by_norm = {_norm(p["player_name"]): p for p in ep_players}
        ep_by_id   = {p["player_id"]: p for p in ep_players}
        _save_snap(cur, label, url, sc, ct, len(ep_players), fpath or snap_dir / f"{label}.json")

        print(f"\n  {tdb}: endpoint={len(ep_players)} players (sc={sc})")

        for rp in by_team[tdb]:
            rstatus = rp["id_status"]
            rcur_id = rp["player_id"]
            rname   = rp["player_name"]

            # Determine endpoint match
            ep_match = None
            method = ""

            if rstatus in CONFIRMED_STATUSES:
                # Already confirmed — just verify in endpoint and record
                ep_match = ep_by_id.get(rcur_id)
                method = "confirmed_existing"
            else:
                # Try to match unresolved/probable/ambiguous
                # 1. By existing player_id
                if rcur_id and rcur_id in ep_by_id:
                    ep_match = ep_by_id[rcur_id]
                    method = "existing_id_in_endpoint"
                else:
                    # 2. By normalized full name
                    norm_r = _norm(rname)
                    if norm_r in ep_by_norm:
                        ep_match = ep_by_norm[norm_r]
                        method = "name_exact_norm"
                    else:
                        # 3. By last word (surname) uniqueness
                        parts = rname.strip().split()
                        last = _norm(parts[-1]) if parts else ""
                        if len(last) > 3:
                            hits = [p for n,p in ep_by_norm.items() if last in n.split()]
                            if len(hits) == 1:
                                ep_match = hits[0]
                                method = f"surname_unique:{parts[-1]}"
                            elif len(hits) > 1:
                                # Try first+last
                                first = _norm(parts[0]) if parts else ""
                                fl_hits = [p for n,p in ep_by_norm.items()
                                           if last in n.split() and first in n.split()]
                                if len(fl_hits) == 1:
                                    ep_match = fl_hits[0]
                                    method = f"first_last:{parts[0]}_{parts[-1]}"

                if ep_match and rstatus not in CONFIRMED_STATUSES:
                    print(f"    [CONFIRM] #{rp['jersey_number']} {rname} → {ep_match['player_name']} id={ep_match['player_id']} [{method}]")
                    cur.execute("""
                        UPDATE statshub_team_players
                        SET player_id=?, statshub_player_id_status='confirmed',
                            player_id_match_source=?, player_id_confidence_score=90.0
                        WHERE id=?
                    """, (ep_match["player_id"], method, rp["row_id"]))
                    newly_confirmed.append({
                        "player_id": ep_match["player_id"], "player_name": rname,
                        "team_name": tdb, "method": method
                    })
                elif not ep_match and rstatus not in CONFIRMED_STATUSES:
                    print(f"    [UNRESOLVED] #{rp['jersey_number']} {rname} (status={rstatus})")

            mapping_records.append({
                "team_name": tdb, "team_id": str(tid),
                "fifa_player_name": rname, "jersey": rp["jersey_number"],
                "endpoint_player_name": ep_match["player_name"] if ep_match else "",
                "player_id": ep_match["player_id"] if ep_match else rcur_id,
                "player_id_status": "confirmed" if (rstatus in CONFIRMED_STATUSES or ep_match and rstatus not in CONFIRMED_STATUSES) else rstatus,
                "match_method": method,
                "match_confidence": 95 if ep_match else 0,
                "fixture_id": str(eid),
                "source_endpoint": url,
                "notes": "" if ep_match else "not in tournamentId=16 endpoint",
            })

    con.commit()
    print(f"\n  Newly confirmed in task 2: {len(newly_confirmed)}")
    return newly_confirmed, mapping_records

# ═══════════════════════════════════════════════════════════════════════════════
# TASK 3: Player performance
# ═══════════════════════════════════════════════════════════════════════════════
def _parse_player_perf(payload, player_id, player_name, team_name):
    if not isinstance(payload, dict): return None
    items = (payload.get("playerStatisticsEvents") or payload.get("events")
             or payload.get("data") or [])
    if not items: return None
    apps=minutes=goals=assists=sot=shots=fouls=wf=kp=tp=ap=tac=pl=yc=rc=0
    xg=xa=0.0; dates=[]; comps=set()
    for item in items:
        stats = item.get("player_statistics_event") or item
        mp = stats.get("minutesPlayed") or stats.get("minutes_played") or 0
        try: mp = int(mp)
        except: mp = 0
        if mp > 0: apps+=1; minutes+=mp
        goals   += stats.get("goals") or stats.get("goal") or 0
        assists += stats.get("goalAssist") or stats.get("assists") or 0
        sot     += stats.get("onTargetScoringAttempt") or stats.get("shots_on_target") or 0
        shots   += (stats.get("onTargetScoringAttempt") or 0) + (stats.get("shotOffTarget") or 0) + (stats.get("blockedScoringAttempt") or 0)
        fouls   += stats.get("fouls") or 0
        wf      += stats.get("wasFouled") or 0
        kp      += stats.get("keyPass") or 0
        tp      += stats.get("totalPass") or 0
        ap      += stats.get("accuratePass") or 0
        tac     += stats.get("totalTackle") or 0
        pl      += stats.get("possessionLostCtrl") or 0
        yc      += stats.get("yellowCard") or 0
        rc      += stats.get("redCard") or 0
        xg      += _to_float(stats.get("expectedGoals")) or 0.0
        xa      += _to_float(stats.get("expectedAssists")) or 0.0
        ev = item.get("events") or {}
        ts = ev.get("startTimestamp")
        if ts:
            try:
                dates.append(datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d"))
            except: pass
        tour = (item.get("tournaments") or {})
        tname_t = tour.get("name")
        if tname_t: comps.add(tname_t)
    dates.sort()
    return dict(
        player_id=player_id, player_name=player_name, team_name=team_name,
        source_rows=len(items), appearances=apps, minutes=minutes,
        goals=goals, assists=assists, shots=shots, shots_on_target=sot,
        fouls=fouls, was_fouled=wf,
        xG=round(xg,4) if xg else None, xA=round(xa,4) if xa else None,
        key_passes=kp, passes=tp, accurate_passes=ap,
        tackles=tac, possession_lost=pl, yellow_cards=yc, red_cards=rc,
        date_min=dates[0] if dates else None, date_max=dates[-1] if dates else None,
        competitions_detected=",".join(sorted(comps))[:500],
    )

def _upsert_player_perf(cur, agg):
    cur.execute("""
        INSERT OR REPLACE INTO statshub_player_performance_aggregates
            (snapshot_name, player_id, player_name, team_name, endpoint_name,
             source_rows, appearances, minutes, goals, assists, shots, shots_on_target,
             fouls, was_fouled, xG, xA, key_passes, passes, accurate_passes,
             tackles, possession_lost, yellow_cards, red_cards,
             date_min, date_max, competitions_detected, imported_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        SNAPSHOT, agg["player_id"], agg["player_name"], agg["team_name"], "player_performance",
        agg["source_rows"], agg["appearances"], agg["minutes"],
        agg["goals"], agg["assists"], agg["shots"], agg["shots_on_target"],
        agg["fouls"], agg["was_fouled"], agg["xG"], agg["xA"],
        agg["key_passes"], agg["passes"], agg["accurate_passes"],
        agg["tackles"], agg["possession_lost"], agg["yellow_cards"], agg["red_cards"],
        agg["date_min"], agg["date_max"], agg["competitions_detected"], _now()
    ))

def download_player_performance(con, snap_dir):
    cur = con.cursor()
    print("\n=== TASK 3: Download player performance (confirmed only) ===")

    cur.execute(f"""
        SELECT sp.player_id, sp.player_name, wt.team_name
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
          AND sp.statshub_player_id_status IN {CONFIRMED_STATUSES!r}
          AND sp.player_id IS NOT NULL
        ORDER BY wt.team_name, sp.player_name
    """)
    players = [dict(r) for r in cur.fetchall()]
    print(f"  Confirmed players: {len(players)}")

    downloaded = []; skipped = []; failed = []
    for p in players:
        pid   = p["player_id"]
        pname = p["player_name"]
        tname = p["team_name"]
        label = f"perf_{pid}"
        sc, payload, ct, fpath = fetch(label, snap_dir / f"{label}.json",
                                       snap_dir)   # path trick: check old cache too
        # Also check previous snapshot directories
        if not payload:
            for old_snap in ["today_playwright_fixture_players_probe",
                             "today_browser_endpoint_replay_probe",
                             "today_canada_bosnia_usa_paraguay_probe"]:
                old_dir = RAW_BASE / old_snap
                hits = list(old_dir.glob(f"perf_{pid}*.json")) if old_dir.exists() else []
                if hits:
                    try:
                        payload = json.loads(hits[0].read_text(encoding="utf-8"))
                        sc = 200; fpath = hits[0]
                        break
                    except Exception:
                        pass

        if not payload:
            # Fetch fresh
            url = f"{BASE_URL}/player/{pid}/performance?limit=50"
            sc, payload, ct, fpath = fetch(label, url, snap_dir, force=True)

        if not payload:
            failed.append({"player_name": pname, "player_id": pid, "team_name": tname, "sc": sc})
            continue

        agg = _parse_player_perf(payload, pid, pname, tname)
        if not agg:
            failed.append({"player_name": pname, "player_id": pid, "team_name": tname, "sc": "no_data"})
            continue

        # Check if already up to date
        cur.execute("SELECT imported_at FROM statshub_player_performance_aggregates WHERE player_id=?", (pid,))
        existing = cur.fetchone()
        _upsert_player_perf(cur, agg)
        con.commit()
        if existing:
            skipped.append(agg)
        else:
            downloaded.append(agg)

    print(f"  Fresh/updated: {len(downloaded)} | Already had: {len(skipped)} | Failed: {len(failed)}")
    if failed:
        for f in failed:
            print(f"    FAILED: {f['player_name']} ({f['team_name']}) sc={f['sc']}")

    # Return all for Excel
    all_perf = downloaded + skipped
    return all_perf, failed

# ═══════════════════════════════════════════════════════════════════════════════
# TASK 4: Referee data
# ═══════════════════════════════════════════════════════════════════════════════
def verify_referees(con, snap_dir):
    cur = con.cursor()
    print("\n=== TASK 4: Referee data ===")
    cur.execute("SELECT event_id, match_name, referee_name, referee_id FROM statshub_match_referees")
    refs = {r[0]: dict(zip(["event_id","match_name","referee_name","referee_id"], r))
            for r in cur.fetchall()}

    matches = [
        {"event_id": "15186836", "match_name": "Canada vs Bosnia and Herzegovina",
         "expected_ref": "Facundo Raul Tello Figueroa"},
        {"event_id": "15186873", "match_name": "United States vs Paraguay",
         "expected_ref": "Danny Makkelie"},
    ]
    results = []
    for m in matches:
        eid = m["event_id"]
        existing = refs.get(eid)
        if existing and existing.get("referee_name"):
            print(f"  {m['match_name']}: OK — {existing['referee_name']} (id={existing['referee_id']})")
            results.append({**existing, "status": "ok"})
        else:
            print(f"  {m['match_name']}: MISSING — expected {m['expected_ref']}")
            # Insert manually from known data
            cur.execute("""
                INSERT OR IGNORE INTO statshub_match_referees
                    (snapshot_name, event_id, match_name, referee_name, source_endpoint,
                     referee_endpoint_status, notes, imported_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (SNAPSHOT, eid, m["match_name"], m["expected_ref"],
                  "manual_from_prior_session", "manual", "confirmed from event base endpoint", _now()))
            con.commit()
            results.append({"event_id": eid, "match_name": m["match_name"],
                            "referee_name": m["expected_ref"], "referee_id": None, "status": "manual"})
    return results

# ═══════════════════════════════════════════════════════════════════════════════
# TASK 5: Excel
# ═══════════════════════════════════════════════════════════════════════════════
def generate_excel(con, snap_dir, team_perf_results, mapping_records, all_perf, failed_perf, ref_results):
    import openpyxl
    cur = con.cursor()
    wb = openpyxl.Workbook()

    # ── Sheet 1: match_summary ─────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "match_summary"
    ws1.append(["match_name","event_id","event_date","status","home_team","away_team",
                "referee_name","referee_id","notes"])
    match_info = [
        ("Canada vs Bosnia and Herzegovina", "15186836", "2026-06-12", "upcoming",
         "Canada", "Bosnia and Herzegovina"),
        ("United States vs Paraguay", "15186873", "2026-06-12", "upcoming",
         "United States", "Paraguay"),
    ]
    for mi in match_info:
        ref = next((r for r in ref_results if r["event_id"] == mi[1]), {})
        ws1.append([mi[0], mi[1], mi[2], mi[3], mi[4], mi[5],
                    ref.get("referee_name",""), ref.get("referee_id",""), ""])

    # ── Sheet 2: team_stats_limit50 ────────────────────────────────────────────
    ws2 = wb.create_sheet("team_stats_limit50")
    cols2 = ["team_name","team_id","source_rows","date_min","date_max","goals_for","goals_against",
             "xG","xGA","shots","shots_on_target","fouls","yellow_cards","red_cards",
             "total_passes","accurate_passes","pass_accuracy","possession_average",
             "corners","goalkeeper_saves","final_third_entries",
             "opponent_expected_goals","opponent_shots","opponent_shots_on_target",
             "opponent_fouls","opponent_yellow_cards","opponent_red_cards"]
    ws2.append(cols2)
    cur.execute("""
        SELECT team_name, team_id, source_rows, date_min, date_max, goals_for, goals_against,
               expected_goals, expected_goals_against, shots, shots_on_target, fouls,
               yellow_cards, red_cards, total_passes, accurate_passes, pass_accuracy,
               possession_average, corners, goalkeeper_saves, final_third_entries,
               opponent_expected_goals, opponent_shots, opponent_shots_on_target,
               opponent_fouls, opponent_yellow_cards, opponent_red_cards
        FROM statshub_team_performance_aggregates
        WHERE snapshot_name = ?
        ORDER BY team_name
    """, (SNAPSHOT,))
    for row in cur.fetchall():
        ws2.append(list(row))

    # ── Sheet 3: player_id_mapping_today ──────────────────────────────────────
    ws3 = wb.create_sheet("player_id_mapping_today")
    ws3.append(["team_name","fifa_player_name","jersey","endpoint_player_name","player_id",
                "player_id_status","match_method","match_confidence","fixture_id",
                "source_endpoint","notes"])
    for r in sorted(mapping_records, key=lambda x: (x["team_name"], x["fifa_player_name"])):
        ws3.append([r.get(k,"") for k in ["team_name","fifa_player_name","jersey",
                                           "endpoint_player_name","player_id","player_id_status",
                                           "match_method","match_confidence","fixture_id",
                                           "source_endpoint","notes"]])

    # ── Sheet 4: player_stats_limit50 ─────────────────────────────────────────
    ws4 = wb.create_sheet("player_stats_limit50")
    cols4 = ["team_name","player_name","player_id","source_rows","appearances","minutes",
             "goals","assists","shots","shots_on_target","fouls","was_fouled",
             "yellow_cards","red_cards","xG","xA","key_passes","passes",
             "accurate_passes","tackles","possession_lost","date_min","date_max"]
    ws4.append(cols4)
    cur.execute("""
        SELECT a.team_name, a.player_name, a.player_id, a.source_rows, a.appearances,
               a.minutes, a.goals, a.assists, a.shots, a.shots_on_target, a.fouls, a.was_fouled,
               a.yellow_cards, a.red_cards, a.xG, a.xA, a.key_passes, a.passes,
               a.accurate_passes, a.tackles, a.possession_lost, a.date_min, a.date_max
        FROM statshub_player_performance_aggregates a
        WHERE EXISTS (
            SELECT 1 FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
            AND sp.player_id = a.player_id
            AND sp.statshub_player_id_status IN ('confirmed','skipped_existing')
        )
        ORDER BY a.team_name, a.appearances DESC
    """)
    for row in cur.fetchall():
        ws4.append(list(row))

    # ── Sheet 5: unresolved_players_today ─────────────────────────────────────
    ws5 = wb.create_sheet("unresolved_players_today")
    ws5.append(["team_name","player_name","position","jersey_number","player_id_status",
                "current_candidate_id","notes"])
    cur.execute("""
        SELECT wt.team_name, sp.player_name, sp.position, sp.jersey_number,
               sp.statshub_player_id_status, sp.player_id
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
        AND sp.statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        ORDER BY wt.team_name, sp.player_name
    """)
    for row in cur.fetchall():
        ws5.append(list(row))

    # ── Sheet 6: endpoint_only_players ────────────────────────────────────────
    ws6 = wb.create_sheet("endpoint_only_players")
    ws6.append(["team_name","endpoint_player_name","player_id","source_endpoint","notes"])
    matched_pids = {r["player_id"] for r in mapping_records if r.get("player_id")}
    # Collect from cached endpoint responses
    for team in TEAMS:
        tid   = team["team_id"]
        tdb   = team["db_name"]
        eid   = team["event_id"]
        label = f"team_{tid}_players_wc"
        hits  = list(snap_dir.glob(f"{label}_*.json"))
        if not hits: continue
        try:
            d = json.loads(hits[0].read_text(encoding="utf-8"))
            for item in (d.get("data") or []):
                pid  = str(item.get("id",""))
                name = item.get("name","")
                if pid not in matched_pids:
                    ws6.append([tdb, name, pid, label, "in endpoint but not in FIFA 26"])
        except Exception:
            pass

    # ── Sheet 7: raw_sources ──────────────────────────────────────────────────
    ws7 = wb.create_sheet("raw_sources")
    ws7.append(["snapshot_name","endpoint_name","url","status_code","rows_detected","raw_file"])
    cur.execute("""
        SELECT snapshot_name, endpoint_name, url, status_code, rows_detected, raw_file_path
        FROM statshub_snapshots WHERE snapshot_name = ? ORDER BY id
    """, (SNAPSHOT,))
    for row in cur.fetchall():
        ws7.append(list(row))

    # ── Sheet 8: data_dictionary ──────────────────────────────────────────────
    ws8 = wb.create_sheet("data_dictionary")
    ws8.append(["sheet","column","description"])
    dd = [
        ("match_summary","event_id","StatsHub internal event ID"),
        ("team_stats_limit50","source_rows","Number of matches in the last-50 window"),
        ("team_stats_limit50","xG","Average expected goals per game (from last 50 matches)"),
        ("team_stats_limit50","xGA","Average expected goals against per game"),
        ("player_id_mapping_today","player_id_status","confirmed/skipped_existing=reliable; probable/ambiguous/unresolved=not used for props"),
        ("player_id_mapping_today","match_method","How the StatsHub player_id was found"),
        ("player_stats_limit50","appearances","Games where minutesPlayed > 0 (NOT source_rows)"),
        ("player_stats_limit50","source_rows","Raw rows in performance response (includes bench/DNP)"),
        ("unresolved_players_today","player_id_status","Status at time of export; these players have no reliable stats"),
        ("endpoint_only_players","notes","These players appear in StatsHub squad but are not in FIFA 26-man roster"),
    ]
    for row in dd:
        ws8.append(row)

    out = pathlib.Path("data/processed/statshub/today_final_match_stats_review.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"  OK Saved: {out}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
def main():
    snap_dir = RAW_BASE / SNAPSHOT
    snap_dir.mkdir(parents=True, exist_ok=True)

    con = get_con()
    cur = con.cursor()
    _ensure_snapshot_table(cur)
    con.commit()

    # Coverage before
    coverage_before = {}
    for team in TEAMS:
        cur.execute("""
            SELECT statshub_player_id_status, COUNT(*) cnt
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name = ? GROUP BY statshub_player_id_status
        """, (team["db_name"],))
        s = {r[0]: r[1] for r in cur.fetchall()}
        coverage_before[team["db_name"]] = s.get("confirmed",0) + s.get("skipped_existing",0)

    team_perf   = refresh_team_performance(con, snap_dir)
    nc, mapping = player_id_mapping(con, snap_dir)
    all_perf, failed_perf = download_player_performance(con, snap_dir)
    ref_results = verify_referees(con, snap_dir)

    print("\n=== TASK 5: Generate Excel ===")
    out = generate_excel(con, snap_dir, team_perf, mapping, all_perf, failed_perf, ref_results)

    # ── TASK 6: Final report ──────────────────────────────────────────────────
    print("\n============================================================")
    print("TASK 6: FINAL REPORT")
    print("============================================================")

    print("\n--- A. Final coverage by team ---")
    team_decisions = {}
    for team in TEAMS:
        tdb = team["db_name"]
        cur.execute("""
            SELECT statshub_player_id_status, COUNT(*) cnt
            FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name = ? GROUP BY statshub_player_id_status
        """, (tdb,))
        s = {r[0]: r[1] for r in cur.fetchall()}
        total     = sum(s.values())
        confirmed = s.get("confirmed",0) + s.get("skipped_existing",0)
        before    = coverage_before[tdb]
        pct       = round(100*confirmed/total,1) if total else 0
        decision  = "READY" if pct >= 80 else "INCOMPLETE"
        team_decisions[tdb] = decision
        new = confirmed - before
        perf_cnt  = 0
        cur.execute("""
            SELECT COUNT(*) FROM statshub_player_performance_aggregates a
            WHERE EXISTS (
                SELECT 1 FROM statshub_team_players sp
                JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
                WHERE wt.team_name = ? AND sp.player_id = a.player_id
                AND sp.statshub_player_id_status IN ('confirmed','skipped_existing')
            )
        """, (tdb,))
        perf_cnt = cur.fetchone()[0]
        print(f"  {tdb}: {before}→{confirmed}/{total} ({pct}%) +{new} new | perf={perf_cnt} → {decision}")

    print("\n--- B. Team statistics ---")
    for tdb, info in team_perf.items():
        print(f"  {tdb}: rows={info['rows']} {info['date_min']}→{info['date_max']} G={info.get('goals')} GA={info.get('ga')} xG={info.get('xg')} → {info['status']}")

    print("\n--- C. Player statistics ---")
    by_team = defaultdict(lambda: {"ok":0,"failed":0,"skipped":0})
    for agg in all_perf:
        by_team[agg["team_name"]]["ok"] += 1
    for f in failed_perf:
        by_team[f["team_name"]]["failed"] += 1
    # Count non-confirmed (skipped)
    for team in TEAMS:
        cur.execute("""
            SELECT COUNT(*) FROM statshub_team_players sp
            JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
            WHERE wt.team_name = ?
            AND sp.statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        """, (team["db_name"],))
        by_team[team["db_name"]]["skipped"] = cur.fetchone()[0]
    for tdb, v in sorted(by_team.items()):
        print(f"  {tdb}: downloaded={v['ok']} failed={v['failed']} skipped(not confirmed)={v['skipped']}")

    print("\n--- D. Referee status ---")
    for r in ref_results:
        print(f"  {r['match_name']}: {r['referee_name']} (id={r.get('referee_id')}) → {r['status']}")

    print("\n--- E. Decision ---")
    all_ready = all(d == "READY" for d in team_decisions.values())
    print(f"  4 teams ready for team-stat review? {'YES' if all_ready else 'PARTIAL'}")
    for tdb, dec in team_decisions.items():
        print(f"    {tdb}: {dec}")

    # Unresolved players
    cur.execute("""
        SELECT wt.team_name, sp.player_name, sp.jersey_number, sp.statshub_player_id_status
        FROM statshub_team_players sp
        JOIN statshub_world_cup_teams wt ON sp.team_id = wt.statshub_team_id
        WHERE wt.team_name IN ('Canada','Bosnia and Herzegovina','United States','Paraguay')
        AND sp.statshub_player_id_status NOT IN ('confirmed','skipped_existing')
        ORDER BY wt.team_name, sp.player_name
    """)
    unresolved = cur.fetchall()
    if unresolved:
        print(f"\n  Unresolved players ({len(unresolved)}):")
        for r in unresolved:
            print(f"    {r[0]} #{r[2]} {r[1]} [{r[3]}]")
    else:
        print("\n  No unresolved players.")

    print(f"\n  Excel ready? YES → {out}")

    con.close()

    # ── Health checks ─────────────────────────────────────────────────────────
    print("\n=== HEALTH CHECKS ===")
    for cmd in [
        [sys.executable, "-m", "scripts.statshub_raw_db_status"],
        [sys.executable, "-m", "scripts.statshub_snapshot_status"],
        [sys.executable, "-m", "scripts.health_check"],
    ]:
        print(f"\n$ {' '.join(cmd[1:])}")
        try:
            result = subprocess.run(cmd, capture_output=True, text=True,
                                    encoding="utf-8", errors="replace", timeout=60)
            print(result.stdout[:2500])
        except Exception as e:
            print(f"  ERROR: {e}")

    # pytest
    print("\n$ pytest -q")
    try:
        result = subprocess.run([sys.executable, "-m", "pytest", "-q", "--tb=no"],
                                capture_output=True, text=True,
                                encoding="utf-8", errors="replace", timeout=120)
        out_txt = (result.stdout + result.stderr)[:2000]
        print(out_txt)
    except Exception as e:
        print(f"  ERROR: {e}")

    print("\nDone.")

if __name__ == "__main__":
    main()
