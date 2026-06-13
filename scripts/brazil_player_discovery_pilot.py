"""
Brazil StatsHub player discovery pilot.

Problem: Brazil has 50 team-performance rows but 0 confirmed StatsHub player IDs.
Root cause: The parser was looking for 'playerId'/'playerName' but the endpoint
            returns 'id'/'name' at the top-level item.  The 890 KB response from
            team_4748_players_wc26_tournamentId16.json already contains the data.

Strategy:
  1. Parse the existing 890 KB local file   (0 API calls)
  2. Try /api/team/4748/players             (live if --execute)
  3. Try /api/team/4748/players/performance (live if --execute)
  4. Match FIFA roster → StatsHub IDs (token + slug + alias table)
  5. Cross-match Bet365 player names        (from betting_odds_normalized)
  6. Update statshub_team_players
  7. Download player performance for confirmed IDs
  8. Write diagnostic workbook

Usage:
    python -m scripts.brazil_player_discovery_pilot            # parse local, no API
    python -m scripts.brazil_player_discovery_pilot --execute  # + live endpoints
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests

from app.betting.odds_driven import connect
from app.db.queries import utc_now

# ── Constants ─────────────────────────────────────────────────────────────────

SNAPSHOT_NAME  = "brazil_player_discovery_pilot"
BRAZIL_TEAM_ID = "4748"
BRAZIL_SLUG    = "brazil"
BASE           = "https://www.statshub.com"

# Existing 890 KB file from previous run (0 bytes would have failed)
EXISTING_SQUAD_FILE = (
    ROOT / "data" / "raw" / "statshub" / "snapshots"
    / "today_4_matches_statshub_coverage"
    / "team_4748_players_wc26_tournamentId16.json"
)

RAW_DIR = ROOT / "data" / "raw" / "statshub" / "snapshots" / SNAPSHOT_NAME
OUT_DIR = ROOT / "data" / "processed" / "statshub"
OUT_XLSX = OUT_DIR / "brazil_player_discovery_pilot.xlsx"

RATE_DELAY = 1.5

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.statshub.com/",
    "Origin":  "https://www.statshub.com",
}

# ── Manual aliases: FIFA full name fragment → (common_name, statshub_slug) ───
# Used when token/slug matching fails.
# Key = distinctive substring of the FIFA name (lower, no accents).
# Value = (common_name_for_display, statshub_slug)
BRAZIL_ALIASES: list[tuple[str, str, str]] = [
    # (fifa_name_fragment, common_name, statshub_slug)
    ("belloli",             "Raphinha",        "raphinha"),
    ("tolentino coelho",    "Lucas Paqueta",   "lucas-paqueta"),
    ("aoas correa",         "Marquinhos",      "marquinhos"),
    ("tavares",             "Fabinho",         "fabinho"),
    ("paixao",              "Vinicius Junior", "vinicius-junior"),
    ("vinicius jose",       "Vinicius Junior", "vinicius-junior"),
    ("cunha",               "Matheus Cunha",   "matheus-cunha"),
    ("carneiro",            "Matheus Cunha",   "matheus-cunha"),
    ("endrick",             "Endrick",         "endrick"),
    ("moreira de sousa",    "Endrick",         "endrick"),
    ("simplicio rocha",     "Rayan Vitor",     "rayan-vitor"),
    ("rayan vitor",         "Rayan Vitor",     "rayan-vitor"),
    ("igor thiago",         "Igor Thiago",     "igor-thiago"),
    ("luiz henrique",       "Luiz Henrique",   "luiz-henrique"),
    ("gabriel magalhaes",   "Gabriel Magalhaes", "gabriel-magalhaes"),
    ("dos santos magalh",   "Gabriel Magalhaes", "gabriel-magalhaes"),
    ("leo pereira",         "Leo Pereira",     "leo-pereira"),
    ("leonardo pereira",    "Leo Pereira",     "leo-pereira"),
    ("roger ibanez",        "Roger Ibanez",    "roger-ibanez"),
    ("ibanez da silva",     "Roger Ibanez",    "roger-ibanez"),
    ("douglas dos santos",  "Douglas Santos",  "douglas-luiz"),
    ("justino de melo",     "Douglas Santos",  "douglas-luiz"),
    ("danilo dos santos",   "Danilo dos Santos", "danilo-dos-santos"),
    ("santos de oliveira",  "Danilo dos Santos", "danilo-dos-santos"),
    ("gleison bremer",      "Bremer",          "bremer"),
    ("alex sandro",         "Alex Sandro",     "alex-sandro"),
    ("alisson",             "Alisson",         "alisson"),
    ("ederson santana",     "Ederson",         "ederson"),
    ("weverton",            "Weverton",        "weverton"),
    ("neymar",              "Neymar",          "neymar"),
    ("casimiro",            "Casemiro",        "casemiro"),
    ("danilo luiz",         "Danilo",          "danilo"),
    ("gabriel martinelli",  "Gabriel Martinelli", "gabriel-martinelli"),
    ("teodoro martinelli",  "Gabriel Martinelli", "gabriel-martinelli"),
    ("bruno guimaraes",     "Bruno Guimaraes", "bruno-guimaraes"),
    ("rodriguez moura",     "Bruno Guimaraes", "bruno-guimaraes"),
    ("eder militao",        "Eder Militao",    "eder-militao"),
    ("militao",             "Eder Militao",    "eder-militao"),
    ("fabio henrique",      "Fabinho",         "fabinho"),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", " ", s.lower()).strip()


def _slug(s: str) -> str:
    return re.sub(r"\s+", "-", _norm(s))


def _norm_key(s: str) -> str:
    return " ".join(_norm(s).split())


def _to_float(v: Any) -> float | None:
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def fetch(url: str, label: str, execute: bool = True) -> tuple[Any | None, dict]:
    target = RAW_DIR / f"{label}.json"
    txt    = RAW_DIR / f"{label}.txt"
    if target.exists():
        try:
            return json.loads(target.read_text(encoding="utf-8")), {"status": "cached", "file": str(target)}
        except Exception:
            pass
    if txt.exists():
        return None, {"status": "cached_error", "file": str(txt)}
    if not execute:
        return None, {"status": "dry_run", "file": ""}
    time.sleep(RATE_DELAY)
    try:
        r = requests.get(url, headers=HEADERS, timeout=25)
        try:
            p = json.loads(r.text)
            target.write_text(json.dumps(p), encoding="utf-8")
            return p, {"status": f"http_{r.status_code}", "file": str(target)}
        except Exception:
            txt.write_text(r.text[:5000], encoding="utf-8")
            return None, {"status": f"http_{r.status_code}_nonjson", "file": str(txt)}
    except Exception as exc:
        txt.write_text(f"error: {exc}", encoding="utf-8")
        return None, {"status": "error", "file": str(txt)}


# ── Parse any team-players endpoint ──────────────────────────────────────────

def parse_team_players(payload: Any, source: str) -> list[dict]:
    """
    Extract player records from various StatsHub endpoint shapes.
    Always look for items with a numeric id and a name string.
    """
    if not payload:
        return []

    candidates: list[dict] = []

    def _try_items(items: Any) -> None:
        if not isinstance(items, list):
            return
        for item in items:
            if not isinstance(item, dict):
                continue
            # Shape A: {id, name, slug, position} — top-level player list
            pid  = item.get("id") or item.get("playerId") or item.get("player_id")
            name = item.get("name") or item.get("playerName") or item.get("player_name")
            if pid and name and isinstance(name, str) and len(name) > 1:
                slug = item.get("slug", "")
                pos  = item.get("position", "")
                candidates.append({
                    "player_id": str(pid),
                    "player_name": name,
                    "slug": slug,
                    "position": pos,
                    "source": source,
                })
                continue
            # Shape B: nested player object
            p = item.get("player") or item.get("playerInfo")
            if isinstance(p, dict):
                pid2  = p.get("id") or p.get("playerId")
                name2 = p.get("name") or p.get("playerName")
                if pid2 and name2:
                    candidates.append({
                        "player_id": str(pid2),
                        "player_name": name2,
                        "slug": p.get("slug", ""),
                        "position": p.get("position", ""),
                        "source": source,
                    })

    if isinstance(payload, list):
        _try_items(payload)
    elif isinstance(payload, dict):
        for key in ("data", "players", "squad", "roster", "results", "items"):
            _try_items(payload.get(key, []))
        # Also try top-level values that are lists
        for v in payload.values():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                _try_items(v)

    # Deduplicate by player_id
    seen: set[str] = set()
    result: list[dict] = []
    for c in candidates:
        pid = c["player_id"]
        if pid not in seen and _norm(c["player_name"]).strip():
            seen.add(pid)
            result.append(c)
    return result


# ── Phase 0: Load Brazil roster from DB ──────────────────────────────────────

def phase0_load_roster(con: Any) -> list[dict]:
    print("\n=== PHASE 0: Load Brazil roster from DB ===")
    rows = [dict(r) for r in con.execute("""
        SELECT id as row_id, player_name, player_id, statshub_player_id_status,
               jersey_number, position, player_name_canonical
        FROM statshub_team_players
        WHERE team_name='Brazil'
        ORDER BY CAST(jersey_number AS INTEGER)
    """).fetchall()]
    print(f"  Brazil roster in DB: {len(rows)} players")
    for r in rows:
        print(f"    #{r['jersey_number']:>2} {r['position'] or '':2}  {r['player_name']}  "
              f"[id={r['player_id']} status={r['statshub_player_id_status']}]")
    return rows


# ── Phase 1: Parse endpoint sources ──────────────────────────────────────────

def phase1_parse_sources(execute: bool) -> list[dict]:
    """
    Collect players from ALL available sources:
    - The existing 890 KB file (always, no API call)
    - /api/team/4748/players        (live if --execute)
    - /api/team/4748/players/performance  (live if --execute)
    """
    print("\n=== PHASE 1: Parse team player sources ===")
    all_ep: list[dict] = []
    endpoint_tests: list[dict] = []

    def _add_source(payload: Any, label: str, url: str, status: str) -> None:
        players = parse_team_players(payload, label)
        endpoint_tests.append({
            "endpoint_label": label,
            "url": url,
            "status": status,
            "players_extracted": len(players),
        })
        print(f"  [{label}] status={status} players={len(players)}")
        all_ep.extend(players)

    # Source 1: existing 890 KB file
    ep1_file = EXISTING_SQUAD_FILE
    ep1_label = "team_4748_players_wc26_tournamentId16_existing"
    if ep1_file.exists():
        try:
            payload1 = json.loads(ep1_file.read_text(encoding="utf-8"))
            _add_source(payload1, ep1_label,
                        f"{BASE}/api/team/{BRAZIL_TEAM_ID}/players/performance?tournamentId=16&location=both",
                        "local_file")
        except Exception as e:
            endpoint_tests.append({"endpoint_label": ep1_label, "url": "", "status": f"parse_error:{e}", "players_extracted": 0})
    else:
        print(f"  WARNING: existing file not found: {ep1_file}")

    # Source 2: /api/team/{id}/players (no tournament filter)
    url2   = f"{BASE}/api/team/{BRAZIL_TEAM_ID}/players"
    label2 = f"team_{BRAZIL_TEAM_ID}_players_raw"
    p2, m2 = fetch(url2, label2, execute=execute)
    _add_source(p2, label2, url2, m2["status"])

    # Source 3: /api/team/{id}/players/performance (no filter)
    url3   = f"{BASE}/api/team/{BRAZIL_TEAM_ID}/players/performance"
    label3 = f"team_{BRAZIL_TEAM_ID}_players_perf_nofilt"
    p3, m3 = fetch(url3, label3, execute=execute)
    _add_source(p3, label3, url3, m3["status"])

    # Source 4: /api/team/{id}/squad
    url4   = f"{BASE}/api/team/{BRAZIL_TEAM_ID}/squad"
    label4 = f"team_{BRAZIL_TEAM_ID}_squad"
    p4, m4 = fetch(url4, label4, execute=execute)
    _add_source(p4, label4, url4, m4["status"])

    # Deduplicate across all sources (keep first occurrence)
    seen: set[str] = set()
    deduped: list[dict] = []
    for ep in all_ep:
        if ep["player_id"] not in seen:
            seen.add(ep["player_id"])
            deduped.append(ep)

    # Build lookups
    by_slug = {ep["slug"]: ep for ep in deduped if ep.get("slug")}
    by_norm = {_norm_key(ep["player_name"]): ep for ep in deduped}

    print(f"\n  Total unique players discovered: {len(deduped)}")
    print(f"  By slug: {len(by_slug)}, by norm name: {len(by_norm)}")
    return deduped, by_slug, by_norm, endpoint_tests


# ── Phase 2: Match FIFA roster → StatsHub IDs ────────────────────────────────

def _alias_lookup(fifa_name: str, by_slug: dict[str, dict]) -> dict | None:
    """Try known aliases for Brazilian players who go by nicknames."""
    norm = _norm(fifa_name)
    for fragment, common_name, slug in BRAZIL_ALIASES:
        if fragment in norm:
            if slug in by_slug:
                return {**by_slug[slug], "match_method": f"alias:{fragment}→{slug}",
                        "common_name": common_name}
    return None


def phase2_match_players(
    roster: list[dict],
    ep_players: list[dict],
    by_slug: dict[str, dict],
    by_norm: dict[str, dict],
) -> list[dict]:
    print("\n=== PHASE 2: Match FIFA roster → StatsHub IDs ===")
    results: list[dict] = []

    for r in roster:
        fifa_name   = r["player_name"] or ""
        current_id  = r["player_id"]
        status      = r["statshub_player_id_status"] or ""
        jersey      = r["jersey_number"]
        row_id      = r["row_id"]
        norm_fifa   = _norm_key(fifa_name)
        slug_guess  = _slug(norm_fifa)

        if status in ("confirmed", "skipped_existing") and current_id:
            results.append({
                "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                "matched_statshub_name": "already_confirmed",
                "player_id": current_id, "match_method": "existing_confirmed",
                "confidence": 1.0, "status": "existing_confirmed",
            })
            continue

        match = None
        method = ""

        # Pass 1: exact normalized name
        if norm_fifa in by_norm:
            match = by_norm[norm_fifa]
            method = "exact_norm"

        # Pass 2: slug match
        if not match and slug_guess in by_slug:
            match = by_slug[slug_guess]
            method = "slug_exact"

        # Pass 3: alias table
        if not match:
            alias = _alias_lookup(norm_fifa, by_slug)
            if alias:
                match = alias
                method = alias.get("match_method", "alias")

        # Pass 4: token subset (wanted ⊆ db or db ⊆ wanted)
        if not match:
            wanted_tokens = set(norm_fifa.split())
            for ep_norm, ep in by_norm.items():
                ep_tokens = set(ep_norm.split())
                if wanted_tokens and (wanted_tokens <= ep_tokens or ep_tokens <= wanted_tokens):
                    match = ep
                    method = f"token_subset:{ep_norm}"
                    break

        # Pass 5: slug substring (norm_slug in ep_slug)
        if not match:
            for ep in ep_players:
                ep_slug = ep.get("slug", "")
                if ep_slug and (slug_guess in ep_slug or ep_slug in slug_guess) and len(slug_guess) >= 5:
                    match = ep
                    method = f"slug_substring:{slug_guess}↔{ep_slug}"
                    break

        if match:
            pid = match["player_id"]
            pname = match.get("player_name", "")
            confidence = 0.95 if method in ("exact_norm", "slug_exact", "alias") else 0.85
            results.append({
                "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                "matched_statshub_name": pname,
                "player_id": pid, "match_method": method,
                "confidence": confidence, "status": "matched",
            })
            print(f"  [MATCH] #{jersey:>2} {fifa_name[:40]:40s} → {pname} (id={pid}) [{method}]")
        else:
            results.append({
                "row_id": row_id, "jersey": jersey, "fifa_name": fifa_name,
                "matched_statshub_name": "", "player_id": None,
                "match_method": "unresolved", "confidence": 0.0, "status": "unresolved",
            })
            print(f"  [UNRESOLVED] #{jersey:>2} {fifa_name}")

    matched = sum(1 for r in results if r["status"] not in ("unresolved",))
    print(f"\n  Matched: {matched}/{len(results)}")
    return results


# ── Phase 3: Commit matches to DB ────────────────────────────────────────────

def phase3_commit_matches(con: Any, matches: list[dict]) -> int:
    print("\n=== PHASE 3: Commit player IDs to statshub_team_players ===")
    newly_confirmed = 0
    for m in matches:
        if m["status"] in ("existing_confirmed",):
            continue
        if m["status"] == "matched" and m["player_id"]:
            con.execute("""
                UPDATE statshub_team_players
                SET player_id=?, statshub_player_id_status='confirmed',
                    player_id_match_method=?, player_id_confidence_score=?,
                    player_id_match_source=?, updated_at=?
                WHERE id=?
            """, (m["player_id"], m["match_method"], m["confidence"] * 100,
                  "brazil_player_discovery_pilot", utc_now(), m["row_id"]))
            newly_confirmed += 1
    con.commit()
    print(f"  Newly confirmed: {newly_confirmed}")
    return newly_confirmed


# ── Phase 4: Cross-match Bet365 player names ─────────────────────────────────

def phase4_bet365_crossmatch(con: Any, by_slug: dict, by_norm: dict) -> list[dict]:
    """
    Get all Brazil player names from betting_odds_normalized for Brazil vs Morocco.
    For each, attempt to find the StatsHub player_id.
    """
    print("\n=== PHASE 4: Bet365 player name cross-match ===")

    # All unique player names in the match (team=None since Brazil props aren't tagged)
    b365_names = [r[0] for r in con.execute("""
        SELECT DISTINCT player_name
        FROM betting_odds_normalized
        WHERE match_name='Brazil vs Morocco'
          AND player_name IS NOT NULL AND player_name != ''
          AND player_name NOT LIKE '%(Score)%'
          AND player_name NOT LIKE '%(Assist)%'
          AND player_name NOT LIKE '%(Score or Assist)%'
          AND player_name NOT IN (
              'No Goalscorer (Brazil)','No Goalscorer (Morocco)',
              'Brazil','Morocco','Draw'
          )
        ORDER BY player_name
    """).fetchall()]

    print(f"  Total unique player-like names in Brazil vs Morocco: {len(b365_names)}")

    # Get confirmed Brazil players from DB
    confirmed = {r["player_id"]: r["player_name"] for r in con.execute("""
        SELECT player_id, player_name FROM statshub_team_players
        WHERE team_name='Brazil' AND statshub_player_id_status='confirmed'
          AND player_id IS NOT NULL
    """).fetchall()}
    confirmed_by_norm = {_norm_key(v): k for k, v in confirmed.items()}
    confirmed_by_id   = {k: v for k, v in confirmed.items()}

    # Also get Morocco players (to separate them)
    morocco_norm = {_norm_key(r[0]) for r in con.execute("""
        SELECT DISTINCT player_name FROM statshub_team_players
        WHERE team_name='Morocco'
    """).fetchall()}

    results: list[dict] = []
    for b365_name in b365_names:
        b365_norm = _norm_key(b365_name)

        # Check if it looks like a Morocco player
        is_morocco = any(tok in b365_norm for tok in ("hakimi","mazraoui","amrabat","bounou",
            "saibari","bouaddi","ziyech","ezzalzouli","bounou","tagnaouti","halhal",
            "riad","el kaabi","amaimouni","mourabat","saadane","ouahdi","bellamm",
            "khannouss","talbi","sbai","diop","ounahi","aguerd","rahimi","yassine",
            "saleh","neil","anass","gessime","munir","brahim","marwan"))

        # Try to find StatsHub match
        pid = None
        ep_name = ""
        method = ""

        # Pass 1: exact on confirmed DB names
        if b365_norm in confirmed_by_norm:
            pid = confirmed_by_norm[b365_norm]
            ep_name = confirmed_by_id.get(pid, "")
            method = "exact_confirmed_db"
        else:
            # Pass 2: token subset on confirmed
            b365_tokens = set(b365_norm.split())
            for db_norm, db_pid in confirmed_by_norm.items():
                db_tokens = set(db_norm.split())
                if b365_tokens and (b365_tokens <= db_tokens or db_tokens <= b365_tokens):
                    pid = db_pid
                    ep_name = confirmed_by_id.get(pid, "")
                    method = "token_subset_confirmed"
                    break

        # Pass 3: direct endpoint lookup (useful for uncollected players)
        if not pid and b365_norm in by_norm:
            ep = by_norm[b365_norm]
            pid = ep["player_id"]
            ep_name = ep["player_name"]
            method = "endpoint_exact"
        if not pid:
            b365_tokens = set(b365_norm.split())
            for en, ep in by_norm.items():
                et = set(en.split())
                if b365_tokens and (b365_tokens <= et or et <= b365_tokens):
                    pid = ep["player_id"]
                    ep_name = ep["player_name"]
                    method = f"endpoint_token:{en}"
                    break

        # Pass 4: alias table
        if not pid:
            alias = _alias_lookup(b365_name, by_slug)
            if alias:
                pid = alias["player_id"]
                ep_name = alias["player_name"]
                method = alias.get("match_method", "alias")

        team_guess = "Morocco" if is_morocco else "Brazil"
        status_str = "matched" if pid else ("unmatched_morocco" if is_morocco else "unmatched_brazil")

        results.append({
            "bet365_player_name": b365_name,
            "normalized_bet365_name": b365_norm,
            "matched_statshub_name": ep_name,
            "statshub_player_id": pid or "",
            "match_method": method,
            "confidence": 0.9 if method.endswith("exact") or method == "alias" else (0.8 if pid else 0.0),
            "team_guess": team_guess,
            "status": status_str,
        })

    unmatched_brazil = [r for r in results if r["status"] == "unmatched_brazil"]
    unmatched_morocco = [r for r in results if r["status"] == "unmatched_morocco"]
    matched = [r for r in results if r["status"] == "matched"]
    print(f"  Matched: {len(matched)} | Unmatched-Brazil: {len(unmatched_brazil)} | Unmatched-Morocco: {len(unmatched_morocco)}")
    if unmatched_brazil:
        print("  UNMATCHED Brazil players:")
        for r in unmatched_brazil:
            print(f"    {r['bet365_player_name']!r}")
    return results


# ── Phase 5: Download player performance ─────────────────────────────────────

def _parse_player_perf(payload: Any, pid: str, pname: str, tname: str, ep: str) -> list[dict]:
    rows: list[dict] = []
    if not payload:
        return rows
    items = payload if isinstance(payload, list) else payload.get("data", payload.get("events", []))
    if not isinstance(items, list):
        return rows
    for item in items:
        if not isinstance(item, dict):
            continue
        stats  = item.get("player_statistics_event") or item
        evt    = item.get("events") or {}
        row_pid = str(stats.get("playerId") or stats.get("player_id") or "")
        if row_pid and row_pid != str(pid):
            continue
        ts = evt.get("startTimestamp")
        date_v = (datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d")
                  if ts else evt.get("date") or item.get("date") or "")
        tourn = evt.get("tournament") or item.get("tournament") or {}
        tn = tourn.get("name", "") if isinstance(tourn, dict) else str(tourn or "")
        rows.append({
            "player_id": pid, "player_name": pname, "team_name": tname,
            "endpoint_name": ep, "event_date": str(date_v)[:10],
            "tournament_name": tn,
            "minutes_played":  _to_float(stats.get("minutesPlayed") or stats.get("minutes")),
            "goals":           _to_float(stats.get("goals") or stats.get("goalsScored")),
            "assists":         _to_float(stats.get("goalAssist") or stats.get("assists")),
            "shots":           _to_float(stats.get("shots") or stats.get("totalShots")),
            "shots_on_target": _to_float(stats.get("onTargetScoringAttempt") or stats.get("shotsOnTarget")),
            "fouls":           _to_float(stats.get("fouls") or stats.get("foulsCommitted")),
            "was_fouled":      _to_float(stats.get("wasFouled") or stats.get("foulsSuffered")),
            "yellow_cards":    _to_float(stats.get("yellowCard") or stats.get("yellowCards")),
            "red_cards":       _to_float(stats.get("redCard") or stats.get("redCards")),
            "xG":              _to_float(stats.get("expectedGoals") or stats.get("xG")),
            "xA":              _to_float(stats.get("expectedAssists") or stats.get("xA")),
            "key_passes":      _to_float(stats.get("keyPass") or stats.get("keyPasses")),
            "passes":          _to_float(stats.get("totalPass") or stats.get("passes") or stats.get("totalPasses")),
            "accurate_passes": _to_float(stats.get("accuratePass") or stats.get("accuratePasses")),
            "tackles":         _to_float(stats.get("totalTackle") or stats.get("tackles")),
            "possession_lost": _to_float(stats.get("possessionLostCtrl") or stats.get("possessionLost")),
            "raw_json": json.dumps(item)[:1000],
        })
    return rows


def phase5_download_performance(con: Any, execute: bool) -> list[dict]:
    print("\n=== PHASE 5: Download player performance for confirmed Brazil players ===")
    players = [dict(r) for r in con.execute("""
        SELECT player_id, player_name
        FROM statshub_team_players
        WHERE team_name='Brazil'
          AND statshub_player_id_status IN ('confirmed','skipped_existing')
          AND player_id IS NOT NULL AND player_id != ''
        ORDER BY player_name
    """).fetchall()]
    print(f"  Confirmed Brazil players to download: {len(players)}")

    results: list[dict] = []
    ep_suffix = "brazil_pilot"

    for p in players:
        pid   = p["player_id"]
        pname = p["player_name"]
        ep    = f"player_{pid}_performance_limit50_{ep_suffix}"
        url   = f"{BASE}/api/player/{pid}/performance?limit=50"

        existing = con.execute(
            "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=? AND endpoint_name=?",
            (pid, ep),
        ).fetchone()[0]
        if existing > 0:
            results.append({
                "player_name": pname, "player_id": pid, "status": "db_cached",
                "events_inserted": 0, "existing_events": existing,
                "valid_min15": 0, "has_shots": False, "has_goals": False,
            })
            print(f"  {pname} (id={pid}): db_cached ({existing} events)")
            continue

        payload, meta = fetch(url, ep, execute=execute)
        if meta["status"] == "dry_run":
            results.append({
                "player_name": pname, "player_id": pid, "status": "dry_run",
                "events_inserted": 0, "existing_events": 0,
                "valid_min15": 0, "has_shots": False, "has_goals": False,
            })
            continue

        perf_rows = _parse_player_perf(payload, pid, pname, "Brazil", ep) if payload else []
        if perf_rows:
            con.execute(
                "DELETE FROM statshub_player_performance_events WHERE player_id=? AND endpoint_name=?",
                (pid, ep),
            )
            for row in perf_rows:
                con.execute("""
                    INSERT INTO statshub_player_performance_events
                        (snapshot_name, player_id, player_name, team_name, endpoint_name,
                         event_date, tournament_name, minutes_played, goals, assists,
                         shots, shots_on_target, fouls, was_fouled, yellow_cards, red_cards,
                         xG, xA, key_passes, passes, accurate_passes, tackles,
                         possession_lost, raw_json, imported_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (SNAPSHOT_NAME, pid, pname, "Brazil", ep,
                      row["event_date"], row["tournament_name"], row["minutes_played"],
                      row["goals"], row["assists"], row["shots"], row["shots_on_target"],
                      row["fouls"], row["was_fouled"], row["yellow_cards"], row["red_cards"],
                      row["xG"], row["xA"], row["key_passes"], row["passes"],
                      row["accurate_passes"], row["tackles"], row["possession_lost"],
                      row["raw_json"], utc_now()))
            con.commit()

        valid15 = sum(1 for row in perf_rows if (row["minutes_played"] or 0) >= 15)
        has_shots = any((row["shots"] or 0) > 0 for row in perf_rows)
        has_goals = any((row["goals"] or 0) > 0 for row in perf_rows)
        has_sot   = any((row["shots_on_target"] or 0) > 0 for row in perf_rows)

        results.append({
            "player_name": pname, "player_id": pid, "status": meta["status"],
            "events_inserted": len(perf_rows), "existing_events": 0,
            "valid_min15": valid15, "has_shots": has_shots, "has_goals": has_goals,
            "has_sot": has_sot,
        })
        print(f"  {pname} (id={pid}): events={len(perf_rows)} valid_min15={valid15} "
              f"[{meta['status']}]")

    total_events = sum(r["events_inserted"] for r in results)
    total_valid = sum(r["valid_min15"] for r in results)
    print(f"\n  Total: {len(results)} players, {total_events} events, {total_valid} with min15")
    return results


# ── Phase 6: Write workbook ───────────────────────────────────────────────────

def phase6_workbook(
    con: Any,
    roster: list[dict],
    ep_players: list[dict],
    endpoint_tests: list[dict],
    matches: list[dict],
    b365_results: list[dict],
    perf_results: list[dict],
    newly_confirmed: int,
) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import warnings
        warnings.filterwarnings("ignore", "Title is more than")
    except ImportError:
        print("  openpyxl not available — skipping workbook")
        return OUT_XLSX

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    GRN  = PatternFill("solid", fgColor="C6EFCE")
    RED  = PatternFill("solid", fgColor="FFC7CE")
    YEL  = PatternFill("solid", fgColor="FFEB9C")

    def hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = HDR_FONT
            c.fill = HDR_FILL

    def fit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    # Sheet 1: run_summary
    ws1 = wb.create_sheet("run_summary")
    hdr(ws1, ["key", "value"])
    conf_count = sum(1 for r in matches if r["status"] not in ("unresolved","existing_confirmed") and r.get("player_id"))
    ex_conf_count = sum(1 for r in matches if r["status"] == "existing_confirmed")
    for k, v in [
        ("run_date", utc_now()),
        ("brazil_team_id", BRAZIL_TEAM_ID),
        ("endpoint_sources_tried", len(endpoint_tests)),
        ("total_ep_players_discovered", len(ep_players)),
        ("fifa_roster_size", len(roster)),
        ("existing_confirmed", ex_conf_count),
        ("newly_confirmed_this_run", newly_confirmed),
        ("total_now_confirmed", ex_conf_count + newly_confirmed),
        ("still_unresolved", len(roster) - ex_conf_count - newly_confirmed),
        ("bet365_players_matched", sum(1 for r in b365_results if r["status"]=="matched")),
        ("bet365_players_unmatched_brazil", sum(1 for r in b365_results if r["status"]=="unmatched_brazil")),
        ("perf_events_inserted", sum(r["events_inserted"] for r in perf_results)),
        ("perf_players_valid_min15", sum(1 for r in perf_results if r.get("valid_min15",0) > 0)),
    ]:
        ws1.append([k, str(v)])
    fit(ws1)

    # Sheet 2: brazil_team_resolution
    ws2 = wb.create_sheet("brazil_team_resolution")
    hdr(ws2, ["key", "value"])
    for r in con.execute("""
        SELECT statshub_team_id, team_name, statshub_team_slug, group_name,
               source_confidence, statshub_match_status
        FROM statshub_world_cup_teams WHERE team_name='Brazil'
    """).fetchall():
        cols = ["statshub_team_id","team_name","statshub_team_slug","group_name","source_confidence","statshub_match_status"]
        for k, v in zip(cols, r):
            ws2.append([k, str(v or "")])
    n_perf = con.execute("SELECT COUNT(*) FROM statshub_team_performance_events WHERE team_name='Brazil'").fetchone()[0]
    ws2.append(["team_perf_events", str(n_perf)])
    ws2.append(["player_id_discovery_method", "team_4748_players_wc26_tournamentId16 + alias table + token matching"])
    fit(ws2)

    # Sheet 3: endpoint_tests
    ws3 = wb.create_sheet("endpoint_tests")
    hdr(ws3, ["endpoint_label", "url", "status", "players_extracted"])
    for et in endpoint_tests:
        ws3.append([et.get(k,"") for k in ["endpoint_label","url","status","players_extracted"]])
    fit(ws3)

    # Sheet 4: players_extracted_from_team_page
    ws4 = wb.create_sheet("players_extracted_from_team_page")
    hdr(ws4, ["player_id", "player_name", "slug", "position", "source"])
    for ep in ep_players:
        ws4.append([ep.get(k,"") for k in ["player_id","player_name","slug","position","source"]])
    fit(ws4)

    # Sheet 5: bet365_brazil_player_names
    ws5 = wb.create_sheet("bet365_brazil_player_names")
    hdr(ws5, ["bet365_player_name", "normalized_bet365_name", "team_guess",
               "matched_statshub_name", "statshub_player_id", "match_method",
               "confidence", "status"])
    for r in sorted(b365_results, key=lambda x: x["status"]):
        ws5.append([r.get(k,"") for k in ["bet365_player_name","normalized_bet365_name","team_guess",
                                            "matched_statshub_name","statshub_player_id","match_method",
                                            "confidence","status"]])
        fill = GRN if r["status"]=="matched" else (RED if r["status"]=="unmatched_brazil" else YEL)
        for c in ws5[ws5.max_row]: c.fill = fill
    fit(ws5)

    # Sheet 6: player_name_matching
    ws6 = wb.create_sheet("player_name_matching")
    hdr(ws6, ["jersey","fifa_name","matched_statshub_name","player_id",
               "match_method","confidence","status"])
    for m in matches:
        ws6.append([m.get(k,"") for k in ["jersey","fifa_name","matched_statshub_name","player_id",
                                            "match_method","confidence","status"]])
        fill = GRN if m["status"] in ("matched","existing_confirmed") else RED
        for c in ws6[ws6.max_row]: c.fill = fill
    fit(ws6)

    # Sheet 7: player_performance_downloads
    ws7 = wb.create_sheet("player_performance_downloads")
    hdr(ws7, ["player_name","player_id","status","events_inserted","existing_events",
               "valid_min15","has_shots","has_goals"])
    for r in perf_results:
        ws7.append([r.get(k,"") for k in ["player_name","player_id","status","events_inserted",
                                            "existing_events","valid_min15","has_shots","has_goals"]])
    fit(ws7)

    # Sheet 8: unresolved_or_ambiguous
    ws8 = wb.create_sheet("unresolved_or_ambiguous")
    hdr(ws8, ["jersey","fifa_name","status","note"])
    unresolved_bets = [r for r in b365_results if r["status"]=="unmatched_brazil"]
    for m in matches:
        if m["status"] == "unresolved":
            ws8.append([m["jersey"], m["fifa_name"], "unresolved_fifa",
                        "Not found in any endpoint or alias table"])
            for c in ws8[ws8.max_row]: c.fill = RED
    for r in unresolved_bets:
        ws8.append(["?", r["bet365_player_name"], "unmatched_bet365_brazil",
                    "Bet365 Brazil prop player has no StatsHub match"])
        for c in ws8[ws8.max_row]: c.fill = YEL
    fit(ws8)

    # Sheet 9: raw_sources
    ws9 = wb.create_sheet("raw_sources")
    hdr(ws9, ["endpoint_label","url","status","players_extracted","raw_file"])
    for et in endpoint_tests:
        ws9.append([et.get(k,"") for k in ["endpoint_label","url","status","players_extracted"]]
                   + [str(RAW_DIR / (et.get("endpoint_label","") + ".json"))])
    fit(ws9)

    wb.save(OUT_XLSX)
    print(f"\n  Workbook saved: {OUT_XLSX}")
    return OUT_XLSX


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true",
                        help="Make live StatsHub API calls (requires STATSHUB_ENABLED=true)")
    args = parser.parse_args()

    if args.execute:
        if os.environ.get("STATSHUB_ENABLED", "").lower() not in ("true","1","yes"):
            print("ERROR: --execute requires STATSHUB_ENABLED=true", file=sys.stderr)
            sys.exit(1)
        print("Mode: LIVE")
    else:
        print("Mode: DRY RUN (local files only)")

    RAW_DIR.mkdir(parents=True, exist_ok=True)

    with connect() as con:
        roster = phase0_load_roster(con)
        ep_players, by_slug, by_norm, ep_tests = phase1_parse_sources(execute=args.execute)
        matches = phase2_match_players(roster, ep_players, by_slug, by_norm)
        newly_confirmed = phase3_commit_matches(con, matches)
        b365 = phase4_bet365_crossmatch(con, by_slug, by_norm)
        perf = phase5_download_performance(con, execute=args.execute)
        wb_path = phase6_workbook(con, roster, ep_players, ep_tests,
                                   matches, b365, perf, newly_confirmed)

    # ── Final report ──────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("BRAZIL PLAYER DISCOVERY — FINAL REPORT")
    print("=" * 70)
    print(f"Brazil team_id: {BRAZIL_TEAM_ID}")
    print(f"Working endpoint: team_4748_players_wc26_tournamentId16.json")
    print(f"                  (tournamentId=16, data[].id + data[].name)")
    print(f"Endpoint players discovered: {len(ep_players)}")
    print(f"FIFA roster size: 26")
    conf = sum(1 for m in matches if m.get("player_id"))
    unres = sum(1 for m in matches if not m.get("player_id"))
    print(f"Newly confirmed this run: {newly_confirmed}")
    print(f"Total confirmed (incl. pre-existing): {conf}")
    print(f"Still unresolved: {unres}")

    b365_matched = sum(1 for r in b365 if r["status"]=="matched")
    b365_unmatched_br = [r["bet365_player_name"] for r in b365 if r["status"]=="unmatched_brazil"]
    print(f"Bet365 players matched: {b365_matched}/{len([r for r in b365 if r['team_guess']=='Brazil'])}")
    if b365_unmatched_br:
        print(f"Unresolved Bet365 Brazil names: {b365_unmatched_br}")

    valid15 = sum(1 for r in perf if r.get("valid_min15", 0) > 0)
    total_ev = sum(r["events_inserted"] for r in perf)
    print(f"Player performance: {total_ev} events, {valid15} players with min15 appearances")

    unres_match = [m for m in matches if not m.get("player_id")]
    if unres_match:
        print(f"\nUnresolved FIFA players (need new endpoint):")
        for m in unres_match:
            print(f"  #{m['jersey']:>2}  {m['fifa_name']}")

    print(f"\nNext steps for remaining teams:")
    print(f"  Qatar/Switzerland/Morocco/Haiti/Scotland/Australia/Turkey:")
    print(f"  Same approach: parse team_{{id}}_players_wc26_tournamentId16.json,")
    print(f"  fix parser (use data[].id / data[].name), apply alias table.")
    print(f"  Turkey: already confirmed 10 players. Scotland 6. Morocco 17.")
    print(f"  Remaining gap: Qatar (3 confirmed), Australia (5), Switzerland (8).")
    print(f"\nWorkbook: {wb_path}")


if __name__ == "__main__":
    main()
