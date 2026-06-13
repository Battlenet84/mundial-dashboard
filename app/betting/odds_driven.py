from __future__ import annotations

import csv
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.betting.market_taxonomy import evaluate_market_mapping


DB_PATH = Path("data/mundial.db")
OUT_DIR = Path("data/processed/betting")
ODDS_API_CACHE = Path(
    "data/processed/odds_api_io/target_matches/canada_bosnia_world_cup/"
    "canada_bosnia_bet365_odds_normalized.csv"
)
TARGET_MATCHES = {
    "Canada vs Bosnia and Herzegovina": ("Canada", "Bosnia and Herzegovina"),
    "United States vs Paraguay": ("United States", "Paraguay"),
    "Qatar vs Switzerland": ("Qatar", "Switzerland"),
    "Brazil vs Morocco": ("Brazil", "Morocco"),
    "Haiti vs Scotland": ("Haiti", "Scotland"),
    "Australia vs Turkey": ("Australia", "Turkey"),
}
SNAP_DIRS = [
    Path("data/raw/statshub/snapshots/today_final_match_stats_probe"),
    Path("data/raw/statshub/snapshots/today_playwright_fixture_players_probe"),
    Path("data/raw/statshub/snapshots/today_browser_endpoint_replay_probe"),
]

RAW_COLUMNS = [
    "run_name", "source_name", "bookmaker", "match_name", "event_id",
    "api_event_id", "statshub_event_id", "raw_market_group", "raw_market_name",
    "raw_selection_name", "raw_line", "raw_odds", "odds_format", "captured_at",
    "raw_payload", "source_url", "request_id", "raw_file", "status", "notes",
]
TEMPLATE_COLUMNS = [
    "bookmaker", "match_name", "raw_market_group", "raw_market_name",
    "raw_selection_name", "raw_line", "raw_odds", "odds_format",
    "source_url", "captured_at", "notes",
]

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS betting_event_candidates (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_name TEXT, api_event_id TEXT, event_name TEXT,
    home_team TEXT, away_team TEXT, start_time TEXT,
    sport TEXT, league TEXT, confidence REAL, selected INTEGER,
    notes TEXT, captured_at TEXT
);

CREATE TABLE IF NOT EXISTS betting_odds_raw (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_name TEXT, run_name TEXT, bookmaker TEXT, match_name TEXT,
    event_id TEXT, api_event_id TEXT, statshub_event_id TEXT,
    raw_market_group TEXT, raw_market_name TEXT, raw_selection_name TEXT,
    raw_line REAL, raw_odds REAL, odds_format TEXT, captured_at TEXT,
    raw_payload TEXT, source_url TEXT, request_id TEXT, raw_file TEXT,
    status TEXT, notes TEXT,
    created_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS betting_odds_normalized (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    raw_id INTEGER, source_name TEXT, run_name TEXT, bookmaker TEXT,
    match_name TEXT, event_id TEXT, api_event_id TEXT, statshub_event_id TEXT,
    market_type TEXT, market_name TEXT, selection_type TEXT, side TEXT,
    team_name TEXT, team_id TEXT, player_name TEXT, player_id TEXT,
    line REAL, odds_decimal REAL, raw_market_name TEXT, raw_selection_name TEXT,
    raw_player_name TEXT, raw_side TEXT, raw_line REAL, raw_odds_decimal REAL,
    raw_event_id TEXT, raw_bookmaker TEXT, raw_market_index INTEGER,
    raw_outcome_index INTEGER, raw_source_file TEXT,
    market_mapping_status TEXT, market_mapping_reason TEXT,
    exact_market_match INTEGER, canonical_market_type TEXT,
    statshub_field_used TEXT, market_contract_version TEXT,
    model_uses_proxy INTEGER, field_mapping_status TEXT,
    side_line_status TEXT, data_completeness_status TEXT,
    normalized_status TEXT, match_confidence REAL, captured_at TEXT, notes TEXT
);
"""

VALUE_SCORE_SQL = """
CREATE TABLE IF NOT EXISTS betting_value_scores_new (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    rank INTEGER, source_name TEXT, bookmaker TEXT, match_name TEXT,
    market_type TEXT, raw_market_name TEXT, raw_selection_name TEXT,
    selection TEXT, team_name TEXT, player_name TEXT, player_id TEXT,
    side TEXT, line REAL, odds_decimal REAL, implied_probability REAL,
    model_probability REAL, edge REAL, expected_value REAL,
    probability_method TEXT, sample_size INTEGER, probability_status TEXT,
    normalized_status TEXT, data_quality_status TEXT, verdict TEXT,
    notes TEXT, computed_at TEXT,
    min_minutes_filter INTEGER, valid_appearance_count INTEGER,
    excluded_zero_minutes_count INTEGER, excluded_low_minutes_count INTEGER,
    minutes_filter_status TEXT,
    priority_class TEXT, market_scope TEXT, bet_description TEXT,
    market_mapping_status TEXT, market_mapping_reason TEXT,
    exact_market_match INTEGER, canonical_market_type TEXT,
    statshub_field_used TEXT, market_contract_version TEXT,
    model_uses_proxy INTEGER, field_mapping_status TEXT,
    side_line_status TEXT, data_completeness_status TEXT
);
"""

# ---------------------------------------------------------------------------
# Team alias normalization
# ---------------------------------------------------------------------------
TEAM_CANONICAL_MAP: dict[str, str] = {
    "united states": "United States", "usa": "United States",
    "usmnt": "United States", "eeuu": "United States",
    "estados unidos": "United States", "u s": "United States",
    "us": "United States",
    "paraguay": "Paraguay", "par": "Paraguay",
    "canada": "Canada", "can": "Canada",
    "bosnia": "Bosnia and Herzegovina", "bih": "Bosnia and Herzegovina",
    "bosnia and herzegovina": "Bosnia and Herzegovina",
    "bosnia herzegovina": "Bosnia and Herzegovina",
    # Today's 4 matches
    "brazil": "Brazil", "brasil": "Brazil", "bra": "Brazil",
    "morocco": "Morocco", "mar": "Morocco", "maroc": "Morocco",
    "qatar": "Qatar", "qat": "Qatar",
    "switzerland": "Switzerland", "sui": "Switzerland",
    "suisse": "Switzerland", "schweiz": "Switzerland",
    "haiti": "Haiti", "hai": "Haiti",
    "scotland": "Scotland", "sco": "Scotland",
    "australia": "Australia", "aus": "Australia", "socceroos": "Australia",
    "turkey": "Turkey", "tur": "Turkey", "turkiye": "Turkey",
    "türkiye": "Turkey",
}

PRIORITY_CLASS_MAP: dict[str, str] = {
    "player_total_shots": "hard_data_priority",
    "player_shots_on_target": "hard_data_priority",
    "player_fouls_committed": "hard_data_priority",
    "player_fouled": "hard_data_priority",
    "player_cards": "hard_data_priority",
    "player_to_be_booked": "hard_data_priority",
    "goalkeeper_saves": "hard_data_priority",
    "team_corners": "hard_data_priority",
    "total_corners": "hard_data_priority",
    "team_cards": "hard_data_priority",
    "total_cards": "hard_data_priority",
    "team_total_shots": "hard_data_priority",
    "team_shots_on_target": "hard_data_priority",
    "team_goals": "hard_data_priority",
    "over_under_goals": "hard_data_priority",
    "anytime_goalscorer": "medium_priority",
    "player_goals": "medium_priority",
    "player_assists": "medium_priority",
    "player_passes": "medium_priority",
    "player_tackles": "medium_priority",
    "both_teams_to_score": "medium_priority",
}

MARKET_SCOPE_MAP: dict[str, str] = {
    "player_total_shots": "player",
    "player_shots_on_target": "player",
    "player_fouls_committed": "player",
    "player_fouled": "player",
    "player_cards": "player",
    "player_goals": "player",
    "player_assists": "player",
    "player_tackles": "player",
    "player_passes": "player",
    "anytime_goalscorer": "player",
    "goalkeeper_saves": "player",
    "team_corners": "team",
    "team_cards": "team",
    "team_total_shots": "team",
    "team_shots_on_target": "team",
    "team_goals": "team",
    "total_corners": "match",
    "total_cards": "match",
    "over_under_goals": "match",
    "both_teams_to_score": "match",
    "match_result": "match",
    "draw_no_bet": "match",
    "double_chance": "match",
    "spread": "match",
    "correct_score": "match",
}

MARKET_LABEL_MAP: dict[str, str] = {
    "player_shots_on_target": "shots on target",
    "player_total_shots": "total shots",
    "player_fouls_committed": "fouls committed",
    "player_fouled": "fouled",
    "player_cards": "cards",
    "player_goals": "goals",
    "player_assists": "assists",
    "player_tackles": "tackles",
    "player_passes": "passes",
    "goalkeeper_saves": "goalkeeper saves",
    "anytime_goalscorer": "anytime goalscorer",
    "team_corners": "corners",
    "total_corners": "total corners",
    "team_cards": "cards",
    "total_cards": "total cards",
    "team_total_shots": "total shots",
    "team_shots_on_target": "shots on target",
    "team_goals": "goals",
    "over_under_goals": "total goals",
    "both_teams_to_score": "BTTS",
    "match_result": "match result",
    "draw_no_bet": "draw no bet",
    "double_chance": "double chance",
    "spread": "handicap",
}

# Player stat keys for raw JSON snapshots (legacy path)
PLAYER_STAT_KEYS = {
    "player_shots_on_target": "onTargetScoringAttempt",
    "player_total_shots": "shots",
    "player_fouls_committed": "fouls",
    "player_fouled": "wasFouled",
    "player_cards": "yellowCard",
    "player_goals": "goals",
    "player_assists": "goalAssist",
    "anytime_goalscorer": "goals",
}

# Player stat columns for statshub_player_performance_events DB table
PLAYER_STAT_COLUMNS_DB: dict[str, str] = {
    "player_shots_on_target": "shots_on_target",
    "player_total_shots": "shots",
    "player_fouls_committed": "fouls",
    "player_fouled": "was_fouled",
    "player_cards": "yellow_cards",
    "player_goals": "goals",
    "player_assists": "assists",
    "anytime_goalscorer": "goals",
    "player_tackles": "tackles",
    "player_passes": "passes",
}

# Team stat columns in statshub_team_performance_events
TEAM_STAT_COLUMNS: dict[str, str] = {
    "team_shots_on_target": "shots_on_target",
    "team_total_shots": "shots",
    "team_cards": "yellow_cards",
    "team_corners": "corners",
    "team_goals": "goals_for",
    "goalkeeper_saves": "goalkeeper_saves",
}

# Markets that need COMBINED (home + away) totals
COMBINED_TEAM_STAT_COLUMNS: dict[str, str] = {
    "total_corners": "corners",
    "total_cards": "yellow_cards",
    "total_shots_on_target": "shots_on_target",
    "total_shots": "shots",
}


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def ensure_schema(con: sqlite3.Connection) -> None:
    con.executescript(SCHEMA_SQL)
    con.executescript(VALUE_SCORE_SQL)
    for col in ("api_event_id", "run_name", "statshub_event_id", "request_id", "raw_file"):
        ensure_column(con, "betting_odds_raw", col, "TEXT")
    for col in ("api_event_id", "run_name", "statshub_event_id"):
        ensure_column(con, "betting_odds_normalized", col, "TEXT")
    for col, typ in (
        ("raw_player_name", "TEXT"),
        ("raw_side", "TEXT"),
        ("raw_line", "REAL"),
        ("raw_odds_decimal", "REAL"),
        ("raw_event_id", "TEXT"),
        ("raw_bookmaker", "TEXT"),
        ("raw_market_index", "INTEGER"),
        ("raw_outcome_index", "INTEGER"),
        ("raw_source_file", "TEXT"),
        ("market_mapping_status", "TEXT"),
        ("market_mapping_reason", "TEXT"),
        ("exact_market_match", "INTEGER"),
        ("canonical_market_type", "TEXT"),
        ("statshub_field_used", "TEXT"),
        ("market_contract_version", "TEXT"),
        ("model_uses_proxy", "INTEGER"),
        ("field_mapping_status", "TEXT"),
        ("side_line_status", "TEXT"),
        ("data_completeness_status", "TEXT"),
    ):
        ensure_column(con, "betting_odds_normalized", col, typ)
    # v3 columns: minutes-filter metadata
    for col, ddl in [
        ("min_minutes_filter", "INTEGER"),
        ("valid_appearance_count", "INTEGER"),
        ("excluded_zero_minutes_count", "INTEGER"),
        ("excluded_low_minutes_count", "INTEGER"),
        ("minutes_filter_status", "TEXT"),
        ("priority_class", "TEXT"),
        ("market_scope", "TEXT"),
        ("bet_description", "TEXT"),
        ("market_mapping_status", "TEXT"),
        ("market_mapping_reason", "TEXT"),
        ("exact_market_match", "INTEGER"),
        ("canonical_market_type", "TEXT"),
        ("statshub_field_used", "TEXT"),
        ("market_contract_version", "TEXT"),
        ("model_uses_proxy", "INTEGER"),
        ("field_mapping_status", "TEXT"),
        ("side_line_status", "TEXT"),
        ("data_completeness_status", "TEXT"),
    ]:
        ensure_column(con, "betting_value_scores_new", col, ddl)
    con.commit()


def ensure_column(con: sqlite3.Connection, table: str, column: str, ddl_type: str) -> None:
    existing = {row[1] for row in con.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in existing:
        con.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl_type}")


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def canonical(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()


def normalize_team_name(raw: Any) -> str | None:
    """Map bookmaker team name (e.g. 'USA') to canonical DB name ('United States')."""
    if not raw:
        return None
    return TEAM_CANONICAL_MAP.get(canonical(str(raw)))


def extract_label_from_raw(raw_payload_json: str | None) -> str | None:
    """
    Extract the player/selection label from stored raw_payload JSON.
    Handles bookmaker formats like 'Christian Pulisic (1)' and 'Timothy Weah (First)'.
    """
    if not raw_payload_json:
        return None
    try:
        payload = json.loads(raw_payload_json)
        outcome = payload.get("outcome", {}) if isinstance(payload, dict) else {}
        label = outcome.get("label") or outcome.get("name") or outcome.get("participant")
        if label:
            # Strip trailing parenthetical suffixes like "(1)", "(First)", "(Anytime)"
            clean = re.sub(r"\s*\([^)]{0,30}\)\s*$", "", str(label)).strip()
            return clean if clean else None
    except Exception:
        return None
    return None


def extract_raw_payload_meta(raw_payload_json: str | None) -> dict[str, Any]:
    if not raw_payload_json:
        return {}
    try:
        payload = json.loads(raw_payload_json)
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    outcome = payload.get("outcome")
    if not isinstance(outcome, dict):
        outcome = {}
    raw_player = outcome.get("label") or outcome.get("name") or outcome.get("participant")
    if raw_player:
        raw_player = re.sub(r"\s*\([^)]{0,30}\)\s*$", "", str(raw_player)).strip() or None
    return {
        "raw_player_name": raw_player,
        "raw_market_index": payload.get("market_index"),
        "raw_outcome_index": payload.get("odds_index"),
    }


def match_name_for_team(team_name: str | None) -> str | None:
    for match_name, teams in TARGET_MATCHES.items():
        if team_name in teams:
            return match_name
    return None


def write_actual_odds_template(path: Path = OUT_DIR / "today_actual_odds_input_template.xlsx") -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ActualOddsInput"
    ws.append(TEMPLATE_COLUMNS)
    examples = [
        ["Bet365", "Canada vs Bosnia and Herzegovina", "Player Props", "Player Shots on Target",
         "Alphonso Davies Over 0.5", 0.5, "", "decimal", "", "", "Only markets visible on bookmaker page"],
        ["Bet365", "United States vs Paraguay", "Team Props", "United States Total Shots on Target",
         "Over 3.5", 3.5, "", "decimal", "", "", ""],
    ]
    for row in examples:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [16, 34, 20, 34, 34, 10, 12, 12, 28, 24, 36]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path


def insert_raw_rows(con: sqlite3.Connection, rows: list[dict[str, Any]], replace: bool = False) -> int:
    ensure_schema(con)
    if replace:
        con.execute("DELETE FROM betting_odds_raw")
    sql = f"""
        INSERT INTO betting_odds_raw ({','.join(RAW_COLUMNS)})
        VALUES ({','.join(['?'] * len(RAW_COLUMNS))})
    """
    con.executemany(sql, [[row.get(col) for col in RAW_COLUMNS] for row in rows])
    con.commit()
    return len(rows)


def raw_rows_from_manual_file(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            rows = list(csv.DictReader(handle))
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["ActualOddsInput"] if "ActualOddsInput" in wb.sheetnames else wb.active
        values = list(ws.values)
        headers = [str(v) for v in values[0]]
        rows = [dict(zip(headers, row)) for row in values[1:]]
    out: list[dict[str, Any]] = []
    for row in rows:
        odds = to_float(row.get("raw_odds"))
        if odds is None or odds <= 1:
            continue
        out.append({
            "source_name": "manual_actual_odds", "run_name": "manual_actual_odds",
            "bookmaker": row.get("bookmaker"), "match_name": row.get("match_name"),
            "event_id": None, "api_event_id": None, "statshub_event_id": None,
            "raw_market_group": row.get("raw_market_group"),
            "raw_market_name": row.get("raw_market_name"),
            "raw_selection_name": row.get("raw_selection_name"),
            "raw_line": to_float(row.get("raw_line")), "raw_odds": odds,
            "odds_format": row.get("odds_format") or "decimal",
            "captured_at": row.get("captured_at") or now_utc(),
            "raw_payload": json.dumps(row, ensure_ascii=False),
            "source_url": row.get("source_url"), "request_id": None,
            "raw_file": None, "status": "raw", "notes": row.get("notes"),
        })
    return out


def raw_rows_from_odds_api_cache(path: Path = ODDS_API_CACHE) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    out: list[dict[str, Any]] = []
    for row in rows:
        odds = to_float(row.get("odds_decimal"))
        if odds is None or odds <= 1:
            continue
        out.append({
            "source_name": "odds_api_io_cache", "run_name": "odds_api_io_cache",
            "bookmaker": row.get("bookmaker") or "Bet365",
            "match_name": f"{row.get('home')} vs {row.get('away')}",
            "event_id": row.get("event_id"), "api_event_id": row.get("event_id"),
            "statshub_event_id": None,
            "raw_market_group": row.get("market_name"),
            "raw_market_name": row.get("market_name"),
            "raw_selection_name": row.get("selection_name"),
            "raw_line": to_float(row.get("line")), "raw_odds": odds,
            "odds_format": "decimal",
            "captured_at": row.get("fetched_at_utc") or now_utc(),
            "raw_payload": json.dumps(row, ensure_ascii=False),
            "source_url": row.get("source_file"), "request_id": None,
            "raw_file": row.get("source_file"), "status": "raw",
            "notes": "Imported from cached Odds-API.io normalized CSV",
        })
    return out


# ---------------------------------------------------------------------------
# Player / team lookup
# ---------------------------------------------------------------------------

def lookup_team_id(con: sqlite3.Connection, team_name: str | None) -> str | None:
    if not team_name:
        return None
    row = con.execute(
        "SELECT statshub_team_id FROM statshub_world_cup_teams WHERE team_name=? LIMIT 1",
        (team_name,),
    ).fetchone()
    return str(row[0]) if row and row[0] is not None else None


def lookup_player(
    con: sqlite3.Connection,
    player_name: str | None,
    match_name: str | None,
) -> tuple[str | None, str | None, float]:
    if not player_name:
        return None, None, 0.0
    teams = TARGET_MATCHES.get(match_name or "", ())
    candidates = con.execute(
        """
        SELECT player_name, player_id, team_name
        FROM statshub_team_players
        WHERE statshub_player_id_status IN ('confirmed','skipped_existing')
          AND player_id IS NOT NULL
        """,
    ).fetchall()

    wanted = canonical(player_name)
    wanted_tokens = set(wanted.split())
    wanted_collapsed = wanted.replace(" ", "")

    # Pass 1: exact canonical match
    for row in candidates:
        if teams and row["team_name"] not in teams:
            continue
        if canonical(row["player_name"]) == wanted:
            return row["player_name"], str(row["player_id"]), 1.0

    # Pass 2: token-subset match (all bookmaker tokens appear in DB name tokens)
    for row in candidates:
        if teams and row["team_name"] not in teams:
            continue
        db_tokens = set(canonical(row["player_name"]).split())
        if wanted_tokens and (wanted_tokens <= db_tokens or db_tokens <= wanted_tokens):
            return row["player_name"], str(row["player_id"]), 0.9

    # Pass 3: space-collapsed match ("mckennie" in "westonjamesearlmckennie")
    if len(wanted_collapsed) >= 5:
        for row in candidates:
            if teams and row["team_name"] not in teams:
                continue
            db_collapsed = canonical(row["player_name"]).replace(" ", "")
            if wanted_collapsed in db_collapsed or db_collapsed in wanted_collapsed:
                return row["player_name"], str(row["player_id"]), 0.85

    # Pass 4: substring match (original fallback)
    for row in candidates:
        if teams and row["team_name"] not in teams:
            continue
        db_can = canonical(row["player_name"])
        if wanted and (wanted in db_can or db_can in wanted):
            return row["player_name"], str(row["player_id"]), 0.75

    return player_name, None, 0.0


# ---------------------------------------------------------------------------
# Market parsing helpers
# ---------------------------------------------------------------------------

# Keywords that indicate a player-prop market where outcome.label = player name
_PLAYER_LABEL_MARKET_KEYWORDS = frozenset({
    "goalscorer", "scorer", "player", "booked", "tackle", "foul",
    "pass", "save", "headed", "assist", "goal", "shot", "card",
})


def _is_player_label_market(market_name: str) -> bool:
    """True when the market uses outcome.label as player name (not as Over/Under prefix)."""
    ml = canonical(market_name)
    return any(kw in ml for kw in _PLAYER_LABEL_MARKET_KEYWORDS)


def parse_side_and_line(selection: str | None, raw_line: float | None) -> tuple[str | None, float | None]:
    text = str(selection or "")
    line = raw_line
    side = None
    sel_l = text.lower()
    if "yes" in sel_l and "no" not in sel_l:
        return "yes", line
    if "no" in sel_l and "yes" not in sel_l:
        return "no", line
    if re.search(r"\bover\b", text, re.I):
        side = "over"
    elif re.search(r"\bunder\b", text, re.I):
        side = "under"
    m = re.search(r"(\d+(?:\.\d+)?)\s*\+", text)
    if m:
        side = "over"
        line = float(m.group(1)) - 0.5
    elif line is None:
        nums = re.findall(r"[-+]?\d+(?:\.\d+)?", text)
        if nums:
            line = float(nums[-1])
    return side, line


def _classify_market(market_l: str, teams: tuple[str, ...]) -> tuple[str, bool, bool]:
    """
    Returns (market_type, is_team_market, is_player_market).
    is_team_market: team-level stat needing team_name
    is_player_market: player-level stat needing player_name
    """
    # --- Match result ---
    if market_l in ("ml", "1x2", "match result"):
        return "match_result", False, False
    if "draw no bet" in market_l:
        return "draw_no_bet", False, False
    if "double chance" in market_l:
        return "double_chance", False, False
    if "half time result" in market_l or market_l == "ht result":
        return "match_result_ht", False, False

    # --- Spread / handicap ---
    if market_l in ("spread", "asian handicap", "handicap") or market_l.startswith("spread "):
        return "spread", False, False
    if "european handicap" in market_l:
        return "spread", False, False
    if "alternative asian handicap" in market_l:
        return "spread", False, False
    if "1st half handicap" in market_l:
        return "spread_ht", False, False
    if "spread ht" in market_l:
        return "spread_ht", False, False

    # --- Goals / totals ---
    if market_l in ("totals", "goals over under") or (
        ("over under" in market_l or "totals" in market_l)
        and "corner" not in market_l and "card" not in market_l
        and "booking" not in market_l and "ht" not in market_l
    ):
        return "over_under_goals", False, False
    if "totals ht" in market_l:
        return "totals_ht", False, False
    if "alternative total goals" in market_l:
        return "over_under_goals", False, False
    if "alternative goal line" in market_l:
        return "over_under_goals", False, False
    if "number of goals" in market_l:
        return "over_under_goals", False, False
    if "exact total goals" in market_l:
        return "over_under_goals", False, False
    if "team total goals" in market_l:
        return "team_goals", True, False
    if "correct score" in market_l:
        return "correct_score", False, False
    if "goal method" in market_l:
        return "goal_method", False, False

    # --- BTTS ---
    if "both teams to score" in market_l or market_l in ("btts",):
        return "both_teams_to_score", False, False

    # --- Corners ---
    if "corners spread" in market_l or "corner handicap" in market_l:
        return "team_corners", True, False
    if ("corners" in market_l or "corner" in market_l) and (
        "totals" in market_l or "total" in market_l or "2 way" in market_l
        or "ht" in market_l or market_l == "corners"
    ):
        return "total_corners", False, False
    if "alternative corners" in market_l:
        return "team_corners", True, False
    if "corners race" in market_l:
        return "corners_race", False, False
    if "total corners" in market_l:
        return "total_corners", False, False
    if "team corners" in market_l:
        return "team_corners", True, False

    # --- Cards / bookings ---
    if "bookings spread" in market_l:
        return "team_cards", True, False
    if "bookings totals" in market_l or "number of cards" in market_l:
        return "total_cards", False, False
    if "card handicap" in market_l:
        return "team_cards", True, False
    if "team cards" in market_l:
        return "team_cards", True, False

    # --- Shots ---
    if "team shots on target" in market_l:
        return "team_shots_on_target", True, False
    if "team shots" in market_l:
        return "team_total_shots", True, False
    if "match shots on target" in market_l:
        return "total_shots_on_target", False, False
    if "match shots" in market_l:
        return "total_shots", False, False

    # --- Tackles / passes (team/match level) ---
    if "match tackles" in market_l:
        return "total_tackles", False, False
    if "team tackles" in market_l:
        return "team_tackles", True, False
    if "match offsides" in market_l or "team offsides" in market_l:
        return "offsides", False, False

    # --- First N minutes ---
    if "first 10 minutes" in market_l or "first 10 min" in market_l:
        return "first_10_min", False, False

    # --- Goalkeeper ---
    if "goalkeeper saves" in market_l or "gk saves" in market_l:
        return "goalkeeper_saves", False, True

    # --- Player props ---
    if "shots on target outside" in market_l or "outside box" in market_l:
        return "player_shots_on_target_outside_box", False, True
    if "headed shots on target" in market_l:
        return "unsupported_market", False, True
    if "player shots on target" in market_l:
        return "player_shots_on_target", False, True
    if "player shots" in market_l:
        return "player_total_shots", False, True
    if "player to be fouled" in market_l or "player fouled" in market_l:
        return "player_fouled", False, True
    if "player fouls committed" in market_l or "player fouls" in market_l:
        return "player_fouls_committed", False, True
    if "player cards" in market_l:
        return "player_cards", False, True
    if "player to be booked" in market_l or "player booked" in market_l:
        return "player_cards", False, True
    if "player tackles" in market_l:
        return "player_tackles", False, True
    if "player passes" in market_l:
        return "player_passes", False, True
    if "player to score or assist" in market_l:
        return "player_score_or_assist", False, True
    if "player of the match" in market_l:
        return "player_of_the_match", False, True

    # --- Anytime / team goalscorer ---
    if "anytime goalscorer" in market_l or "team goalscorer" in market_l:
        return "anytime_goalscorer", False, True

    # --- Goals (player level) ---
    if "goal" in market_l and "over under" not in market_l and "method" not in market_l:
        return "player_goals", False, True
    if "assist" in market_l:
        return "player_assists", False, True

    # --- Multi-player / specials ---
    if "multi scorer" in market_l or "multi score" in market_l:
        return "multi_scorers", False, False
    if "specials" in market_l:
        return "specials", False, False

    return "unsupported_market", False, False


def parse_market(con: sqlite3.Connection, raw: sqlite3.Row) -> dict[str, Any]:
    market = str(raw["raw_market_name"] or "")
    selection = str(raw["raw_selection_name"] or "")
    match_name = raw["match_name"]
    market_l = canonical(market)
    selection_l = canonical(selection)
    teams: tuple[str, ...] = TARGET_MATCHES.get(match_name or "", ())

    side, line = parse_side_and_line(selection, raw["raw_line"])
    status = "ok"
    notes: list[str] = []
    selection_type = None
    team_name: str | None = None
    player_name: str | None = None
    player_id: str | None = None
    confidence: float = 1.0

    market_type, is_team, is_player = _classify_market(market_l, teams)

    # --- Player name: try label from raw_payload first, then parse selection ---
    raw_meta = extract_raw_payload_meta(raw["raw_payload"])
    raw_label = raw_meta.get("raw_player_name") or extract_label_from_raw(raw["raw_payload"])

    if is_player:
        if raw_label:
            # Use label directly (e.g. "Christian Pulisic")
            candidate = raw_label
        else:
            # Parse player name from selection string (e.g. "Christian Pulisic Over 1.5")
            candidate = re.sub(r"\b(over|under)\b.*$", "", selection, flags=re.I).strip(" -")
            candidate = re.sub(r"\s*\(\d+\)\s*$", "", candidate).strip()
        player_name, player_id, confidence = lookup_player(con, candidate, match_name)
        if not player_id:
            status = "unmatched_selection"
            notes.append(f"Player not matched: '{candidate}'")

    # --- Team name: normalize aliases (USA → United States) ---
    if is_team:
        # First: check if selection itself is a team alias
        team_from_sel = normalize_team_name(selection)
        if team_from_sel and team_from_sel in teams:
            team_name = team_from_sel
        else:
            # Second: check if any team's aliases appear in selection_l or market_l
            for t in teams:
                t_can = canonical(t)
                all_aliases = {t_can}
                if t == "United States":
                    all_aliases |= {"united states", "usa", "usmnt", "eeuu", "us"}
                elif t == "Paraguay":
                    all_aliases |= {"paraguay", "par"}
                elif t == "Canada":
                    all_aliases |= {"canada", "can"}
                elif t == "Bosnia and Herzegovina":
                    all_aliases |= {"bosnia", "bih", "bosnia and herzegovina"}
                elif t == "Brazil":
                    all_aliases |= {"brazil", "brasil", "bra"}
                elif t == "Morocco":
                    all_aliases |= {"morocco", "mar", "maroc", "marrocos"}
                elif t == "Qatar":
                    all_aliases |= {"qatar", "qat"}
                elif t == "Switzerland":
                    all_aliases |= {"switzerland", "sui", "suisse", "schweiz"}
                elif t == "Haiti":
                    all_aliases |= {"haiti", "haïti", "hai"}
                elif t == "Scotland":
                    all_aliases |= {"scotland", "sco"}
                elif t == "Australia":
                    all_aliases |= {"australia", "aus", "socceroos"}
                elif t == "Turkey":
                    all_aliases |= {"turkey", "tur", "turkiye", "türkiye"}
                if any(a in selection_l or a in market_l for a in all_aliases):
                    team_name = t
                    break
        if not team_name and status == "ok":
            status = "unmatched_selection"
            notes.append(f"Team not matched for '{market}' sel='{selection}'")

    # --- Match result / draw handling ---
    if market_type == "match_result":
        selection_type = "draw" if selection_l == "draw" else "team"
        if selection_type == "draw":
            team_name = None
        else:
            team_name = normalize_team_name(selection) or team_name
        side = selection_type

    # --- BTTS: side = yes/no ---
    if market_type == "both_teams_to_score":
        sel_l2 = selection_l
        if "yes" in sel_l2:
            side = "yes"
        elif "no" in sel_l2:
            side = "no"

    # --- Mark unsupported calculators ---
    _no_calculator = {
        "spread", "spread_ht", "draw_no_bet", "double_chance", "match_result",
        "match_result_ht", "totals_ht", "correct_score", "goal_method",
        "multi_scorers", "specials", "first_10_min", "player_of_the_match",
        "corners_race", "offsides", "total_offsides", "team_offsides",
        "total_tackles", "team_tackles", "player_score_or_assist",
    }
    if market_type in _no_calculator and status == "ok":
        status = "unsupported_market"

    if market_type == "player_tackles" and status == "ok":
        # Check if DB has tackle data
        if player_id:
            cnt = con.execute(
                "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=? AND tackles IS NOT NULL",
                (str(player_id),),
            ).fetchone()[0]
            if cnt < 3:
                status = "unsupported_market"
                notes.append("unsupported_calculator: insufficient tackles data")

    if market_type == "player_passes" and status == "ok":
        if player_id:
            cnt = con.execute(
                "SELECT COUNT(*) FROM statshub_player_performance_events WHERE player_id=? AND passes IS NOT NULL",
                (str(player_id),),
            ).fetchone()[0]
            if cnt < 3:
                status = "unsupported_market"
                notes.append("unsupported_calculator: insufficient passes data")

    if market_type == "unsupported_market":
        status = "unsupported_market"

    team_id = lookup_team_id(con, team_name)
    market_mapping = evaluate_market_mapping(market_type, market)
    data_completeness_status = "COMPLETE" if match_name in {
        "Haiti vs Scotland", "Australia vs Turkey", "Brazil vs Morocco"
    } else "PARTIAL"
    return {
        "raw_id": raw["id"],
        "source_name": raw["source_name"],
        "run_name": raw["run_name"],
        "bookmaker": raw["bookmaker"],
        "match_name": match_name,
        "event_id": raw["event_id"],
        "api_event_id": raw["api_event_id"],
        "statshub_event_id": raw["statshub_event_id"],
        "market_type": market_type,
        "market_name": market,
        "selection_type": selection_type,
        "side": side,
        "team_name": team_name,
        "team_id": team_id,
        "player_name": player_name,
        "player_id": player_id,
        "line": line,
        "odds_decimal": raw["raw_odds"],
        "raw_market_name": market,
        "raw_selection_name": selection,
        "raw_player_name": raw_meta.get("raw_player_name"),
        "raw_side": side,
        "raw_line": raw["raw_line"],
        "raw_odds_decimal": raw["raw_odds"],
        "raw_event_id": raw["event_id"],
        "raw_bookmaker": raw["bookmaker"],
        "raw_market_index": raw_meta.get("raw_market_index"),
        "raw_outcome_index": raw_meta.get("raw_outcome_index"),
        "raw_source_file": raw["raw_file"],
        "market_mapping_status": market_mapping["market_mapping_status"],
        "market_mapping_reason": market_mapping["market_mapping_reason"],
        "exact_market_match": 1 if market_mapping["exact_market_match"] else 0,
        "canonical_market_type": market_mapping["canonical_market_type"],
        "statshub_field_used": market_mapping["statshub_field_used"],
        "market_contract_version": market_mapping["market_contract_version"],
        "model_uses_proxy": 1 if market_mapping["model_uses_proxy"] else 0,
        "field_mapping_status": market_mapping["field_mapping_status"],
        "side_line_status": market_mapping["side_line_status"] if side and line is not None else "MISSING_SIDE_OR_LINE",
        "data_completeness_status": data_completeness_status,
        "normalized_status": status,
        "match_confidence": confidence,
        "captured_at": raw["captured_at"],
        "notes": "; ".join(notes) if notes else (raw["notes"] or ""),
    }


def normalize_raw_odds(con: sqlite3.Connection, replace: bool = True) -> dict[str, int]:
    ensure_schema(con)
    if replace:
        con.execute("DELETE FROM betting_odds_normalized")
    rows = con.execute("SELECT * FROM betting_odds_raw ORDER BY id").fetchall()
    parsed = [parse_market(con, row) for row in rows]
    if parsed:
        cols = list(parsed[0].keys())
        con.executemany(
            f"INSERT INTO betting_odds_normalized ({','.join(cols)}) VALUES ({','.join(['?'] * len(cols))})",
            [[row[col] for col in cols] for row in parsed],
        )
    con.commit()
    return {
        "raw": len(rows),
        "normalized": len(parsed),
        "supported": sum(1 for r in parsed if r["normalized_status"] == "ok"),
        "unsupported": sum(1 for r in parsed if r["normalized_status"] == "unsupported_market"),
        "unmatched": sum(1 for r in parsed if r["normalized_status"] == "unmatched_selection"),
    }


# ---------------------------------------------------------------------------
# Player event loaders
# ---------------------------------------------------------------------------

MIN_MINUTES_PLAYER = 15  # minimum appearance length counted in player prop samples


def load_player_events_from_db(
    con: sqlite3.Connection, player_id: str, stat_col: str,
    min_minutes: int = MIN_MINUTES_PLAYER,
) -> tuple[list[float], dict]:
    """
    Load player historical stat values filtered to valid appearances (minutes >= min_minutes).

    Returns (values, meta) where meta contains appearance diagnostics.
    """
    all_rows = con.execute(
        f"""
        SELECT {stat_col}, minutes_played
        FROM statshub_player_performance_events
        WHERE player_id = ?
          AND {stat_col} IS NOT NULL
        ORDER BY event_date DESC, id DESC
        LIMIT 50
        """,
        (str(player_id),),
    ).fetchall()

    total = len(all_rows)
    excluded_zero = 0
    excluded_low = 0
    valid_values: list[float] = []

    for row in all_rows:
        raw_min = row[1]
        try:
            mp = float(raw_min) if raw_min is not None else 0.0
        except (TypeError, ValueError):
            mp = 0.0

        if mp == 0.0:
            excluded_zero += 1
        elif mp < min_minutes:
            excluded_low += 1
        else:
            valid_values.append(float(row[0]))

    meta = {
        "total_db_rows": total,
        "valid_appearance_count": len(valid_values),
        "excluded_zero_minutes_count": excluded_zero,
        "excluded_low_minutes_count": excluded_low,
        "min_minutes_filter": min_minutes,
        "minutes_filter_status": "ok" if valid_values else "no_valid_appearances",
    }
    return valid_values, meta


def load_raw_player_events(player_id: str) -> list[dict[str, Any]]:
    """Load player events from raw JSON snapshot files (fallback)."""
    for directory in SNAP_DIRS:
        hits = list(directory.glob(f"perf_{player_id}*.json"))
        if not hits:
            continue
        try:
            payload = json.loads(hits[0].read_text(encoding="utf-8"))
            items = (
                payload.get("playerStatisticsEvents")
                or payload.get("events")
                or payload.get("data")
                or []
            )
            events = []
            for item in items:
                stats = item.get("player_statistics_event") if isinstance(item, dict) else None
                stats = stats or item
                if int(stats.get("minutesPlayed") or 0) > 0:
                    events.append(stats)
            return events
        except Exception:
            return []
    return []


def probability_from_values(
    values: list[float], side: str | None, line: float | None
) -> float | None:
    if not values or side not in {"over", "under"} or line is None:
        return None
    hits = sum(1 for v in values if (v > line if side == "over" else v < line))
    return round(hits / len(values), 4)


# ---------------------------------------------------------------------------
# EV Calculators
# ---------------------------------------------------------------------------

def calculate_probability(con: sqlite3.Connection, row: sqlite3.Row) -> dict[str, Any]:  # noqa: C901
    mtype = row["market_type"]
    side = row["side"]
    line = row["line"]
    notes: list[str] = []
    values: list[float] = []
    method: str = "historical_frequency"
    # Default appearance metadata for non-player markets
    extra: dict[str, Any] = {
        "valid_appearance_count": None,
        "excluded_zero_minutes_count": None,
        "excluded_low_minutes_count": None,
        "min_minutes_filter": None,
        "minutes_filter_status": "not_applicable",
    }

    if row["market_mapping_status"] not in ("OK",):
        verdict = "REVIEW" if row["market_mapping_status"] == "REVIEW" else "BLOCKED"
        return {
            "status": row["market_mapping_status"],
            "verdict": verdict,
            "data_quality_status": row["market_mapping_status"],
            "notes": row["market_mapping_reason"],
            **extra,
        }

    if row["exact_market_match"] != 1:
        return {
            "status": "BLOCKED_UNVERIFIED_MARKET",
            "verdict": "BLOCKED",
            "data_quality_status": "BLOCKED_UNVERIFIED_MARKET",
            "notes": "exact_market_match_false",
            **extra,
        }

    if row["model_uses_proxy"] or row["field_mapping_status"] in ("WRONG", "MISSING_FIELD") or row["side_line_status"] != "OK":
        return {
            "status": "BLOCKED_MAPPING_GUARDRAIL",
            "verdict": "BLOCKED",
            "data_quality_status": "BLOCKED_MAPPING_GUARDRAIL",
            "notes": "strict_actionable_guardrail_failed",
            **extra,
        }

    if row["normalized_status"] not in ("ok",):
        verdict = (
            "UNMATCHED" if row["normalized_status"] == "unmatched_selection"
            else "BLOCKED" if row["normalized_status"] == "BLOCKED_ODDS_MISMATCH"
            else "UNSUPPORTED"
        )
        return {"status": row["normalized_status"], "verdict": verdict, **extra}

    match_teams = TARGET_MATCHES.get(row["match_name"] or "", (None, None))
    home_team, away_team = match_teams

    # --- Over/under goals ---
    if mtype == "over_under_goals":
        events = con.execute(
            "SELECT goals_for, goals_against FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (home_team,),
        ).fetchall()
        values = [float((e["goals_for"] or 0) + (e["goals_against"] or 0)) for e in events]
        notes.append("Match total from home-team historical totals")
        method = "historical_frequency_match_total"

    # --- Team-level stats ---
    elif mtype in TEAM_STAT_COLUMNS and row["team_name"]:
        col = TEAM_STAT_COLUMNS[mtype]
        events = con.execute(
            f"SELECT {col} AS stat_value FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (row["team_name"],),
        ).fetchall()
        values = [float(e["stat_value"] or 0) for e in events]
        method = "historical_frequency"

    # --- Goalkeeper saves (team-level proxy) ---
    elif mtype == "goalkeeper_saves":
        # Find GK's team from player_id, use team-level historical saves
        player_team = None
        if row["player_id"]:
            pt = con.execute(
                "SELECT team_name FROM statshub_team_players WHERE player_id=? LIMIT 1",
                (str(row["player_id"]),),
            ).fetchone()
            if pt:
                player_team = pt["team_name"]
        if not player_team:
            player_team = home_team
        events = con.execute(
            "SELECT goalkeeper_saves AS stat_value FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (player_team,),
        ).fetchall()
        values = [float(e["stat_value"] or 0) for e in events]
        method = "team_historical_gk_saves"
        notes.append(f"Team-level GK saves for {player_team}")

    # --- Total corners (combined home + away historical) ---
    elif mtype == "total_corners":
        h_events = con.execute(
            "SELECT corners FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (home_team,),
        ).fetchall()
        a_events = con.execute(
            "SELECT corners FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (away_team,),
        ).fetchall()
        n = min(len(h_events), len(a_events))
        if n < 5:
            return {"status": "insufficient_data", "verdict": "UNSUPPORTED",
                    "sample_size": n, "notes": "insufficient team corner history"}
        values = [float(h_events[i]["corners"] or 0) + float(a_events[i]["corners"] or 0)
                  for i in range(n)]
        method = "historical_frequency_combined_teams"
        notes.append("Estimated from independent pairing of home+away historical corners")

    # --- Total cards (combined) ---
    elif mtype == "total_cards":
        h_events = con.execute(
            "SELECT yellow_cards FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (home_team,),
        ).fetchall()
        a_events = con.execute(
            "SELECT yellow_cards FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (away_team,),
        ).fetchall()
        n = min(len(h_events), len(a_events))
        if n < 5:
            return {"status": "insufficient_data", "verdict": "UNSUPPORTED",
                    "sample_size": n, "notes": "insufficient team card history"}
        values = [float(h_events[i]["yellow_cards"] or 0) + float(a_events[i]["yellow_cards"] or 0)
                  for i in range(n)]
        method = "historical_frequency_combined_teams"
        notes.append("Estimated from independent pairing of home+away historical yellow cards")

    # --- Both teams to score ---
    elif mtype == "both_teams_to_score":
        h_evs = con.execute(
            "SELECT goals_for FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (home_team,),
        ).fetchall()
        a_evs = con.execute(
            "SELECT goals_for FROM statshub_team_performance_events "
            "WHERE team_name=? ORDER BY event_date DESC LIMIT 50",
            (away_team,),
        ).fetchall()
        if len(h_evs) < 5 or len(a_evs) < 5:
            return {"status": "insufficient_data", "verdict": "UNSUPPORTED",
                    "notes": "insufficient team goal history for BTTS"}
        p_home = sum(1 for e in h_evs if float(e["goals_for"] or 0) > 0) / len(h_evs)
        p_away = sum(1 for e in a_evs if float(e["goals_for"] or 0) > 0) / len(a_evs)
        p_yes = round(p_home * p_away, 4)
        sel_l = canonical(row["raw_selection_name"] or "")
        prob = round(1 - p_yes, 4) if ("no" in sel_l and "yes" not in sel_l) else p_yes
        sample_size = min(len(h_evs), len(a_evs))
        dq = "ok" if sample_size >= 20 else "low_sample"
        return {
            "model_probability": prob,
            "probability_method": "btts_historical_frequency",
            "sample_size": sample_size,
            "data_quality_status": dq,
            "status": "ok",
            "notes": f"P(home scores)={p_home:.3f} × P(away scores)={p_away:.3f}",
        }

    # --- Player stats (DB-backed primary) ---
    elif mtype in PLAYER_STAT_COLUMNS_DB:
        if not row["player_id"]:
            return {
                "status": "INCOMPLETE_PLAYER_DATA",
                "verdict": "BLOCKED",
                "data_quality_status": "INCOMPLETE_PLAYER_DATA",
                "notes": "player prop missing stable player_id",
                **extra,
            }
        col = PLAYER_STAT_COLUMNS_DB[mtype]
        values, mp_meta = load_player_events_from_db(con, str(row["player_id"]), col)
        if not values:
            # Fallback to raw JSON files (also filters minutes > 0 by construction)
            events_raw = load_raw_player_events(str(row["player_id"]))
            json_key = PLAYER_STAT_KEYS.get(mtype)
            if json_key:
                values = [float(e.get(json_key) or 0) for e in events_raw]
            mp_meta = {
                "valid_appearance_count": len(values),
                "excluded_zero_minutes_count": 0,
                "excluded_low_minutes_count": 0,
                "min_minutes_filter": MIN_MINUTES_PLAYER,
                "minutes_filter_status": "fallback_raw_json" if values else "no_valid_appearances",
            }
        if mp_meta.get("minutes_filter_status") == "no_valid_appearances":
            return {
                "status": "INCOMPLETE_PLAYER_DATA", "verdict": "BLOCKED",
                "data_quality_status": "INCOMPLETE_PLAYER_DATA",
                "sample_size": 0,
                "valid_appearance_count": 0,
                "excluded_zero_minutes_count": mp_meta.get("excluded_zero_minutes_count", 0),
                "excluded_low_minutes_count": mp_meta.get("excluded_low_minutes_count", 0),
                "min_minutes_filter": MIN_MINUTES_PLAYER,
                "minutes_filter_status": "no_valid_appearances",
                "notes": f"no appearances with minutes >= {MIN_MINUTES_PLAYER}",
            }
        method = "historical_frequency"
        # Propagate appearance metadata
        extra = {
            "valid_appearance_count": mp_meta.get("valid_appearance_count", len(values)),
            "excluded_zero_minutes_count": mp_meta.get("excluded_zero_minutes_count", 0),
            "excluded_low_minutes_count": mp_meta.get("excluded_low_minutes_count", 0),
            "min_minutes_filter": MIN_MINUTES_PLAYER,
            "minutes_filter_status": "ok",
        }
    # --- Unsupported (no branch matched or values still empty) ---
    else:
        return {"status": "unsupported_market", "verdict": "UNSUPPORTED", **extra}

    if not values:
        return {"status": "unsupported_market", "verdict": "UNSUPPORTED",
                "sample_size": 0, "notes": "; ".join(notes), **extra}

    if mtype in PLAYER_STAT_COLUMNS_DB and len(values) < 5:
        return {
            "status": "INCOMPLETE_PLAYER_DATA",
            "verdict": "BLOCKED",
            "data_quality_status": "INCOMPLETE_PLAYER_DATA",
            "sample_size": len(values),
            "notes": f"player sample_size {len(values)} < 5",
            **extra,
        }

    prob = probability_from_values(values, side, line)
    if prob is None:
        return {
            "status": "unsupported_market", "verdict": "UNSUPPORTED",
            "sample_size": len(values), "notes": "; ".join(notes), **extra,
        }
    dq = "ok" if len(values) >= 20 else "low_sample" if len(values) >= 5 else "insufficient_data"
    return {
        "model_probability": prob,
        "probability_method": method,
        "sample_size": len(values),
        "data_quality_status": dq,
        "status": "ok" if dq != "insufficient_data" else "insufficient_data",
        "notes": "; ".join(notes),
        **extra,
    }


def _compute_bet_description(
    market_type: str, scope: str,
    player_name: str | None, team_name: str | None, match_name: str | None,
    side: str | None, line: float | None,
) -> str:
    label = MARKET_LABEL_MAP.get(market_type, market_type.replace("_", " "))
    if side in ("over", "under") and line is not None:
        side_str = f"{side.capitalize()} {line}"
    elif side:
        side_str = side.capitalize()
    else:
        side_str = ""
    subject = (
        player_name if scope == "player" and player_name
        else team_name if scope == "team" and team_name
        else match_name or "Match"
    )
    parts = [subject, "—"]
    if side_str:
        parts.append(side_str)
    parts.append(label)
    return " ".join(parts)


def calculate_ev(con: sqlite3.Connection, replace: bool = True) -> dict[str, Any]:
    ensure_schema(con)
    if replace:
        con.execute("DELETE FROM betting_value_scores_new")
    rows = con.execute("SELECT * FROM betting_odds_normalized ORDER BY id").fetchall()
    score_rows: list[dict[str, Any]] = []
    computed_at = now_utc()

    for row in rows:
        odds = row["odds_decimal"]
        calc = calculate_probability(con, row)
        model = calc.get("model_probability")
        implied = round(1 / odds, 4) if odds else None
        edge = round(model - implied, 4) if (model is not None and implied is not None) else None
        ev = round(model * odds - 1, 4) if (model is not None and odds) else None
        prob_status = calc.get("status")
        if prob_status == "ok":
            dq = calc.get("data_quality_status")
            verdict = (
                "REVIEW" if dq in {"low_sample", "partial"}
                else "VALUE" if ev and ev > 0
                else "NO_VALUE"
            )
        elif prob_status == "unmatched_selection":
            verdict = "UNMATCHED"
        elif prob_status == "REVIEW" or prob_status == "INCOMPLETE_DATA":
            verdict = "REVIEW"
        elif prob_status == "INCOMPLETE_PLAYER_DATA":
            verdict = "BLOCKED"
        elif str(prob_status or "").startswith("BLOCKED"):
            verdict = "BLOCKED"
        else:
            verdict = "UNSUPPORTED"
        notes = "; ".join(p for p in [row["notes"], calc.get("notes")] if p)
        mtype = row["market_type"]
        scope = MARKET_SCOPE_MAP.get(mtype, "unknown")
        priority = PRIORITY_CLASS_MAP.get(mtype, "ignored_complex")
        bet_desc = _compute_bet_description(
            mtype, scope, row["player_name"], row["team_name"],
            row["match_name"], row["side"], row["line"],
        )
        score_rows.append({
            "rank": None,
            "source_name": row["source_name"],
            "bookmaker": row["bookmaker"],
            "match_name": row["match_name"],
            "market_type": mtype,
            "raw_market_name": row["raw_market_name"],
            "raw_selection_name": row["raw_selection_name"],
            "selection": row["raw_selection_name"],
            "team_name": row["team_name"],
            "player_name": row["player_name"],
            "player_id": row["player_id"],
            "side": row["side"],
            "line": row["line"],
            "odds_decimal": odds,
            "implied_probability": implied,
            "model_probability": model,
            "edge": edge,
            "expected_value": ev,
            "probability_method": calc.get("probability_method"),
            "sample_size": calc.get("sample_size"),
            "probability_status": prob_status,
            "normalized_status": row["normalized_status"],
            "data_quality_status": calc.get("data_quality_status"),
            "verdict": verdict,
            "notes": notes,
            "computed_at": computed_at,
            "min_minutes_filter": calc.get("min_minutes_filter"),
            "valid_appearance_count": calc.get("valid_appearance_count"),
            "excluded_zero_minutes_count": calc.get("excluded_zero_minutes_count"),
            "excluded_low_minutes_count": calc.get("excluded_low_minutes_count"),
            "minutes_filter_status": calc.get("minutes_filter_status"),
            "priority_class": priority,
            "market_scope": scope,
            "bet_description": bet_desc,
            "market_mapping_status": row["market_mapping_status"],
            "market_mapping_reason": row["market_mapping_reason"],
            "exact_market_match": row["exact_market_match"],
            "canonical_market_type": row["canonical_market_type"],
            "statshub_field_used": row["statshub_field_used"],
            "market_contract_version": row["market_contract_version"],
            "model_uses_proxy": row["model_uses_proxy"],
            "field_mapping_status": row["field_mapping_status"],
            "side_line_status": row["side_line_status"],
            "data_completeness_status": row["data_completeness_status"],
        })

    ranked = [r for r in score_rows if r["expected_value"] is not None]
    ranked.sort(key=lambda r: r["expected_value"], reverse=True)
    rank_map = {id(r): idx for idx, r in enumerate(ranked, start=1)}
    for r in score_rows:
        r["rank"] = rank_map.get(id(r))

    if score_rows:
        cols = list(score_rows[0].keys())
        con.executemany(
            f"INSERT INTO betting_value_scores_new ({','.join(cols)}) VALUES ({','.join(['?'] * len(cols))})",
            [[r[col] for col in cols] for r in score_rows],
        )
    con.commit()
    return {
        "normalized": len(rows),
        "ev_rows": len([r for r in score_rows if r["expected_value"] is not None]),
        "supported": len([r for r in score_rows if r["probability_status"] == "ok"]),
        "unsupported": len([r for r in score_rows if r["verdict"] == "UNSUPPORTED"]),
        "unmatched": len([r for r in score_rows if r["verdict"] == "UNMATCHED"]),
        "top20": ranked[:20],
    }


def write_scores_workbook(
    con: sqlite3.Connection,
    path: Path = OUT_DIR / "today_odds_driven_value_scores.xlsx",
) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = [
        ("EV Ranking",
         "SELECT * FROM betting_value_scores_new WHERE expected_value IS NOT NULL ORDER BY rank LIMIT 200"),
        ("Raw Odds", "SELECT * FROM betting_odds_raw ORDER BY id"),
        ("Normalized Markets", "SELECT * FROM betting_odds_normalized ORDER BY id"),
        ("Unsupported Unmatched",
         "SELECT * FROM betting_value_scores_new WHERE verdict IN ('UNSUPPORTED','UNMATCHED') ORDER BY id"),
    ]
    for title, sql in sheets:
        ws = wb.create_sheet(title[:31])
        rows = con.execute(sql).fetchall()
        if rows:
            headers = list(rows[0].keys())
        else:
            cursor = con.execute(sql)
            headers = [d[0] for d in cursor.description]
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(42, len(str(header)) + 4))
    wb.save(path)
    return path
